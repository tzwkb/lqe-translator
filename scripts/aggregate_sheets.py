#!/usr/bin/env python3
"""多工作表任务汇总：把已完成的各子任务合成跨工作表交付物。

产出（父 job 目录）：
  <label>_corrected.xlsx  保留原始 workbook 的 sheet/公式/样式/合并单元格，
                          仅将程序生成的非空 corrected 写回原译文列。
  <label>_lqe.xlsx         汇总 sheet（各子表分数 + 按词数加权总分）+ 各子表 LQE Results 明细。

子 job 发现：父 job 目录下含 state.json 的直接子目录即一个 sheet 子 job。
段→行映射：优先使用 segment.row_index；旧 state 缺该字段时才回退 segment.id。

用法：
  python scripts/aggregate_sheets.py --job LQE测试用 [--sheets 剧情,功能,社媒] [--threshold 98]
"""
import argparse
from contextlib import ExitStack
from copy import copy, deepcopy
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lqe_engine import (  # noqa: E402
    _SKILL_ROOT,
    current_target,
    read_json,
    requires_bound_artifacts,
    validate_scope_entries,
)
from lqe_corrections import CheckFormatError, verify_results  # noqa: E402
from lqe_chunk import verification_generation_lease  # noqa: E402
from lqe_excel_diff import build_rich_diff  # noqa: E402
from lqe_paths import (  # noqa: E402
    file_sha256,
    paths_alias,
    publish_replacement_transaction,
    state_reference_paths,
    validate_artifact_paths,
)
from lqe_scoring import (  # noqa: E402
    resolve_scoring_policy,
    score_errors,
    score_totals,
)
from lqe_report_contract import validate_report_contract  # noqa: E402
from lqe_result_contract import (  # noqa: E402
    result_contract_path,
    validate_result_contract,
)
from lqe_split_contract import canonical_digest  # noqa: E402

_NON_TEXT = object()
_POLICY_COMPATIBILITY_KEYS = (
    "scorecard_profile",
    "severity_scale",
    "critical_gate",
    "repeat_dedup",
)


def _copy_report_cell(source_cell, target_cell) -> None:
    value = source_cell.value
    target_cell.value = deepcopy(value) if isinstance(value, CellRichText) else value
    if source_cell.data_type in {"s", "f"}:
        target_cell.data_type = source_cell.data_type
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
    if source_cell.comment is not None:
        target_cell.comment = copy(source_cell.comment)
    if source_cell.hyperlink is not None:
        target_cell._hyperlink = copy(source_cell.hyperlink)


def _copy_report_layout(source_sheet, target_sheet) -> None:
    target_sheet.freeze_panes = source_sheet.freeze_panes
    for index, source_dimension in source_sheet.row_dimensions.items():
        target_dimension = target_sheet.row_dimensions[index]
        for attribute in ("height", "hidden", "outlineLevel", "collapsed"):
            setattr(target_dimension, attribute, getattr(source_dimension, attribute))
    for key, source_dimension in source_sheet.column_dimensions.items():
        target_dimension = target_sheet.column_dimensions[key]
        for attribute in (
            "width",
            "hidden",
            "bestFit",
            "outlineLevel",
            "collapsed",
        ):
            setattr(target_dimension, attribute, getattr(source_dimension, attribute))


def _report_text(value):
    if isinstance(value, CellRichText):
        return str(value)
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return _NON_TEXT


def _label(p: Path) -> str:
    """父任务和子任务标签：jobs/ 下的子路径用下划线连接。"""
    parts = p.resolve().parts
    if "jobs" in parts:
        sub = parts[parts.index("jobs") + 1:]
        if sub:
            return "_".join(sub)
    return p.resolve().name


def _state_no_header(state: dict) -> bool:
    if "no_header" in state:
        value = state["no_header"]
        if not isinstance(value, bool):
            raise ValueError("state.no_header must be a boolean")
        return value

    source_column = state.get("source_col")
    return isinstance(source_column, int) or (
        isinstance(source_column, str) and source_column.isdigit()
    )


def _column_idx(state: dict, field: str, default: int) -> int:
    value = state.get(field, default)
    if _state_no_header(state) or isinstance(value, int):
        try:
            return int(value)
        except (TypeError, ValueError):
            return default
    headers = state.get("headers", [])
    return headers.index(value) if value in headers else default


def _target_idx(state) -> int:
    return _column_idx(state, "target_col", 1)


def _source_idx(state) -> int:
    return _column_idx(state, "source_col", 0)


def _first_data_row(state: dict) -> int:
    return 1 if _state_no_header(state) else 2


def _cell_text(value: object) -> str:
    return "" if value is None else str(value)


def _policy_signature(policy: dict) -> tuple:
    return tuple(policy[key] for key in _POLICY_COMPATIBILITY_KEYS)


def _discover(job_dir: Path):
    return sorted([d for d in job_dir.iterdir()
                   if d.is_dir() and (d / "state.json").exists()],
                  key=lambda d: d.name)


def _validated_results(
    sj: Path,
    state: dict,
) -> tuple[list[dict], dict[int, dict], dict]:
    bound = requires_bound_artifacts(state)
    try:
        with verification_generation_lease(
            sj / "state.json",
            exclusive=False,
            require_generation=bound,
        ) as (live_state, segments, manifest, _):
            if live_state != state or (bound and manifest is None):
                raise ValueError("state changed while loading child results")
            raw_results = read_json(sj / "errors.json")
            if bound:
                contract_path = result_contract_path(sj / "errors.json")
                if not contract_path.is_file():
                    raise ValueError("bound result contract is required")
                validate_result_contract(
                    read_json(contract_path),
                    manifest,
                    raw_results,
                    label=f"{sj.name}/errors.json",
                )
            results = verify_results(
                segments,
                raw_results,
                f"{sj.name}/errors.json",
                allow_internal_provenance=bound,
                require_internal_provenance=bound,
            )
            validate_scope_entries(
                state,
                results,
                issues_key="errors",
                label=f"{sj.name}/errors.json",
            )
            snapshot = {
                "state_digest": canonical_digest(state),
                "manifest_digest": (
                    manifest["manifest_digest"] if manifest is not None else None
                ),
                "results_digest": canonical_digest(raw_results),
            }
    except (CheckFormatError, OSError, ValueError) as exc:
        raise ValueError(f"{sj.name}: {exc}") from exc
    return (
        results,
        {segment["id"]: segment for segment in segments},
        snapshot,
    )


def _child_report_path(sj: Path) -> Path:
    path = sj / f"{_label(sj)}_lqe.xlsx"
    if not path.is_file():
        raise ValueError(f"missing child report: {path}")
    return path


def _revalidate_child_snapshot(item: dict, lease: tuple) -> None:
    state, _, manifest, _ = lease
    snapshot = item["generation_snapshot"]
    if canonical_digest(state) != snapshot["state_digest"]:
        raise ValueError(f"{item['job'].name}: state changed during aggregate")
    current_manifest_digest = (
        manifest["manifest_digest"] if manifest is not None else None
    )
    if current_manifest_digest != snapshot["manifest_digest"]:
        raise ValueError(
            f"{item['job'].name}: chunk generation changed during aggregate"
        )
    errors_path = item["job"] / "errors.json"
    raw_results = read_json(errors_path)
    if canonical_digest(raw_results) != snapshot["results_digest"]:
        raise ValueError(f"{item['job'].name}: errors changed during aggregate")
    if requires_bound_artifacts(state):
        contract_path = result_contract_path(errors_path)
        if not contract_path.is_file():
            raise ValueError(
                f"{item['job'].name}: bound result contract is missing"
            )
        validate_result_contract(
            read_json(contract_path),
            manifest,
            raw_results,
            label=f"{item['job'].name}/errors.json",
        )
    if file_sha256(item["report_path"]) != item["report_digest"]:
        raise ValueError(
            f"{item['job'].name}: child report changed during aggregate"
        )


def _source_workbook_path(sj: Path, state: dict) -> Path:
    configured = state.get("input_path")
    if isinstance(configured, str) and configured.strip():
        path = Path(configured)
        if path.is_file():
            return path
    fallback = sj / "src.xlsx"
    if fallback.is_file():
        return fallback
    raise ValueError(f"source workbook not found for child job: {sj}")


def _apply_corrections(
    source_path: Path,
    validated: list[dict],
    staged_output: Path,
) -> int:
    workbook = openpyxl.load_workbook(source_path, data_only=False)
    total = 0
    try:
        for item in validated:
            sj = item["job"]
            state = item["state"]
            sheet_name = state.get("sheet_name") or sj.name
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            elif len(validated) == 1:
                sheet = workbook.active
            else:
                raise ValueError(
                    f"source workbook missing sheet: {sheet_name}"
                )
            first_data_row = _first_data_row(state)
            if state.get("input_sha256") is None:
                source_index = _source_idx(state)
                for segment in state.get("segments", []):
                    row_index = int(
                        segment.get("row_index", segment.get("id", 0))
                    )
                    row_number = first_data_row + row_index
                    if row_number < first_data_row or row_number > sheet.max_row:
                        raise ValueError(
                            f"{sj.name}: source row for segment "
                            f"{segment.get('id')} is missing"
                        )
                    source = _cell_text(
                        sheet.cell(row=row_number, column=source_index + 1).value
                    )
                    target = _cell_text(
                        sheet.cell(
                            row=row_number,
                            column=item["target_index"] + 1,
                        ).value
                    )
                    if source != segment.get("source", "") or target != segment.get(
                        "target", ""
                    ):
                        raise ValueError(
                            f"{sj.name}: source row for segment "
                            f"{segment.get('id')} changed after read"
                        )
            replacements = 0
            for row_index, corrected in item["corrections"].items():
                cell = sheet.cell(
                    row=row_index + first_data_row,
                    column=item["target_index"] + 1,
                )
                if cell.value != corrected:
                    replacements += 1
                cell.value = corrected
                if isinstance(corrected, str) and corrected.startswith("="):
                    cell.data_type = "s"
            item["delivery_replacements"] = replacements
            total += replacements
        workbook.save(staged_output)
    finally:
        workbook.close()
    return total


def _append_child_results(
    aggregate_workbook: openpyxl.Workbook,
    sj: Path,
    report_path: Path,
    state: dict,
    results: list[dict],
) -> None:
    workbook = openpyxl.load_workbook(
        report_path,
        rich_text=True,
        data_only=False,
    )
    try:
        validate_report_contract(workbook, state, results)
        if "LQE Results" not in workbook.sheetnames:
            raise ValueError(
                f"child report missing LQE Results sheet: {report_path}"
            )
        source_sheet = workbook["LQE Results"]
        target_sheet = aggregate_workbook.create_sheet(f"{sj.name} Results")
        _copy_report_layout(source_sheet, target_sheet)
        for row in source_sheet.iter_rows():
            for source_cell in row:
                target_cell = target_sheet.cell(
                    source_cell.row,
                    source_cell.column,
                )
                _copy_report_cell(source_cell, target_cell)

        headers = [cell.value for cell in target_sheet[1]]
        suggested_header = next(
            (
                candidate
                for candidate in ("AI/建议译文", "建议译文")
                if candidate in headers
            ),
            None,
        )
        if "原译" in headers and suggested_header is not None:
            original_column = headers.index("原译") + 1
            suggested_column = headers.index(suggested_header) + 1
            for row_number in range(2, target_sheet.max_row + 1):
                original_cell = target_sheet.cell(row_number, original_column)
                suggested_cell = target_sheet.cell(row_number, suggested_column)
                original_text = _report_text(original_cell.value)
                suggested_text = _report_text(suggested_cell.value)
                if (
                    original_cell.data_type != "f"
                    and suggested_cell.data_type != "f"
                    and isinstance(original_text, str)
                    and isinstance(suggested_text, str)
                    and suggested_text
                    and original_text != suggested_text
                ):
                    original_cell.value, suggested_cell.value = build_rich_diff(
                        original_text,
                        suggested_text,
                    )
                    for cell in (original_cell, suggested_cell):
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            cell.data_type = "s"
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

        if "LQA Scorecard" in workbook.sheetnames:
            source_scorecard = workbook["LQA Scorecard"]
            target_scorecard = aggregate_workbook.create_sheet(
                f"{sj.name} Scorecard"[:31]
            )
            _copy_report_layout(source_scorecard, target_scorecard)
            for row in source_scorecard.iter_rows():
                for source_cell in row:
                    _copy_report_cell(
                        source_cell,
                        target_scorecard.cell(
                            source_cell.row,
                            source_cell.column,
                        ),
                    )
            for merged_range in source_scorecard.merged_cells.ranges:
                target_scorecard.merge_cells(str(merged_range))
    finally:
        workbook.close()


def _build_aggregate_report(
    staged_output: Path,
    *,
    label: str,
    validated: list[dict],
    summary: list[list],
    overall: dict,
    threshold: float,
    total_segments: int,
    total_errors: int,
    total_fixes: int,
) -> None:
    first_state = validated[0]["state"]
    workbook = openpyxl.Workbook()
    try:
        sheet = workbook.active
        sheet.title = "汇总"
        sheet.append([f"LQE 质检汇总报告 · {label}"])
        sheet.append(
            [
                f"阈值 {threshold:g}  "
                f"语言对 {first_state.get('language_pair', '')}  "
                f"项目 {first_state.get('project', '')}"
            ]
        )
        sheet.append([])
        sheet.append(
            ["子表", "段数", "词数", "错误数", "Critical", "分数", "结果", "建议修改数"]
        )
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)
        for row in summary:
            sheet.append(row)
        sheet.append(
            [
                "合计",
                total_segments,
                overall["wordcount"],
                total_errors,
                overall["critical"],
                overall["score"],
                overall["status"],
                total_fixes,
            ]
        )
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)
        sheet.append([])
        sheet.append(
            ["注：总分=按词数加权 (1-ΣL/Σwordcount)×100；各子表词数在首次读取后固定。"]
        )

        for item in validated:
            _append_child_results(
                workbook,
                item["job"],
                item["report_path"],
                item["state"],
                item["results"],
            )
        workbook.save(staged_output)
    finally:
        workbook.close()


def _aggregate(a) -> None:
    job_dir = Path(a.job)
    if not job_dir.is_absolute():
        job_dir = _SKILL_ROOT / "jobs" / a.job
    if not job_dir.is_dir():
        raise ValueError(f"job dir not found: {job_dir}")

    if a.sheets:
        subs = [job_dir / s.strip() for s in a.sheets.split(",")]
        for s in subs:
            if not (s / "state.json").exists():
                raise ValueError(f"sub-job missing state.json: {s}")
    else:
        subs = _discover(job_dir)
    if not subs:
        raise ValueError(f"no sub-jobs (state.json) under {job_dir}")

    label = _label(job_dir)
    corrected_output = job_dir / f"{label}_corrected.xlsx"
    report_output = job_dir / f"{label}_lqe.xlsx"

    states = {}
    policies = {}
    for sj in subs:
        state = read_json(sj / "state.json")
        if state.get("input_format") == "sdlxliff":
            raise ValueError("SDLXLIFF jobs are not multi-sheet workbooks")
        try:
            policy = resolve_scoring_policy(
                state,
                {"threshold": a.threshold},
            )
        except ValueError as exc:
            raise ValueError(
                f"{sj.name}: invalid scoring policy: {exc}"
            ) from exc
        states[sj] = state
        policies[sj] = policy

    first_policy = policies[subs[0]]
    first_signature = _policy_signature(first_policy)
    incompatible = [
        sj.name
        for sj in subs[1:]
        if _policy_signature(policies[sj]) != first_signature
    ]
    if incompatible:
        dimensions = ", ".join(_POLICY_COMPATIBILITY_KEYS)
        raise ValueError(
            "scoring policy mismatch across child jobs "
            f"({dimensions}): {subs[0].name}, {', '.join(incompatible)}"
        )
    effective_threshold = max(policy["threshold"] for policy in policies.values())

    total_weighted = 0.0
    total_wordcount = 0.0
    total_errors = 0
    total_critical = 0
    total_segments = 0
    validated: list[dict] = []
    protected_inputs: dict[str, Path] = {}

    first_source = _source_workbook_path(subs[0], states[subs[0]])
    source_digest = file_sha256(first_source)

    for sj in subs:
        state = states[sj]
        results, seg_by_id, generation_snapshot = _validated_results(sj, state)
        computation = score_errors(state, results, policies[sj])
        score = computation["output"]
        target_index = _target_idx(state)
        segment_rows = {
            segment_id: int(segment.get("row_index", segment_id))
            for segment_id, segment in seg_by_id.items()
        }
        corrections = {}
        for entry in results:
            segment = seg_by_id[entry["id"]]
            corrected = entry["corrected"]
            if corrected is None and current_target(segment) != segment.get("target", ""):
                corrected = current_target(segment)
            if (
                corrected is not None
                and not segment.get("protected")
                and not any(
                    issue.get("protected") for issue in entry.get("errors", [])
                )
            ):
                corrections[segment_rows[entry["id"]]] = corrected

        source_path = _source_workbook_path(sj, state)
        if not paths_alias(first_source, source_path):
            raise ValueError(
                "sub-jobs must share one source workbook: "
                f"{first_source}, {source_path}"
            )
        current_source_digest = file_sha256(source_path)
        if current_source_digest != source_digest:
            raise ValueError(
                f"source workbook changed during aggregate: {source_path}"
            )
        expected_source_digest = state.get("input_sha256")
        if (
            expected_source_digest is not None
            and current_source_digest != expected_source_digest
        ):
            raise ValueError(
                f"{sj.name}: source workbook changed after read: {source_path}"
            )
        report_path = _child_report_path(sj)
        validated.append(
            {
                "job": sj,
                "state": state,
                "source_path": source_path,
                "report_path": report_path,
                "target_index": target_index,
                "corrections": corrections,
                "score": score,
                "segments": len(state.get("segments", [])),
                "results": results,
                "generation_snapshot": generation_snapshot,
                "report_digest": file_sha256(report_path),
                "delivery_replacements": 0,
            }
        )
        total_weighted += computation["total_weighted"]
        total_wordcount += score["wordcount"]
        total_errors += score["errors"]
        total_critical += score["critical"]
        total_segments += len(state.get("segments", []))

        protected_inputs[f"{sj.name}/state.json"] = sj / "state.json"
        protected_inputs[f"{sj.name}/errors.json"] = sj / "errors.json"
        if requires_bound_artifacts(state):
            protected_inputs[f"{sj.name}/errors contract"] = (
                result_contract_path(sj / "errors.json")
            )
        protected_inputs[f"{sj.name}/report"] = report_path
        protected_inputs[f"{sj.name}/source"] = source_path
        for reference_label, reference_path in state_reference_paths(state).items():
            protected_inputs[
                f"{sj.name}/state.{reference_label}"
            ] = reference_path

    validate_artifact_paths(
        {
            "corrected workbook": corrected_output,
            "aggregate report": report_output,
        },
        protected_inputs,
        context="aggregate",
    )

    aggregate_policy = {
        **first_policy,
        "threshold": effective_threshold,
    }
    overall = score_totals(
        total_weighted,
        total_wordcount,
        total_critical,
        aggregate_policy,
    )
    if any(item["score"]["status"] != "PASS" for item in validated):
        overall["status"] = "FAIL"

    with tempfile.TemporaryDirectory(
        dir=job_dir,
        prefix=".aggregate.",
    ) as staging_dir_name:
        staging_dir = Path(staging_dir_name)
        staged_corrected = staging_dir / corrected_output.name
        staged_report = staging_dir / report_output.name
        total_fixes = _apply_corrections(
            first_source,
            validated,
            staged_corrected,
        )
        summary = [
            [
                item["job"].name,
                item["segments"],
                item["score"]["wordcount"],
                item["score"]["errors"],
                item["score"]["critical"],
                item["score"]["score"],
                item["score"]["status"],
                item["delivery_replacements"],
            ]
            for item in validated
        ]
        _build_aggregate_report(
            staged_report,
            label=label,
            validated=validated,
            summary=summary,
            overall=overall,
            threshold=effective_threshold,
            total_segments=total_segments,
            total_errors=total_errors,
            total_fixes=total_fixes,
        )
        with ExitStack() as leases:
            for item in sorted(
                validated,
                key=lambda value: str(value["job"].resolve()),
            ):
                lease = leases.enter_context(
                    verification_generation_lease(
                        item["job"] / "state.json",
                        exclusive=False,
                        require_generation=requires_bound_artifacts(
                            item["state"]
                        ),
                    )
                )
                _revalidate_child_snapshot(item, lease)
            if file_sha256(first_source) != source_digest:
                raise ValueError(
                    f"source workbook changed during aggregate: {first_source}"
                )
            publish_replacement_transaction(
                [
                    (staged_corrected, corrected_output),
                    (staged_report, report_output),
                ]
            )

    print(f"[aggregate] {len(subs)} sheets: {', '.join(s.name for s in subs)}")
    print(f"[aggregate] 建议译文 -> {corrected_output}")
    print(f"[aggregate] 汇总报告 -> {report_output}")
    print(
        f"[aggregate] 总分 {overall['score']:.2f} ({overall['status']}) "
        f"segs={total_segments} wc={overall['wordcount']} "
        f"errors={total_errors} critical={overall['critical']} fixes={total_fixes}"
    )
    for row in summary:
        print(
            f"  {row[0]}: {row[5]} {row[6]} "
            f"(segs {row[1]}, wc {row[2]}, err {row[3]}, "
            f"crit {row[4]}, fix {row[7]})"
        )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--job",
        required=True,
        help="父任务目录：jobs/ 下的子路径（如 LQE测试用）或绝对路径",
    )
    parser.add_argument(
        "--sheets",
        default=None,
        help="逗号分隔的子任务目录名，用于指定顺序或子集；默认按名称读取全部",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=None,
        help="显式覆盖所有子任务阈值；省略时各自继承 state.scoring_policy",
    )
    args = parser.parse_args()
    try:
        _aggregate(args)
    except Exception as exc:
        raise SystemExit(f"[aggregate] {exc}") from exc


if __name__ == "__main__":
    main()
