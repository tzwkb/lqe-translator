"""Bind generated LQE workbooks to the state and verified results they render."""

from __future__ import annotations

from datetime import date, datetime, time
import json

from openpyxl.cell.rich_text import CellRichText

from lqe_engine import requires_bound_artifacts
from lqe_provenance import AUDIT_HEADER_BASES, issue_detail, issue_review_columns
from lqe_split_contract import canonical_digest


SHEET_NAME = "_LQE_CONTRACT"
SCHEMA = "lqe.report-contract"
VERSION = 5


def _audit_column(headers: list[object], base: str) -> int:
    candidates = [
        index
        for index, header in enumerate(headers, start=1)
        if header == base
        or (
            isinstance(header, str)
            and header.startswith(f"{base}（审计 ")
        )
    ]
    if not candidates:
        raise ValueError(f"LQE Results is missing required audit column: {base}")
    return candidates[-1]


def _validate_current_results_shape(workbook, state: dict, results: list[dict]) -> None:
    if not requires_bound_artifacts(state):
        return
    if "LQE Results" not in workbook.sheetnames:
        raise ValueError("child report is missing LQE Results")
    sheet = workbook["LQE Results"]
    headers = [cell.value for cell in sheet[1]]
    if not headers or headers[-1] != "LQE_Iter":
        raise ValueError(
            "LQE Results must place LQE_Iter in the last column"
        )
    columns = {
        key: _audit_column(headers, base)
        for key, base in AUDIT_HEADER_BASES.items()
    }
    if "LQA Scorecard" not in workbook.sheetnames:
        raise ValueError("current report is missing LQA Scorecard")
    try:
        detail_column = headers.index("错误详情") + 1
    except ValueError as exc:
        raise ValueError(
            "LQE Results is missing required audit column: 错误详情"
        ) from exc

    state_ids = [segment.get("id") for segment in state.get("segments", [])]
    result_ids = [entry.get("id") for entry in results if isinstance(entry, dict)]
    if (
        any(type(value) is not int for value in state_ids + result_ids)
        or len(result_ids) != len(results)
        or len(set(state_ids)) != len(state_ids)
        or len(set(result_ids)) != len(result_ids)
        or set(result_ids) != set(state_ids)
    ):
        raise ValueError("report results do not cover the current state exactly")
    by_id = {entry["id"]: entry for entry in results}

    expected = []
    for segment_id in state_ids:
        errors = by_id[segment_id].get("errors")
        if not isinstance(errors, list):
            raise ValueError("report result errors must be arrays")
        for issue_number, issue in enumerate(errors or [None], start=1):
            if issue is not None and not isinstance(
                issue.get("review_provenance"), dict
            ):
                raise ValueError(
                    "current report errors require explicit review_provenance"
                )
            review_status, edit_status, source = issue_review_columns(
                issue,
                segment_id,
            )
            expected.append(
                (
                    segment_id,
                    issue_number if issue is not None else None,
                    review_status,
                    edit_status,
                    source,
                    issue_detail(issue),
                )
            )

    actual = []
    for row in range(2, sheet.max_row + 1):
        segment_id = sheet.cell(row, columns["segment_id"]).value
        if segment_id in (None, ""):
            continue
        issue_number = sheet.cell(row, columns["issue_number"]).value
        if issue_number == "":
            issue_number = None
        actual.append(
            (
                segment_id,
                issue_number,
                sheet.cell(row, columns["review_status"]).value,
                sheet.cell(row, columns["edit_status"]).value,
                sheet.cell(row, columns["check_source"]).value,
                sheet.cell(row, detail_column).value or "",
            )
        )
    if actual != expected:
        raise ValueError(
            "LQE Results rows do not match current segment/error provenance"
        )


def _cell_content(cell) -> dict:
    value = cell.value
    if value in (None, ""):
        return {"data_type": "blank", "value": None}
    if isinstance(value, CellRichText):
        return {"data_type": "text", "value": str(value)}
    elif isinstance(value, (date, datetime, time)):
        return {"data_type": "date", "value": value.isoformat()}
    elif isinstance(value, bool):
        return {"data_type": "boolean", "value": value}
    elif isinstance(value, (int, float)):
        normalized = (
            int(value)
            if isinstance(value, float) and value.is_integer()
            else value
        )
        return {"data_type": "number", "value": normalized}
    elif isinstance(value, str):
        return {
            "data_type": "formula" if cell.data_type == "f" else "text",
            "value": value,
        }
    return {"data_type": "text", "value": str(value)}


def _visible_sheet_digest(workbook, sheet_name: str) -> str:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"child report is missing {sheet_name}")
    sheet = workbook[sheet_name]
    payload = {
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "merged_ranges": sorted(str(cell_range) for cell_range in sheet.merged_cells.ranges),
        "cells": [
            [_cell_content(cell) for cell in row]
            for row in sheet.iter_rows(
                min_row=1,
                max_row=sheet.max_row,
                min_col=1,
                max_col=sheet.max_column,
            )
        ],
    }
    return canonical_digest(payload)


def _visible_results_digest(workbook) -> str:
    return _visible_sheet_digest(workbook, "LQE Results")


def _visible_scorecard_digest(workbook) -> str:
    return _visible_sheet_digest(workbook, "LQA Scorecard")


def build_report_contract(workbook, state: dict, results: list[dict]) -> dict:
    _validate_current_results_shape(workbook, state, results)
    payload = {
        "schema": SCHEMA,
        "version": VERSION,
        "state_digest": canonical_digest(state),
        "results_digest": canonical_digest(results),
        "visible_results_digest": _visible_results_digest(workbook),
        "visible_scorecard_digest": (
            _visible_scorecard_digest(workbook)
            if "LQA Scorecard" in workbook.sheetnames
            else None
        ),
    }
    payload["contract_digest"] = canonical_digest(payload)
    return payload


def attach_report_contract(workbook, state: dict, results: list[dict]) -> None:
    encoded = json.dumps(
        build_report_contract(workbook, state, results),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    if SHEET_NAME in workbook.sheetnames:
        del workbook[SHEET_NAME]
    sheet = workbook.create_sheet(SHEET_NAME)
    sheet["A1"] = encoded
    sheet.sheet_state = "veryHidden"


def validate_report_contract(workbook, state: dict, results: list[dict]) -> None:
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"child report is missing {SHEET_NAME}")
    raw = workbook[SHEET_NAME]["A1"].value
    if not isinstance(raw, str):
        raise ValueError("child report contract is invalid")
    try:
        actual = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError("child report contract is invalid") from exc
    expected = build_report_contract(workbook, state, results)
    if actual != expected:
        fields = sorted(
            key
            for key in set(actual) | set(expected)
            if actual.get(key) != expected.get(key)
        )
        raise ValueError(
            "child report is stale for current state/errors "
            f"(mismatch: {', '.join(fields)})"
        )
