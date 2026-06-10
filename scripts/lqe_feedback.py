"""
导出客户反馈表（QAFeedback 模板格式）。
仅列出有 actionable 错误的段（errors 非空且非纯 Neutral），跳过无错段与 locked。

表头（固定，客户模板）：
  Source | Current translation | Suggest translation | Reason | 如果需要修改，请改到在线memoQ文件中，此处敲1
"""
import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

_HEADER = ["Source", "Current translation", "Suggest translation",
           "Reason", "如果需要修改，请改到在线memoQ文件中，此处敲1"]


def cmd(args):
    job = Path(args.job)
    state = json.loads((job / "state.json").read_text(encoding="utf-8"))
    errs = {r["id"]: r for r in json.loads((job / "errors.json").read_text(encoding="utf-8"))}
    segs = {s["id"]: s for s in state["segments"]}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QA Feedback"
    ws.append(_HEADER)
    hfill = PatternFill("solid", fgColor="4472C4")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(wrap_text=True, vertical="center")

    n = 0
    for sid, s in segs.items():
        r = errs.get(sid)
        if not r:
            continue
        actionable = [e for e in r["errors"] if e.get("severity") != "Neutral"]
        if not actionable:
            continue
        parts = []
        for e in actionable:
            c = e.get("comment", "").strip()
            if not c:
                c = f"{e.get('category','?')} issue (no detail provided by reviewer)"
            parts.append(f"[{e.get('category','?')}/{e.get('severity','?')}] {c}")
        reason = "; ".join(parts)
        suggest = r.get("corrected") or ""
        ws.append([s["source"], s["target"], suggest, reason, ""])
        n += 1

    for col, w in zip("ABCDE", (42, 42, 42, 60, 18)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    stem = Path(state["input_path"]).stem
    out = job / f"{stem}_QAFeedback.xlsx"
    wb.save(out)
    print(f"[feedback] {n} flagged segments → {out}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--job", required=True)
    cmd(ap.parse_args())
