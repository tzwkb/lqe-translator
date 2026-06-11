from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUT = os.path.join(os.path.dirname(__file__), "..", "docs", "质量检查项清单.xlsx")

TITLE_F = Font(bold=True, size=14, color="1F4E79")
SEC_F = Font(bold=True, size=12, color="1F4E79")
H_F = Font(bold=True, size=10, color="FFFFFF")
H_FILL = PatternFill("solid", fgColor="4472C4")
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")
WRAP = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="B0B0B0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def put_table(ws, start_row, headers, rows, widths=None, highlight_rows=None):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font, c.fill, c.alignment, c.border = H_F, H_FILL, WRAP, BORDER
    for i, row in enumerate(rows, 1):
        for j, v in enumerate(row, 1):
            c = ws.cell(row=start_row + i, column=j, value=v)
            c.alignment, c.border = WRAP, BORDER
            if highlight_rows and i in highlight_rows:
                c.fill = NEW_FILL
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + j)].width = w
    return start_row + len(rows) + 1

# ---------- Sheet 1 ----------
ws = wb.active
ws.title = "概览与计分规则"
ws["A1"] = "质量检查项清单（含评估关注点）"
ws["A1"].font = TITLE_F
meta = [
    "更新 2026-06-11",
    "证据① 0512《【AI】【英】globaltrunk【0511新增】_LQE Report》＝WWM 中→英 AI 译文人工 LQE（5,314 词，82.78 FAIL，计分错误 101＋重复 34）",
    "证据② 客户《LQA template to evaluate cooperating translators》（译员评估模板；全文提取 projects/nrc-common/LQA_template_extract.txt）",
    "基线：SKILL.md 现行 pre-check＋17 子类；docs/调研-行业标准证据简报.md；docs/差距分析.md",
    "用法：Step 2 评估前与项目 adjudications.md 同读；源文档 docs/质量检查项清单.md（更新以 md 为准，本表由 scripts/gen_checklist_xlsx.py 重导）",
]
r = 3
for m in meta:
    ws.cell(row=r, column=1, value=m).alignment = WRAP
    r += 1

r += 1
ws.cell(row=r, column=1, value="规则来源代号（全表通用，〔〕标注）").font = SEC_F
r += 1
r = put_table(ws, r, ["代号", "出处文档"], [
    ["SKILL", "SKILL.md（现行 skill 主文档；R 号=docs/LQE质量标准优化报告.html 的 R1-R30 整改项，上游分析=差距分析 GAP-06）"],
    ["SG", "projects/wwm/sg.txt（WWM 权威风格指南）"],
    ["ADJ", "projects/wwm/adjudications.md 0512 前旧段（历史 LQE 经验）"],
    ["GAP", "docs/差距分析.md（GAP-01~08）"],
    ["调研", "docs/调研-行业标准证据简报.md"],
    ["0512", "0512 人工 LQE 报告·错误明细＋LQA Scorecard sheet（本次新证据）"],
    ["0512def", "同报告 Error Definition sheet（客户权威错误定义表，含逐类三档严重度例）"],
    ["LQA", "客户 LQA 译员评估模板（How to Fill / Score-for evaluator sheet，本次新证据）"],
])

r += 1
ws.cell(row=r, column=1, value="0512 失分画像（优先级依据）").font = SEC_F
r += 1
r = put_table(ws, r, ["类别", "Minor", "Major", "Critical", "重复(不罚分)", "加权罚分", "占比"], [
    ["Mistranslation", 2, 22, 41, 13, 783, "85.6%"],
    ["Terminology", 0, 12, 0, 8, 90, "9.8%"],
    ["Company style", 12, 1, 0, 0, 25.5, "2.8%"],
    ["Unidiomatic", 6, 0, 0, 11, 9, "1.0%"],
    ["Inconsistency", 5, 0, 0, 2, 7.5, "0.8%"],
    ["其余 12 类", 0, 0, 0, 0, 0, "0"],
    ["TOTAL", 25, 35, 41, 34, 915, "100%"],
])
notes = [
    "41 个 Critical 构成：对联 26＋对联流对话 3＋玩法规则文本 7＋活动名/考勤 UI 5",
    "教训① 85% 失分来自 Mistranslation——重灾区是成组文本拆单句直译与规则文本语义错",
    "教训② 确定性类（Markup/Length/标点/拼写/漏多译）零错，pre-check 路线有效",
    "教训③ 下一批输入《玩法数据表_题目表_皇宫以外内容.xlsx》与对联块同型，最高危，必须成组评估",
]
for n in notes:
    ws.cell(row=r, column=1, value=n).alignment = WRAP
    r += 1

r += 1
ws.cell(row=r, column=1, value="评估元规则（每条错误先过）").font = SEC_F
r += 1
r = put_table(ws, r, ["规则", "内容", "规则来源"], [
    ["三点法则", "准确＋合规（SG/指令/术语）＋合语法→视为正确，偏好性改写不计错（至多 Neutral 建议）。防误报第一道闸", "LQA How to Fill（新增）"],
    ["单一归属＋存疑取重", "按归类决策表取最具体一类；Minor/Major 拿不准取 Major", "SKILL §归类决策规则←调研 §1.4 J2450 元规则（沿用）"],
    ["Word Choice 边界", "意思错→Mistranslation；意思对但有更优词→Unidiomatic(Min) 或 Neutral", "LQA 模板明文（新增）"],
    ["严重度三档锚", "Minor=可察觉不误导；Major=很可能误导玩家/损公信；Critical=投诉/法务/功能破坏/高曝光面", "SKILL 严重度判定（沿用）＋0512def 逐类细化"],
])

r += 1
ws.cell(row=r, column=1, value="计分与流程规则").font = SEC_F
r += 1
r = put_table(ws, r, ["规则", "内容", "规则来源", "证据"], [
    ["重复错误", "记录（Repeated=YES）不罚分，首次全额", "0512 Scorecard 实证（新增）；问题背景=GAP-07", "34 条 rep 罚分=0"],
    ["阈值分层", "TEP/MTPE=98；润色/二审=99", "数值=0512def（新增）；分层概念上游=调研 §3 DQF", "客户 Error Definition 表；profile.threshold 已支持"],
    ["严重度分值", "0/1/5/10（客户=LISA 制，非 MQM 25）——GAP-03 的 MQM 调参不采", "SKILL 沿用；与 0512def 一致", "两报告一致"],
    ["公式", "score=1−Σ(加权罚分)/词数；词数口径=target-words", "SKILL 沿用；与 0512def 完全一致", "0512 复算吻合（5,314 词）"],
    ["成组文本", "对联/题目/谜题按组评估：上下联对仗＋押韵＋题目-答案可配对；拆单句直译=组级 Critical", "0512 错误明细（全新）", "对联块 26 Cri"],
    ["内容桶倾向", "规则说明→错译默认 Cri；UI 短串→Unidiomatic/style 主导；诗句 banner→Company style 重写权", "概念=调研 §2.3/GAP-05；具体倾向=0512（新增）", "0512 判级分布"],
    ["报告解析", "corrected 列可为 #N/A（审校未给改稿）；Repeated 列 YES/NO", "0512 实表（新增）", "r129 等"],
])

r += 1
ws.cell(row=r, column=1, value="待落地变更").font = SEC_F
r += 1
r = put_table(ws, r, ["#", "变更", "目标", "状态"], [
    [1, "重复错误去重计分（N4）", "scripts/lqe_calc.py", "待 go"],
    [2, "拼音残留（N1）＋同源异译（N2）", "scripts/lqe_io.py pre-check", "待 go"],
    [3, "罗马数字 custom（N3）", "projects/wwm/checks.json", "待 go"],
    [4, "0512 术语/风格裁决注入", "projects/wwm/adjudications.md", "已完成"],
    [5, "Step 2 评估提示引用本清单", "SKILL.md", "待 go"],
    [6, "题目表/对联输入带组上下文", "lqe_io.py read / Step 2 流程", "待 go（下批题目表前急需）"],
])
for col, w in zip("ABCDEFG", [24, 60, 46, 26, 14, 12, 10]):
    ws.column_dimensions[col].width = w

# ---------- Sheet 2 ----------
ws2 = wb.create_sheet("确定性pre-check项")
rows2 = [
    ["1", "untranslated_cjk", "target 含中文", "Untranslated", "Major", "现行沿用", "SKILL Step1.5 初版", "指令允许保留项除外", "内置"],
    ["2", "empty_target", "空译文", "Untranslated", "Major", "现行沿用", "SKILL Step1.5", "口径与词数基准联动（source-chars 时仍计分）", "内置"],
    ["3", "em_dash", "破折号 —", "Punctuation", "Minor", "现行沿用", "SKILL 初版＋SG 标点", "项目可关（nrc-en 已关）", "内置"],
    ["4", "color_tags", "#G/#C/#Y…#E 配对", "Markup", "Major", "现行沿用", "SKILL 初版＋SG Markup", "相对位置也须一致，非仅数量", "内置"],
    ["5", "variables", "{} / %s 缺失多余", "Markup", "Major", "现行沿用", "SKILL 初版＋SG Markup", "0512 实证：「{}天后领取{}」丢一个占位符被人工判 Mistranslation Critical——占位符错按客户口径是最重级〔0512〕", "内置"],
    ["6", "pos_placeholder", "%s/%d 顺序错位", "Markup", "Major", "现行沿用", "SKILL R1 整改←GAP-06", "命名/带索引占位符允许重排", "内置"],
    ["7", "newline_count", "\\n 数量不匹配", "Markup", "Major", "现行沿用", "SKILL 初版＋SG", "", "内置"],
    ["8", "numbers_consistency", "数值漏译/改值", "Mistranslation", "Major", "现行沿用", "SKILL R6 整改←GAP-06", "仅源含阿拉伯数字触发；中文数字不误报；语境误报可移除", "内置"],
    ["9", "length", "max-length / 1.5× 超长", "Length", "Major", "现行沿用", "max-length=SKILL R3←GAP-06；1.5× 回退=初版", "max-length 列优先；无列时仅非 CJK 源回退", "内置"],
    ["10", "locale_numbers", "千位分隔符缺失", "Locale convention", "Minor", "现行沿用", "SKILL 初版＋SG 数字", "", "内置"],
    ["11", "whitespace", "首尾空白/双空格", "Punctuation", "Minor", "现行沿用", "SKILL R5 整改←GAP-06", "", "内置"],
    ["12", "fullwidth_punct", "EN 译文全角标点", "Punctuation", "Minor", "现行沿用", "SKILL R5 整改＋SG 半角标点", "", "内置"],
    ["13", "terminology", "术语命中缺译", "Terminology", "Major", "现行沿用", "SKILL 初版；TB:status 机制=项目档案", "[TB:status]：Approved 硬判；New/WorkingTB 语境甄别；泛词命中≠错误〔ADJ〕", "内置"],
    ["N1", "pinyin_residue", "拼音残留", "Mistranslation", "Critical", "⊕新增", "0512 全新发现", "target 含 2+ 连续拼音音节大写词且不在官方拼音白名单（Kaifeng/Qinghe/Jianghu/Fu Shen/Xuanyu…）。0512：画卯→Mark Mao、平安→Ping'an 均 Cri。半确定：regex 初筛＋AI 复核", "pre-check 待落地"],
    ["N2", "intra_consistency", "同源异译", "Inconsistency", "Minor", "⊕新增", "GAP-06「文件内一致性」首次落清单；0512 实证支持", "文件内相同 source 不同 target；反向（相同 target 不同 source）一并报", "pre-check 待落地"],
    ["N3", "roman_numeral", "罗马数字风格", "Company style", "Minor", "⊕新增", "0512 改稿风格", "序号/卷号用 Unicode Ⅰ Ⅱ Ⅲ，非 ASCII I/II/III（0512：其一→Ⅰ、Volume Ⅱ）", "wwm checks.json 待落地"],
    ["N4", "repeat_dedup", "重复错误去重计分", "（计分规则）", "—", "⊕新增", "问题=GAP-07 死列；策略=0512 实证落定", "同源＋同译＋同错仅首次计分，其余标 Repeated=YES 不罚分（0512：34 条 rep 罚分=0）", "lqe_calc.py 待落地"],
]
put_table(ws2, 1, ["序号", "键/ID", "检查项", "类别", "严重度", "状态", "规则来源", "评估关注点", "落地"], rows2,
          widths=[6, 20, 22, 16, 10, 10, 34, 62, 18], highlight_rows=[14, 15, 16, 17])
ws2.freeze_panes = "A2"

# ---------- Sheet 3 ----------
ws3 = wb.create_sheet("AI评估项17子类")
ws3["A1"] = "子类体系/权重/强制 Major/严重度 0-1-5-10 全部沿用〔SKILL §LQE错误分类；七父维度=GAP-01/02 整改〕；「严重度梯度」列统一取自〔0512def〕；关注点逐条〔〕标来源"
ws3["A1"].font = Font(bold=True, size=10, color="1F4E79")

def fp(*items):
    return "• " + "\n• ".join(items)

rows3 = [
    ["Accuracy", "Mistranslation 错译", 1.5, "—", "Min2·Maj22(+8rep)·Cri41(+5rep)＝最大失分源",
     fp("专名↔通名双向误判：拼音残留（画卯→Roll Call，×Mark Mao；平安→Peace，×Ping'an）；通名实体化幻觉（年年≠Niannian；岁月长安≠Chang'an；紫色团花=purple flower，×Epic Flower）〔0512〕",
        "玩法规则文本逐条核：数值/上限/门槛（铜筹折算上限2000）、阶段流程序、机制词（多轮循环=Round-robin，×Rotating）、奖励归属与返还条件——客户一律 Critical〔0512〕",
        "操作指引动作语义：技能效果的对象与方式（复原巨石使其停下≠restore…to a halt；招式=any attack，×a Move）〔0512〕",
        "价格/概率语义词不可丢：折扣价、有机会获得（Cri 例：源「有机会获得[68金币]」译 contains）；回復≠receive〔0512＋0512def〕",
        "升 Cri 信号：误导玩家决策的规则文/题目-答案失配/占位符语义错位/经济敏感〔0512〕"),
     "Min=误读但贴近源｜Maj=译错（关卡→Close Card）｜Cri=日期/规则/内容错致投诉法务"],
    ["Accuracy", "Omission 漏译", 1.5, "—", "0",
     fp("限定词清点：任意/1次/仅/同时/最高/每（0512 近失：a Sword Trial 丢「任意」）〔0512〕",
        "进度/资格条件（充值/达标/解锁）丢失即 Cri〔0512def〕"),
     "Min=丢衔接词｜Maj=丢改义成分（范围内的敌人→all enemies）｜Cri=丢进度资格关键信息"],
    ["Accuracy", "Addition 多译", 1.5, "—", "0",
     fp("贴源过译与偏离区分〔0512def〕", "勿引入不存在的玩法元素（单体伤害→all enemies）〔0512def〕"),
     "Min=过译仍贴源｜Maj=偏离源｜Cri=引入不存在玩法元素/招致投诉破坏沉浸"],
    ["Accuracy", "Untranslated 未翻译", 1.5, "始终Maj", "0（pre-check 拦截）",
     fp("拼音输出≠已翻译——按 Mistranslation Cri 处理（N1）〔0512〕", "英文源原样复制也算未译（不止中文残留）〔GAP-06 未译扩展〕"),
     "始终 Major"],
    ["Terminology", "Terminology 术语", 1.5, "始终Maj", "Maj12(+8rep)＝第二失分源",
     fp("活动/玩法/界面入口名逐字对官方库与已上线译名：皇宫寻宝=Imperial Palace Treasure Hunt（自创=Maj）；御前练兵=Imperial Drill〔0512〕",
        "成对称号体系成对取：新锐|新兵=Recruit／老将|老兵=Veteran（×New Edge/Old General）〔0512〕",
        "人名代号查角色档：青=Halcyon（×Qing 直拼）〔0512〕",
        "同物跨段同名：手札统一 Journal（×Note）〔0512〕",
        "过度术语化=错：通名勿造专名（异色灵蝶=strangely colored butterflies，×Spectral Butterfly）〔0512〕",
        "系列名格式：赋神·乘桴归梦=Fu Shen - Rippling Dream（系列前缀保留〔0512〕；·→\" - \"=SG 既有规则〔SG〕）",
        "强制定译 19 条见 projects/wwm/adjudications.md《0512 裁决》〔0512〕；泛词命中按语境甄别〔ADJ 沿用〕"),
     "始终 Major"],
    ["Fluency", "Inconsistency 一致性", 1.5, "—", "Min5(+2rep)",
     fp("平行句族统一句型：任务/成就列表同模板（Complete any X once with a Veteran/Recruit 族）〔0512〕",
        "任务名引用格式统一（Lost Chapter quest: X）〔0512〕",
        "涉术语表词条的冲突归 Terminology，其余归此〔SKILL 归类决策表沿用〕"),
     "Min=拼写大小写语气不一（approx./approximately 混用）｜Maj=术语混用致误解（hero/protagonist/main character）"],
    ["Fluency", "Grammar 语法", 1.5, "—", "0",
     fp("主谓一致/时态/冠词〔0512def/LQA 通则〕", "占位符邻接可数名词用「{} day(s)」型复数〔0512〕"),
     "Min=显粗心（Me and my friends are…）｜Maj=损公信（The lego set are nice）｜Cri=灾难性后果"],
    ["Fluency", "Punctuation 标点", 1.0, "—", "0（pre-check 盖大半）",
     fp("半角标点、对齐源标点（pre-check/custom 已盖大半）〔SG＋SKILL 沿用〕", "长句逗号缺失影响可读性〔LQA〕"),
     "Min=错但可懂｜Maj=改句意（Let's eat, Timmy→Let's eat Timmy）｜Cri=金额标点致事实错误"],
    ["Fluency", "Spelling 拼写", 1.0, "—", "0",
     fp("Equipments/Acheive/Entrence 型错拼〔0512def〕", "易混词 accept/except、then/than、your/you're=Maj〔0512def〕"),
     "Min=粗心错拼｜Maj=损公信易混词｜Cri=灾难性（pubic library）"],
    ["Style", "Company style 公司风格", 1.5, "—", "Min12·Maj1",
     fp("句中通名小写：accept a quest / the treasure（×句中 Accept a Quest）〔0512；Title/Sentence Mode 本体=SG 沿用〕",
        "序号/卷号 Unicode 罗马数字（N3）〔0512〕",
        "括号体系：设施/可放置物名 [Stove]（×「」原样、×#Y\"X\"#E 引号式）〔0512〕",
        "诗句/banner 有押韵节奏重写权（九天阊阖联→bold/gold 韵对）；过度直译判此类〔0512〕",
        "禁古英语/网络俚语/UI 语气进剧情〔SG 沿用〕；SG 明文违反才归此，无明文归 Unidiomatic〔SKILL 归类决策表沿用〕"),
     "Min=Oxford comma/大小写规则未循｜Maj=语域错置（古风项目用 Sup, Harry?）"],
    ["Style", "Unidiomatic 不合语言习惯", 1.5, "—", "Min6(+11rep)",
     fp("UI 短句自然化模板：Continue Completing <X>→Continue completing the X quest（去尖括号＋补通名 quest＋句子化大小写）——0512 同一模式重复 12 次〔0512；非代码禁 <> 旧规=ADJ〕",
        "直译腔标志：Warm Tips / Successfully Claimed 型〔0512def〕",
        "三档标尺：母语者难懂／直译但可懂／好文笔——仅前两档计错〔LQA Naturalness，新增〕"),
     "Min=不地道但可懂（Successfully Claimed）｜Maj=不地道且致困惑（Warm Tips）｜Cri=冒犯或彻底破坏沉浸（You no go. I go for you!）"],
    ["Locale convention", "Locale convention 语言环境约定", 1.0, "—", "0",
     fp("日期防歧义拼写月份（6/5/2023→May 6, 2023）〔0512def〕", "货币符号与币种一致〔0512def〕"),
     "Min=格式少见仍可懂｜Maj=影响理解｜Cri=币种错（¥299→$299）致法务/财务风险"],
    ["Audience Appropriateness（客户父类=Verity）", "Culture specific reference 文化特定所指", 1.5, "—", "0",
     fp("源文化梗错置目标受众（520、圣诞吃苹果）〔0512def〕", "冒犯/涉政/禁忌=Cri〔0512def〕"),
     "Min=文化梗略怪｜Maj=过于小众致困惑｜Cri=冒犯/政治敏感"],
    ["Audience Appropriateness", "Audience appropriateness 受众适配", 1.5, "—", "0（本 skill 独有子类）",
     fp("语域/世界观口吻（仙侠敬语→现代俚语）〔SKILL←GAP-01 整改沿用〕", "准确但不合受众期待〔SKILL 沿用〕"),
     "参照 Company style/Unidiomatic 梯度按影响定级"],
    ["Design & Markup", "Markup 标记", 1.5, "始终Maj", "0（pre-check 盖）",
     fp("色标相对位置、{} 数量与顺序、\\n 保留〔SKILL＋SG 沿用〕"),
     "始终 Major"],
    ["Design & Markup", "Length 长度", 1.0, "始终Maj", "0（pre-check 盖）",
     fp("max-length 列优先；超长截断风险〔SKILL R3←GAP-06〕"),
     "始终 Major"],
    ["Other", "Other 其他", 1.0, "—", "0",
     fp("兜底；先过单一归属表再落此〔SKILL 沿用〕"),
     "—"],
]
put_table(ws3, 2, ["父维度", "子类别", "权重", "强制严重度", "0512 计分", "评估关注点（〔〕=规则来源）", "严重度梯度（源:0512def）"],
          rows3, widths=[20, 24, 7, 11, 26, 90, 48])
ws3.freeze_panes = "A3"

wb.save(OUT)
print("OK", os.path.abspath(OUT))
