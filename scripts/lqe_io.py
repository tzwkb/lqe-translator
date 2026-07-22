"""
LQE I/O utilities.

Subcommands:
  read          Excel/CSV/TSV + project profile → state.json
  pre-check     确定性错误自动检测（标点/Markup/术语/长度等）
  protect-segments 把已确认的 TM/100% 匹配段标记为已保护
  apply-fixes   把程序生成的建议译文写回 state.json
  write         state.json + errors.json → *_lqe.xlsx
  ingest-corpus 建议译文回传 AIPE 语料库（接口尚未确定，暂不执行）
"""
import argparse
import csv
import io
import json
import math
import os
import re
import sys
import tempfile
import unicodedata
from copy import copy, deepcopy
from datetime import date
from pathlib import Path

import openpyxl
import regex
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from lqe_engine import (
    read_json, RE_CJK as _RE_CJK, _source_lang, _target_lang, _load_lang, _LANG_DIR, _SKILL_ROOT,
    CATEGORY_ORDER as _ALL_CATS, CATEGORY_PARENT as _PARENT,
    VALID_CATEGORIES as _VALID_CATEGORIES, VALID_SEVERITIES as _VALID_SEVERITIES,
    apply_severity, build_check_scope, get_check_scope,
    current_target,
    load_terms as _load_terms, group_terms as _group_terms,
    requires_bound_artifacts,
    raw_points, weighted_points,
    load_scorecard_profile, normalize_category_for_profile, scorecard_category_order,
    scorecard_category_parent, scorecard_category_weight,
    scorecard_severity_points,
    language_tags_match, normalize_language_tag,
    resolve_language_assets,
    validate_scope_entries,
)
from lqe_corrections import (
    CheckFormatError,
    build_results,
    normalize_check_entries,
    verify_results,
)
from lqe_inputs import SDLXLIFFImportError, detect_input_format, read_sdlxliff
from lqe_inputs.sdlxliff import (
    is_exact_tm,
    validate_options as validate_sdlxliff_options,
)
from lqe_excel_diff import build_rich_diff
from lqe_paths import (
    file_sha256,
    paths_alias as _paths_alias,
    publish_replacement_transaction,
    state_reference_paths,
    validate_artifact_paths,
    write_json_atomic,
)
from lqe_terms import canonicalize_terms, load_canonical_terminology
from lqe_scoring import (
    resolve_scoring_policy,
    score_errors,
    scoring_policy_overrides,
)
from lqe_report_contract import attach_report_contract
from lqe_provenance import (
    AUDIT_HEADER_BASES,
    issue_detail,
    issue_review_columns as _issue_review_columns,
)
from lqe_result_contract import (
    build_result_contract,
    result_contract_path,
    validate_result_contract,
)


def _validate_scope_or_exit(
    state: dict,
    entries: list[dict],
    *,
    issues_key: str,
    label: str,
    command: str,
) -> None:
    try:
        get_check_scope(state)
        validate_scope_entries(
            state, entries, issues_key=issues_key, label=label
        )
    except ValueError as exc:
        raise SystemExit(f"[{command}] {exc}") from exc


def _processing_label(entry: dict) -> str:
    errors = entry.get("errors") or []
    if any(error.get("protected") for error in errors):
        return "已保护，不修改"
    if any(error.get("needs_confirmation") for error in errors):
        return "需要人工确认"
    if entry.get("corrected") is not None:
        return "建议修改"
    if errors:
        return "仅提醒"
    return "无需修改"


def _issue_processing_label(
    issue: dict | None,
    entry: dict,
    *,
    protected: bool,
) -> str:
    if protected or (issue is not None and issue.get("protected")):
        return "已保护，不修改"
    if issue is None:
        return _processing_label(entry)
    if not isinstance(issue.get("review_provenance"), dict):
        return _processing_label(entry)
    if issue.get("needs_confirmation"):
        return "需要人工确认"
    if issue.get("edit") is not None and entry.get("corrected") is not None:
        return "建议修改"
    return "仅提醒"


# ── read ──────────────────────────────────────────────────────────────────────

def _clean_terms(items: list) -> list:
    return canonicalize_terms(items)
_MAXLEN_KEYS = {"maxlen", "max_len", "max length", "maxlength", "max_length",
                "char_limit", "charlimit", "limit", "width", "ui_max",
                "限长", "字符上限", "长度上限", "字数上限"}


def _parse_maxlen(val) -> int | None:
    if val is None or str(val).strip() == "":
        return None
    try:
        n = int(float(str(val).strip()))
        return n if n > 0 else None
    except ValueError:
        return None


def _job_label(state_path) -> str:
    """输出文件名前缀：标注产物来自哪个任务。取 jobs/ 下的子路径用 _ 连接
    （如 jobs/LQE测试用/剧情/ → 'LQE测试用_剧情'），否则退回 job 目录名。
    避免多文件/多 sheet 拆分时所有 job 都叫 src_*，看不出来源。"""
    d = Path(state_path).resolve().parent
    parts = d.parts
    if "jobs" in parts:
        sub = parts[parts.index("jobs") + 1:]
        if sub:
            return "_".join(sub)
    return d.name


def _load_terminology(
    path: str,
    *,
    term_status_map: object = None,
    protected_statuses: object = None,
) -> list:
    return load_canonical_terminology(
        Path(path),
        term_status_map=term_status_map,
        protected_statuses=protected_statuses,
    )


def _load_project(name_or_path: str) -> dict:
    # 项目名 = <game>/<track>（如 nrc/zh-th、wwm/zh-en），在 skill 根 projects/ 下解析（CWD 无关）；
    # 带后缀或绝对路径按字面处理（支持任意位置的 profile.json）。
    p = Path(name_or_path)
    if not p.suffix and not p.is_absolute():
        p = _SKILL_ROOT / "projects" / p
    if p.is_dir():
        p = p / "profile.json"
    if not p.exists():
        print(f"[ERROR] project profile not found: {p}", file=sys.stderr)
        sys.exit(1)
    prof = read_json(p)
    prof["_dir"] = str(p.parent.resolve())
    prof["_path"] = str(p.resolve())
    return prof


def _validate_project_profile(prof: dict):
    required = ("language_pair", "source_lang", "target_lang")
    missing = [k for k in required if not str(prof.get(k, "")).strip()]
    if missing:
        print("[ERROR] project profile must define language_pair, source_lang, and target_lang; "
              f"missing: {', '.join(missing)}", file=sys.stderr)
        sys.exit(1)
    protected_statuses = prof.get("protected_term_statuses")
    if "protected_term_statuses" in prof and (
        not isinstance(protected_statuses, list)
        or any(
            not isinstance(value, str) or not value.strip()
            for value in protected_statuses
        )
    ):
        print(
            "[ERROR] project profile protected_term_statuses must be an "
            "array of non-empty strings",
            file=sys.stderr,
        )
        sys.exit(1)
    if "scoring_policy" in prof and not isinstance(
        prof["scoring_policy"], dict
    ):
        print(
            "[ERROR] project profile scoring_policy must be an object",
            file=sys.stderr,
        )
        sys.exit(1)


def _project_path(prof: dict, val: str) -> str:
    if not val:
        return ""
    q = Path(val)
    return str(q if q.is_absolute() else Path(prof["_dir"]) / q)


def _profile_reference_paths(prof: dict | None) -> dict[str, Path]:
    if not prof:
        return {}
    references = {"--project": Path(prof["_path"])}
    for field in ("style_guide", "terminology", "checks", "confirmed_rules"):
        value = prof.get(field)
        if isinstance(value, str) and value.strip():
            references[f"project.{field}"] = Path(_project_path(prof, value))
    confirmed = references.get("project.confirmed_rules")
    if confirmed is not None:
        references["project.common_confirmed_rules"] = (
            confirmed.parent.parent / "common" / "confirmed_rules_common.md"
        )
    return references


def _load_style_guide(path: str) -> str:
    p = Path(path)
    if not p.exists():
        print(f"[warn] style-guide file not found: {path}", file=sys.stderr)
        return ""
    suffix = p.suffix.lower()
    if suffix == ".docx":
        import docx
        doc = docx.Document(str(p))
        lines = []
        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue
            style = para.style.name
            if style.startswith("Heading 1"):
                lines.append(f"\n# {text}")
            elif style.startswith("Heading 2"):
                lines.append(f"\n## {text}")
            elif style.startswith("Heading 3"):
                lines.append(f"\n### {text}")
            else:
                lines.append(text)
        return "\n".join(lines)
    if suffix in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(str(p), data_only=True)
        out = []
        for ws in wb.worksheets:
            rows = []
            for r in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c).strip() for c in r]
                if any(cells):
                    rows.append(cells)
            if len(rows) < 2:
                continue
            header, data = rows[0], rows[1:]
            out.append(f"\n# {ws.title}")
            category = ""
            for r in data:
                title, body = "", []
                for i, v in enumerate(r):
                    h = header[i] if i < len(header) else ""
                    if i == 0 and not h:
                        if v:
                            category = v
                        continue
                    if not v:
                        continue
                    if not title:
                        title = v
                    else:
                        body.append(f"[{h}] {v}" if h else v)
                if not title and not body:
                    continue
                out.append(f"## {category} — {title}" if category else f"## {title}")
                out.extend(body)
        return "\n".join(out)
    return p.read_text(encoding="utf-8")


def _cell(row, idx):
    return row[idx] if idx is not None and idx < len(row) else None


def _text(v):
    return str(v).strip() if v is not None else ""


def _extract_text_type_marker(src, tgt=None, content_type=None):
    """AIPE CSV may contain text-type header rows; they are context, not segments."""
    s = _text(src)
    if not s:
        return None
    explicit = {"对话类文本", "游戏内侧页文本", "故事类文本"}
    if s in explicit:
        return s
    m = re.match(r"^(?:文本类型|文本类别)\s*[:：]\s*(.+)$", s)
    if m and m.group(1).strip():
        return m.group(1).strip()
    if s in {"文本类型", "文本类别"}:
        return _text(tgt) or _text(content_type) or s
    return None


def _prepare_read_assets(
    args,
    prof: dict | None,
    check_scope: dict,
    segments: list[dict],
    source_lang: str,
    target_lang: str,
    *,
    asset_dir: Path | None = None,
    publish_dir: Path | None = None,
) -> dict:
    output_dir = Path(args.out).parent
    write_dir = Path(asset_dir) if asset_dir is not None else output_dir
    final_dir = Path(publish_dir) if publish_dir is not None else write_dir
    write_dir.mkdir(parents=True, exist_ok=True)

    sg_path = ""
    if args.style_guide:
        sg_text = _load_style_guide(args.style_guide)
        if sg_text:
            staged_file = write_dir / "sg.txt"
            published_file = final_dir / "sg.txt"
            staged_file.write_text(sg_text, encoding="utf-8")
            sg_path = str(published_file)
            print(f"[lqe_io] style_guide: {len(sg_text)} chars → {published_file}")

    terms_path = ""
    if args.terminology:
        protected_statuses = (
            prof.get("protected_term_statuses", []) if prof else []
        )
        terms = _load_terminology(
            args.terminology,
            term_status_map=prof.get("term_status_map") if prof else None,
            protected_statuses=protected_statuses,
        )
        if terms:
            n_protected = sum(
                1
                for term in terms
                for sense in term.get("senses", [term])
                if sense.get("protected") is True
            )
            if n_protected:
                print(f"[lqe_io] protected term senses: {n_protected}")
            staged_file = write_dir / "terms.json"
            published_file = final_dir / "terms.json"
            staged_file.write_text(
                json.dumps(terms, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            terms_path = str(published_file)
            print(
                f"[lqe_io] terminology: {len(terms)} entries → {published_file}"
            )

    requested_asset_lang = (
        _target_lang({"target_lang": getattr(args, "target_lang", None)})
        or _target_lang(prof if prof else {})
        or str(target_lang or "").lower()
    )
    asset_lang, lang_cfg = resolve_language_assets(requested_asset_lang)
    if lang_cfg:
        print(
            f"[lqe_io] target language attributes: "
            f"target_languages/{asset_lang}/attributes.json"
        )

    lang_notes_path = ""
    if asset_lang:
        notes_path = _LANG_DIR / asset_lang / "eval_notes.md"
        if notes_path.exists():
            staged_file = write_dir / "lang_notes.md"
            published_file = final_dir / "lang_notes.md"
            staged_file.write_text(
                notes_path.read_text(encoding="utf-8"), encoding="utf-8"
            )
            lang_notes_path = str(published_file)
            print(f"[lqe_io] language eval notes → {published_file}")

    background_path = ""
    if prof and (prof.get("background") or "").strip():
        staged_file = write_dir / "background.md"
        published_file = final_dir / "background.md"
        staged_file.write_text(
            "# 项目背景\n\n" + prof["background"].strip() + "\n",
            encoding="utf-8",
        )
        background_path = str(published_file)
        print(f"[lqe_io] project background → {published_file}")

    basis = (
        getattr(args, "wordcount_basis", None)
        or (prof.get("wordcount_basis") if prof else None)
        or lang_cfg.get("wordcount_basis")
        or "target-words"
    )
    if lang_cfg.get("word_delim") == "none" and basis == "target-words":
        print(
            "[warn] target language has no word delimiter — 'target-words' basis will "
            "undercount severely; use source-chars",
            file=sys.stderr,
        )
    if basis == "source-chars":
        wordcount = sum(
            len(_RE_CJK.findall(segment.get("source_plain", segment["source"])))
            + len(
                re.findall(
                    r"[A-Za-z0-9]+",
                    segment.get("source_plain", segment["source"]),
                )
            )
            for segment in segments
        )
    else:
        wordcount = sum(
            len(segment.get("target_plain", segment["target"]).split())
            for segment in segments
        )

    checks_path = confirmed_rules_path = ""
    if prof:
        checks = Path(_project_path(prof, prof.get("checks", "checks.json")))
        checks_path = str(checks) if checks.exists() else ""
        confirmed = Path(
            _project_path(prof, prof.get("confirmed_rules", "confirmed_rules.md"))
        )
        common_confirmed = confirmed.parent.parent / "common" / "confirmed_rules_common.md"
        parts = []
        if common_confirmed.exists():
            parts.append(
                f"<!-- ===== 共通确认规则（游戏级）: {common_confirmed.name} ===== -->\n"
                + common_confirmed.read_text(encoding="utf-8")
            )
        if confirmed.exists():
            parts.append(
                f"<!-- ===== 语言专有确认规则: {confirmed} ===== -->\n"
                + confirmed.read_text(encoding="utf-8")
            )
        if parts:
            staged_file = write_dir / "confirmed_rules.md"
            published_file = final_dir / "confirmed_rules.md"
            staged_file.write_text("\n\n".join(parts), encoding="utf-8")
            confirmed_rules_path = str(published_file)
            print(
                f"[lqe_io] confirmed rules: "
                f"{'共通+' if common_confirmed.exists() else ''}语言专有 → "
                f"{published_file}"
            )

    profile_policy = dict((prof or {}).get("scoring_policy", {}))
    for key in (
        "scorecard_profile",
        "severity_scale",
        "critical_gate",
        "repeat_dedup",
    ):
        if prof and key in prof and key not in profile_policy:
            profile_policy[key] = prof[key]
    if prof and "threshold" in prof and "threshold" not in profile_policy:
        profile_policy["threshold"] = prof["threshold"]
    scoring_policy = resolve_scoring_policy({}, profile_policy)

    return {
        "aipe_url": None,
        "check_scope": check_scope,
        "project": prof.get("name", "") if prof else "",
        "language_pair": prof.get("language_pair", "") if prof else (
            f"{source_lang}-{target_lang}" if source_lang and target_lang else ""
        ),
        "source_lang": source_lang,
        "target_lang": target_lang,
        "lang_notes_path": lang_notes_path,
        "background_path": background_path,
        "checks_path": checks_path,
        "confirmed_rules_path": confirmed_rules_path,
        "threshold": scoring_policy["threshold"],
        "scoring_policy": scoring_policy,
        "sg_path": sg_path,
        "terms_path": terms_path,
        "terminology": [],
        "style_guide": "",
        "wordcount": wordcount,
        "wordcount_basis": basis,
        "iteration": 0,
    }


def _language_values_match(first: str, second: str) -> bool:
    return language_tags_match(first, second) or language_tags_match(second, first)


def _validate_sdlxliff_languages(result, args, prof: dict | None) -> tuple[str, str]:
    declarations = result.manifest.get("languages", [])
    if not declarations:
        raise SDLXLIFFImportError("SDLXLIFF input has no language declarations")
    normalized = []
    for index, declaration in enumerate(declarations):
        source = normalize_language_tag(declaration.get("source_language"))
        target = normalize_language_tag(declaration.get("target_language"))
        if not source or not target:
            raise SDLXLIFFImportError(
                f"language declaration {index} must include source-language and target-language"
            )
        normalized.append((source, target))
    expected_source, expected_target = normalized[0]
    for index, (source, target) in enumerate(normalized[1:], start=1):
        if source != expected_source or target != expected_target:
            raise SDLXLIFFImportError(
                "conflicting SDLXLIFF language declarations: "
                f"declaration 0={expected_source}->{expected_target}, "
                f"declaration {index}={source}->{target}"
            )

    profile_source = normalize_language_tag(prof.get("source_lang")) if prof else ""
    profile_target = normalize_language_tag(prof.get("target_lang")) if prof else ""
    cli_source = normalize_language_tag(getattr(args, "source_lang", None))
    cli_target = normalize_language_tag(getattr(args, "target_lang", None))
    for label, profile_value, cli_value in (
        ("source", profile_source, cli_source),
        ("target", profile_target, cli_target),
    ):
        if profile_value and cli_value and not _language_values_match(
            profile_value, cli_value
        ):
            raise SDLXLIFFImportError(
                f"profile and CLI {label} language conflict: "
                f"{profile_value!r} != {cli_value!r}"
            )

    for origin, source, target in (
        ("profile", profile_source, profile_target),
        ("CLI", cli_source, cli_target),
    ):
        if source and not language_tags_match(source, expected_source):
            raise SDLXLIFFImportError(
                f"{origin} source language {source!r} does not match "
                f"declared source language {expected_source!r}"
            )
        if target and not language_tags_match(target, expected_target):
            raise SDLXLIFFImportError(
                f"{origin} target language {target!r} does not match "
                f"declared target language {expected_target!r}"
            )
    return expected_source, expected_target


def _publish_sdlxliff_job(
    state_path: Path,
    *,
    manifest: dict,
    tm_candidates: dict,
    scope: dict,
    state: dict,
    staged_assets: dict[Path, Path] | None = None,
) -> None:
    job_dir = state_path.parent
    reserved_names = {
        "source_manifest.json",
        "tm_candidates.json",
        "scope.json",
    }
    if state_path.name.casefold() in reserved_names:
        raise ValueError(
            f"SDLXLIFF state path conflicts with reserved helper artifact: {state_path}"
        )
    paths = {
        "manifest": job_dir / "source_manifest.json",
        "candidates": job_dir / "tm_candidates.json",
        "scope": job_dir / "scope.json",
        "state": state_path,
    }
    assets = {
        Path(destination): Path(staged)
        for destination, staged in (staged_assets or {}).items()
    }
    destinations = [*paths.values(), *assets]
    canonical_destinations: dict[str, Path] = {}
    for destination in destinations:
        canonical = str(destination.resolve()).casefold()
        previous = canonical_destinations.get(canonical)
        if previous is not None:
            raise ValueError(
                "SDLXLIFF job artifacts resolve to the same path: "
                f"{previous}, {destination}"
            )
        canonical_destinations[canonical] = destination

    missing_staged = [path for path in assets.values() if not path.is_file()]
    if missing_staged:
        raise FileNotFoundError(
            "staged SDLXLIFF job asset is missing: "
            + ", ".join(str(path) for path in missing_staged)
        )
    existing = [path for path in destinations if os.path.lexists(path)]
    if existing:
        raise FileExistsError(
            "SDLXLIFF job artifact already exists; use a new job directory: "
            + ", ".join(str(path) for path in existing)
        )

    values = {
        "manifest": manifest,
        "candidates": tm_candidates,
        "scope": scope,
        "state": state,
    }
    serialized = {
        key: json.dumps(value, ensure_ascii=False, indent=2)
        for key, value in values.items()
    }
    job_dir.mkdir(parents=True, exist_ok=True)
    staged: dict[str, Path] = {}

    try:
        for key, payload in serialized.items():
            with tempfile.NamedTemporaryFile(
                mode="w",
                encoding="utf-8",
                dir=job_dir,
                prefix=f".{paths[key].name}.",
                suffix=".tmp",
                delete=False,
            ) as handle:
                staged[key] = Path(handle.name)
                handle.write(payload)
        replacements = [
            (source, destination)
            for destination, source in sorted(
                assets.items(), key=lambda item: str(item[0])
            )
        ]
        replacements.extend(
            (staged[key], paths[key])
            for key in ("manifest", "candidates", "scope", "state")
        )
        try:
            publish_replacement_transaction(replacements, overwrite=False)
        except FileExistsError as exc:
            raise FileExistsError(
                "SDLXLIFF job artifact appeared during publication"
            ) from exc
    finally:
        for path in staged.values():
            path.unlink(missing_ok=True)


def _read_sdlxliff_job(args, prof: dict | None, check_scope: dict) -> None:
    state_path = Path(args.out)
    job_dir = state_path.parent
    helper_paths = (
        job_dir / "source_manifest.json",
        job_dir / "tm_candidates.json",
        job_dir / "scope.json",
    )
    if state_path.name.casefold() in {
        helper_path.name.casefold() for helper_path in helper_paths
    } or any(
        state_path.resolve() == helper_path.resolve() for helper_path in helper_paths
    ):
        raise ValueError(
            f"SDLXLIFF --out path conflicts with reserved helper artifact: {state_path}"
        )
    formal_paths = (
        state_path,
        *helper_paths,
    )
    existing = [path for path in formal_paths if path.exists()]
    if existing:
        raise FileExistsError(
            "SDLXLIFF job artifact already exists; use a new job directory: "
            + ", ".join(str(path) for path in existing)
        )

    raw_options = prof.get("sdlxliff", {}) if prof else {}
    options = validate_sdlxliff_options(
        raw_options,
        cli_protect_exact_tm=getattr(args, "protect_exact_tm", False),
    )
    result = read_sdlxliff(Path(args.input), options=options)
    source_lang, target_lang = _validate_sdlxliff_languages(result, args, prof)

    for segment in result.segments:
        metadata = segment["metadata"]["sdlxliff"]
        metadata["content_type"] = segment.get("content_type")
        segment["context_note"] = metadata.get("comment") or None
        segment["iter"] = 0

    job_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(
        dir=job_dir, prefix=".read-assets."
    ) as asset_staging_dir:
        staging_dir = Path(asset_staging_dir)
        common = _prepare_read_assets(
            args,
            prof,
            check_scope,
            result.segments,
            source_lang,
            target_lang,
            asset_dir=staging_dir,
            publish_dir=job_dir,
        )
        manifest_path = job_dir / "source_manifest.json"
        candidates_path = job_dir / "tm_candidates.json"
        candidates = {
            "candidate_ids": list(result.tm_candidates.get("candidate_ids", [])),
            "segments": [
                {
                    "id": item["segment_id"],
                    "evidence": item.get("evidence", {}),
                    "source_ref": item.get("source_ref", {}),
                }
                for item in result.tm_candidates.get("segments", [])
            ],
        }
        state = {
            "artifact_contract_version": 1,
            "input_format": "sdlxliff",
            "input_path": str(Path(args.input).resolve()),
            "input_paths": result.input_paths,
            "source_manifest_path": str(manifest_path),
            "tm_candidates_path": str(candidates_path),
            "source_col": "原文",
            "target_col": "译文",
            "headers": result.headers,
            "rows_raw": result.rows_raw,
            "text_type_markers": [],
            **common,
            "segments": result.segments,
        }
        staged_assets = {
            job_dir / path.name: path
            for path in staging_dir.iterdir()
            if path.is_file()
        }
        _publish_sdlxliff_job(
            state_path,
            manifest=result.manifest,
            tm_candidates=candidates,
            scope=check_scope,
            state=state,
            staged_assets=staged_assets,
        )
    print(
        f"[lqe_io] {len(result.segments)} segments → {args.out}  "
        f"wordcount={state['wordcount']}"
    )


def _cmd_read_locked(args):
    check_scope = build_check_scope(getattr(args, "no_terminology", False))
    out_path = Path(args.out)
    job_dir = out_path.parent
    scope_path = job_dir / "scope.json"
    if (
        out_path.name.casefold() == "scope.json"
        or out_path.resolve() == scope_path.resolve()
    ):
        print(
            f"[ERROR] --out path conflicts with reserved scope artifact: {scope_path}",
            file=sys.stderr,
        )
        sys.exit(2)
    if _paths_alias(out_path, Path(args.input)):
        print(
            f"[ERROR] --out path conflicts with --input: {args.input}",
            file=sys.stderr,
        )
        sys.exit(2)
    generated_asset_names = {
        "sg.txt",
        "terms.json",
        "lang_notes.md",
        "background.md",
        "confirmed_rules.md",
    }
    if out_path.name.casefold() in generated_asset_names:
        print(
            f"[ERROR] --out path conflicts with generated asset: {out_path.name}",
            file=sys.stderr,
        )
        sys.exit(2)
    try:
        input_format = detect_input_format(
            Path(args.input), getattr(args, "input_format", "auto")
        )
    except ValueError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        sys.exit(1)
    if input_format == "sdlxliff" and out_path.name.casefold() in {
        "source_manifest.json",
        "tm_candidates.json",
        "scope.json",
    }:
        print(
            f"[ERROR] SDLXLIFF --out path conflicts with reserved helper artifact: {out_path}",
            file=sys.stderr,
        )
        sys.exit(2)
    prof = _load_project(args.project) if getattr(args, "project", None) else None
    if prof:
        _validate_project_profile(prof)
        if not args.style_guide and prof.get("style_guide"):
            args.style_guide = _project_path(prof, prof["style_guide"])
        if prof.get("terminology"):
            if not check_scope["terminology_enabled"]:
                print("[lqe_io] profile terminology overridden by --no-terminology")
            elif not args.terminology:
                args.terminology = _project_path(prof, prof["terminology"])
        print(f"[lqe_io] project: {prof.get('name', '?')} ({prof.get('language_pair', '?')})")

    protected_inputs = {
        "--input": Path(args.input),
        **_profile_reference_paths(prof),
    }
    if args.style_guide:
        protected_inputs["--style-guide"] = Path(args.style_guide)
    if args.terminology:
        protected_inputs["--terminology"] = Path(args.terminology)
    planned_outputs = {
        "state": out_path,
        "scope": scope_path,
        "style guide copy": job_dir / "sg.txt",
        "terminology copy": job_dir / "terms.json",
        "language notes copy": job_dir / "lang_notes.md",
        "background copy": job_dir / "background.md",
        "confirmed rules copy": job_dir / "confirmed_rules.md",
    }
    if input_format == "sdlxliff":
        planned_outputs.update(
            {
                "source manifest": job_dir / "source_manifest.json",
                "TM candidates": job_dir / "tm_candidates.json",
            }
        )
    try:
        validate_artifact_paths(
            planned_outputs,
            protected_inputs,
            context="read",
        )
    except ValueError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        sys.exit(2)

    if input_format == "sdlxliff":
        try:
            _read_sdlxliff_job(args, prof, check_scope)
        except (OSError, ValueError) as exc:
            print(f"[ERROR] {exc}", file=sys.stderr)
            sys.exit(1)
        return

    if getattr(args, "protect_exact_tm", False):
        print(
            "[ERROR] --protect-exact-tm is only valid for SDLXLIFF input",
            file=sys.stderr,
        )
        sys.exit(2)
    if args.source_col is None or args.target_col is None:
        print(
            "[ERROR] tabular input requires --source-col and --target-col",
            file=sys.stderr,
        )
        sys.exit(2)

    no_header = getattr(args, "no_header", False)
    input_path = Path(args.input)
    input_sha256 = file_sha256(input_path)
    suffix = input_path.suffix.lower()

    if suffix in (".csv", ".tsv"):
        delim = "\t" if suffix == ".tsv" else ","
        raw_rows = list(csv.reader(io.StringIO(input_path.read_bytes().decode("utf-8-sig")), delimiter=delim))
        if not raw_rows:
            print("[ERROR] No data rows found.", file=sys.stderr)
            sys.exit(1)
        width = max(len(r) for r in raw_rows)
        if no_header:
            default_headers = ["Key", "Source", "Target", "Status", "Comment", "Scope", "File", "Reviewer Note"]
            headers = [default_headers[j] if j < len(default_headers) else f"col{j}" for j in range(width)]
            data_rows = raw_rows
        else:
            headers = [str(h).strip() if h is not None else "" for h in raw_rows[0]]
            data_rows = raw_rows[1:]
    else:
        wb = openpyxl.load_workbook(args.input)
        ws = wb.active
        if no_header:
            default_headers = ["Key", "Source", "Target", "Status", "Comment", "Scope", "File", "Reviewer Note"]
            headers = [default_headers[j] if j < len(default_headers) else f"col{j}" for j in range(ws.max_column)]
            data_rows = list(ws.iter_rows(min_row=1, values_only=True))
        else:
            headers = [cell.value for cell in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))

    if no_header:
        # 列参数为整数索引（0-based）
        try:
            si = int(args.source_col)
            ti = int(args.target_col)
        except ValueError:
            print("[ERROR] --no-header mode requires integer column indices for --source-col and --target-col", file=sys.stderr)
            sys.exit(1)
    else:
        for col in [args.source_col, args.target_col]:
            if col not in headers:
                print(f"[ERROR] Column '{col}' not found. Available: {headers}", file=sys.stderr)
                sys.exit(1)
        si = headers.index(args.source_col)
        ti = headers.index(args.target_col)

    # R3: 自动识别 max-length 列（UI 字段宽度上限），用于逐元素截断检查
    mi = None
    for idx, h in enumerate(headers):
        if h is not None and str(h).strip().lower() in _MAXLEN_KEYS:
            mi = idx
            break
    if mi is not None:
        print(f"[lqe_io] max-length column detected: '{headers[mi]}' (col {mi})")

    gi = None
    if getattr(args, "group_col", None):
        g = args.group_col
        gi = int(g) if str(g).isdigit() else (headers.index(g) if g in headers else None)
        if gi is None:
            print(f"[warn] group column '{g}' not found; grouping disabled", file=sys.stderr)
        else:
            print(f"[lqe_io] group column: '{headers[gi]}' (col {gi})")

    ci = None
    for idx, h in enumerate(headers):
        if h is not None and str(h).strip().lower() in {"content_type", "text_type", "文本类型", "文本类别"}:
            ci = idx
            break

    segments, rows_raw, text_type_markers = [], [], []
    text_type_context = None
    for i, row in enumerate(data_rows):
        if any(_text(c) for c in row):
            src = _text(_cell(row, si))
            tgt = _text(_cell(row, ti))
            row_content_type = _text(_cell(row, ci)) if ci is not None else ""
            marker = _extract_text_type_marker(src, tgt, row_content_type)
            if marker:
                text_type_context = marker
                text_type_markers.append({
                    "row_index": i,
                    "source": src,
                    "target": tgt,
                    "content_type": row_content_type or None,
                })
                continue
            seg_id = len(segments)
            segments.append({
                "id": seg_id,
                "row_index": i,
                "source": src,
                "target": tgt,
                "corrected": None,
                "content_type": row_content_type or None,
                "text_type_context": text_type_context,
                "max_len": _parse_maxlen(row[mi]) if mi is not None and mi < len(row) else None,
                "group": (str(row[gi]).strip() if gi is not None and gi < len(row) and row[gi] is not None and str(row[gi]).strip() else None),
                "iter": 0,
            })
            rows_raw.append(list(row))

    if not segments:
        print("[ERROR] No data rows found.", file=sys.stderr)
        sys.exit(1)

    source_lang = _source_lang({"source_lang": getattr(args, "source_lang", None)}) \
        or _source_lang(prof if prof else {})
    lang = _target_lang({"target_lang": getattr(args, "target_lang", None)}) \
        or _target_lang(prof if prof else {})
    try:
        job_dir.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(
            dir=job_dir, prefix=".read-assets."
        ) as asset_staging_dir:
            staging_dir = Path(asset_staging_dir)
            common = _prepare_read_assets(
                args,
                prof,
                check_scope,
                segments,
                source_lang,
                lang,
                asset_dir=staging_dir,
                publish_dir=job_dir,
            )

            state = {
                "artifact_contract_version": 1,
                "input_format": "tabular",
                "input_path": str(Path(args.input).resolve()),
                "input_sha256": input_sha256,
                "no_header": bool(no_header),
                "source_col": args.source_col,
                "target_col": args.target_col,
                "headers": headers,
                "rows_raw": rows_raw,
                "text_type_markers": text_type_markers,
                **common,
                "segments": segments,
            }
            staged_scope = staging_dir / "scope.json"
            staged_state = staging_dir / out_path.name
            staged_scope.write_text(
                json.dumps(check_scope, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            staged_state.write_text(
                json.dumps(state, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            asset_replacements = [
                (path, job_dir / path.name)
                for path in staging_dir.iterdir()
                if path.is_file() and path not in {staged_scope, staged_state}
            ]
            input_paths = [Path(args.input)]
            for configured in (args.style_guide, args.terminology):
                if configured:
                    input_paths.append(Path(configured))
            for _, destination in asset_replacements:
                for input_source in input_paths:
                    if _paths_alias(destination, input_source):
                        raise ValueError(
                            "generated job asset conflicts with input: "
                            f"{destination} == {input_source}"
                        )
            if file_sha256(input_path) != input_sha256:
                raise ValueError(
                    f"tabular input changed while it was being read: {input_path}"
                )
            publish_replacement_transaction(
                [
                    *sorted(asset_replacements, key=lambda item: str(item[1])),
                    (staged_scope, scope_path),
                    (staged_state, out_path),
                ]
            )
    except (OSError, ValueError) as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        sys.exit(1)
    print(
        f"[lqe_io] {len(segments)} segments → {args.out}  "
        f"wordcount={state['wordcount']}"
    )


def cmd_read(args):
    from lqe_split_contract import generation_lock

    state_path = Path(args.out)
    job_dir = state_path.parent
    read_lock_target = job_dir.parent / f"{job_dir.name}.lqe-read"
    try:
        with generation_lock(read_lock_target, exclusive=True):
            _cmd_read_locked(args)
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[read] {exc}") from exc

# ── lookup-terms ──────────────────────────────────────────────────────────────

def cmd_lookup_terms(args):
    state = read_json(args.state)
    terms = _load_terms(state)
    if not terms:
        print("[lqe_io] no terminology available.", file=sys.stderr)
        return

    term_map = _group_terms(terms)

    segs = state["segments"]
    if args.ids:
        id_set = set(int(x) for x in args.ids.split(","))
        segs = [s for s in segs if s["id"] in id_set]

    # 逐段匹配，避免跨段拼接产生误命中
    hits: dict[str, dict] = {}  # term_source → {senses, seg_ids}
    for seg in segs:
        src_text = seg["source"]
        for term_src, senses in term_map.items():
            if term_src in src_text:
                if term_src not in hits:
                    hits[term_src] = {"senses": senses, "seg_ids": []}
                hits[term_src]["seg_ids"].append(seg["id"])

    if not hits:
        print("[lookup-terms] no terminology matches found.")
        return

    print(f"[lookup-terms] {len(hits)} matches:\n")
    for src, info in sorted(hits.items(), key=lambda x: -len(x[0])):
        seg_ids = info["seg_ids"]
        id_str = f"  (segs: {seg_ids})" if len(seg_ids) <= 5 else f"  ({len(seg_ids)} segs)"
        tgt_str = " | ".join(s["target"] for s in info["senses"])
        print(f"  {src} → {tgt_str}{id_str}")


# ── apply-fixes ───────────────────────────────────────────────────────────────

def _protected_ids(args, *, allow_candidates: bool = False) -> set[int]:
    ids: set[int] = set()
    if getattr(args, "protected_ids", None):
        ids.update(int(x.strip()) for x in args.protected_ids.split(",") if x.strip())
    if getattr(args, "protected_file", None):
        data = read_json(args.protected_file)
        if isinstance(data, dict):
            if "protected_ids" in data:
                data = data["protected_ids"]
            elif "candidate_ids" in data:
                data = data["candidate_ids"] if allow_candidates else []
            else:
                data = data.get("segments") or []
        for item in data:
            if isinstance(item, int):
                ids.add(item)
            elif isinstance(item, dict):
                sid = item.get("id", item.get("seg_id", item.get("segment_id")))
                if sid is not None:
                    ids.add(int(sid))
    return ids


def _validated_tm_candidate_ids(
    state: dict, candidate_path: Path, payload: object
) -> set[int]:
    expected_path = state.get("tm_candidates_path")
    if not isinstance(expected_path, str) or not expected_path:
        raise ValueError("state has no tm_candidates_path")
    if not _paths_alias(candidate_path, Path(expected_path)):
        raise ValueError(
            "candidate file does not match state tm_candidates_path: "
            f"{candidate_path} != {expected_path}"
        )
    if not isinstance(payload, dict):
        raise ValueError("TM candidate payload must be an object")
    candidate_ids = payload.get("candidate_ids")
    candidates = payload.get("segments")
    if not isinstance(candidate_ids, list) or not all(
        type(segment_id) is int for segment_id in candidate_ids
    ):
        raise ValueError("TM candidate_ids must be an integer array")
    if len(candidate_ids) != len(set(candidate_ids)):
        raise ValueError("TM candidate_ids contains duplicates")
    if not isinstance(candidates, list):
        raise ValueError("TM candidate segments must be an array")

    candidate_by_id = {}
    for index, candidate in enumerate(candidates):
        if not isinstance(candidate, dict) or type(candidate.get("id")) is not int:
            raise ValueError(f"TM candidate segments[{index}] has an invalid id")
        segment_id = candidate["id"]
        if segment_id in candidate_by_id:
            raise ValueError(f"TM candidate segments contains duplicate id {segment_id}")
        candidate_by_id[segment_id] = candidate
    if set(candidate_ids) != set(candidate_by_id):
        raise ValueError(
            "TM candidate_ids must exactly match candidate segment evidence"
        )

    state_by_id = {segment.get("id"): segment for segment in state.get("segments", [])}
    for segment_id in candidate_ids:
        candidate = candidate_by_id[segment_id]
        state_segment = state_by_id.get(segment_id)
        if state_segment is None:
            raise ValueError(f"TM candidate id {segment_id} is not in state")
        if candidate.get("source_ref") != state_segment.get("source_ref"):
            raise ValueError(f"TM candidate id {segment_id} source_ref mismatch")
        evidence = candidate.get("evidence")
        metadata = (state_segment.get("metadata") or {}).get("sdlxliff") or {}
        if not is_exact_tm(evidence) or not is_exact_tm(metadata):
            raise ValueError(f"TM candidate id {segment_id} lacks exact-match evidence")
        for key in ("origin", "match_percent", "text_match"):
            if evidence.get(key) != metadata.get(key):
                raise ValueError(
                    f"TM candidate id {segment_id} evidence mismatch for {key}"
                )
    return set(candidate_ids)


def _state_protected_ids(state) -> set[int]:
    return {s["id"] for s in state.get("segments", []) if s.get("protected")}


def _stage_json_replacement(path: Path, value: object) -> Path:
    payload = json.dumps(value, ensure_ascii=False, indent=2)
    staged = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=path.parent,
            prefix=f".{path.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            staged = Path(handle.name)
            handle.write(payload)
        return staged
    except BaseException:
        if staged is not None and staged.exists():
            staged.unlink()
        raise


def _assert_json_unchanged(path: Path, expected: object, *, label: str) -> None:
    if read_json(path) != expected:
        raise ValueError(f"{label} changed during artifact publication: {path}")


def _stage_bound_result_replacements(
    errors_path: Path,
    errors_data: list[dict],
    state: dict,
    manifest: dict | None,
) -> tuple[list[Path], list[tuple[Path, Path]]]:
    staged_errors = _stage_json_replacement(errors_path, errors_data)
    staged = [staged_errors]
    replacements = [(staged_errors, errors_path)]
    if requires_bound_artifacts(state):
        if manifest is None:
            staged_errors.unlink(missing_ok=True)
            raise ValueError("bound result publication requires a generation")
        contract_path = result_contract_path(errors_path)
        staged_contract = _stage_json_replacement(
            contract_path,
            build_result_contract(manifest, errors_data),
        )
        staged.append(staged_contract)
        replacements.append((staged_contract, contract_path))
    return staged, replacements


def _publish_bound_result_update(
    errors_path: Path,
    errors_data: list[dict],
    state: dict,
    manifest: dict | None,
) -> None:
    staged: list[Path] = []
    try:
        staged, replacements = _stage_bound_result_replacements(
            errors_path,
            errors_data,
            state,
            manifest,
        )
        publish_replacement_transaction(replacements)
    finally:
        for path in staged:
            path.unlink(missing_ok=True)


def _publish_protection_transaction(
    state_path: Path,
    state: dict,
    output_path: Path,
    payload: dict,
) -> None:
    staged: list[Path] = []
    try:
        staged_output = _stage_json_replacement(output_path, payload)
        staged_state = _stage_json_replacement(state_path, state)
        staged.extend((staged_output, staged_state))
        publish_replacement_transaction(
            [
                (staged_output, output_path),
                (staged_state, state_path),
            ]
        )
    finally:
        for path in staged:
            path.unlink(missing_ok=True)


def _publish_write_transaction(
    state_path: Path,
    state: dict,
    errors_path: Path,
    errors_data: list,
    output_path: Path,
    staged_output: Path,
    *,
    publish_errors: bool,
    manifest: dict | None,
) -> None:
    staged: list[Path] = [staged_output]
    try:
        replacements = []
        if publish_errors:
            result_staged, result_replacements = _stage_bound_result_replacements(
                errors_path,
                errors_data,
                state,
                manifest,
            )
            staged.extend(result_staged)
            replacements.extend(result_replacements)
        staged_state = _stage_json_replacement(state_path, state)
        staged.append(staged_state)
        replacements.extend(
            [
                (staged_output, output_path),
                (staged_state, state_path),
            ]
        )
        publish_replacement_transaction(replacements)
    finally:
        for path in staged:
            path.unlink(missing_ok=True)


def _scrub_protected_entries(errors_data: list, protected_ids: set[int]) -> int:
    changed = 0
    for entry in errors_data:
        if entry.get("id") in protected_ids:
            if entry.get("errors") or entry.get("corrected") is not None:
                changed += len(entry.get("errors", [])) or 1
            entry["errors"] = []
            entry["corrected"] = None
    return changed


def _correction_candidates(errors_data: list[dict]) -> dict[int, str]:
    return {
        entry["id"]: entry["corrected"]
        for entry in errors_data
        if entry.get("corrected") is not None
        and not any(
            error.get("protected") for error in (entry.get("errors") or [])
        )
    }


def _cmd_protect_segments_locked(args, state_path: Path, state: dict):
    protected_file = getattr(args, "protected_file", None)
    protected_payload = read_json(protected_file) if protected_file else None
    try:
        if isinstance(protected_payload, dict) and "candidate_ids" in protected_payload:
            ids = _validated_tm_candidate_ids(
                state, Path(protected_file), protected_payload
            )
            if getattr(args, "protected_ids", None):
                ids.update(
                    int(value.strip())
                    for value in args.protected_ids.split(",")
                    if value.strip()
                )
        else:
            ids = _protected_ids(args, allow_candidates=False)
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[protect-segments] {exc}") from exc
    if not ids:
        print("[protect-segments] no ids supplied; state unchanged.")
        return

    out_path = Path(args.out) if args.out else state_path.parent / "tm_protected.json"
    try:
        validate_artifact_paths(
            {"protection decision": out_path},
            {
                "state": state_path,
                **state_reference_paths(state),
                **(
                    {"protected file": Path(protected_file)}
                    if protected_file
                    else {}
                ),
            },
            context="protect-segments",
        )
    except ValueError as exc:
        raise SystemExit(f"[protect-segments] {exc}") from exc

    seg_by_id = {s["id"]: s for s in state.get("segments", [])}
    valid = sorted(sid for sid in ids if sid in seg_by_id)
    unknown = sorted(sid for sid in ids if sid not in seg_by_id)
    for sid in valid:
        seg = seg_by_id[sid]
        seg["protected"] = True
        if seg.get("protected_reason") != "SOURCE_LOCKED":
            seg["protected_reason"] = args.reason
        working_target = current_target(seg)
        if working_target != seg.get("target", ""):
            seg["current_target"] = working_target
        seg["corrected"] = None

    payload = {
        "protected_ids": valid,
        "reason": args.reason,
        "source": "agent_decision",
    }
    if unknown:
        payload["unknown_ids"] = unknown
    try:
        _publish_protection_transaction(state_path, state, out_path, payload)
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[protect-segments] {exc}") from exc
    print(f"[protect-segments] protected {len(valid)} segment(s) → {state_path}")
    print(f"[protect-segments] protected file → {out_path}")
    if unknown:
        print(f"[protect-segments] ignored unknown ids: {unknown[:20]}")


def cmd_protect_segments(args):
    from lqe_split_contract import generation_lock

    state_path = Path(args.state).resolve()
    try:
        with generation_lock(state_path.parent / "chunks", exclusive=True):
            state = read_json(state_path)
            _cmd_protect_segments_locked(args, state_path, state)
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[protect-segments] {exc}") from exc


def _cmd_build_results_locked(args, state_path: Path, state: dict):
    checks_path = Path(args.checks)
    out_path = Path(args.out)
    entries = normalize_check_entries(
        json.loads(checks_path.read_text(encoding="utf-8")),
        label=args.checks,
    )
    _validate_scope_or_exit(
        state,
        entries,
        issues_key="issues",
        label=Path(args.checks).name,
        command="build-results",
    )
    segments = deepcopy(state["segments"])
    state_ids = {segment["id"] for segment in segments}
    check_ids = {entry["id"] for entry in entries}
    missing = sorted(state_ids - check_ids)
    extra = sorted(check_ids - state_ids)
    if missing or extra:
        sys.exit(
            f"[build-results] check ids must match state segment ids: "
            f"missing={missing} extra={extra}"
        )
    results = build_results(segments, entries)
    if requires_bound_artifacts(state):
        raise SystemExit(
            "[build-results] unbound checks cannot publish into a current job; "
            "use lqe_chunk.py merge"
        )
    try:
        validate_artifact_paths(
            {"results": out_path},
            {
                "state": state_path,
                "checks": checks_path,
                **state_reference_paths(state),
            },
            context="build-results",
        )
        write_json_atomic(out_path, results)
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[build-results] {exc}") from exc


def cmd_build_results(args):
    from lqe_split_contract import generation_lock

    state_path = Path(args.state)
    try:
        with generation_lock(state_path.parent / "chunks", exclusive=True):
            state = read_json(state_path)
            _cmd_build_results_locked(args, state_path, state)
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[build-results] {exc}") from exc


def _cmd_apply_fixes_locked(
    args,
    state_path: Path,
    state: dict,
    segments: list[dict],
    manifest: dict | None,
    revalidate,
):
    errors_data = read_json(args.errors)
    original_errors_data = deepcopy(errors_data)
    _validate_scope_or_exit(
        state,
        errors_data,
        issues_key="errors",
        label=Path(args.errors).name,
        command="apply-fixes",
    )
    protected_ids = _protected_ids(args) | _state_protected_ids(state)
    raw_attempted = _correction_candidates(errors_data)

    scrubbed = _scrub_protected_entries(errors_data, protected_ids)
    try:
        scoring_policy = resolve_scoring_policy(
            state,
            scoring_policy_overrides(args),
        )
    except (CheckFormatError, ValueError) as exc:
        raise SystemExit(f"[apply-fixes] {exc}") from exc
    scorecard_profile = load_scorecard_profile(
        scoring_policy["scorecard_profile"]
    )
    seg_ids = {segment["id"] for segment in segments}
    issues = _validate_errors(errors_data, seg_ids, scorecard_profile)
    for msg in issues:
        print(f"[validate] {msg}")
    try:
        verified = _verify_result_payload_with_segments(
            state,
            segments,
            manifest,
            errors_data,
            Path(args.errors),
            command="apply-fixes",
        )
        computation = score_errors(
            state,
            verified,
            scoring_policy,
            protected_ids=protected_ids,
        )
    except (CheckFormatError, ValueError) as exc:
        raise SystemExit(f"[apply-fixes] {exc}") from exc
    errors_data = computation["annotated_errors"]
    attempted = _correction_candidates(errors_data)

    corrections = {sid: text for sid, text in attempted.items() if sid not in protected_ids}
    segment_by_id = {segment["id"]: segment for segment in state["segments"]}
    protected_skipped = [
        {
            "id": sid,
            "reason": _protection_reason(segment_by_id[sid], protected_ids),
            "evidence": segment_by_id[sid].get("protection_evidence"),
            "attempted": text,
        }
        for sid, text in raw_attempted.items()
        if sid in protected_ids and sid in segment_by_id
    ]
    if not corrections and not protected_skipped:
        if scrubbed or computation["annotations_changed"]:
            revalidate()
            _assert_json_unchanged(
                Path(args.errors),
                original_errors_data,
                label="errors input",
            )
            _publish_bound_result_update(
                Path(args.errors),
                errors_data,
                state,
                manifest,
            )
        if scrubbed:
            print(
                f"[apply-fixes] scrubbed {scrubbed} protected-segment "
                f"issue(s) from {args.errors}"
            )
        print("[lqe_io] apply-fixes: no corrections found, state unchanged.")
        print(
            json.dumps(
                {"applied_count": 0, "lifecycle": "review_required"},
                ensure_ascii=False,
            )
        )
        return

    cur_iter = state.get("iteration", 0)
    history = state.get("error_history", [])
    score_result = computation["output"]
    score = score_result["score"]
    supplied_score = getattr(args, "score", None)
    if supplied_score is not None and not math.isclose(
        float(supplied_score), score, abs_tol=0.005
    ):
        print(
            f"[apply-fixes] supplied score {float(supplied_score):g} differs "
            f"from recomputed score {score:g}; using recomputed score",
            file=sys.stderr,
        )
    cur_entry = {
        "iteration": cur_iter,
        "score": score,
        "status": score_result["status"],
        "errors": errors_data,
        "corrections_count": len(corrections),
        "protected_ids": sorted(protected_ids),
        "skipped_corrections": protected_skipped,
    }
    history.append(cur_entry)
    state["error_history"] = history

    next_iter = cur_iter + (1 if corrections else 0)
    for seg in state["segments"]:
        if seg["id"] in protected_ids:
            seg["protected"] = True
            if not seg.get("protected_reason"):
                seg["protected_reason"] = "TM_100_MATCH"
            working_target = current_target(seg)
            if working_target != seg.get("target", ""):
                seg["current_target"] = working_target
            seg["corrected"] = None
            continue
        if seg["id"] in corrections:
            seg["current_target"] = corrections[seg["id"]]
            seg["corrected"] = corrections[seg["id"]]
            seg["iter"] = next_iter

    state["iteration"] = next_iter
    state["threshold"] = scoring_policy["threshold"]
    state["scoring_policy"] = scoring_policy
    state["pending_recheck"] = bool(corrections)

    archived = state_path.parent / f"errors_iter{cur_iter}.json"
    iter_out = state_path.parent / (_job_label(state_path) + f"_lqe_iter{cur_iter}.xlsx")
    try:
        validate_artifact_paths(
            {
                "iteration error archive": archived,
                "iteration report": iter_out,
            },
            {
                "state": state_path,
                "errors": Path(args.errors),
                **state_reference_paths(state),
            },
            context="apply-fixes",
        )
    except ValueError as exc:
        raise SystemExit(f"[apply-fixes] {exc}") from exc
    with tempfile.NamedTemporaryFile(
        prefix=f".{iter_out.stem}.",
        suffix=iter_out.suffix,
        dir=iter_out.parent,
        delete=False,
    ) as staging_file:
        staged_report = Path(staging_file.name)
    staged_paths = []
    try:
        _build_xlsx(
            state,
            [cur_entry],
            score,
            scoring_policy["threshold"],
            staged_report,
            scoring_policy["scorecard_profile"],
            announce=False,
            scoring_policy=scoring_policy,
            scoring_computation=computation,
        )
        replacements = []
        if scrubbed or computation["annotations_changed"]:
            result_staged, result_replacements = _stage_bound_result_replacements(
                Path(args.errors),
                errors_data,
                state,
                manifest,
            )
            staged_paths.extend(result_staged)
            replacements.extend(result_replacements)
        staged_archive = _stage_json_replacement(archived, errors_data)
        staged_state = _stage_json_replacement(state_path, state)
        staged_paths.extend((staged_archive, staged_state))
        replacements.extend(
            [
                (staged_archive, archived),
                (staged_report, iter_out),
                (staged_state, state_path),
            ]
        )
        revalidate()
        _assert_json_unchanged(
            Path(args.errors),
            original_errors_data,
            label="errors input",
        )
        publish_replacement_transaction(replacements)
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[apply-fixes] {exc}") from exc
    finally:
        staged_report.unlink(missing_ok=True)
        for staged_path in staged_paths:
            staged_path.unlink(missing_ok=True)

    if scrubbed:
        print(
            f"[apply-fixes] scrubbed {scrubbed} protected-segment "
            f"issue(s) from {args.errors}"
        )
    print(f"[lqe_io] Applied {len(corrections)} corrections → iteration {next_iter}")
    print(f"[lqe_io] Errors archived → {archived}")
    print(f"[lqe_io] Output → {iter_out}")
    print(
        json.dumps(
            {
                "applied_count": len(corrections),
                "lifecycle": (
                    "pending_recheck" if corrections else "review_required"
                ),
            },
            ensure_ascii=False,
        )
    )


def cmd_apply_fixes(args):
    from lqe_chunk import verification_generation_lease

    state_path = Path(args.state)
    try:
        with verification_generation_lease(
            state_path,
            exclusive=True,
        ) as (state, segments, manifest, revalidate):
            _cmd_apply_fixes_locked(
                args,
                state_path,
                state,
                segments,
                manifest,
                revalidate,
            )
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[apply-fixes] {exc}") from exc


# ── write ─────────────────────────────────────────────────────────────────────

_DARK_BLUE   = PatternFill("solid", fgColor="073763")
_LIGHT_BLUE  = PatternFill("solid", fgColor="CFE2F3")
_ORANGE      = PatternFill("solid", fgColor="FCE5CD")
_RED         = PatternFill("solid", fgColor="CC0000")
_GREEN       = PatternFill("solid", fgColor="006600")
_GREEN_LIGHT = PatternFill("solid", fgColor="D9EAD3")
_WHITE_FONT = Font(color="FFFFFF")
_BOLD       = Font(bold=True)
_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT_TOP   = Alignment(horizontal="left", vertical="top", wrap_text=True)
_LEFT_MID   = Alignment(horizontal="left", vertical="center")


def _validate_errors(errors_data: list, seg_ids: set, scorecard_profile: dict | None = None) -> list[str]:
    issues = []
    valid_categories = set(scorecard_category_order(scorecard_profile))
    for entry in errors_data:
        sid = entry.get("id")
        if sid not in seg_ids:
            issues.append(f"[seg {sid}] 未知 segment id")
            continue
        errs = entry.get("errors", [])
        for e in errs:
            raw_cat = e.get("category", "")
            cat = normalize_category_for_profile(raw_cat, scorecard_profile)
            sev = e.get("severity", "")
            if cat not in valid_categories:
                issues.append(f"[seg {sid}] 非法 category: '{raw_cat}'")
            if sev not in _VALID_SEVERITIES:
                issues.append(f"[seg {sid}] 非法 severity: '{sev}'")
            new_sev = apply_severity(cat, sev, scorecard_profile)
            if new_sev != sev:
                issues.append(f"[seg {sid}] {cat} severity {sev}→{new_sev} (auto-corrected)")
                e["severity"] = new_sev
    return issues


def _s(cell, fill=None, font=None, align=None):
    if fill:  cell.fill  = fill
    if font:  cell.font  = font
    if align: cell.alignment = align


def _set_excel_text(cell, value) -> None:
    cell.value = value
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"


_GRAPHEME_PATTERN = regex.compile(r"\X")
_EMOJI_WIDTH_PATTERN = regex.compile(
    r"\p{Emoji_Presentation}|\p{Regional_Indicator}|\u20e3"
)
_EMOJI_BASE_PATTERN = regex.compile(r"\p{Emoji}")
_MAX_EXCEL_ROW_HEIGHT = 409.0


def _grapheme_units(grapheme: str) -> int:
    if grapheme == "\t":
        return 4
    if _EMOJI_WIDTH_PATTERN.search(grapheme) or (
        "\ufe0f" in grapheme and _EMOJI_BASE_PATTERN.search(grapheme)
    ):
        return 2
    units = 0
    for char in grapheme:
        if char in "\r\n":
            continue
        if unicodedata.combining(char) or unicodedata.category(char) in {"Mn", "Me", "Cf"}:
            continue
        char_units = 2 if unicodedata.east_asian_width(char) in {"W", "F"} else 1
        units = max(units, char_units)
    return units


def _display_units(value) -> int:
    text = "" if value is None else str(value)
    return sum(
        _grapheme_units(grapheme)
        for grapheme in _GRAPHEME_PATTERN.findall(text)
    )


def _wrapped_line_count(value, column_width) -> int:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    capacity = max(1, math.floor(float(column_width) * 0.88))
    total_lines = 0

    def place_characters(characters, used_units):
        additional_lines = 0
        for grapheme in _GRAPHEME_PATTERN.findall(characters):
            grapheme_units = _grapheme_units(grapheme)
            if not grapheme_units:
                continue
            if used_units and used_units + grapheme_units > capacity:
                additional_lines += 1
                used_units = 0
            used_units += grapheme_units
        return additional_lines, used_units

    for explicit_line in text.split("\n"):
        matches = list(re.finditer(r"\S+", explicit_line))
        if not matches:
            additional_lines, _ = place_characters(explicit_line, 0)
            total_lines += 1 + additional_lines
            continue

        wrapped_lines = 1
        additional_lines, used_units = place_characters(
            explicit_line[:matches[0].start()],
            0,
        )
        wrapped_lines += additional_lines
        previous_end = matches[0].start()
        for index, match in enumerate(matches):
            separator = (
                explicit_line[previous_end:match.start()]
                if index
                else ""
            )
            token_units = _display_units(match.group())
            separator_units = _display_units(separator)
            if used_units and used_units + separator_units + token_units > capacity:
                wrapped_lines += 1
                used_units = 0
                separator = ""
            additional_lines, used_units = place_characters(
                separator + match.group(),
                used_units,
            )
            wrapped_lines += additional_lines
            previous_end = match.end()
        additional_lines, _ = place_characters(
            explicit_line[previous_end:],
            used_units,
        )
        wrapped_lines += additional_lines
        total_lines += wrapped_lines

    return total_lines


def _wrapped_row_height(cells, minimum=15.75, *, context="wrapped row") -> float:
    maximum_lines = max(
        (_wrapped_line_count(value, column_width) for value, column_width in cells),
        default=1,
    )
    required_height = max(float(minimum), 2.0 + 16.5 * maximum_lines)
    if required_height > _MAX_EXCEL_ROW_HEIGHT:
        raise ValueError(
            f"{context}: {maximum_lines} wrapped lines require {required_height:g} pt; "
            f"Excel/WPS row height is limited to {_MAX_EXCEL_ROW_HEIGHT:g} pt"
        )
    return required_height


def _fit_wrapped_row(cells, minimum=15.75, *, context="wrapped row") -> tuple[float, float]:
    maximum_lines = max(
        (_wrapped_line_count(value, column_width) for value, column_width in cells),
        default=1,
    )
    for font_size, line_height in ((11.0, 16.5), (10.0, 15.0), (9.0, 13.5)):
        required_height = max(float(minimum), 2.0 + line_height * maximum_lines)
        if required_height <= _MAX_EXCEL_ROW_HEIGHT:
            return required_height, font_size
    raise ValueError(
        f"{context}: {maximum_lines} wrapped lines still require {required_height:g} pt "
        f"at {font_size:g} pt; Excel/WPS row height is limited to "
        f"{_MAX_EXCEL_ROW_HEIGHT:g} pt"
    )


def _fit_or_span_wrapped_row(
    cells,
    minimum=15.75,
    *,
    context="wrapped row",
) -> tuple[float, float, int]:
    try:
        height, font_size = _fit_wrapped_row(
            cells,
            minimum,
            context=context,
        )
        return height, font_size, 1
    except ValueError:
        maximum_lines = max(
            (_wrapped_line_count(value, column_width) for value, column_width in cells),
            default=1,
        )
        total_height = max(float(minimum), 2.0 + 13.5 * maximum_lines)
        row_span = math.ceil(total_height / _MAX_EXCEL_ROW_HEIGHT)
        return total_height / row_span, 9.0, row_span


def _set_row_font_size(worksheet, row: int, columns: int, font_size: float) -> None:
    if font_size >= 11.0:
        return
    for column in range(1, columns + 1):
        cell = worksheet.cell(row=row, column=column)
        font = copy(cell.font)
        font.sz = font_size
        cell.font = font


def _segment_filename(state: dict, segment: dict, source_row=None) -> str:
    if state.get("input_format") == "sdlxliff":
        metadata = segment.get("metadata") or {}
        sdl_metadata = metadata.get("sdlxliff") or {}
        file_original = _text(sdl_metadata.get("file_original"))
        if file_original:
            return file_original
        source_ref = segment.get("source_ref") or {}
        relative_path = _text(source_ref.get("relative_path"))
        if relative_path:
            return relative_path
        return Path(state.get("input_path") or "").name
    fallback = Path(state.get("input_path") or "").stem
    headers = state.get("headers") or []
    try:
        source_path_index = headers.index("来源相对路径")
    except ValueError:
        return fallback
    if source_row is None:
        segments = state.get("segments") or []
        rows = state.get("rows_raw") or []
        if len(rows) == len(segments):
            segment_id = segment.get("id")
            for index, candidate in enumerate(segments):
                if candidate.get("id") == segment_id:
                    source_row = rows[index]
                    break
    if not source_row or source_path_index >= len(source_row):
        return fallback
    value = source_row[source_path_index]
    return _text(value) or fallback


def _report_source_table(state: dict) -> tuple[list[str], list[list[object]]]:
    segments = state.get("segments") or []
    if state.get("input_format") == "sdlxliff":
        headers = ["来源文件", "TU ID", "SDL Segment ID", "原文", "译文"]
        rows = []
        for segment in segments:
            source_ref = segment.get("source_ref") or {}
            rows.append(
                [
                    _segment_filename(state, segment),
                    source_ref.get("tu_id") or "",
                    source_ref.get("sdl_segment_id") or "",
                    segment.get("source", ""),
                    segment.get("target", ""),
                ]
            )
        return headers, rows

    rows_raw = state.get("rows_raw") or []
    if len(rows_raw) != len(segments):
        raise ValueError(
            "tabular report rows_raw/segments length mismatch: "
            f"rows_raw={len(rows_raw)} segments={len(segments)}"
        )
    return list(state.get("headers") or []), [list(row) for row in rows_raw]


def _protection_reason(segment: dict, protected_ids: set[int]) -> str:
    if segment.get("id") not in protected_ids:
        return ""
    return _text(segment.get("protected_reason")) or "TM_100_MATCH"


def _protection_evidence(segment: dict, protected_ids: set[int]) -> str:
    reason = _protection_reason(segment, protected_ids)
    if not reason:
        return ""
    evidence = segment.get("protection_evidence")
    if evidence is None:
        return reason
    return json.dumps(
        {"reason": reason, "evidence": evidence},
        ensure_ascii=False,
        sort_keys=True,
    )


def _build_xlsx(
    state,
    history,
    score,
    threshold,
    out_path,
    scorecard_profile_id="legacy",
    *,
    announce=True,
    scoring_policy=None,
    scoring_computation=None,
    report_contract_results=None,
):
    if scoring_policy is not None:
        scorecard_profile_id = scoring_policy["scorecard_profile"]
        threshold = scoring_policy["threshold"]
    scorecard_profile = load_scorecard_profile(scorecard_profile_id)
    severity_points = scorecard_severity_points(
        scorecard_profile,
        (scoring_policy or {}).get("severity_scale", "lisa"),
    )
    categories = scorecard_category_order(scorecard_profile)
    check_scope = get_check_scope(state)
    terminology_status = (
        "Enabled"
        if check_scope["terminology_enabled"]
        else "Disabled by runtime request"
    )
    enabled_modules = ", ".join(check_scope["enabled_modules"])
    scope_summary = (
        f"Terminology check: {terminology_status}; "
        f"Enabled modules: {enabled_modules}"
    )
    segments = state["segments"]
    seg_map = {s["id"]: s for s in segments}
    cat_counts: dict[str, dict[str, int]] = {
        cat: {"Neutral": 0, "Minor": 0, "Major": 0, "Critical": 0}
        for cat in categories
    }
    rep_counts: dict[str, dict[str, int]] = {
        cat: {"Neutral": 0, "Minor": 0, "Major": 0, "Critical": 0}
        for cat in categories
    }
    detail_rows: list[dict] = []
    all_protected_ids = set()
    for entry in history:
        all_protected_ids.update(entry.get("protected_ids", []))
        all_protected_ids.update(
            result["id"]
            for result in entry.get("errors", [])
            if any(issue.get("protected") for issue in result.get("errors", []))
        )
    for seg in segments:
        if seg.get("protected"):
            all_protected_ids.add(seg["id"])
    max_iter = max((entry["iteration"] for entry in history), default=0)
    latest_entry = history[-1] if history else None

    for entry in history:
        fixed = entry["iteration"] < max_iter
        for e_seg in entry["errors"]:
            seg = seg_map.get(e_seg["id"])
            if not seg:
                continue
            if seg["id"] in all_protected_ids:
                continue
            corrected = e_seg.get("corrected")
            if (
                corrected is None
                and entry["iteration"] == max_iter
                and current_target(seg) != seg.get("target", "")
            ):
                corrected = current_target(seg)
            for e in e_seg.get("errors", []):
                cat = normalize_category_for_profile(e.get("category", "Other"), scorecard_profile)
                sev = apply_severity(cat, e.get("severity", "Minor"), scorecard_profile)
                if entry is latest_entry:
                    if e.get("repeated"):
                        if cat in rep_counts:
                            rep_counts[cat][sev] = rep_counts[cat].get(sev, 0) + 1
                    elif cat in cat_counts:
                        cat_counts[cat][sev] = cat_counts[cat].get(sev, 0) + 1
                review_status, edit_status, check_source = _issue_review_columns(
                    e,
                    seg["id"],
                )
                detail_rows.append({
                    "filename": _segment_filename(state, seg),
                    "seg_id":   seg["id"],
                    "source":   seg["source"],
                    "original": seg["target"],
                    "corrected": corrected,
                    "parent":   scorecard_category_parent(cat, scorecard_profile),
                    "category": cat,
                    "severity": sev,
                    "iteration": f"Iter {entry['iteration']}",
                    "comment":  ("[Repeated] " if e.get("repeated") else "") + e.get("comment", ""),
                    "fixed":    fixed,
                    "processing": _issue_processing_label(
                        e,
                        {**e_seg, "corrected": corrected},
                        protected=False,
                    ),
                    "review_status": review_status,
                    "edit_status": edit_status,
                    "check_source": check_source,
                })

    if scoring_computation is not None:
        for category in categories:
            cat_counts[category] = {
                severity: int(
                    scoring_computation.get("category_counts", {})
                    .get(category, {})
                    .get(severity, 0)
                )
                for severity in ("Neutral", "Minor", "Major", "Critical")
            }
            rep_counts[category] = {
                severity: int(
                    scoring_computation.get("repeated_counts", {})
                    .get(category, {})
                    .get(severity, 0)
                )
                for severity in ("Neutral", "Minor", "Major", "Critical")
            }

    total_counts = {"Neutral": 0, "Minor": 0, "Major": 0, "Critical": 0}
    total_rep    = {"Neutral": 0, "Minor": 0, "Major": 0, "Critical": 0}
    for c in cat_counts.values():
        for sev, n in c.items():
            total_counts[sev] += n
    for c in rep_counts.values():
        for sev, n in c.items():
            total_rep[sev] += n
    total_raw = sum(
        raw_points(counts, scorecard_profile, severity_points)
        for counts in cat_counts.values()
    )
    total_weighted = (
        scoring_computation["total_weighted"]
        if scoring_computation is not None
        else sum(
            weighted_points(cat, counts, scorecard_profile, severity_points)
            for cat, counts in cat_counts.items()
        )
    )

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "LQA Scorecard"
    latest_status = history[-1].get("status") if history else None
    critical_gate_fail = bool(
        (scoring_policy or {}).get("critical_gate")
        and total_counts.get("Critical", 0)
    )
    status = latest_status or (
        "FAIL" if critical_gate_fail or score < threshold else "PASS"
    )

    # ── 导读 sheet（插最前）：说明后两个 sheet 的用处 + 建议使用思路 ──
    intro = wb.create_sheet("说明·导读", 0)
    _wrap = Alignment(wrap_text=True, vertical="top")
    for r in [
        ("LQE 质检报告 · 导读", "", "", ""),
        ("", "", "", ""),
        (f"本报告 3 个 sheet：本「说明·导读」+ 后两个正文。阈值 {threshold}；本次 {status}（{score:.2f} 分）。", "", "", ""),
        ("Check scope", scope_summary, "", ""),
        ("Sheet", "是什么", "给谁看", "怎么用"),
        ("LQA Scorecard（第 2 个）", "计分卡：过/不过 + 分数 + 错误(类别×严重度)分布 + 罚分", "PM / 客户", f"先看这里拿整体判定：是否达标(阈值 {threshold})、差在哪类"),
        ("LQE Results（第 3 个）", "错误明细：同段连续、每个错误一行，并标注 AI 复核与编辑状态", "审校 / 译员", "按 LQE Segment ID 查看同段错误，并核对 AI 复核与编辑状态"),
        ("", "", "", ""),
        ("建议使用思路：", "", "", ""),
        ("1. PM 先看「LQA Scorecard」拿整体判定（分数 / 状态 / 错误分布）。", "", "", ""),
        ("2. 审校/译员到「LQE Results」，同一 LQE Segment ID 的多个错误会连续列出，每个错误一行。", "", "", ""),
        ("3. 「AI 模块记录」是本地内容绑定证据，不是外部身份签名；「已生成并验证建议」不表示已写回 state。", "", "", ""),
        ("4. 「建议译文」留空时保留原译；「处理方式」说明是否仅提醒或需要人工确认。", "", "", ""),
        ("5. 修正若要落地，按项目流程（如改在线 memoQ）。", "", "", ""),
        ("6. 已保护内容不修改；需要人工确认的问题由审校人员判断。", "", "", ""),
    ]:
        intro.append(r)
        for c in intro[intro.max_row]:
            c.alignment = _wrap
    intro["A1"].font = Font(bold=True, size=13, color="073763")
    intro["A4"].font = _BOLD
    for c in intro[5]:                       # 表头行
        c.fill = _DARK_BLUE; c.font = _WHITE_FONT
    intro["A9"].font = Font(bold=True)
    for col, w in zip("ABCD", [22, 50, 12, 46]):
        intro.column_dimensions[col].width = w
    wb.active = intro

    def _db_row(row, height=14.25):
        ws.row_dimensions[row].height = height
        for col in range(1, 11):
            _s(ws.cell(row=row, column=col), fill=_DARK_BLUE, font=_WHITE_FONT)

    for r in [1, 2, 3]:
        ws.row_dimensions[r].height = 23.25
        _db_row(r, height=23.25)
    ws.merge_cells("A1:J3")
    c = ws["A1"]
    c.value = "LQA Scorecard"
    c.font  = Font(color="FFFFFF", size=16, bold=True)
    c.alignment = _CENTER

    source_lang = state.get("source_lang") or "-"
    target_lang = state.get("target_lang") or "-"
    info = [
        ("File", Path(state["input_path"]).name, "Wordcount", state.get("wordcount", 0)),
        ("Source language", source_lang, "Target language", target_lang),
        ("Total iterations", len(history), "Threshold", threshold),
        ("Date", date.today().isoformat(), "Terminology check", terminology_status),
    ]
    for ri, (l1, v1, l2, v2) in enumerate(info, start=4):
        _db_row(ri, height=15.0)
        ws.cell(row=ri, column=1, value=l1)
        ws.merge_cells(f"B{ri}:D{ri}")
        c = ws.cell(row=ri, column=2)
        _set_excel_text(c, v1)
        c.fill = _DARK_BLUE; c.font = Font(color="FFFFFF")
        ws.cell(row=ri, column=5, value=l2)
        ws.merge_cells(f"F{ri}:H{ri}")
        c = ws.cell(row=ri, column=6)
        _set_excel_text(c, v2)
        c.fill = _DARK_BLUE; c.font = Font(color="FFFFFF")

    ws.row_dimensions[8].height = 6

    _db_row(9)
    ws.merge_cells("A9:J9")
    c = ws["A9"]
    c.value = "LQA results"
    c.alignment = _LEFT_MID

    for row, height in [(10, 34.0), (11, 34.0), (12, 34.0)]:
        ws.row_dimensions[row].height = height

    for row, label, val, val_fill, val_font in [
        (10, "Status",      status,          _RED if status == "FAIL" else _GREEN,
                                             Font(color="FFFFFF", bold=True)),
        (11, "Final score", round(score, 4), _ORANGE, Font(bold=True)),
        (12, "Threshold",   threshold,       _ORANGE, Font(bold=True)),
    ]:
        c = ws.cell(row=row, column=1, value=label)
        _s(c, fill=_LIGHT_BLUE, align=_CENTER)
        c = ws.cell(row=row, column=2, value=val)
        _s(c, fill=val_fill, font=val_font, align=_CENTER)

    ws.merge_cells("C10:J12")
    c = ws["C10"]
    c.value = "Overall feedback"
    _s(c, fill=_LIGHT_BLUE, align=_CENTER)

    cur_row = 13
    if len(history) > 1:
        ws.row_dimensions[cur_row].height = 6
        cur_row += 1
        _db_row(cur_row)
        ws.merge_cells(f"A{cur_row}:J{cur_row}")
        c = ws.cell(row=cur_row, column=1, value="Iteration log")
        c.alignment = _LEFT_MID
        cur_row += 1
        for col, hdr in [(1,"Iteration"),(2,"Score"),(3,"Errors found"),(4,"Corrections applied")]:
            c = ws.cell(row=cur_row, column=col, value=hdr)
            _s(c, fill=_LIGHT_BLUE, align=_CENTER, font=_BOLD)
        ws.row_dimensions[cur_row].height = 14.25
        cur_row += 1
        for entry in history:
            s  = entry.get("score")
            ec = sum(len(e.get("errors", [])) for e in entry["errors"])
            for col, val in [
                (1, f"Iter {entry['iteration']}"),
                (2, round(s, 2) if s is not None else ""),
                (3, ec),
                (4, entry.get("corrections_count", 0)),
            ]:
                c = ws.cell(row=cur_row, column=col, value=val)
                _s(c, fill=_ORANGE, align=_CENTER)
            ws.row_dimensions[cur_row].height = 14.25
            cur_row += 1

    ws.row_dimensions[cur_row].height = 6
    cur_row += 1

    _db_row(cur_row)
    ws.merge_cells(f"A{cur_row}:L{cur_row}")
    c = ws.cell(row=cur_row, column=1, value="Error summary")
    c.alignment = _LEFT_MID
    cur_row += 1

    ws.row_dimensions[cur_row].height = 15.0
    ws.merge_cells(f"A{cur_row}:A{cur_row+1}")
    ws.merge_cells(f"B{cur_row}:B{cur_row+1}")
    ws.merge_cells(f"C{cur_row}:J{cur_row}")
    ws.merge_cells(f"K{cur_row}:L{cur_row}")
    for col, val in [(1,"Error category"),(2,"Weight"),(3,"Error severity"),(11,"Penalty points")]:
        c = ws.cell(row=cur_row, column=col, value=val)
        _s(c, font=_BOLD, align=_CENTER)
    cur_row += 1

    ws.row_dimensions[cur_row].height = 14.25
    for col, val in enumerate(
        ["Neutral","Neutral – repeated","Minor","Minor – repeated",
         "Major","Major – repeated","Critical","Critical – repeated","Raw","Weighted"], start=3):
        c = ws.cell(row=cur_row, column=col, value=val)
        _s(c, fill=_LIGHT_BLUE, align=_CENTER)
    cur_row += 1

    ws.row_dimensions[cur_row].height = 14.25
    for col, val in [
        (1,"TOTAL"),(2,None),
        (3,total_counts.get("Neutral",0)),(4,total_rep.get("Neutral",0)),
        (5,total_counts.get("Minor",0)),  (6,total_rep.get("Minor",0)),
        (7,total_counts.get("Major",0)),  (8,total_rep.get("Major",0)),
        (9,total_counts.get("Critical",0)),(10,total_rep.get("Critical",0)),
        (11,total_raw),(12,round(total_weighted,2)),
    ]:
        c = ws.cell(row=cur_row, column=col, value=val)
        _s(c, fill=_ORANGE, align=_CENTER)
    cur_row += 1

    for cat in categories:
        counts = cat_counts[cat]
        r = raw_points(counts, scorecard_profile, severity_points)
        w = weighted_points(cat, counts, scorecard_profile, severity_points)
        ws.row_dimensions[cur_row].height = 14.25
        rep = rep_counts[cat]
        for col, val in [
            (1,cat),(2,scorecard_category_weight(cat, scorecard_profile)),
            (3,counts.get("Neutral",0)),(4,rep.get("Neutral",0)),
            (5,counts.get("Minor",0)),  (6,rep.get("Minor",0)),
            (7,counts.get("Major",0)),  (8,rep.get("Major",0)),
            (9,counts.get("Critical",0)),(10,rep.get("Critical",0)),
            (11,r),(12,round(w,2)),
        ]:
            c = ws.cell(row=cur_row, column=col, value=val)
            _s(c, align=_CENTER)
        cur_row += 1

    ws.row_dimensions[cur_row].height = 6
    cur_row += 1

    is_sdlxliff_report = state.get("input_format") == "sdlxliff"
    filename_width = (
        45
        if is_sdlxliff_report
        else 70
        if "来源相对路径" in (state.get("headers") or [])
        else 22
    )
    scorecard_widths = [
        filename_width, 8, 80, 80, 80,
        14, 22, 10, 10, 45, 12, 18, 18, 25, 12, 45,
    ]
    for column, width in enumerate(scorecard_widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width

    ws.row_dimensions[cur_row].height = 14.25
    for col, hdr in enumerate(
        ["File name","Segment #","Source text","原译",
         "建议译文","Error category","Error sub-category",
         "Error severity","Iteration","Reviewer's comment","处理方式",
         "AI 复核状态","AI 编辑状态","检查来源",
         "Protected","Protection Evidence"], start=1):
        c = ws.cell(row=cur_row, column=col, value=hdr)
        _s(c, fill=_DARK_BLUE, font=_WHITE_FONT, align=_CENTER)
    cur_row += 1

    for dr in detail_rows:
        row_fill = _GREEN_LIGHT if dr["fixed"] else None
        rich_pair = (
            build_rich_diff(dr["original"], dr["corrected"])
            if isinstance(dr["corrected"], str)
            else None
        )
        detail_values = [
            dr["filename"], dr["seg_id"], dr["source"], dr["original"],
            dr["corrected"], dr["parent"], dr["category"], dr["severity"],
            dr["iteration"], dr["comment"], dr["processing"],
            dr["review_status"], dr["edit_status"], dr["check_source"],
            "Yes" if dr["seg_id"] in all_protected_ids else "No",
            _protection_evidence(seg_map[dr["seg_id"]], all_protected_ids),
        ]
        if rich_pair:
            detail_values[3], detail_values[4] = rich_pair
        for col, val in enumerate(detail_values, start=1):
            c = ws.cell(row=cur_row, column=col)
            _set_excel_text(c, val)
            _s(
                c,
                fill=row_fill,
                align=(
                    _CENTER
                    if col in (2, 8, 9, 11, 12, 13, 15)
                    else _LEFT_TOP
                ),
            )
        row_height, row_font_size, row_span = _fit_or_span_wrapped_row(
            [
                (
                    value,
                    ws.column_dimensions[get_column_letter(column)].width,
                )
                for column, value in enumerate(detail_values, start=1)
            ],
            context=f"LQA Scorecard row {cur_row}",
        )
        for physical_row in range(cur_row, cur_row + row_span):
            ws.row_dimensions[physical_row].height = row_height
        _set_row_font_size(ws, cur_row, len(detail_values), row_font_size)
        if row_span > 1:
            for column in range(1, len(detail_values) + 1):
                ws.merge_cells(
                    start_row=cur_row,
                    start_column=column,
                    end_row=cur_row + row_span - 1,
                    end_column=column,
                )
        cur_row += row_span

    ws2 = wb.create_sheet("LQE Results")

    def _fmt_errors(errs):
        return "\n".join(issue_detail(error) for error in errs)

    _WRAP_TOP = Alignment(wrap_text=True, vertical="top")

    current_entries = {e["id"]: e for e in history[-1].get("errors", [])} if history else {}
    report_headers, report_rows = _report_source_table(state)
    target_col = state.get("target_col")
    if target_col is None:
        target_col = "译文" if "译文" in report_headers else 1
    try:
        target_index = int(target_col)
    except (ValueError, TypeError):
        target_index = report_headers.index(target_col) if target_col in report_headers else 1
    if 0 <= target_index < len(report_headers):
        report_headers[target_index] = "原译"

    used_headers = {str(header) for header in report_headers}

    def _audit_header(base: str) -> str:
        candidate = base
        suffix = 2
        while candidate in used_headers:
            candidate = f"{base}（审计 {suffix}）"
            suffix += 1
        used_headers.add(candidate)
        return candidate

    segment_audit_header = _audit_header(AUDIT_HEADER_BASES["segment_id"])
    issue_number_header = _audit_header(AUDIT_HEADER_BASES["issue_number"])
    review_status_header = _audit_header(AUDIT_HEADER_BASES["review_status"])
    edit_status_header = _audit_header(AUDIT_HEADER_BASES["edit_status"])
    check_source_header = _audit_header(AUDIT_HEADER_BASES["check_source"])

    ws2_headers = report_headers + [
        "建议译文",
        "处理方式",
        segment_audit_header,
        issue_number_header,
        review_status_header,
        edit_status_header,
        check_source_header,
        "错误详情",
        "Protected",
        "Protection Evidence",
        "LQE_Iter",
    ]
    results_widths = []
    source_header = state.get("source_col")
    for header in ws2_headers:
        if header in {"原文", "原译", "建议译文", source_header}:
            width = 80
        elif header == "错误详情":
            width = 80
        elif header == "Protection Evidence":
            width = 45
        elif is_sdlxliff_report and header == "来源文件":
            width = 35
        elif is_sdlxliff_report and header == "TU ID":
            width = 38
        elif is_sdlxliff_report and header == "SDL Segment ID":
            width = 14
        elif header == "处理方式":
            width = 18
        elif header == segment_audit_header:
            width = 12
        elif header == issue_number_header:
            width = 10
        elif header in {review_status_header, edit_status_header}:
            width = 18
        elif header == check_source_header:
            width = 24
        elif header == "LQE_Iter":
            width = 10
        elif header == "Protected":
            width = 12
        else:
            width = 20
        results_widths.append(width)
    for column, width in enumerate(results_widths, start=1):
        ws2.column_dimensions[get_column_letter(column)].width = width

    for ci, h in enumerate(ws2_headers, start=1):
        c = ws2.cell(row=1, column=ci)
        _set_excel_text(c, h)
        _s(c, fill=_DARK_BLUE, font=_WHITE_FONT, align=_CENTER)
    ws2.row_dimensions[1].height = 15.0

    ri = 2
    for segment_index, seg in enumerate(segments):
        raw_row = report_rows[segment_index]
        is_protected = seg["id"] in all_protected_ids
        current_entry = current_entries.get(seg["id"])
        baseline = (
            current_target(seg)
            if current_target(seg) != seg.get("target", "")
            else None
        )
        if current_entry is None:
            entry = {"errors": [], "corrected": baseline}
        else:
            entry = {
                **current_entry,
                "corrected": (
                    current_entry.get("corrected")
                    if current_entry.get("corrected") is not None
                    else baseline
                ),
            }
        errs = [] if is_protected else entry.get("errors", [])
        has_error = bool(errs)
        segment_processing = _issue_processing_label(
            None,
            entry,
            protected=is_protected,
        )
        corrected = entry.get("corrected")
        suggestion = (
            "" if segment_processing == "已保护，不修改" else (corrected or "")
        )
        rich_pair = (
            build_rich_diff(seg["target"], suggestion)
            if corrected is not None and segment_processing != "已保护，不修改"
            else None
        )
        suggestion_column = len(report_headers) + 1
        row_fill = _ORANGE if has_error else _GREEN_LIGHT if (is_protected or suggestion) else None
        row_issues = errs or [None]
        for issue_index, issue in enumerate(row_issues, start=1):
            review_status, edit_status, check_source = _issue_review_columns(
                issue,
                seg["id"],
            )
            processing = _issue_processing_label(
                issue,
                entry,
                protected=is_protected,
            )
            row_data = list(raw_row) + [
                suggestion,
                processing,
                seg["id"],
                issue_index if issue is not None else "",
                review_status,
                edit_status,
                check_source,
                _fmt_errors([issue]) if issue is not None else "",
                "Yes" if is_protected else "No",
                _protection_evidence(seg, all_protected_ids),
                seg.get("iter", 0),
            ]
            if rich_pair and 0 <= target_index < len(report_headers):
                row_data[target_index] = rich_pair[0]
            if rich_pair:
                row_data[suggestion_column - 1] = rich_pair[1]
            for ci, val in enumerate(row_data, start=1):
                c = ws2.cell(row=ri, column=ci)
                _set_excel_text(c, val)
                c.alignment = _WRAP_TOP
                if row_fill:
                    c.fill = row_fill
            row_height, row_font_size, row_span = _fit_or_span_wrapped_row(
                [
                    (
                        value,
                        ws2.column_dimensions[get_column_letter(column)].width,
                    )
                    for column, value in enumerate(row_data, start=1)
                ],
                context=f"LQE Results row {ri}",
            )
            for physical_row in range(ri, ri + row_span):
                ws2.row_dimensions[physical_row].height = row_height
            _set_row_font_size(ws2, ri, len(row_data), row_font_size)
            if row_span > 1:
                for column in range(1, len(row_data) + 1):
                    ws2.merge_cells(
                        start_row=ri,
                        start_column=column,
                        end_row=ri + row_span - 1,
                        end_column=column,
                    )
            ri += row_span

    if report_contract_results is not None:
        attach_report_contract(wb, state, report_contract_results)
    wb.save(str(out_path))
    if announce:
        print(f"[lqe_io] Output → {out_path}")


def _cmd_write_locked(
    args,
    state_path: Path,
    state: dict,
    segments: list[dict],
    manifest: dict | None,
    revalidate,
):
    errors_path = Path(args.errors)
    final_errors_data = read_json(errors_path)
    original_errors_data = deepcopy(final_errors_data)
    _validate_scope_or_exit(
        state,
        final_errors_data,
        issues_key="errors",
        label=errors_path.name,
        command="write",
    )
    protected_ids = _state_protected_ids(state)
    scrubbed = _scrub_protected_entries(final_errors_data, protected_ids)
    try:
        scoring_policy = resolve_scoring_policy(
            state,
            scoring_policy_overrides(args),
        )
        scorecard_profile = load_scorecard_profile(
            scoring_policy["scorecard_profile"]
        )
        validation_messages = _validate_errors(
            final_errors_data,
            {segment["id"] for segment in state.get("segments", [])},
            scorecard_profile,
        )
        for message in validation_messages:
            print(f"[validate] {message}")
        final_errors_data = _verify_result_payload_with_segments(
            state,
            segments,
            manifest,
            final_errors_data,
            errors_path,
            command="write",
        )
        computation = score_errors(
            state,
            final_errors_data,
            scoring_policy,
            protected_ids=protected_ids,
        )
    except (CheckFormatError, ValueError) as exc:
        raise SystemExit(f"[write] {exc}") from exc
    final_errors_data = computation["annotated_errors"]
    score_result = computation["output"]
    score = score_result["score"]
    supplied_score = float(args.score)
    if not math.isclose(supplied_score, score, abs_tol=0.005):
        print(
            f"[write] supplied score {supplied_score:g} differs from "
            f"recomputed score {score:g}; using recomputed score",
            file=sys.stderr,
        )
    state["threshold"] = scoring_policy["threshold"]
    state["scoring_policy"] = scoring_policy
    state["pending_recheck"] = False

    final_entry = {
        "iteration": state.get("iteration", 0),
        "score": score,
        "status": score_result["status"],
        "errors": final_errors_data,
        "corrections_count": 0,
        "protected_ids": sorted(protected_ids),
    }

    history = state.get("error_history", [])
    final_iter = state.get("iteration", 0)
    skipped_corrections = [
        skipped
        for entry in history
        if entry.get("iteration") == final_iter
        for skipped in entry.get("skipped_corrections", [])
    ]
    if skipped_corrections:
        final_entry["skipped_corrections"] = skipped_corrections
    history = [
        entry for entry in history if entry.get("iteration") != final_iter
    ]
    history.append(final_entry)
    state["error_history"] = history

    out_path = state_path.parent / (_job_label(state_path) + "_lqe.xlsx")
    try:
        validate_artifact_paths(
            {"LQE report": out_path},
            {
                "state": state_path,
                "errors": errors_path,
                **state_reference_paths(state),
            },
            context="write",
        )
    except ValueError as exc:
        raise SystemExit(f"[write] {exc}") from exc
    with tempfile.NamedTemporaryFile(
        prefix=f".{out_path.stem}.",
        suffix=out_path.suffix,
        dir=out_path.parent,
        delete=False,
    ) as staging_file:
        staging_path = Path(staging_file.name)
    try:
        try:
            _build_xlsx(
                state,
                history,
                score,
                scoring_policy["threshold"],
                staging_path,
                scoring_policy["scorecard_profile"],
                announce=False,
                scoring_policy=scoring_policy,
                scoring_computation=computation,
                report_contract_results=final_errors_data,
            )
        except ValueError as exc:
            raise SystemExit(f"[write] {exc}") from exc
        revalidate()
        _assert_json_unchanged(
            errors_path,
            original_errors_data,
            label="errors input",
        )
        _publish_write_transaction(
            state_path,
            state,
            errors_path,
            final_errors_data,
            out_path,
            staging_path,
            publish_errors=final_errors_data != original_errors_data,
            manifest=manifest,
        )
    finally:
        staging_path.unlink(missing_ok=True)

    if scrubbed:
        print(f"[write] scrubbed {scrubbed} protected-segment issue(s) from {errors_path}")
    print(f"[lqe_io] Output → {out_path}")


def cmd_write(args):
    from lqe_chunk import verification_generation_lease

    state_path = Path(args.state)
    try:
        with verification_generation_lease(
            state_path,
            exclusive=True,
        ) as (state, segments, manifest, revalidate):
            _cmd_write_locked(
                args,
                state_path,
                state,
                segments,
                manifest,
                revalidate,
            )
    except ValueError as exc:
        raise SystemExit(f"[write] {exc}") from exc


# ── pre-check（实现在 lqe_checks.py）─────────────────────────────────────────

def cmd_pre_check(args):
    from lqe_checks import run_pre_check
    from lqe_split_contract import generation_lock

    state_path = Path(args.state)
    try:
        with generation_lock(state_path.parent / "chunks", exclusive=True):
            run_pre_check(
                state_path,
                Path(args.out) if args.out else None,
            )
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[pre-check] {exc}") from exc


# ── export ───────────────────────────────────────────────────────────────────

def _export_sdlxliff_xlsx(
    state_path: Path,
    state: dict,
    result_entries: dict,
    *,
    out_path: Path | None = None,
) -> Path:
    headers, rows = _report_source_table(state)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Corrected"
    worksheet.append(headers)
    for index, segment in enumerate(state.get("segments") or []):
        source_row = rows[index]
        entry = result_entries[segment["id"]]
        protected = bool(segment.get("protected")) or (
            _processing_label(entry) == "已保护，不修改"
        )
        corrected = entry.get("corrected")
        output_row = list(source_row)
        if not protected and corrected is not None:
            output_row[4] = corrected
        row_number = worksheet.max_row + 1
        for column, value in enumerate(output_row, start=1):
            _set_excel_text(
                worksheet.cell(row=row_number, column=column),
                value,
            )
    out_path = out_path or state_path.parent / (
        _job_label(state_path) + "_corrected.xlsx"
    )
    workbook.save(out_path)
    workbook.close()
    return out_path


def _verification_segments(state_path: Path, state: dict) -> list[dict]:
    from lqe_chunk import load_verification_segments

    try:
        return load_verification_segments(
            state_path,
            state=state,
        )
    except (OSError, ValueError) as exc:
        raise CheckFormatError(str(exc)) from exc


def _verify_result_payload_with_segments(
    state: dict,
    segments: list[dict],
    manifest: dict | None,
    errors_data: list,
    errors_path: Path,
    *,
    command: str,
) -> list[dict]:
    _validate_scope_or_exit(
        state,
        errors_data,
        issues_key="errors",
        label=errors_path.name,
        command=command,
    )
    try:
        if requires_bound_artifacts(state):
            contract_path = result_contract_path(errors_path)
            if manifest is None or not contract_path.is_file():
                raise CheckFormatError(
                    f"{errors_path.name}: bound result contract is required"
                )
            validate_result_contract(
                read_json(contract_path),
                manifest,
                errors_data,
                label=errors_path.name,
            )
        bound = requires_bound_artifacts(state)
        return verify_results(
            segments,
            errors_data,
            str(errors_path),
            allow_internal_provenance=bound,
            require_internal_provenance=bound,
        )
    except (CheckFormatError, OSError, ValueError) as exc:
        raise SystemExit(f"[{command}] {exc}") from exc


def _verify_result_payload(
    state_path: Path,
    state: dict,
    errors_data: list,
    errors_path: Path,
    *,
    command: str,
) -> tuple[list[dict], list[dict]]:
    from lqe_chunk import verification_generation_lease

    try:
        with verification_generation_lease(
            state_path,
            exclusive=False,
        ) as (live_state, segments, manifest, _):
            if live_state != state:
                raise CheckFormatError("state changed while loading results")
            verified = _verify_result_payload_with_segments(
                live_state,
                segments,
                manifest,
                errors_data,
                errors_path,
                command=command,
            )
            return segments, verified
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[{command}] {exc}") from exc


def _state_no_header(state: dict) -> bool:
    if "no_header" in state:
        value = state["no_header"]
        if not isinstance(value, bool):
            raise ValueError("state.no_header must be a boolean")
        return value

    source_column = state.get("source_col")
    return isinstance(source_column, int) or (
        isinstance(source_column, str) and source_column.isdigit()
    )


def _state_column_index(state: dict, field: str) -> int:
    value = state.get(field)
    if _state_no_header(state) or isinstance(value, int):
        try:
            return int(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"cannot locate {field} {value!r}") from exc
    headers = state.get("headers") or []
    if value not in headers:
        raise ValueError(f"cannot locate {field} {value!r}")
    return headers.index(value)


def _validate_legacy_tabular_rows(
    state: dict,
    segments: list[dict],
    row_at,
) -> None:
    source_index = _state_column_index(state, "source_col")
    target_index = _state_column_index(state, "target_col")
    for segment in segments:
        row = row_at(segment)
        if row is None:
            raise ValueError(
                f"source row for segment {segment['id']} is missing"
            )
        source = _text(_cell(row, source_index))
        target = _text(_cell(row, target_index))
        if source != segment.get("source", "") or target != segment.get(
            "target", ""
        ):
            raise ValueError(
                f"source row for segment {segment['id']} changed after read"
            )


def _validate_export_source_digest(state: dict, source_path: Path) -> str:
    digest = file_sha256(source_path)
    expected = state.get("input_sha256")
    if expected is not None:
        if not isinstance(expected, str) or not re.fullmatch(
            r"[0-9a-f]{64}", expected
        ):
            raise ValueError("state.input_sha256 is invalid")
        if digest != expected:
            raise ValueError(f"source input changed after read: {source_path}")
    return digest


def _validate_sdl_source_snapshot(state: dict) -> dict[Path, str]:
    manifest_path = state.get("source_manifest_path")
    if not isinstance(manifest_path, str) or not manifest_path.strip():
        raise ValueError("SDLXLIFF state has no source_manifest_path")
    manifest = read_json(manifest_path)
    manifest_files = manifest.get("files") if isinstance(manifest, dict) else None
    if not isinstance(manifest_files, list):
        raise ValueError("SDLXLIFF source manifest files must be an array")

    input_root_value = state.get("input_path")
    if not isinstance(input_root_value, str) or not input_root_value.strip():
        raise ValueError("SDLXLIFF state has no input_path")
    input_root = Path(input_root_value)
    if input_root.is_file():
        live_paths = [input_root]
        relative_paths = [input_root.name]
    elif input_root.is_dir():
        live_paths = sorted(
            (
                path
                for path in input_root.rglob("*")
                if path.is_file() and path.suffix.casefold() == ".sdlxliff"
            ),
            key=lambda path: path.relative_to(input_root).as_posix(),
        )
        relative_paths = [
            path.relative_to(input_root).as_posix() for path in live_paths
        ]
    else:
        raise ValueError(f"SDLXLIFF input path is missing: {input_root}")

    recorded_paths = state.get("input_paths")
    if not isinstance(recorded_paths, list) or not all(
        isinstance(value, str) and value.strip() for value in recorded_paths
    ):
        raise ValueError("SDLXLIFF state.input_paths must be a string array")
    if [str(path.resolve()) for path in live_paths] != [
        str(Path(value).resolve()) for value in recorded_paths
    ]:
        raise ValueError("SDLXLIFF source file set changed after read")

    manifest_by_path = {}
    for index, item in enumerate(manifest_files):
        if not isinstance(item, dict):
            raise ValueError(
                f"SDLXLIFF source manifest files[{index}] must be an object"
            )
        relative_path = item.get("relative_path")
        digest = item.get("sha256")
        if not isinstance(relative_path, str) or not relative_path:
            raise ValueError(
                f"SDLXLIFF source manifest files[{index}].relative_path is invalid"
            )
        if not isinstance(digest, str) or not re.fullmatch(r"[0-9a-f]{64}", digest):
            raise ValueError(
                f"SDLXLIFF source manifest files[{index}].sha256 is invalid"
            )
        if relative_path in manifest_by_path:
            raise ValueError(
                f"SDLXLIFF source manifest has duplicate path: {relative_path}"
            )
        manifest_by_path[relative_path] = digest
    if set(manifest_by_path) != set(relative_paths):
        raise ValueError("SDLXLIFF source manifest file set is inconsistent")

    snapshot = {}
    for path, relative_path in zip(live_paths, relative_paths):
        digest = file_sha256(path)
        if digest != manifest_by_path[relative_path]:
            raise ValueError(f"SDLXLIFF source changed after read: {path}")
        snapshot[path.resolve()] = digest
    return snapshot


def _recheck_sdl_source_snapshot(
    state: dict,
    snapshot: dict[Path, str],
) -> None:
    try:
        current = _validate_sdl_source_snapshot(state)
    except ValueError as exc:
        raise ValueError(
            f"SDLXLIFF source changed during export: {exc}"
        ) from exc
    if current != snapshot:
        raise ValueError("SDLXLIFF source changed during export")


def _validate_export_paths(
    state_path: Path,
    state: dict,
    out_path: Path,
    errors_path: Path | None,
) -> None:
    protected_inputs = {
        "state": state_path,
        **state_reference_paths(state),
    }
    if errors_path is not None:
        protected_inputs["errors"] = errors_path
        if requires_bound_artifacts(state):
            protected_inputs["errors contract"] = result_contract_path(
                errors_path
            )
    validate_artifact_paths(
        {"corrected export": out_path},
        protected_inputs,
        context="export",
    )


def _cmd_export_locked(
    args,
    state_path: Path,
    state: dict,
    segments: list[dict],
    manifest: dict | None,
    revalidate,
):
    errors_path = Path(args.errors) if getattr(args, "errors", None) else None
    overlay_entries = None
    original_overlay = None
    if errors_path is not None:
        raw_entries = read_json(errors_path)
        original_overlay = deepcopy(raw_entries)
        _validate_scope_or_exit(
            state,
            raw_entries,
            issues_key="errors",
            label=errors_path.name,
            command="export",
        )
        _scrub_protected_entries(raw_entries, _state_protected_ids(state))
        overlay_entries = _verify_result_payload_with_segments(
            state,
            segments,
            manifest,
            raw_entries,
            errors_path,
            command="export",
        )

    def revalidate_inputs():
        revalidate()
        if errors_path is not None:
            _assert_json_unchanged(
                errors_path,
                original_overlay,
                label="errors input",
            )
    seg_map = {s["id"]: s for s in segments}
    result_entries = {
        segment["id"]: {
            "errors": [],
            "corrected": (
                current_target(segment)
                if current_target(segment) != segment.get("target", "")
                else None
            ),
        }
        for segment in segments
    }

    if overlay_entries is not None:
        for e in overlay_entries:
            seg = seg_map.get(e["id"])
            if seg is None:
                continue
            baseline = result_entries[e["id"]]["corrected"]
            result_entries[e["id"]] = {
                **e,
                "corrected": (
                    e.get("corrected")
                    if e.get("corrected") is not None
                    else baseline
                ),
            }

    counts = {
        "建议修改": 0,
        "需要人工确认": 0,
        "保持原译": 0,
        "已保护": 0,
    }

    def export_kind(segment):
        if segment.get("protected"):
            return "已保护"
        label = _processing_label(result_entries[segment["id"]])
        if label == "已保护，不修改":
            return "已保护"
        if label in ("建议修改", "需要人工确认"):
            return label
        return "保持原译"

    def print_summary(out_path):
        print(
            f"[export] 建议修改 {counts['建议修改']} / "
            f"需要人工确认 {counts['需要人工确认']} / "
            f"保持原译 {counts['保持原译']} / "
            f"已保护 {counts['已保护']} → {out_path}"
        )

    if state.get("input_format") == "sdlxliff":
        for segment in segments:
            counts[export_kind(segment)] += 1
        out_path = state_path.parent / (
            _job_label(state_path) + "_corrected.xlsx"
        )
        staged = None
        try:
            _validate_export_paths(
                state_path, state, out_path, errors_path
            )
            source_snapshot = _validate_sdl_source_snapshot(state)
            with tempfile.NamedTemporaryFile(
                dir=out_path.parent,
                prefix=f".{out_path.stem}.",
                suffix=out_path.suffix,
                delete=False,
            ) as handle:
                staged = Path(handle.name)
            _export_sdlxliff_xlsx(
                state_path,
                state,
                result_entries,
                out_path=staged,
            )
            _recheck_sdl_source_snapshot(state, source_snapshot)
            revalidate_inputs()
            publish_replacement_transaction([(staged, out_path)])
        except (OSError, ValueError) as exc:
            raise SystemExit(f"[export] {exc}") from exc
        finally:
            if staged is not None:
                staged.unlink(missing_ok=True)
        print_summary(out_path)
        return

    try:
        no_header = _state_no_header(state)
        ti = _state_column_index(state, "target_col")
    except ValueError as exc:
        print(f"[export] {exc}", file=sys.stderr)
        sys.exit(1)

    src_path = Path(state["input_path"])
    if src_path.suffix.lower() in (".csv", ".tsv"):
        delim = "\t" if src_path.suffix.lower() == ".tsv" else ","
        out_path = state_path.parent / (
            _job_label(state_path) + "_corrected" + src_path.suffix.lower()
        )
        enc = "utf-8-sig" if src_path.suffix.lower() == ".csv" else "utf-8"
        staged = None
        try:
            _validate_export_paths(
                state_path, state, out_path, errors_path
            )
            source_digest = _validate_export_source_digest(state, src_path)
            raw_rows = list(
                csv.reader(
                    io.StringIO(src_path.read_bytes().decode("utf-8-sig")),
                    delimiter=delim,
                )
            )
            offset = 0 if no_header else 1
            if state.get("input_sha256") is None:
                _validate_legacy_tabular_rows(
                    state,
                    segments,
                    lambda segment: (
                        raw_rows[
                            offset
                            + int(
                                segment.get(
                                    "row_index", segment.get("id", 0)
                                )
                            )
                        ]
                        if 0
                        <= offset
                        + int(
                            segment.get("row_index", segment.get("id", 0))
                        )
                        < len(raw_rows)
                        else None
                    ),
                )
            for seg in segments:
                row_idx = offset + int(
                    seg.get("row_index", seg.get("id", 0))
                )
                if row_idx < 0 or row_idx >= len(raw_rows):
                    raise ValueError(
                        f"source row for segment {seg['id']} is missing"
                    )
                row = raw_rows[row_idx]
                kind = export_kind(seg)
                corrected = result_entries[seg["id"]].get("corrected")
                if (
                    kind != "已保护"
                    and corrected is not None
                    and ti < len(row)
                ):
                    row[ti] = corrected
                counts[kind] += 1
            with tempfile.NamedTemporaryFile(
                mode="w",
                newline="",
                encoding=enc,
                dir=out_path.parent,
                prefix=f".{out_path.name}.",
                suffix=".tmp",
                delete=False,
            ) as handle:
                staged = Path(handle.name)
                csv.writer(handle, delimiter=delim).writerows(raw_rows)
            if file_sha256(src_path) != source_digest:
                raise ValueError(
                    f"source input changed during export: {src_path}"
                )
            revalidate_inputs()
            publish_replacement_transaction([(staged, out_path)])
        except (OSError, ValueError) as exc:
            raise SystemExit(f"[export] {exc}") from exc
        finally:
            if staged is not None:
                staged.unlink(missing_ok=True)
        print_summary(out_path)
        return

    out_path = state_path.parent / (_job_label(state_path) + "_corrected.xlsx")
    staged = None
    workbook = None
    try:
        _validate_export_paths(state_path, state, out_path, errors_path)
        source_digest = _validate_export_source_digest(state, src_path)
        workbook = openpyxl.load_workbook(str(src_path))
        sheet_name = state.get("sheet_name")
        worksheet = (
            workbook[sheet_name]
            if sheet_name in workbook.sheetnames
            else workbook.active
        )
        start_row = 1 if no_header else 2
        if state.get("input_sha256") is None:
            _validate_legacy_tabular_rows(
                state,
                segments,
                lambda segment: (
                    [
                        worksheet.cell(
                            row=start_row
                            + int(
                                segment.get(
                                    "row_index", segment.get("id", 0)
                                )
                            ),
                            column=column,
                        ).value
                        for column in range(1, worksheet.max_column + 1)
                    ]
                    if start_row
                    <= start_row
                    + int(
                        segment.get("row_index", segment.get("id", 0))
                    )
                    <= worksheet.max_row
                    else None
                ),
            )
        for seg in segments:
            row_num = start_row + int(
                seg.get("row_index", seg.get("id", 0))
            )
            if row_num < start_row or row_num > worksheet.max_row:
                raise ValueError(
                    f"source row for segment {seg['id']} is missing"
                )
            kind = export_kind(seg)
            corrected = result_entries[seg["id"]].get("corrected")
            if kind != "已保护" and corrected is not None:
                _set_excel_text(
                    worksheet.cell(row=row_num, column=ti + 1), corrected
                )
            counts[kind] += 1
        with tempfile.NamedTemporaryFile(
            dir=out_path.parent,
            prefix=f".{out_path.stem}.",
            suffix=out_path.suffix,
            delete=False,
        ) as handle:
            staged = Path(handle.name)
        workbook.save(str(staged))
        workbook.close()
        workbook = None
        if file_sha256(src_path) != source_digest:
            raise ValueError(f"source input changed during export: {src_path}")
        revalidate_inputs()
        publish_replacement_transaction([(staged, out_path)])
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[export] {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()
        if staged is not None:
            staged.unlink(missing_ok=True)
    print_summary(out_path)


def cmd_export(args):
    from lqe_chunk import verification_generation_lease
    from lqe_split_contract import generation_lock, state_fingerprint

    state_path = Path(args.state)
    errors_path = Path(args.errors) if getattr(args, "errors", None) else None
    try:
        if errors_path is not None:
            with verification_generation_lease(
                state_path,
                exclusive=False,
            ) as (state, segments, manifest, revalidate):
                _cmd_export_locked(
                    args,
                    state_path,
                    state,
                    segments,
                    manifest,
                    revalidate,
                )
            return

        chunks_dir = state_path.parent / "chunks"
        with generation_lock(chunks_dir, exclusive=False):
            state = read_json(state_path)
            expected_state_fingerprint = state_fingerprint(state)

            def revalidate_state():
                if state_fingerprint(read_json(state_path)) != expected_state_fingerprint:
                    raise ValueError("state changed during export")

            _cmd_export_locked(
                args,
                state_path,
                state,
                deepcopy(state["segments"]),
                None,
                revalidate_state,
            )
    except (OSError, ValueError) as exc:
        raise SystemExit(f"[export] {exc}") from exc


# ── ingest-corpus (stub) ──────────────────────────────────────────────────────

def cmd_ingest_corpus(args):
    # TODO: 接口格式待确认（JSON 直传 vs 文件上传）
    print("[lqe_io] ingest-corpus: AIPE RAG ingest interface TBD, skipping.")


# ── main ──────────────────────────────────────────────────────────────────────

def _add_scoring_policy_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--threshold", type=float, default=None)
    parser.add_argument(
        "--scorecard-profile",
        default=None,
        dest="scorecard_profile",
        help="评分卡 profile；省略时继承 state.scoring_policy",
    )
    parser.add_argument(
        "--severity-scale",
        choices=["lisa", "mqm"],
        default=None,
        dest="severity_scale",
        help="严重度乘数档；省略时继承 state.scoring_policy",
    )
    critical = parser.add_mutually_exclusive_group()
    critical.add_argument(
        "--critical-gate", action="store_true", dest="critical_gate"
    )
    critical.add_argument(
        "--no-critical-gate", action="store_false", dest="critical_gate"
    )
    repeat = parser.add_mutually_exclusive_group()
    repeat.add_argument(
        "--repeat-dedup", action="store_true", dest="repeat_dedup"
    )
    repeat.add_argument(
        "--no-repeat-dedup", action="store_false", dest="repeat_dedup"
    )
    parser.set_defaults(critical_gate=None, repeat_dedup=None)


def main():
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest="cmd", required=True)

    r = sub.add_parser("read")
    r.add_argument("--input", required=True)
    r.add_argument(
        "--input-format",
        choices=["auto", "tabular", "sdlxliff"],
        default="auto",
        dest="input_format",
    )
    r.add_argument(
        "--protect-exact-tm",
        action="store_true",
        dest="protect_exact_tm",
        help="保护同时满足 origin=TM、100%% 和 SourceAndTarget 的 SDLXLIFF 段",
    )
    r.add_argument("--project", default=None, help="项目档案：projects/<名>/profile.json 或目录/文件路径；提供 SG/术语/词数基准/checks/confirmed_rules 默认值，显式参数优先")
    r.add_argument("--source-col", default=None, dest="source_col", help="列名或列索引（0-based，配合 --no-header）；表格输入必填")
    r.add_argument("--target-col", default=None, dest="target_col", help="列名或列索引（0-based，配合 --no-header）；表格输入必填")
    r.add_argument("--no-header", action="store_true", dest="no_header", help="文件无表头行，source-col/target-col 为整数索引")
    r.add_argument("--group-col", default=None, dest="group_col", help="成组文本（对联/题目）的组标识列名或索引；同组段落 Step 2 合并评估")
    terminology = r.add_mutually_exclusive_group()
    terminology.add_argument("--terminology", default=None, help="术语表文件路径（.csv/.tsv/.json/.xlsx）")
    terminology.add_argument("--no-terminology", action="store_true", dest="no_terminology",
                             help="禁用术语、专名和术语审计检查")
    r.add_argument("--style-guide", default=None, dest="style_guide", help="风格指南文件路径（.txt/.md/.docx/.xlsx）")
    r.add_argument("--target-lang", default=None, dest="target_lang",
                   help="目标语言代码（en/th/zh 等）；挂载 target_languages/<code>/ 目标语言属性。项目 profile 必须显式写 target_lang")
    r.add_argument("--source-lang", default=None, dest="source_lang",
                   help="源语言代码（zh/en 等）；项目 profile 必须显式写 source_lang，显式参数优先")
    r.add_argument("--wordcount-basis", default=None, choices=["target-words", "source-chars"],
                   dest="wordcount_basis",
                   help="词数基准：target-words=译文空格分词（EN 等）；source-chars=源文 CJK 字符数+拉丁词数（泰语等无空格译文用）")
    r.add_argument("--out", default="lqe_state.json")

    af = sub.add_parser("apply-fixes")
    af.add_argument("--state",     required=True)
    af.add_argument("--errors",    required=True)
    af.add_argument("--score",     default=None, help="本轮分数（来自 lqe_calc.py 输出）")
    _add_scoring_policy_args(af)
    af.add_argument("--protected-ids", default=None, help="逗号分隔的 TM 100%% match segment ids")
    af.add_argument("--protected-file", default=None, help="TM 100%% match protected ids JSON 文件")

    ps = sub.add_parser("protect-segments")
    ps.add_argument("--state", required=True)
    ps.add_argument("--protected-ids", default=None, help="逗号分隔的已确认段 id")
    ps.add_argument("--protected-file", default=None, help="已确认段 id JSON 文件")
    ps.add_argument("--reason", default="TM_100_MATCH")
    ps.add_argument("--out", default=None, help="输出 protected ids JSON，默认 {job}/tm_protected.json")

    br = sub.add_parser("build-results")
    br.add_argument("--state", required=True)
    br.add_argument("--checks", required=True)
    br.add_argument("--out", required=True)

    w = sub.add_parser("write")
    w.add_argument("--state",     required=True)
    w.add_argument("--errors",    required=True)
    w.add_argument("--score",     required=True)
    _add_scoring_policy_args(w)

    pc = sub.add_parser("pre-check")
    pc.add_argument("--state", required=True)
    pc.add_argument("--out", default=None, help="输出路径（默认 {job_dir}/errors_precheck.json）")

    lt = sub.add_parser("lookup-terms")
    lt.add_argument("--state", required=True)
    lt.add_argument("--ids", default=None, help="逗号分隔的 seg id，不传则扫描全部段落")

    ex = sub.add_parser("export")
    ex.add_argument("--state", required=True)
    ex.add_argument("--errors", default=None, help="可选 errors.json：state.corrected 为空时用其 corrected 填充建议修正（单轮 FAIL 导出用）")

    ic = sub.add_parser("ingest-corpus")
    ic.add_argument("--state",    required=True)
    ic.add_argument("--aipe-url", required=True, dest="aipe_url")

    args = p.parse_args()
    {
        "read":           cmd_read,
        "pre-check":      cmd_pre_check,
        "protect-segments": cmd_protect_segments,
        "build-results":  cmd_build_results,
        "apply-fixes":    cmd_apply_fixes,
        "write":          cmd_write,
        "export":         cmd_export,
        "lookup-terms":   cmd_lookup_terms,
        "ingest-corpus":  cmd_ingest_corpus,
    }[args.cmd](args)


if __name__ == "__main__":
    main()
