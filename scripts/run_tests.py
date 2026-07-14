"""Self-contained regression suite for the LQE skill.

Run:  python scripts/run_tests.py
Covers corrected ownership, SDLXLIFF, and no-terminology regression suites;
all 23 builtin pre-checks; project profiles; counting, dedup, and wordcount
guards; and a smoke test for lqe_batch.
Fixtures are built in a temp dir; nothing is written into the repo.
"""
import json
import re
import shutil
import subprocess
import sys
import tempfile
import csv
from pathlib import Path

import openpyxl

SCRIPTS = Path(__file__).resolve().parent
TMP = Path(tempfile.mkdtemp(prefix="lqe_tests_"))
PASS, FAIL = [], []


def run(script, *argv, cwd=None):
    return subprocess.run([sys.executable, str(SCRIPTS / script), *argv],
                          capture_output=True, text=True, cwd=cwd or TMP)


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    if not cond:
        print(f"  ✗ {name}  {detail}")


def make_xlsx(path, rows, headers=("原文", "译文")):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    wb.save(path)


def load_errs(path):
    return {r["id"]: [e["comment"] for e in r["issues"]] for r in json.loads(Path(path).read_text(encoding="utf-8"))}


def load_full_errs(path):
    return {r["id"]: r["issues"] for r in json.loads(Path(path).read_text(encoding="utf-8"))}


def has(res, i, kw):
    return any(kw in c for c in res.get(i, []))


def has_err(res, i, category=None, severity=None, kw=None):
    for e in res.get(i, []):
        if category and e.get("category") != category:
            continue
        if severity and e.get("severity") != severity:
            continue
        if kw and kw not in e.get("comment", ""):
            continue
        return True
    return False


# ── T1: EN main fixture — N5-N9 / #3 / #7 / #10 / R5 ─────────────────────────
def t1():
    rows = [
        ('你好。', 'Hello'),                          # 0 N5 src-terminal missing
        ('你好', 'Hello.'),                           # 1 N5 tgt adds .
        ('开始！', 'Start!'),                         # 2 N5 ok
        ('你有三次机会。', 'You have chances.'),       # 3 N6 missing
        ('你有三次机会。', 'You have three chances.'), # 4 N6 ok word
        ('一起出发吧。', 'Let us set off together.'),  # 5 N6 no classifier
        ('第五章开启。', 'Chapter 5 unlocked.'),       # 6 N6 ordinal ok
        ('获得两个金币。', 'Got 2 coins.'),            # 7 N6 两 ok
        ('那只猫。', 'The the cat.'),                  # 8 N7 repeat
        ('我受够了。', 'I had had enough.'),           # 9 N7 whitelist
        ('苹果酱。', 'AppLe sauce.'),                  # 10 N8 mixed
        ('我的手机。', 'Check my iPhone.'),            # 11 N8 exempt
        ('对战模式。', 'PvP mode.'),                   # 12 N8 exempt
        ('（测试）完成。', '(test done.'),             # 13 N9 unbalanced
        ('没事。', "It's fine."),                      # 14 N9 apostrophe ok
        ('恢复生命值。', 'Restore hp.'),               # 15 #7 acronym case
        ('恢复生命值。', 'Restore HP.'),               # 16 #7 ok
        ('<b>你</b>好。', 'Hello there.'),             # 17 #3 angle tag lost
        ('等等…', 'Wait…'),                           # 18 #10 majority
        ('还有…', 'And…'),                            # 19 #10 majority
        ('走吧…', 'Go...'),                           # 20 #10 minority
        ('他说：「好」', 'He said 『fine』'),           # 21 R5 corner quotes
    ]
    make_xlsx(TMP / "en_main.xlsx", rows)
    (TMP / "tb.json").write_text(json.dumps(
        [{"source": "生命值", "target": "HP", "status": "Approved"}], ensure_ascii=False), encoding="utf-8")
    r = run("lqe_io.py", "read", "--input", str(TMP / "en_main.xlsx"),
            "--source-col", "原文", "--target-col", "译文", "--target-lang", "en",
            "--terminology", str(TMP / "tb.json"), "--out", str(TMP / "j1/state.json"))
    check("T1 read rc", r.returncode == 0, r.stderr[-200:])
    r = run("lqe_io.py", "pre-check", "--state", str(TMP / "j1/state.json"), "--out", str(TMP / "j1/pc.json"))
    check("T1 pre-check rc", r.returncode == 0, r.stderr[-200:])
    res = load_errs(TMP / "j1/pc.json")
    for name, cond in [
        ("N5 src-terminal", has(res, 0, "target does not")),
        ("N5 tgt adds", has(res, 1, "adds terminal")),
        ("N5 clean", not has(res, 2, "terminal")),
        ("N6 missing", has(res, 3, "三次")),
        ("N6 word ok", not has(res, 4, "Chinese numeral")),
        ("N6 no classifier", not has(res, 5, "Chinese numeral")),
        ("N6 ordinal ok", not has(res, 6, "Chinese numeral")),
        ("N6 两 ok", not has(res, 7, "Chinese numeral")),
        ("N7 repeat", has(res, 8, "Repeated word")),
        ("N7 whitelist", not has(res, 9, "Repeated word")),
        ("N8 mixed", has(res, 10, "Mixed case")),
        ("N8 iPhone", not has(res, 11, "Mixed case")),
        ("N8 PvP", not has(res, 12, "Mixed case")),
        ("N9 paren", has(res, 13, "Unbalanced")),
        ("N9 apostrophe", not has(res, 14, "Unbalanced") and not has(res, 14, "quote")),
        ("#7 hp", has(res, 15, "Acronym case")),
        ("#7 HP ok", not has(res, 16, "Acronym")),
        ("#3 tag", has(res, 17, "angle-bracket tag")),
        ("#10 minority", has(res, 20, "Ellipsis style")),
        ("#10 majority clean", not has(res, 18, "Ellipsis") and not has(res, 19, "Ellipsis")),
        ("R5 corner quotes", has(res, 21, "『")),
        ("N2 catches 3/4 divergence", has(res, 4, "Same source translated differently")),
    ]:
        check(f"T1 {name}", cond, str(res.get(0, ''))[:80])


# ── T2: N1 pinyin / N2 consistency / group-col ───────────────────────────────
def t2():
    rows = [
        ('前往张家口。', 'Go to Zhangjiakou.', ''),
        ('他叫青河。', 'His name is Qinghe.', ''),
        ('平安街到了。', "Ping'an Street ahead.", ''),
        ('坐出租车。', 'Take a Taxi today.', ''),
        ('改变战局。', 'Change the battle.', ''),
        ('点击开始', 'Click Start', ''),
        ('点击开始', 'Tap Begin', ''),
        ('这是一段超过二十个字符的长句子用于测试异源同译规则。', 'This long sentence is reused verbatim.', ''),
        ('另一段完全不同但同样超过二十字符的源文本内容在此。', 'This long sentence is reused verbatim.', ''),
        ('确定', 'OK', ''),
        ('好的', 'OK', ''),
        ('上联内容', 'First line', 'G1'),
    ]
    make_xlsx(TMP / "n1n2.xlsx", rows, headers=("原文", "译文", "组"))
    run("lqe_io.py", "read", "--input", str(TMP / "n1n2.xlsx"), "--source-col", "原文",
        "--target-col", "译文", "--group-col", "组", "--target-lang", "en",
        "--out", str(TMP / "j2/state.json"))
    run("lqe_io.py", "pre-check", "--state", str(TMP / "j2/state.json"), "--out", str(TMP / "j2/pc.json"))
    res = load_errs(TMP / "j2/pc.json")
    for name, cond in [
        ("N1 zh", has(res, 0, "pinyin residue: 'Zhangjiakou'")),
        ("N1 q", has(res, 1, "pinyin residue: 'Qinghe'")),
        ("N1 apostrophe", has(res, 2, "Ping'an")),
        ("N1 whitelist", not has(res, 3, "pinyin")),
        ("N1 ch-not-strong", not has(res, 4, "pinyin")),
        ("N2 base clean", not has(res, 5, "Same source")),
        ("N2 divergent", has(res, 6, "Same source translated differently")),
        ("N2 conv base clean", not has(res, 7, "reused")),
        ("N2 conv flagged", has(res, 8, "reused for different sources")),
        ("N2 short exempt", not has(res, 10, "reused")),
    ]:
        check(f"T2 {name}", cond)
    state = json.loads((TMP / "j2/state.json").read_text(encoding="utf-8"))
    check("T2 group stored", state["segments"][11].get("group") == "G1")


# ── T3: thai attribute derivation silences latin/sentence checks ─────────────
def t3():
    rows = [('你好。', 'สวัสดี'), ('那只猫。', 'แมว the the'), ('苹果。', 'AppLe แอปเปิ้ล')]
    make_xlsx(TMP / "th.xlsx", rows)
    r = run("lqe_io.py", "read", "--input", str(TMP / "th.xlsx"), "--source-col", "原文",
            "--target-col", "译文", "--target-lang", "th", "--out", str(TMP / "j3/state.json"))
    state = json.loads((TMP / "j3/state.json").read_text(encoding="utf-8"))
    check("T3 basis from lang attrs", state["wordcount_basis"] == "source-chars")
    check("T3 lang notes copied", bool(state["lang_notes_path"]) and "ครับ" in Path(state["lang_notes_path"]).read_text(encoding="utf-8"))
    r = run("lqe_io.py", "pre-check", "--state", str(TMP / "j3/state.json"), "--out", str(TMP / "j3/pc.json"))
    check("T3 derivation printed", "terminal_punct" in r.stdout and "word_repeat" in r.stdout)
    cs = [c for v in load_errs(TMP / "j3/pc.json").values() for c in v]
    check("T3 N5/N7/N8 silent", not any(k in c for c in cs for k in ("terminal", "Repeated word", "Mixed case")))
    # guard: explicit target-words on no-delim language warns
    r = run("lqe_io.py", "read", "--input", str(TMP / "th.xlsx"), "--source-col", "原文",
            "--target-col", "译文", "--target-lang", "th", "--wordcount-basis", "target-words",
            "--out", str(TMP / "j3b/state.json"))
    check("T3 word_delim guard", "no word delimiter" in r.stderr)


# ── T4: method C (project profiles) + wwm N3 custom ──────────────────────────
def t4():
    make_xlsx(TMP / "tiny.xlsx", [('他说：「你好」', 'เขาพูดว่า สวัสดี')])
    for proj, lang, basis in [("nrc/zh-th", "th", "source-chars"), ("nrc/zh-en", "en", "target-words"),
                              ("wwm/zh-en", "en", "target-words")]:
        slug = proj.replace("/", "-")
        r = run("lqe_io.py", "read", "--input", str(TMP / "tiny.xlsx"), "--source-col", "原文",
                "--target-col", "译文", "--project", proj, "--out", str(TMP / f"j4-{slug}/state.json"))
        ok = r.returncode == 0
        check(f"T4 {proj} read", ok, r.stderr[-200:])
        if not ok:
            continue
        s = json.loads((TMP / f"j4-{slug}/state.json").read_text(encoding="utf-8"))
        check(f"T4 {proj} lang/basis", s["target_lang"] == lang and s["wordcount_basis"] == basis)
        check(f"T4 {proj} source lang", s.get("source_lang") == "zh")
        check(f"T4 {proj} checks+rules", bool(s["checks_path"]) and bool(s["confirmed_rules_path"]))
    # wwm N3 roman numeral custom
    sp = TMP / "j4-wwm-zh-en/state.json"
    if sp.exists():
        s = json.loads(sp.read_text(encoding="utf-8"))
        s["segments"] = [{"id": 0, "source": "第二章。", "target": "Chapter II begins.",
                          "corrected": None, "max_len": None, "iter": 0}]
        sp.write_text(json.dumps(s, ensure_ascii=False), encoding="utf-8")
        run("lqe_io.py", "pre-check", "--state", str(sp), "--out", str(TMP / "j4-n3.json"))
        check("T4 N3 roman numeral", has(load_errs(TMP / "j4-n3.json"), 0, "罗马数字"))
    else:
        check("T4 N3 roman numeral", False, "missing wwm/zh-en state")
    r = run("lqe_io.py", "read", "--input", str(TMP / "tiny.xlsx"), "--source-col", "原文",
            "--target-col", "译文", "--project", "wwm/en", "--out", str(TMP / "j4-old/state.json"))
    check("T4 old target-only path removed", r.returncode != 0 and "project profile not found" in r.stderr)


# ── T5: custom count_match ────────────────────────────────────────────────────
def t5():
    (TMP / "cm_checks.json").write_text(json.dumps({"builtin": {}, "custom": [
        {"id": "cm-probe", "type": "count_match", "pattern": "#P\\d+#",
         "category": "Markup", "severity": "Major", "comment": "tag #Pn# count"}]}), encoding="utf-8")
    state = {"wordcount": 10, "language_pair": "zh-en", "checks_path": str(TMP / "cm_checks.json"),
             "segments": [{"id": 0, "source": "按#P1#键和#P2#键。", "target": "Press #P1#.",
                           "corrected": None, "max_len": None, "iter": 0}]}
    (TMP / "cm_state.json").write_text(json.dumps(state, ensure_ascii=False), encoding="utf-8")
    run("lqe_io.py", "pre-check", "--state", str(TMP / "cm_state.json"), "--out", str(TMP / "cm_pc.json"))
    check("T5 count_match", has(load_errs(TMP / "cm_pc.json"), 0, "source=2, target=1"))


# ── T6: N4 repeat dedup in calc ───────────────────────────────────────────────
def t6():
    state = {"wordcount": 100, "segments": [
        {"id": i, "source": "重复句。", "target": "Dup sentence.", "corrected": None} for i in range(3)]}
    errors = [{"id": i, "errors": [{"category": "Mistranslation", "severity": "Major",
                                    "comment": "wrong"}], "corrected": None} for i in range(3)]
    (TMP / "n4_state.json").write_text(json.dumps(state), encoding="utf-8")
    (TMP / "n4_errors.json").write_text(json.dumps(errors), encoding="utf-8")
    r = run("lqe_calc.py", "--state", str(TMP / "n4_state.json"), "--errors", str(TMP / "n4_errors.json"))
    check("T6 score dedup", "SCORE=92.50" in r.stdout and "REPEATED=2" in r.stdout, r.stdout[:120])
    flags = [e.get("repeated", False) for x in json.loads((TMP / "n4_errors.json").read_text(encoding="utf-8"))
             for e in x["errors"]]
    check("T6 repeated written back", flags == [False, True, True], str(flags))
    r = run("lqe_calc.py", "--state", str(TMP / "n4_state.json"), "--errors", str(TMP / "n4_errors.json"),
            "--no-repeat-dedup")
    check("T6 dedup off", "REPEATED=0" in r.stdout and "SCORE=77.50" in r.stdout, r.stdout[:120])


# ── T7: lqe_batch plan/merge smoke ────────────────────────────────────────────
def t7():
    job = TMP / "j1"
    shutil.copy(job / "pc.json", job / "errors_precheck.json")
    r = run("lqe_batch.py", "plan", "--job", str(job), "--output-budget", "400")
    check("T7 plan rc", r.returncode == 0, r.stderr[-200:])
    check("T7 manifest", (job / "manifest.json").exists() and len(list((job / "batches").glob("batch_*.txt"))) >= 2)
    evals = job / "evals"
    evals.mkdir(exist_ok=True)
    state = json.loads((job / "state.json").read_text(encoding="utf-8"))
    entries = [{"id": segment["id"], "issues": []} for segment in state["segments"]]
    entries[0]["issues"] = [
        {
            "category": "Punctuation",
            "severity": "Minor",
            "comment": "Add required terminal punctuation.",
            "needs_confirmation": False,
            "edit": {
                "from": "Hello",
                "to": "Hello.",
                "start": 0,
                "end": 5,
                "evidence": None,
            },
        }
    ]
    (evals / "eval_00.json").write_text(json.dumps(entries), encoding="utf-8")
    r = run("lqe_batch.py", "merge", "--job", str(job))
    check("T7 merge complete", r.returncode == 0 and "complete" in r.stdout, (r.stdout + r.stderr)[-300:])
    merged = json.loads((job / "errors.json").read_text(encoding="utf-8"))
    check("T7 merge content", merged[0]["corrected"] == "Hello." and merged[1]["errors"] == [])


# ── T8: lqe_engine term_senses / group_terms ──────────────────────────────
def t8():
    sys.path.insert(0, str(SCRIPTS))
    from lqe_engine import term_senses, group_terms

    singleton = {"source": "马尔文", "target": "มาร์วิน", "status": "New"}
    check("T8 singleton term_senses",
          term_senses(singleton) == [{"target": "มาร์วิน", "status": "New",
                                      "confirmed": False, "protected": False}])

    multi = {"source": "里奥", "senses": [
        {"target": "ลีโอ", "category": "Creature Individual"},
        {"target": "ไลเอล", "category": "Creature Species"},
    ]}
    check("T8 multi term_senses flags", term_senses(multi) == [
        {"target": "ลีโอ", "category": "Creature Individual", "confirmed": False, "protected": False},
        {"target": "ไลเอล", "category": "Creature Species", "confirmed": False, "protected": False},
    ])

    grouped = group_terms([singleton, multi])
    check("T8 group_terms keys", set(grouped.keys()) == {"马尔文", "里奥"})
    check("T8 group_terms multi count", len(grouped["里奥"]) == 2)
    check("T8 group_terms singleton count", len(grouped["马尔文"]) == 1)


# ── T9: lqe_checks pre-check 多义术语命中 ─────────────────────────────────────
def t9():
    rows = [
        ('看到一只里奥。', 'Saw a ลีโอ.'),      # 0 命中 Individual 候选 -> 不报
        ('看到一只里奥。', 'Saw a ไลเอล.'),      # 1 命中 Species 候选 -> 不报
        ('看到一只里奥。', 'Saw a Rio.'),        # 2 两个候选都不匹配 -> 报错，列出两个候选
        ('马尔文来了。', 'มาร์วิน is here.'),    # 3 单义词条回归检查
    ]
    make_xlsx(TMP / "t9.xlsx", rows)
    (TMP / "t9_tb.json").write_text(json.dumps([
        {"source": "里奥", "senses": [
            {"target": "ลีโอ", "category": "Creature Individual"},
            {"target": "ไลเอล", "category": "Creature Species"},
        ]},
        {"source": "马尔文", "target": "มาร์วิน", "status": "New"},
    ], ensure_ascii=False), encoding="utf-8")
    r = run("lqe_io.py", "read", "--input", str(TMP / "t9.xlsx"),
            "--source-col", "原文", "--target-col", "译文", "--target-lang", "th",
            "--wordcount-basis", "source-chars",
            "--terminology", str(TMP / "t9_tb.json"), "--out", str(TMP / "j9/state.json"))
    check("T9 read rc", r.returncode == 0, r.stderr[-200:])
    r = run("lqe_io.py", "pre-check", "--state", str(TMP / "j9/state.json"), "--out", str(TMP / "j9/pc.json"))
    check("T9 pre-check rc", r.returncode == 0, r.stderr[-200:])
    res = load_errs(TMP / "j9/pc.json")
    full = load_full_errs(TMP / "j9/pc.json")
    check("T9 sense A no Terminology error", not has_err(full, 0, category="Terminology", kw="里奥"))
    check("T9 sense B no Terminology error", not has_err(full, 1, category="Terminology", kw="里奥"))
    check("T9 sense A review-only term hit", has_err(full, 0, category="Other", severity="Neutral", kw="TERM REVIEW"))
    check("T9 sense B review-only term hit", has_err(full, 1, category="Other", severity="Neutral", kw="TERM REVIEW"))
    check("T9 neither-sense reports both candidates", has(res, 2, "ลีโอ") and has(res, 2, "ไลเอล"))
    check("T9 singleton no Terminology error", not has_err(full, 3, category="Terminology", kw="马尔文"))
    check("T9 singleton review-only term hit", has_err(full, 3, category="Other", severity="Neutral", kw="TERM REVIEW"))


# ── T10: lqe_chunk split 多义 term_hits ───────────────────────────────────────
def t10():
    job = TMP / "j10"
    job.mkdir(parents=True, exist_ok=True)
    state = {"segments": [
        {"id": 0, "source": "看到一只里奥。", "target": "Saw a ลีโอ.",
         "content_type": "剧情", "text_type_context": "故事类文本"},
        {"id": 1, "source": "马尔文来了。", "target": "มาร์วิน is here."},
    ]}
    (job / "state.json").write_text(json.dumps(state, ensure_ascii=False), encoding="utf-8")
    (job / "errors_precheck.json").write_text(json.dumps(
        [{"id": 0, "issues": []}, {"id": 1, "issues": []}], ensure_ascii=False), encoding="utf-8")
    (job / "terms.json").write_text(json.dumps([
        {"source": "里奥", "senses": [
            {"target": "ลีโอ", "category": "Creature Individual"},
            {"target": "ไลเอล", "category": "Creature Species"},
        ]},
        {"source": "马尔文", "target": "มาร์วิน", "status": "New"},
    ], ensure_ascii=False), encoding="utf-8")
    outdir = job / "chunks"
    r = run("lqe_chunk.py", "split", "--state", str(job / "state.json"),
            "--errors", str(job / "errors_precheck.json"),
            "--terms", str(job / "terms.json"),
            "--outdir", str(outdir), "--size", "10")
    check("T10 split rc", r.returncode == 0, r.stderr[-300:])
    chunk = json.loads((outdir / "chunk_00.json").read_text(encoding="utf-8"))
    seg0 = next(s for s in chunk["segments"] if s["id"] == 0)
    seg1 = next(s for s in chunk["segments"] if s["id"] == 1)
    hits0 = [h for h in seg0["term_hits"] if h["source"] == "里奥"]
    check("T10 multi-sense flattened to 2 hits", len(hits0) == 2)
    check("T10 multi-sense categories present",
          {hit.get("category") for hit in hits0} == {"Creature Individual", "Creature Species"})
    check("T10 multi-sense explicit flags",
          all(hit["confirmed"] is False and hit["protected"] is False for hit in hits0))
    check("T10 content context preserved", seg0.get("content_type") == "剧情" and
          seg0.get("text_type_context") == "故事类文本")
    hits1 = [h for h in seg1["term_hits"] if h["source"] == "马尔文"]
    check("T10 singleton flattened to 1 hit",
          len(hits1) == 1 and hits1[0]["target"] == "มาร์วิน" and
          hits1[0]["confirmed"] is False and hits1[0]["protected"] is False)


# ── T11: lookup-terms 多义展示 ────────────────────────────────────────────────
def t11():
    rows = [('里奥出现了。', 'placeholder')]
    make_xlsx(TMP / "t11.xlsx", rows)
    (TMP / "t11_tb.json").write_text(json.dumps([
        {"source": "里奥", "senses": [
            {"target": "ลีโอ", "category": "Creature Individual"},
            {"target": "ไลเอล", "category": "Creature Species"},
        ]},
    ], ensure_ascii=False), encoding="utf-8")
    r = run("lqe_io.py", "read", "--input", str(TMP / "t11.xlsx"),
            "--source-col", "原文", "--target-col", "译文", "--target-lang", "th",
            "--wordcount-basis", "source-chars",
            "--terminology", str(TMP / "t11_tb.json"), "--out", str(TMP / "j11/state.json"))
    check("T11 read rc", r.returncode == 0, r.stderr[-200:])
    r = run("lqe_io.py", "lookup-terms", "--state", str(TMP / "j11/state.json"))
    check("T11 lookup rc", r.returncode == 0, r.stderr[-200:])
    check("T11 shows both candidates", "ลีโอ" in r.stdout and "ไลเอล" in r.stdout)


# ── T12: mastertb_to_terms 多义输出 + 去重键改 (source,target) ────────────────
def t12():
    job = TMP / "j12"
    job.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["术语类别 Category", "术语 ZHCN", "术语定义 Definition", "TH"])
    for r in [
        ("NPC", "张三", "张三定义", "ซาน"),
        ("Species", "里奥", "物种定义", "ไลเอล"),
        ("Individual", "里奥", "个体定义", "ลีโอ"),
        ("NPC", "李四", "", ""),
        ("NPC", "王五", "", ""),
        ("NPC", "张三", "张三定义", "ซาน"),  # 完全重复行，应合并不产生第二候选
    ]:
        ws.append(list(r))
    wb.save(job / "master.xlsx")
    (job / "backfill.json").write_text(json.dumps(
        [{"source": "李四", "target": "หลี่ซื่อ"}], ensure_ascii=False), encoding="utf-8")

    r = run("mastertb_to_terms.py", "--input", str(job / "master.xlsx"),
            "--target-col", "TH", "--backfill", str(job / "backfill.json"),
            "--out", str(job / "terms.json"))
    check("T12 rc", r.returncode == 0, r.stderr[-300:])

    terms = {t["source"]: t for t in json.loads((job / "terms.json").read_text(encoding="utf-8"))}
    check("T12 sources (王五 blank+无回填 -> 丢弃)", set(terms.keys()) == {"张三", "里奥", "李四"})
    check("T12 singleton shape 不带 category", terms["张三"] == {
        "source": "张三", "target": "ซาน", "confirmed": False, "protected": False})
    check("T12 回填出的单义 shape", terms["李四"] == {
        "source": "李四", "target": "หลี่ซื่อ", "confirmed": False, "protected": False})
    senses = terms["里奥"]["senses"]
    check("T12 多义候选数", len(senses) == 2)
    check("T12 多义候选 target", {s["target"] for s in senses} == {"ไลเอล", "ลีโอ"})
    check("T12 多义候选 category 带出", {s["category"] for s in senses} == {"Species", "Individual"})
    check("T12 多义候选显式 flags",
          all(s["confirmed"] is False and s["protected"] is False for s in senses))

    multisense = json.loads((job / "terms.multisense.json").read_text(encoding="utf-8"))
    check("T12 multisense.json 只列里奥", len(multisense) == 1 and multisense[0]["source"] == "里奥")


# ── T13: scorecard profile 解耦 legacy 行为不变 ───────────────────────────────
def t13():
    sys.path.insert(0, str(SCRIPTS))
    try:
        from lqe_engine import load_scorecard_profile
        profile = load_scorecard_profile("legacy")
    except Exception as exc:
        check("T13 legacy profile loads", False, repr(exc))
        return

    check("T13 legacy id", profile["id"] == "legacy")
    check("T13 legacy severity minor", profile["severity_points"]["Minor"] == 1)
    check("T13 legacy weight audience", profile["category_weights"]["Audience appropriateness"] == 1.5)
    check("T13 legacy forced length", profile["forced_severity"]["Length"] == "Major")

    state = {"wordcount": 100, "segments": [
        {"id": 0, "source": "甲。", "target": "A.", "corrected": None},
        {"id": 1, "source": "乙。", "target": "B.", "corrected": None},
    ]}
    errors = [
        {"id": 0, "errors": [{"category": "Mistranslation", "severity": "Major", "comment": "wrong"}], "corrected": None},
        {"id": 1, "errors": [{"category": "Punctuation", "severity": "Minor", "comment": "punct"}], "corrected": None},
    ]
    (TMP / "scorecard_state.json").write_text(json.dumps(state, ensure_ascii=False), encoding="utf-8")
    (TMP / "scorecard_errors_default.json").write_text(json.dumps(errors, ensure_ascii=False), encoding="utf-8")
    (TMP / "scorecard_errors_profile.json").write_text(json.dumps(errors, ensure_ascii=False), encoding="utf-8")

    default = run("lqe_calc.py", "--state", str(TMP / "scorecard_state.json"),
                  "--errors", str(TMP / "scorecard_errors_default.json"), "--json")
    profiled = run("lqe_calc.py", "--state", str(TMP / "scorecard_state.json"),
                   "--errors", str(TMP / "scorecard_errors_profile.json"),
                   "--scorecard-profile", "legacy", "--json")
    check("T13 calc profile arg rc", profiled.returncode == 0, profiled.stderr[-200:])
    check("T13 calc legacy parity", default.stdout == profiled.stdout, f"default={default.stdout} profiled={profiled.stdout}")


# ── T14: LQE 2026 scorecard profile ──────────────────────────────────────────
def t14():
    sys.path.insert(0, str(SCRIPTS))
    try:
        from lqe_engine import load_scorecard_profile
        profile = load_scorecard_profile("lqe_2026")
    except Exception as exc:
        check("T14 lqe_2026 profile loads", False, repr(exc))
        return

    check("T14 profile id", profile["id"] == "lqe_2026")
    check("T14 minor severity point", profile["severity_points"]["Minor"] == 2)
    check("T14 unidiomatic weight", profile["category_weights"]["Unidiomatic"] == 3.5)
    check("T14 grammar weight", profile["category_weights"]["Grammar"] == 0.5)
    check("T14 culture parent", profile["category_parent"]["Culture specific reference"] == "Verity")
    check("T14 audience alias", profile["category_aliases"]["Audience appropriateness"] == "Unidiomatic")
    check("T14 template path", profile["report_template"]["path"] == "template.xlsx")
    check("T14 template file exists",
          (SCRIPTS.parent / "scorecard_profiles/lqe_2026" / profile["report_template"]["path"]).exists())

    state = {"wordcount": 1000, "segments": [
        {"id": 0, "source": "甲。", "target": "A.", "corrected": None},
        {"id": 1, "source": "乙。", "target": "B.", "corrected": None},
        {"id": 2, "source": "丙。", "target": "C.", "corrected": None},
    ]}
    errors = [
        {"id": 0, "errors": [{"category": "Punctuation", "severity": "Minor", "comment": "punct"}], "corrected": None},
        {"id": 1, "errors": [{"category": "Unidiomatic", "severity": "Major", "comment": "awkward"}], "corrected": None},
        {"id": 2, "errors": [{"category": "Audience appropriateness", "severity": "Minor", "comment": "tone"}], "corrected": None},
    ]
    (TMP / "lqe2026_state.json").write_text(json.dumps(state, ensure_ascii=False), encoding="utf-8")
    (TMP / "lqe2026_errors.json").write_text(json.dumps(errors, ensure_ascii=False), encoding="utf-8")
    r = run("lqe_calc.py", "--state", str(TMP / "lqe2026_state.json"),
            "--errors", str(TMP / "lqe2026_errors.json"),
            "--scorecard-profile", "lqe_2026", "--json")
    check("T14 calc rc", r.returncode == 0, r.stderr[-200:])
    if r.returncode == 0:
        result = json.loads(r.stdout)
        check("T14 score uses LQE2026 weights", result["score"] == 97.45 and result["status"] == "FAIL", r.stdout)
        check("T14 total errors counted", result["errors"] == 3, r.stdout)


def t15():
    p = TMP / "aipe_export.csv"
    p.write_text(
        "seq,source,translation,status,content_type,rag_references\n"
        "1,对话类文本,Dialogue text,success,未知,[]\n"
        "2,师父来了。,Master has arrived.,success,话术,[]\n"
        "3,游戏内侧页文本,Inner page text,success,UI,[]\n"
        "4,打开背包,Open Inventory,success,UI,[]\n",
        encoding="utf-8-sig"
    )
    r = run("lqe_io.py", "read", "--input", str(p), "--source-col", "source",
            "--target-col", "translation", "--target-lang", "en",
            "--out", str(TMP / "j15/state.json"))
    check("T15 csv read rc", r.returncode == 0, r.stderr[-300:])
    if r.returncode != 0:
        return
    state = json.loads((TMP / "j15/state.json").read_text(encoding="utf-8"))
    check("T15 csv segment count", len(state["segments"]) == 2)
    check("T15 csv source/target", state["segments"][0]["source"] == "师父来了。" and
          state["segments"][0]["target"] == "Master has arrived.")
    check("T15 csv content_type", state["segments"][0]["content_type"] == "话术")
    check("T15 csv text type markers skipped", len(state.get("text_type_markers", [])) == 2 and
          state["segments"][0].get("text_type_context") == "对话类文本" and
          state["segments"][1].get("text_type_context") == "游戏内侧页文本")
    check("T15 csv rows_raw preserves columns", state["headers"][-1] == "rag_references" and
          state["rows_raw"][0][-1] == "[]")
    state["segments"][0]["corrected"] = "Corrected master arrived."
    state["segments"][1]["corrected"] = "Corrected inventory."
    (TMP / "j15/state.json").write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    r = run("lqe_io.py", "export", "--state", str(TMP / "j15/state.json"))
    check("T15 csv export rc", r.returncode == 0, r.stderr[-300:])
    out = TMP / "j15/j15_corrected.csv"
    check("T15 csv export path", out.exists(), r.stdout[-300:])
    if out.exists():
        rows = list(csv.reader(out.open(encoding="utf-8-sig")))
        check("T15 csv marker row preserved", rows[1][1] == "对话类文本" and rows[1][2] == "Dialogue text")
        check("T15 csv corrected rows aligned", rows[2][2] == "Corrected master arrived." and
              rows[4][2] == "Corrected inventory.")


def t16():
    r = run("lqe_io.py", "--help")
    check("T16 help rc", r.returncode == 0, r.stderr[-200:])
    check("T16 from-aipe hidden", "from-aipe" not in r.stdout)
    r = run("lqe_io.py", "from-aipe", "--help")
    check("T16 from-aipe removed", r.returncode != 0 and "invalid choice" in r.stderr, r.stderr[-300:])


def t17():
    make_xlsx(TMP / "lang_pair.xlsx", [("Open the chest.", "打开宝箱。")])

    explicit_dir = TMP / "project-explicit-lang"
    explicit_dir.mkdir(parents=True, exist_ok=True)
    (explicit_dir / "profile.json").write_text(json.dumps({
        "name": "Explicit language fields",
        "language_pair": "zh-en",
        "source_lang": "en",
        "target_lang": "zh",
        "wordcount_basis": "source-chars",
    }, ensure_ascii=False), encoding="utf-8")
    r = run("lqe_io.py", "read", "--input", str(TMP / "lang_pair.xlsx"),
            "--source-col", "原文", "--target-col", "译文",
            "--project", str(explicit_dir / "profile.json"),
            "--out", str(TMP / "j17-explicit/state.json"))
    check("T17 explicit profile read", r.returncode == 0, r.stderr[-300:])
    if r.returncode == 0:
        s = json.loads((TMP / "j17-explicit/state.json").read_text(encoding="utf-8"))
        check("T17 explicit source/target", s.get("source_lang") == "en" and s.get("target_lang") == "zh")
        check("T17 explicit keeps pair", s.get("language_pair") == "zh-en")

    pair_dir = TMP / "project-pair-only"
    pair_dir.mkdir(parents=True, exist_ok=True)
    (pair_dir / "profile.json").write_text(json.dumps({
        "name": "Pair-only language fields",
        "language_pair": "en-th",
        "wordcount_basis": "source-chars",
    }, ensure_ascii=False), encoding="utf-8")
    r = run("lqe_io.py", "read", "--input", str(TMP / "lang_pair.xlsx"),
            "--source-col", "原文", "--target-col", "译文",
            "--project", str(pair_dir / "profile.json"),
            "--out", str(TMP / "j17-pair/state.json"))
    check("T17 pair-only profile rejected", r.returncode != 0 and "source_lang" in r.stderr and "target_lang" in r.stderr,
          r.stderr[-300:])

    r = run("lqe_io.py", "read", "--input", str(TMP / "lang_pair.xlsx"),
            "--source-col", "原文", "--target-col", "译文",
            "--project", str(explicit_dir / "profile.json"),
            "--source-lang", "fr", "--target-lang", "en",
            "--out", str(TMP / "j17-cli/state.json"))
    check("T17 CLI source override read", r.returncode == 0, r.stderr[-300:])
    if r.returncode == 0:
        s = json.loads((TMP / "j17-cli/state.json").read_text(encoding="utf-8"))
        check("T17 CLI source/target override", s.get("source_lang") == "fr" and s.get("target_lang") == "en")


def t18():
    p = TMP / "marker_export.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["source", "translation", "content_type"])
    ws.append(["对话类文本", "Dialogue text", "未知"])
    ws.append(["师父来了。", "Master has arrived.", "话术"])
    ws.append(["游戏内侧页文本", "Inner page text", "UI"])
    ws.append(["打开背包", "Open Inventory", "UI"])
    wb.save(p)

    r = run("lqe_io.py", "read", "--input", str(p), "--source-col", "source",
            "--target-col", "translation", "--target-lang", "en",
            "--out", str(TMP / "j18/state.json"))
    check("T18 xlsx marker read rc", r.returncode == 0, r.stderr[-300:])
    if r.returncode != 0:
        return
    state = json.loads((TMP / "j18/state.json").read_text(encoding="utf-8"))
    state["segments"][0]["corrected"] = "Corrected master arrived."
    state["segments"][1]["corrected"] = "Corrected inventory."
    (TMP / "j18/state.json").write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    r = run("lqe_io.py", "export", "--state", str(TMP / "j18/state.json"))
    check("T18 xlsx marker export rc", r.returncode == 0, r.stderr[-300:])
    out = TMP / "j18/j18_corrected.xlsx"
    check("T18 xlsx marker export path", out.exists(), r.stdout[-300:])
    if out.exists():
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        rows = [tuple(c.value for c in ws[i]) for i in range(2, 6)]
        check("T18 xlsx marker row unchanged", rows[0][1] == "Dialogue text" and rows[2][1] == "Inner page text")
        check("T18 xlsx corrected rows aligned", rows[1][1] == "Corrected master arrived." and
              rows[3][1] == "Corrected inventory.")


def t19():
    make_xlsx(TMP / "enzh.xlsx", [("Open the chest.", "打开宝箱。")], headers=("source", "translation"))
    prof = TMP / "enzh-profile"
    prof.mkdir(parents=True, exist_ok=True)
    (prof / "profile.json").write_text(json.dumps({
        "name": "tmp/en-zh",
        "language_pair": "en-zh",
        "source_lang": "en",
        "target_lang": "zh",
        "wordcount_basis": "source-chars",
    }, ensure_ascii=False), encoding="utf-8")
    r = run("lqe_io.py", "read", "--input", str(TMP / "enzh.xlsx"),
            "--source-col", "source", "--target-col", "translation",
            "--project", str(prof), "--out", str(TMP / "j19/state.json"))
    check("T19 en-zh read rc", r.returncode == 0, r.stderr[-300:])
    if r.returncode != 0:
        return
    r = run("lqe_io.py", "pre-check", "--state", str(TMP / "j19/state.json"), "--out", str(TMP / "j19/pc.json"))
    check("T19 en-zh pre-check rc", r.returncode == 0, r.stderr[-300:])
    full = load_full_errs(TMP / "j19/pc.json")
    comments = [e["comment"] for es in full.values() for e in es]
    check("T19 en-zh no CJK/fullwidth/terminal false positives",
          not any(("Chinese characters" in c or "Full-width punctuation" in c or "terminal punctuation" in c) for c in comments),
          str(comments))
    state = json.loads((TMP / "j19/state.json").read_text(encoding="utf-8"))
    check("T19 zh lang notes copied", bool(state.get("lang_notes_path")) and Path(state["lang_notes_path"]).exists())


def t20():
    make_xlsx(TMP / "report_lang.xlsx", [("你好", "สวัสดี")])
    r = run("lqe_io.py", "read", "--input", str(TMP / "report_lang.xlsx"),
            "--source-col", "原文", "--target-col", "译文", "--target-lang", "th",
            "--source-lang", "zh", "--out", str(TMP / "j20/state.json"))
    check("T20 report read rc", r.returncode == 0, r.stderr[-300:])
    if r.returncode != 0:
        return
    (TMP / "j20/errors.json").write_text(json.dumps([{"id": 0, "errors": [], "corrected": None}],
                                                     ensure_ascii=False), encoding="utf-8")
    r = run("lqe_io.py", "write", "--state", str(TMP / "j20/state.json"),
            "--errors", str(TMP / "j20/errors.json"), "--score", "100", "--threshold", "98")
    check("T20 report write rc", r.returncode == 0, r.stderr[-300:])
    out = TMP / "j20/j20_lqe.xlsx"
    check("T20 report path", out.exists(), r.stdout[-300:])
    if out.exists():
        wb = openpyxl.load_workbook(out)
        vals = [cell.value for ws in wb.worksheets for row in ws.iter_rows() for cell in row if cell.value is not None]
        check("T20 report language dynamic", "Source language" in vals and "zh" in vals and
              "Target language" in vals and "th" in vals, str(vals[:30]))


def t21():
    script = (SCRIPTS / "finalize_job.sh").read_text(encoding="utf-8")
    check("T21 finalize reads threshold from state",
          "state.json" in script and "threshold" in script and "--threshold 98" not in script)


def t22():
    root = SCRIPTS.parent
    missing = []
    for prof_path in sorted(root.glob("projects/*/*/profile.json")):
        prof = json.loads(prof_path.read_text(encoding="utf-8"))
        base = prof_path.parent
        for key in ("style_guide", "terminology", "checks", "confirmed_rules"):
            val = prof.get(key)
            if val and not (base / val).exists():
                missing.append(f"{prof_path.relative_to(root)}:{key}:{val}")
        tm = prof.get("tm") or {}
        for lib in tm.get("libraries") or []:
            if not (base / lib).exists():
                missing.append(f"{prof_path.relative_to(root)}:tm.library:{lib}")
        if tm.get("index") and not (base / tm["index"]).exists():
            missing.append(f"{prof_path.relative_to(root)}:tm.index:{tm['index']}")
    check("T22 project profile referenced files exist", not missing, "; ".join(missing[:5]))


def t23():
    gitignore_path = SCRIPTS.parent / ".gitignore"
    if not gitignore_path.exists():
        check("T23 project profile files not globally ignored", True)
        return
    gitignore = gitignore_path.read_text(encoding="utf-8")
    check("T23 project profile files not globally ignored",
          "projects/" not in {line.strip() for line in gitignore.splitlines()} and
          "!projects/*/*/profile.json" in gitignore)


def t24():
    root = SCRIPTS.parent
    check("T24 target language folder is explicit", (root / "target_languages").is_dir())
    check("T24 legacy languages folder removed", not (root / "languages").exists())
    scanned = [
        root / "SKILL.md",
        root / "README.md",
        root / "README_ZH.md",
        root / "projects/README.md",
        root / "docs/质量检查项清单.md",
        root / "scripts/lqe_engine.py",
        root / "scripts/lqe_io.py",
    ]
    stale = []
    old_paths = (
        re.compile(r"(?<!target_)languages/"),
        re.compile(r"`languages(?:`|/|<)"),
        re.compile(r"(?<!target_)languages<"),
        re.compile(r"/\s*[\"']languages[\"']"),
        re.compile(r"\bPath\(\s*[\"']languages[\"']\s*\)"),
    )
    for path in scanned:
        if path.exists() and any(
            pattern.search(path.read_text(encoding="utf-8"))
            for pattern in old_paths
        ):
            stale.append(str(path.relative_to(root)))
    check("T24 no stale languages path refs", not stale, "; ".join(stale))


def t25():
    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "unittest",
            "-v",
            "tests.test_correction_builder",
            "tests.test_corrected_ownership",
            "tests.test_plain_language",
            "tests.test_no_terminology_mode",
            "tests.test_sdlxliff_input",
            "tests.test_documented_contract",
            "tests.test_mastertb_module_contract",
        ],
        cwd=SCRIPTS.parent,
        capture_output=True,
        text=True,
    )
    output = (result.stdout + result.stderr).strip()
    check(
        "T25 corrected ownership + SDLXLIFF + no-terminology suites",
        result.returncode == 0,
        output[-2000:],
    )


if __name__ == "__main__":
    for t in (t1, t2, t3, t4, t5, t6, t7, t8, t9, t10, t11, t12, t13, t14, t15, t16, t17,
              t18, t19, t20, t21, t22, t23, t24, t25):
        t()
    rag = subprocess.run([sys.executable, str(SCRIPTS / "tm_index_test.py")], capture_output=True, text=True)
    check("TM suite (tm_index_test.py)", rag.returncode == 0,
          (rag.stdout or rag.stderr).strip().splitlines()[-1] if (rag.stdout or rag.stderr) else "")
    total = len(PASS) + len(FAIL)
    print(f"\n{len(PASS)}/{total} passed" + (f"  FAILED: {FAIL}" if FAIL else "  — all green"))
    shutil.rmtree(TMP, ignore_errors=True)
    sys.exit(1 if FAIL else 0)
