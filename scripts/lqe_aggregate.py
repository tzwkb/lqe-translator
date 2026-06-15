#!/usr/bin/env python3
"""Aggregate identical (source,target) segments to ONE verdict, and report the
groups whose independent AI verdicts disagreed → human ruling.

Chunked fan-out judges each row alone, so identical segments can get different
errors/corrected (usually because the TB is missing the term or renders it
inconsistently). This:
  1. collapses each (source,target) group to one representative verdict
     (highest penalty; ties → lowest id) and broadcasts it → consistency;
  2. writes a 待人工裁决 report (xlsx) listing the groups that had disagreed,
     with every competing version, so a human picks the right one and feeds it
     back to the glossary.
"""
import argparse
import json
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

SEV = {"Neutral": 0, "Minor": 1, "Major": 5, "Critical": 10}
LIGHT = {"Punctuation", "Spelling", "Locale convention", "Other"}  # weight 1.0, else 1.5


def penalty(errs):
    return sum((1.0 if e.get("category") in LIGHT else 1.5) * SEV.get(e.get("severity"), 0)
               for e in errs)


def esig(errs):
    return tuple(sorted((e.get("category", ""), e.get("severity", "")) for e in errs))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--state", required=True)
    ap.add_argument("--errors", required=True)
    ap.add_argument("--out", required=True)       # consistent errors.json
    ap.add_argument("--report", required=True)    # 待人工裁决 xlsx
    a = ap.parse_args()

    state = json.loads(Path(a.state).read_text(encoding="utf-8"))
    errs = json.loads(Path(a.errors).read_text(encoding="utf-8"))
    emap = {e["id"]: e for e in errs}
    groups = defaultdict(list)
    for s in state["segments"]:
        groups[(s.get("source", ""), s.get("target", ""))].append(s["id"])

    out_by_id = {}
    divergent = []
    for (src, tgt), ids in groups.items():
        rep = max(ids, key=lambda i: (penalty(emap.get(i, {}).get("errors", [])), -i))
        rep_e = emap.get(rep, {})
        rep_errs, rep_corr = rep_e.get("errors", []), rep_e.get("corrected")
        for i in ids:
            out_by_id[i] = {"id": i, "errors": [dict(e) for e in rep_errs], "corrected": rep_corr}
        sigs = {(esig(emap.get(i, {}).get("errors", [])), emap.get(i, {}).get("corrected") or None)
                for i in ids}
        if len(ids) > 1 and len(sigs) > 1:
            versions = defaultdict(list)
            for i in ids:
                versions[(esig(emap.get(i, {}).get("errors", [])),
                          emap.get(i, {}).get("corrected") or None)].append(i)
            comp = []
            for (sg, corr), vids in sorted(versions.items(), key=lambda kv: -len(kv[1])):
                cats = "; ".join(f"{c}/{s}" for c, s in sg) if sg else "（无错·判为正确）"
                comp.append((cats, corr, len(vids)))
            divergent.append({"source": src, "target": tgt, "count": len(ids),
                              "adopted_corr": rep_corr, "adopted_cats": esig(rep_errs), "versions": comp})

    out = [out_by_id[s["id"]] for s in state["segments"]]
    Path(a.out).write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")

    # ── 待人工裁决 report ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "待人工裁决"
    hdr = ["源文", "原译文", "出现次数", "竞争判定（各版本：类别/严重度 → 修正）",
           "已自动归一采用", "人工裁决（请填最终译法）"]
    ws.append(hdr)
    fill = PatternFill("solid", fgColor="2B3A2E")
    for ci in range(1, len(hdr) + 1):
        c = ws.cell(1, ci); c.font = Font(bold=True, color="FFFFFF"); c.fill = fill
    for d in sorted(divergent, key=lambda x: -x["count"]):
        vers = "\n".join(
            f"[{n}段] " + (("（无错）" if cats.startswith("（无错") else cats) +
                          (f" → {corr}" if corr else ""))
            for cats, corr, n in d["versions"])
        adopted = d["adopted_corr"] or "（判为正确/未改）"
        ws.append([d["source"], d["target"], d["count"], vers, adopted, ""])
    for col, w in zip("ABCDEF", [26, 30, 8, 52, 30, 28]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(a.report)

    print(f"[aggregate] groups={len(groups)} | divergent(需裁决)={len(divergent)} → {a.out}")
    print(f"[aggregate] 待人工裁决报告 → {a.report}")


if __name__ == "__main__":
    main()
