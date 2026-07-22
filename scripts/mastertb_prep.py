#!/usr/bin/env python3
"""MasterTB self-review prep + audit pipeline for LQE.

A glossary/TB audited as the work product (its own target translations are the
thing under review, there is no external TB to diff against). Reproduces the
proven 2024-06 ROCO flow: header-offset clean, EN/category/gender/definition
context injection, four required check modules with an optional proper-name
check, plus a cross-term consistency index handled separately from per-term
checks.

Subcommands
  prep    raw TB xlsx -> clean_input.xlsx + context.json + consistency.json + missing_th.xlsx
  chunks  state.json + context.json + errors_precheck.json -> chunks/chunk_NN.json + _CHECK_CONTEXT.md
  merge   chunks/chunk_NN.out.json + consistency.json -> errors.json (standard format)
  report  state.json + errors.json + context.json -> <label>_审校建议.xlsx

Column mapping is by header name (auto-locates the real header row containing
"术语 ZHCN"); the per-language TH block (TH / TH Comment / status) is resolved
POSITIONALLY because the status header repeats for every language.
"""
import argparse
import json
import re
import sys
import tempfile
from collections import Counter, OrderedDict, defaultdict
from pathlib import Path

import openpyxl

from lqe_chunk import (
    build_trusted_module_outputs,
    load_module_output,
    verification_generation_lease,
    with_precheck_refs,
)
from lqe_corrections import CheckFormatError, build_results, normalize_check_entries
from lqe_engine import (
    get_check_scope,
    load_terms,
    requires_bound_artifacts,
    validate_scope_entries,
)
from lqe_paths import publish_replacement_transaction, write_json_atomic
from lqe_result_contract import build_result_contract, result_contract_path
from lqe_split_contract import (
    add_chunk_payload_digest,
    build_split_manifest,
    build_split_revision,
    make_path_reference,
    generation_lock,
    publish_generation,
)

ZW = {0x200b: None, 0x200c: None, 0x200d: None, 0xfeff: None, 0x2060: None}
SRC_HDR = "术语 ZHCN"
STATUS_HDRS = {"术语状态 status", "status", "术语状态"}
CJK = re.compile(r"[一-鿿㐀-䶿]")
THAI = re.compile(r"[฀-๿]")

CTX_FIELDS = ("zhcn", "en", "definition", "category", "gender", "former",
              "th", "th_comment", "th_status", "scope")


def cl(v) -> str:
    return "" if v is None else str(v).translate(ZW).strip()


def find_header_row(rows):
    for i, r in enumerate(rows):
        if any(cl(c) == SRC_HDR for c in r):
            return i
    sys.exit(f"[err] header row containing {SRC_HDR!r} not found")


def col_map(hdr):
    def need(name):
        if name not in hdr:
            sys.exit(f"[err] column {name!r} not found; headers={hdr}")
        return hdr.index(name)
    m = {
        "src": need(SRC_HDR), "tgt": need("TH"), "en": need("EN"),
        "cat": need("术语类别 Category"), "scope": need("剧情范围"),
        "def": need("术语定义 Definition"), "gender": need("性別 Gender"),
        "former": need("曾用名"),
    }
    m["tgt_comment"] = m["tgt"] + 1
    m["tgt_status"] = next((j for j, c in enumerate(hdr)
                            if j > m["tgt"] and c.lower() in STATUS_HDRS), None)
    return m


def load_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


# ---------------------------------------------------------------- prep
def cmd_prep(a):
    rows = load_rows(a.input)
    h = find_header_row(rows)
    hdr = [cl(c) for c in rows[h]]
    m = col_map(hdr)
    job = Path(a.job_dir)
    job.mkdir(parents=True, exist_ok=True)

    audit, missing = [], []          # audit = filled TH, in order
    scope_ff = ""                    # forward-fill 剧情范围 (merged section marker)
    for r in rows[h + 1:]:
        g = lambda i: (cl(r[i]) if i is not None and i < len(r) else "")
        src = g(m["src"])
        sc = g(m["scope"])
        if sc:
            scope_ff = sc
        if not src or src == SRC_HDR:
            continue
        rec = {
            "zhcn": src, "en": g(m["en"]), "definition": g(m["def"]),
            "category": g(m["cat"]), "gender": g(m["gender"]),
            "former": g(m["former"]), "th": g(m["tgt"]),
            "th_comment": g(m["tgt_comment"]), "th_status": g(m["tgt_status"]),
            "scope": scope_ff,
        }
        (audit if rec["th"] else missing).append(rec)

    # context.json keyed by sequential id == lqe_io.read row order on clean_input
    ctx = OrderedDict((str(i), {k: rec[k] for k in CTX_FIELDS})
                      for i, rec in enumerate(audit))
    (job / "context.json").write_text(
        json.dumps(ctx, ensure_ascii=False, indent=1), encoding="utf-8")

    # clean_input.xlsx: real header at row 1 + audit rows (src + TH only is
    # enough for read; full structure kept so export can restore the TB).
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([SRC_HDR, "TH"])
    for rec in audit:
        ws.append([rec["zhcn"], rec["th"]])
    wb.save(job / "clean_input.xlsx")

    # missing_th.xlsx (record only; excluded from audit)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["术语 ZHCN", "EN", "Category", "Definition", "th_status"])
    for rec in missing:
        ws.append([rec["zhcn"], rec["en"], rec["category"],
                   rec["definition"], rec["th_status"]])
    wb.save(job / "missing_th.xlsx")

    # consistency.json (cross-term; rubric is told NOT to judge this locally)
    by_src, by_th = defaultdict(list), defaultdict(list)
    noth, cjk = [], []
    for i, rec in enumerate(audit):
        by_src[rec["zhcn"]].append({"id": i, "th": rec["th"], "en": rec["en"]})
        by_th[rec["th"]].append({"id": i, "zhcn": rec["zhcn"]})
        if not THAI.search(rec["th"]):
            noth.append({"id": i, "zhcn": rec["zhcn"], "th": rec["th"]})
        if CJK.search(rec["th"]):
            cjk.append({"id": i, "zhcn": rec["zhcn"], "th": rec["th"]})
    inc_src = {s: v for s, v in by_src.items()
              if len({x["th"] for x in v}) > 1}
    multi = {t: v for t, v in by_th.items()
            if len({x["zhcn"] for x in v}) > 1}
    cons = {
        "summary": {
            "total_filled": len(audit), "total_missing_th": len(missing),
            "src_groups_inconsistent_th": len(inc_src),
            "th_groups_multiple_src": len(multi),
            "th_no_thai_char": len(noth), "th_contains_cjk": len(cjk),
        },
        "inconsistent_same_src": inc_src,
        "same_th_multi_src": multi,
        "th_no_thai_char": noth,
        "th_contains_cjk": cjk,
    }
    (job / "consistency.json").write_text(
        json.dumps(cons, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"[ok] audit(filled TH)={len(audit)}  missing TH={len(missing)}")
    print(f"     clean_input.xlsx / context.json / missing_th.xlsx / consistency.json -> {job}")
    print(f"     consistency: {cons['summary']}")


# ---------------------------------------------------------------- chunks
CHECK_CONTEXT = """\
# MasterTB 术语表自检上下文

每个检查任务先读 `references/check_modules/common.md`、自己的模块文件和
`references/check_modules/term_audit.md`，再读任务目录中的项目背景、语言说明、
确认规则和风格指南。

输入是 chunk JSON 的 `segments[]`。`source` 是中文术语，`target` 是待检查的
泰语译文；`en`、`definition`、`category`、`gender`、`former`、
`target_comment` 和 `scope` 提供消歧上下文。`precheck` 是机器预检结果。

四个必需模块分别写 `terminology`、`accuracy`、`grammar` 和 `naturalness`
结果；`proper_names` 仅用于术语表自检中的 name 段，可选。

每个模块先输出覆盖其分配 id 的 JSON 数组草稿：

```json
[{"id": 123, "issues": [{"category": "Mistranslation", "severity": "Major",
  "comment": "说明问题", "needs_confirmation": true, "edit": null}]}]
```

不得输出 `corrected` 或 `review_provenance`。安全局部修改和需要人工确认的规则以模块文档为准。草稿完成后必须按 `references/check_modules/common.md` 使用 `lqe_chunk.py publish-module` 和当前 chunk 指纹发布正式绑定文件，不能直接把裸数组写到正式模块路径。
"""


_NAME_CATS = {"Named NPC", "Creature Species", "Settlement", "Wilderness",
              "Macro Region", "Administrative Region", "Urban Area", "Functional Area"}
_DESC_MARK = ("的", "们")  # descriptive-phrase markers
_REQUIRED_MODULES = ("terminology", "accuracy", "grammar", "naturalness")
_OPTIONAL_MODULES = ("proper_names",)


def _kind(category, zhcn):
    """name=专名；desc=描述或词义内容。Creature Individual 多为描述短语，
    按内容细分（含“的/们”或长度 > 6 时归 desc）；存疑时归 desc。"""
    if category in _NAME_CATS:
        return "name"
    if category == "Creature Individual":
        return "desc" if (any(m in zhcn for m in _DESC_MARK) or len(zhcn) > 6) else "name"
    return "desc"


def cmd_chunks(a):
    job = Path(a.job_dir)
    state_path = job / "state.json"
    precheck_path = job / "errors_precheck.json"
    state = json.loads(state_path.read_text("utf-8"))
    ctx = json.loads((job / "context.json").read_text("utf-8"))
    segs = state["segments"]
    if len(segs) != len(ctx):
        sys.exit(f"[err] state segs {len(segs)} != context {len(ctx)}; re-run read on clean_input")
    # alignment guard: source/target must match context by id
    for s in segs:
        c = ctx[str(s["id"])]
        if c["zhcn"] != s["source"] or c["th"] != s["target"]:
            sys.exit(f"[err] id {s['id']} misaligned: state({s['source']!r}/{s['target']!r}) "
                     f"vs ctx({c['zhcn']!r}/{c['th']!r})")
    if not precheck_path.exists():
        write_json_atomic(
            precheck_path,
            [{"id": segment["id"], "issues": []} for segment in segs],
        )
    precheck_entries = normalize_check_entries(
        json.loads(precheck_path.read_text("utf-8")),
        label=precheck_path.name,
    )
    validate_scope_entries(
        state,
        precheck_entries,
        issues_key="issues",
        label=precheck_path.name,
    )
    precheck = {
        entry["id"]: with_precheck_refs(entry["id"], entry["issues"])
        for entry in precheck_entries
    }
    out = job / "chunks"
    size = a.size
    terms = load_terms(state)
    revision = build_split_revision(
        state,
        precheck_entries,
        terms,
        get_check_scope(state),
        size=size,
        char_budget=0,
    )
    n = (len(segs) + size - 1) // size
    chunk_payloads = []
    dedup_map = {str(segment["id"]): [segment["id"]] for segment in segs}
    for ci in range(n):
        block = segs[ci * size:(ci + 1) * size]
        segments = []
        for s in block:
            c = ctx[str(s["id"])]
            segments.append({
                "id": s["id"], "source": c["zhcn"], "target": c["th"],
                "kind": _kind(c["category"], c["zhcn"]),
                "precheck": precheck.get(s["id"], []),
                "term_hits": [], "term_near": [],
                "protected": bool(s.get("protected")),
                "protected_texts": s.get("protected_texts", []),
                "en": c["en"],
                "definition": c["definition"], "category": c["category"],
                "gender": c["gender"], "former": c["former"],
                "target_comment": c["th_comment"],
                "target_status": c["th_status"], "scope": c["scope"],
            })
        chunk_payloads.append(
            add_chunk_payload_digest(
                {
                    "chunk_id": ci,
                    "iteration": state.get("iteration", 0),
                    "state_fingerprint": revision["state_fingerprint"],
                    "split_fingerprint": revision["split_fingerprint"],
                    "segments": segments,
                }
            )
        )
    manifest = build_split_manifest(
        revision,
        chunks=chunk_payloads,
        dedup_map=dedup_map,
        input_references={
            "state": make_path_reference(state_path, job),
            "precheck": make_path_reference(precheck_path, job),
            "terms": None,
            "terms_mode": "state",
        },
    )
    out.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(
        dir=out.parent,
        prefix=".chunks.generation.",
    ) as staging_name:
        staging = Path(staging_name)
        write_json_atomic(staging / "dedup_map.json", dedup_map)
        write_json_atomic(staging / "split_manifest.json", manifest)
        for payload in chunk_payloads:
            write_json_atomic(
                staging / f"chunk_{payload['chunk_id']:02d}.json",
                payload,
            )
        (staging / "_CHECK_CONTEXT.md").write_text(
            CHECK_CONTEXT,
            encoding="utf-8",
        )
        publish_generation(
            staging,
            out,
            archive_label=f"iter_{state.get('iteration', 0)}_mastertb",
        )
    print(f"[ok] {n} chunks (size {size}, 含 kind 标记) -> {out}")
    print("     必需检查模块：")
    for module in _REQUIRED_MODULES:
        print(f"       chunk_NN.{module}.json")
    print(f"     可选专名检查：chunk_NN.{_OPTIONAL_MODULES[0]}.json")
    print(f"     python scripts/lqe_chunk.py validate-checks --job {job}")
    print(f"     python scripts/lqe_chunk.py merge-checks --job {job}")
    print(f"     python scripts/mastertb_prep.py merge --job-dir {job}")


# ---------------------------------------------------------------- merge
def _load_findings(path, base, state):
    if not path.exists():
        return {}
    bound = requires_bound_artifacts(state)
    entries = load_module_output(
        path,
        base,
        "merged",
        state,
        allow_internal_provenance=bound,
        require_internal_provenance=bound,
    )
    return {entry["id"]: entry for entry in entries}


def _cmd_merge_locked(a, job: Path, state: dict, manifest: dict, revalidate):
    ids = [s["id"] for s in state["segments"]]
    out = job / "chunks"
    chunk_numbers = sorted(
        int(match.group(1))
        for path in out.glob("chunk_*.json")
        if (match := re.fullmatch(r"chunk_(\d+)\.json", path.name))
    )
    nchunks = len(chunk_numbers)
    incomplete = []
    merged = {}
    expected_all = set()
    trusted_chunks = {}
    if requires_bound_artifacts(state):
        trusted_chunks, problems = build_trusted_module_outputs(job, state)
        if problems:
            raise SystemExit(
                "[err] bound module validation failed: "
                + "; ".join(problems[:50])
            )
    for ci in chunk_numbers:
        chunk_path = out / f"chunk_{ci:02d}.json"
        chunk = json.loads(chunk_path.read_text("utf-8"))
        expected = {
            segment["id"]
            for segment in chunk.get("segments", [])
            if isinstance(segment, dict) and type(segment.get("id")) is int
        }
        expected_all.update(expected)
        missing_modules = [
            module
            for module in _REQUIRED_MODULES
            if not (out / f"chunk_{ci:02d}.{module}.json").exists()
        ]
        merged_path = out / f"chunk_{ci:02d}.out.json"
        problem = {"chunk": ci}
        if missing_modules:
            problem["missing_modules"] = missing_modules
        if not merged_path.exists():
            problem["missing_merged_output"] = True
        else:
            try:
                entries = _load_findings(merged_path, chunk, state)
            except (json.JSONDecodeError, CheckFormatError) as exc:
                raise SystemExit(f"[err] invalid {merged_path.name}: {exc}") from exc
            actual = set(entries)
            missing_ids = sorted(expected - actual)
            extra_ids = sorted(actual - expected)
            if missing_ids:
                problem["missing_ids"] = missing_ids
            if extra_ids:
                problem["extra_ids"] = extra_ids
            if not missing_ids and not extra_ids:
                if requires_bound_artifacts(state):
                    trusted = trusted_chunks.get(ci)
                    trusted_entries = trusted[2] if trusted is not None else None
                    trusted_ids = trusted[1] if trusted is not None else []
                    loaded_entries = [entries[key] for key in trusted_ids]
                    if loaded_entries != trusted_entries:
                        raise SystemExit(
                            f"[err] {merged_path.name} differs from current "
                            "bound module outputs"
                        )
                    entries = {
                        entry["id"]: entry for entry in trusted_entries
                    }
                merged.update(entries)
        if len(problem) > 1:
            incomplete.append(problem)

    state_ids = set(ids)
    state_missing = sorted(state_ids - expected_all)
    state_extra = sorted(expected_all - state_ids)
    if state_missing or state_extra:
        problem = {"chunk": "all"}
        if state_missing:
            problem["missing_ids"] = state_missing
        if state_extra:
            problem["extra_ids"] = state_extra
        incomplete.append(problem)

    if incomplete:
        status = {
            "n_chunks": nchunks,
            "checks_complete": False,
            "verdict_allowed": False,
            "incomplete_chunks": incomplete,
            "note": "检查结果不完整；运行 validate-checks 和 merge-checks 后重试",
        }
        (job / "recall_status.json").write_text(
            json.dumps(status, ensure_ascii=False, indent=1), encoding="utf-8"
        )
        errors_path = job / "errors.json"
        if errors_path.exists():
            errors_path.unlink()
        result_contract_path(errors_path).unlink(missing_ok=True)
        details = []
        for problem in incomplete:
            label = f"chunk_{problem['chunk']:02d}" if isinstance(problem["chunk"], int) else "all chunks"
            if problem.get("missing_modules"):
                details.append(f"{label} missing modules={problem['missing_modules']}")
            if problem.get("missing_merged_output"):
                details.append(f"{label} missing merged output; run merge-checks")
            if problem.get("missing_ids"):
                details.append(f"{label} missing={problem['missing_ids']}")
            if problem.get("extra_ids"):
                details.append(f"{label} extra={problem['extra_ids']}")
        raise SystemExit(f"[err] incomplete MasterTB checks: {'; '.join(details)}")

    # fold cross-term consistency (global; not judged locally by the rubric)
    cons_p = job / "consistency.json"
    folded = 0
    if cons_p.exists() and not a.no_consistency:
        cons = json.loads(cons_p.read_text("utf-8"))
        for src, members in cons.get("inconsistent_same_src", {}).items():
            variants = sorted({x["th"] for x in members})
            for x in members:
                sid = x["id"]
                if sid in merged and not any(e.get("category") == "Inconsistency"
                                             for e in merged[sid]["issues"]):
                    merged[sid]["issues"].append({
                        "category": "Inconsistency", "severity": "Minor",
                        "comment": f"[一致性] 同源 '{src}' 跨词条出现多种泰译 {variants}；需统一。",
                        "needs_confirmation": False,
                        "edit": None,
                        **(
                            {
                                "review_provenance": {
                                    "finding_origin": "script_derived",
                                    "ai_reviewed": False,
                                    "ai_edited": False,
                                    "review_module": None,
                                    "reviewed_segment_id": None,
                                    "edit_origin": None,
                                }
                            }
                            if requires_bound_artifacts(state)
                            else {}
                        ),
                    })
                    folded += 1

    checks = [merged[i] for i in ids]
    bound = requires_bound_artifacts(state)
    errors = build_results(
        state["segments"],
        checks,
        allow_internal_provenance=bound,
        require_internal_provenance=bound,
    )
    errors_path = job / "errors.json"
    revalidate()
    if requires_bound_artifacts(state):
        contract_path = result_contract_path(errors_path)
        with tempfile.TemporaryDirectory(
            dir=job,
            prefix=".mastertb-results.",
        ) as staging_name:
            staging = Path(staging_name)
            staged_errors = staging / errors_path.name
            staged_contract = staging / contract_path.name
            write_json_atomic(staged_errors, errors)
            write_json_atomic(
                staged_contract,
                build_result_contract(manifest, errors),
            )
            publish_replacement_transaction(
                [
                    (staged_errors, errors_path),
                    (staged_contract, contract_path),
                ]
            )
    else:
        write_json_atomic(errors_path, errors)
    flagged = sum(1 for e in errors if e["errors"])
    nerr = sum(len(e["errors"]) for e in errors)
    print(f"[ok] merged check results for {nchunks} chunks -> errors.json")
    print(f"     flagged segments={flagged}/{len(ids)}  total errors={nerr}  consistency folded={folded}")
    status = {
        "n_chunks": nchunks,
        "checks_complete": True,
        "verdict_allowed": True,
        "incomplete_chunks": [],
        "note": "各 chunk 的适用检查模块齐全，可以确认结果",
    }
    (job / "recall_status.json").write_text(
        json.dumps(status, ensure_ascii=False, indent=1), encoding="utf-8")


def cmd_merge(a):
    job = Path(a.job_dir)
    initial_state = json.loads((job / "state.json").read_text("utf-8"))
    bound = requires_bound_artifacts(initial_state)
    if not bound:
        with generation_lock(job / "chunks", exclusive=False):
            state = json.loads((job / "state.json").read_text("utf-8"))

            def revalidate():
                live = json.loads((job / "state.json").read_text("utf-8"))
                if live != state:
                    raise ValueError("state changed during MasterTB merge")

            _cmd_merge_locked(a, job, state, None, revalidate)
        return
    with verification_generation_lease(
        job / "state.json",
        exclusive=False,
        require_generation=bound,
    ) as (state, _, manifest, revalidate):
        if bound and manifest is None:
            raise SystemExit("[err] verified chunk generation is required")
        _cmd_merge_locked(a, job, state, manifest, revalidate)


# ---------------------------------------------------------------- report
SEV_ORDER = {"Critical": 0, "Major": 1, "Minor": 2, "Neutral": 3}


def cmd_report(a):
    job = Path(a.job_dir)
    state = json.loads((job / "state.json").read_text("utf-8"))
    errors = json.loads((job / "errors.json").read_text("utf-8"))
    ctx = json.loads((job / "context.json").read_text("utf-8"))
    emap = {e["id"]: e for e in errors}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "审校建议"
    ws.append(["id", "剧情范围", "类别", "中文 ZHCN", "EN", "性别", "定义",
               "原泰译 TH", "建议译文", "错误类别", "严重度", "置信", "说明", "TH状态"])
    rows = []
    for e in errors:
        if not e["errors"]:
            continue
        c = ctx.get(str(e["id"]), {})
        worst = min(e["errors"], key=lambda x: SEV_ORDER.get(x.get("severity"), 9))
        rows.append((SEV_ORDER.get(worst.get("severity"), 9), e["id"], e, c))
    rows.sort(key=lambda t: (t[0], t[1]))
    for _, sid, e, c in rows:
        cats = " / ".join(x.get("category", "") for x in e["errors"])
        sevs = " / ".join(x.get("severity", "") for x in e["errors"])
        cons = " / ".join(x.get("confidence", "") for x in e["errors"])
        cmts = "\n".join(f"[{x.get('severity')}] {x.get('comment','')}" for x in e["errors"])
        ws.append([sid, c.get("scope", ""), c.get("category", ""), c.get("zhcn", ""),
                   c.get("en", ""), c.get("gender", ""), c.get("definition", ""),
                   c.get("th", ""), e.get("corrected") or "", cats, sevs, cons, cmts,
                   c.get("th_status", "")])
    # 汇总页先写检查完整性，避免把临时分数当作最终结果
    sm = wb.create_sheet("汇总")
    rs_p = job / "recall_status.json"
    rs = json.loads(rs_p.read_text("utf-8")) if rs_p.exists() else {}
    if rs.get("verdict_allowed") is False:
        nbad = len(rs.get("incomplete_chunks", []))
        sm.append(["⚠ 检查完整性", f"{nbad}/{rs.get('n_chunks')} 块缺少适用检查模块"])
        sm.append(["⚠ 结果", "当前分数仅供参考；补齐检查模块后重新运行（见 SKILL.md 第 3 步）"])
        sm.append([])
    cat_ct = Counter()
    sev_ct = Counter()
    for e in errors:
        for x in e["errors"]:
            cat_ct[x.get("category", "")] += 1
            sev_ct[x.get("severity", "")] += 1
    sm.append(["指标", "值"])
    sm.append(["审校词条总数", len(state["segments"])])
    sm.append(["有问题词条数", len(rows)])
    sm.append(["错误总数", sum(cat_ct.values())])
    sm.append([])
    sm.append(["严重度", "数量"])
    for k in ("Critical", "Major", "Minor", "Neutral"):
        if sev_ct.get(k):
            sm.append([k, sev_ct[k]])
    sm.append([])
    sm.append(["错误类别", "数量"])
    for k, v in cat_ct.most_common():
        sm.append([k, v])

    label = a.label or job.name
    outp = job / f"{label}_审校建议.xlsx"
    wb.save(outp)
    print(f"[ok] {len(rows)} flagged terms -> {outp}")


# ---------------------------------------------------------------- view
def cmd_view(a):
    """Create compact per-chunk text views for a temporary single-pass check.

    The views are about ten times denser than the raw chunk JSON. A single pass
    is preliminary; all check modules must still run before PASS/FAIL is final.
    """
    job = Path(a.job_dir)
    ctx = json.loads((job / "context.json").read_text("utf-8"))
    out = job / "chunks" / "views"
    out.mkdir(parents=True, exist_ok=True)

    def one(s, n=0):
        s = ("" if s is None else str(s)).replace("\n", " / ").strip()
        return s[:n] if n else s

    ids = sorted(int(k) for k in ctx)
    size = a.size
    n = (len(ids) + size - 1) // size
    for ci in range(n):
        block = ids[ci * size:(ci + 1) * size]
        lines = ["id\tcat\tg\tzhcn\ten\tdef\tth"]
        for i in block:
            c = ctx[str(i)]
            lines.append("\t".join([str(i), one(c["category"]), one(c["gender"]),
                one(c["zhcn"]), one(c["en"]), one(c["definition"], 60), one(c["th"])]))
        (out / f"chunk_{ci:02d}.view.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"[ok] {n} compact views (size {size}) -> {out}")
    print("     view 只用于辅助阅读；正式结果仍写入四个必需模块文件")
    print("     完成后运行 validate-checks 和 merge-checks，再报告 PASS/FAIL")


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    p = sub.add_parser("prep"); p.add_argument("--input", required=True); p.add_argument("--job-dir", required=True); p.set_defaults(fn=cmd_prep)
    p = sub.add_parser("chunks"); p.add_argument("--job-dir", required=True); p.add_argument("--size", type=int, default=300); p.set_defaults(fn=cmd_chunks)
    p = sub.add_parser("merge"); p.add_argument("--job-dir", required=True); p.add_argument("--no-consistency", action="store_true"); p.set_defaults(fn=cmd_merge)
    p = sub.add_parser("report"); p.add_argument("--job-dir", required=True); p.add_argument("--label", default=None); p.set_defaults(fn=cmd_report)
    p = sub.add_parser("view"); p.add_argument("--job-dir", required=True); p.add_argument("--size", type=int, default=300); p.set_defaults(fn=cmd_view)
    a = ap.parse_args()
    a.fn(a)


if __name__ == "__main__":
    main()
