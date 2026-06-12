"""Generate docs/PM反馈检查项修改报告.xlsx — modification report for the 10
check-item suggestions from PM feedback on 质量检查项清单."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUT = os.path.join(os.path.dirname(__file__), "..", "docs", "PM反馈检查项修改报告.xlsx")

TITLE_F = Font(bold=True, size=14, color="1F4E79")
SEC_F = Font(bold=True, size=12, color="1F4E79")
H_F = Font(bold=True, size=10, color="FFFFFF")
H_FILL = PatternFill("solid", fgColor="4472C4")
NOCHANGE_FILL = PatternFill("solid", fgColor="E2EFDA")
WRAP = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="B0B0B0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "修改报告"

ws["A1"] = "质量检查项修改报告（PM 反馈 10 条）"
ws["A1"].font = TITLE_F

meta = [
    "更新日期：2026-06-12",
    "评估对象：PM 审阅《质量检查项清单》后反馈的 10 条补充检查项建议，逐条给出是否修改、修改程度与实现方案",
    "依据：〔PM反馈〕＝建议原文；〔Skill〕＝现行评估体系（SKILL.md 及 scripts/lqe_io.py 代码实测核对）；〔清单〕＝docs/质量检查项清单.md（沿用其 N1–N4 编号，新增项拟编 N5–N9）",
    "结论汇总：新增内置检查 5 项（#1/#2/#6/#8/#9 → 拟编 N5–N9）；扩展现有 2 项（#3 项目自定义检查架构、#7 术语检查大小写校验）；无需改动 2 项（#4 已实现、#5 已含于 N2 设计）；仅项目配置 1 项（#10，零代码）",
    "通用机制：全部新增确定性检查沿用「pre-check 报出 → AI 评估甄别、可移除上下文误报」通道（与现行数值一致性检查相同），保召回、控误报",
    "状态：全部待批准；批准后落码，并同步《质量检查项清单》（md 与 xlsx）",
    "需要 PM 动作的共 5 条（#2/#3/#5/#7/#10，见最后一列）；其余 5 条无需提供任何材料，批准即可实施",
]
r = 3
for m in meta:
    ws.cell(row=r, column=1, value=m).alignment = WRAP
    r += 1

r += 1
ws.cell(row=r, column=1, value="逐条修改报告").font = SEC_F
r += 1

HEADERS = ["序号", "PM 建议", "是否需要修改", "修改程度", "如何修改（实现方案与误报防控）", "涉及文件", "需要 PM 提供 / 拍板"]

ROWS = [
    ["1", "译文句号缺失（建议用 AI 识别源译句号数量是否对应）",
     "部分需要",
     "新增内置检查（拟编 N5，句尾子集）＋ AI 评估关注点（句中数量）",
     "句尾比对不必用 AI 计数，确定性代码即可：源文以「。！？…」结尾而译文句尾无终止标点（.!?…），或源文无句尾标点而译文擅自添加 → Punctuation Minor。"
     "句中句号不做 1:1 数量对账——中英句子合法拆合（一句中文译成两句英文）会大量误报，该部分写入 AI 评估关注点，由 AI 结合语义判断是否漏句。"
     "泰语等无句号体系语言由语言属性（languages/th/：sentence_terminator=none）自动判定不适用，无需逐项目配置。",
     "scripts/lqe_io.py；languages/th/（属性自动判定）；清单文档",
     "无需提供——批准即可实施"],
    ["2", "中文数字对应（如「三次」→ 3/three）",
     "需要",
     "新增内置检查（拟编 N6）",
     "确定性解析即可、无需依赖 AI 识别：解析中文数字（〇一二…十百千万亿、两、第 N），仅「中文数字＋量词/单位」模式触发（次/个/名/位/层/级/天/秒/章/倍等），"
     "防「一起/三思/万一」类虚用成语误报；「一」默认豁免、仅强量词模式（一次/一名）触发。译侧接受阿拉伯数字或英文数词（one~hundred、first 等序数、once/twice）。"
     "报 Mistranslation Major；AI 评估可移除上下文误报（与现行数值检查同机制）。",
     "scripts/lqe_io.py；清单文档",
     "请拍板两件事：① 数字『一』要不要也查——查的话误报会变多，建议只查『一次/一名』这类带量词的；② 泰语版要不要这期一起做"],
    ["3", "标签（TAG）正则按项目单独配置（现内置仅燕云 #G…#E）",
     "需要（方向正确，架构已支持一半）",
     "扩展项目自定义检查架构（约 15 行）＋各项目配置",
     "现行 projects/<game>/<lang>/checks.json 的 custom 正则已实现「按项目喂正则」，但仅支持单边命中报错，无法做源↔译对账。"
     "扩展：custom 增加 type 字段——search（现状，单边命中）／count_match（新增：正则在源、译分别取全部匹配，数量不等报 Markup Major），"
     "即可覆盖 <color=>、[b][/b] 等任意项目标签体系。燕云专属 #G…#E 内置项，其他项目以 \"color_tags\": false 关闭（机制已有）。"
     "标签嵌套、相对位置等复杂校验仍归 AI 评估。",
     "scripts/lqe_io.py；各项目 checks.json",
     "请提供：每个项目的标签长什么样，每种标签贴几行原文例子即可（如 <color=#FF0000>文字</color>、[b]文字[/b]），技术规则我们来写"],
    ["4", "小数点缺失",
     "不需要",
     "无需改动（已实现）",
     "实测现行代码：数值一致性检查的提取正则 \\d[\\d,]*(?:\\.\\d+)? 将小数作为完整数值 token 处理——"
     "源「2.5」译成「25」「2,5」或缺失，均报「Source number missing/changed」（Mistranslation Major）。",
     "无",
     "无需提供——已实现，需要的话可现场演示"],
    ["5", "异源同译（不同源文译成相同译文）",
     "不需要新增设计",
     "无需新增（已含于待批准 N2）；落码时补误报分档",
     "清单 N2「同源异译」设计原文已含双向：「相同源文对应不同译文；以及相同译文对应不同源文」。"
     "落码时补充长度分档防误报：短词条（源 <10 字符）合法收敛（确定/确认→Confirm）不报；"
     "较长文本（≥20 字符）异源同译高度可疑（复制错位/漏改），报 Inconsistency Minor；AI 甄别。",
     "scripts/lqe_io.py（随 N2 实施）",
     "请拍板：多长的句子译文完全相同才算可疑——建议 20 字以上才报，『确定/确认』这类短按钮词不报"],
    ["6", "单词连续重复（如 the the）",
     "需要",
     "新增内置检查（拟编 N7）",
     "正则 \\b(\\w+)\\s+\\1\\b 忽略大小写扫描译文；白名单豁免英文合法重复（had had／that that／no no／so so 等）；"
     "仅空格分词语言生效，由语言属性 word_delim 自动判定（泰语自动跳过——泰文表重复另有 ๆ 标记体系）；Grammar Minor；对话体重复由 AI 甄别。",
     "scripts/lqe_io.py；清单文档",
     "无需提供——批准即可实施"],
    ["7", "特定术语大小写（HP/ATK/人名等必须大写）",
     "需要（实测现版漏检）",
     "扩展现有术语检查",
     "实测确认缺口：现行术语匹配统一转小写比对，HP 译成 hp 不报错。"
     "改法：不区分大小写命中译法后追加大小写校验——全大写缩写词条（≥2 字母：HP/ATK/SSR）及词内混合大小写词条要求精确匹配；首字母大写词条（人名）校验首字母。"
     "归类建议：默认 Company style Minor（Terminology 类被脚本强制 Major，全判将罚分膨胀），全大写缩写误写升 Major；亦可在术语表增加 case_sensitive 列显式控制。",
     "scripts/lqe_io.py；（可选）术语表加列",
     "请拍板两件事：① 大小写写错算多重——建议一般情况算轻微错误，HP/ATK 这类全大写缩写写错算严重错误；② 要不要在术语表加一列，标出哪些词必须严格按表里的大小写"],
    ["8", "同个词内大小写混乱（如 AppLe）",
     "需要",
     "新增内置检查（拟编 N8）",
     "检测词内小写→大写转折（AppLe／heLLo）；豁免：术语表词条、Mc/Mac 前缀、iPhone 型单字母小写头、PvP 型缩写、标签/变量内容；"
     "Spelling Minor；剩余少量品牌词误报由 AI 甄别。",
     "scripts/lqe_io.py；清单文档",
     "无需提供——批准即可实施"],
    ["9", "成对标点不完整（半个引号/括号）",
     "需要",
     "新增内置检查（拟编 N9）",
     "按对类计数比对：()、[]、「」、『』、“”、‘’，直双引号 \" 按奇偶；仅「源侧配对完整而译侧不完整」报错——"
     "源文本身跨段半引号则跳过，防误报；直单引号 ' 默认豁免（英文撇号）；{} 不查（变量检查已覆盖）；Punctuation Minor。",
     "scripts/lqe_io.py；清单文档",
     "无需提供——批准即可实施"],
    ["10", "省略号样式按项目规定（…／...）",
     "不需要改代码",
     "仅项目配置（零代码）",
     "正是现行 custom 正则的标准用例：要求「…」的项目加 pattern \\.{3}（命中提示应使用…），要求「...」的项目加 pattern …；"
     "写入各项目 checks.json 即生效，类别/严重度可配（默认 Company style Minor）。",
     "各项目 checks.json",
     "请告知：每个项目要用哪种省略号——『…』还是『...』，逐项目确认一次即可"],
]

NO_CHANGE = {4, 5, 10}

for j, h in enumerate(HEADERS, 1):
    c = ws.cell(row=r, column=j, value=h)
    c.font, c.fill, c.alignment, c.border = H_F, H_FILL, WRAP, BORDER
for i, row in enumerate(ROWS, 1):
    for j, v in enumerate(row, 1):
        c = ws.cell(row=r + i, column=j, value=v)
        c.alignment, c.border = WRAP, BORDER
        if int(row[0]) in NO_CHANGE:
            c.fill = NOCHANGE_FILL

for j, w in enumerate([5, 22, 15, 22, 64, 26, 26], 1):
    ws.column_dimensions[chr(64 + j)].width = w
ws.freeze_panes = ws.cell(row=r + 1, column=1)

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"saved: {os.path.abspath(OUT)}")
