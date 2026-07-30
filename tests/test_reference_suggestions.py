from pathlib import Path
import sys
import tempfile
import unittest

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

import lqe_io
from lqe_corrections import CheckFormatError, validate_reference_target
from lqe_split_contract import canonical_digest
from lqe_suggestions import (
    ARTIFACT_SCHEMA,
    ARTIFACT_VERSION,
    _normalize_selection,
    _with_digest,
    build_suggestion_packet,
    validate_suggestion_artifact,
)


def issue(
    comment,
    *,
    needs_confirmation=False,
    edit=None,
    category="Unidiomatic",
    severity="Major",
):
    return {
        "category": category,
        "severity": severity,
        "comment": comment,
        "needs_confirmation": needs_confirmation,
        "edit": edit,
    }


class ReferenceSuggestionContractTests(unittest.TestCase):
    def test_packet_uses_major_critical_candidates_and_agent_judgment(self):
        segments = [
            {
                "id": 0,
                "source": "Source 0",
                "target": "Target 0",
                "content_type": "UI/界面文本",
            },
            {"id": 1, "source": "Source 1", "target": "Target 1"},
            {
                "id": 2,
                "source": "Source 2",
                "target": "Target 2",
                "text_type_context": "主线剧情对话",
            },
            {
                "id": 3,
                "source": "Source 3",
                "target": "Target 3",
                "protected": True,
            },
        ]
        results = [
            {
                "id": 0,
                "errors": [
                    issue("Major problem."),
                    issue(
                        "Minor problem.",
                        severity="Minor",
                        needs_confirmation=True,
                    ),
                ],
                "corrected": None,
            },
            {
                "id": 1,
                "errors": [
                    issue(
                        "Minor only.",
                        severity="Minor",
                        needs_confirmation=True,
                    )
                ],
                "corrected": None,
            },
            {
                "id": 2,
                "errors": [issue("Critical problem.", severity="Critical")],
                "corrected": None,
            },
            {
                "id": 3,
                "errors": [issue("Protected problem.")],
                "corrected": None,
            },
        ]

        packet = build_suggestion_packet(segments, None, results)

        self.assertEqual(
            packet["selection"],
            {
                "categories": [],
                "severities": ["Critical", "Major"],
                "only_missing": False,
            },
        )
        self.assertEqual(packet["reviewed_ids"], [0, 2])
        self.assertEqual(
            [error["severity"] for error in packet["segments"][0]["errors"]],
            ["Major"],
        )
        self.assertEqual(
            packet["segments"][0]["content_type"],
            "UI/界面文本",
        )
        self.assertEqual(
            packet["segments"][1]["text_type_context"],
            "主线剧情对话",
        )
        self.assertTrue(
            packet["instructions"]["sparse_suggestions_allowed"]
        )
        self.assertTrue(packet["instructions"]["agent_decides_reliability"])

        with self.assertRaisesRegex(ValueError, "only support Major/Critical"):
            _normalize_selection(
                {
                    "categories": [],
                    "severities": ["Minor"],
                    "only_missing": False,
                }
            )

    def test_full_reference_target_preserves_protected_signature(self):
        segment = {
            "id": 7,
            "target": "Use {name}<b>\nKeep RAW",
            "protected_texts": ["RAW"],
        }
        accepted = "请使用 {name}<b>\n并保留 RAW"
        self.assertEqual(
            validate_reference_target(segment, accepted),
            accepted,
        )
        for rejected in (
            "请使用名字<b>\n并保留 RAW",
            "请使用 {name}<b> 并保留 RAW",
            "请使用 {name}<b>\n并保留",
        ):
            with self.subTest(rejected=rejected):
                with self.assertRaises(CheckFormatError):
                    validate_reference_target(segment, rejected)

        protected = {**segment, "protected": True}
        with self.assertRaises(CheckFormatError):
            validate_reference_target(protected, accepted)

    def test_packet_binding_ignores_score_only_repeated_flag(self):
        segments = [{"id": 0, "source": "Source", "target": "Target"}]
        results = [{
            "id": 0,
            "errors": [issue("Awkward.")],
            "corrected": None,
        }]
        manifest = {
            "manifest_digest": "manifest",
            "state_fingerprint": "state",
        }
        first = build_suggestion_packet(segments, manifest, results)
        repeated = [{
            **results[0],
            "errors": [{**results[0]["errors"][0], "repeated": True}],
        }]
        second = build_suggestion_packet(segments, manifest, repeated)
        self.assertEqual(first["packet_digest"], second["packet_digest"])

    def test_artifact_rejects_duplicate_and_unsafe_suggestions(self):
        segments = [{
            "id": 0,
            "source": "Source",
            "target": "Target {0}",
        }]
        results = [{
            "id": 0,
            "errors": [issue("Awkward.")],
            "corrected": None,
        }]
        packet = build_suggestion_packet(segments, None, results)

        def artifact(entries):
            payload = {
                "schema": ARTIFACT_SCHEMA,
                "version": ARTIFACT_VERSION,
                "packet_digest": packet["packet_digest"],
                "manifest_digest": packet["manifest_digest"],
                "selection": packet["selection"],
                "results_basis_digest": packet["results_basis_digest"],
                "reviewed_ids": packet["reviewed_ids"],
                "entries": entries,
                "entries_digest": canonical_digest(entries),
            }
            return _with_digest(payload, "artifact_digest")

        valid = artifact([
            {"id": 0, "reference_target": "Better target {0}"},
        ])
        self.assertEqual(
            validate_suggestion_artifact(valid, packet, segments),
            {0: "Better target {0}"},
        )
        legacy = {**valid, "version": 1}
        with self.assertRaisesRegex(ValueError, "version"):
            validate_suggestion_artifact(legacy, packet, segments)
        duplicate = artifact([
            {"id": 0, "reference_target": "Better target {0}"},
            {"id": 0, "reference_target": "Another target {0}"},
        ])
        with self.assertRaises(ValueError):
            validate_suggestion_artifact(duplicate, packet, segments)
        unsafe = artifact([
            {"id": 0, "reference_target": "Better target"},
        ])
        with self.assertRaises(CheckFormatError):
            validate_suggestion_artifact(unsafe, packet, segments)


class ReviewerWorkbookTests(unittest.TestCase):
    def test_minor_only_segment_never_displays_suggested_translation(self):
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "minor-review.xlsx"
            state = {
                "input_path": str(Path(tempdir) / "source.xlsx"),
                "headers": ["原文", "译文"],
                "rows_raw": [["Source", "Target"]],
                "target_col": 1,
                "segments": [
                    {"id": 0, "source": "Source", "target": "Target"}
                ],
                "wordcount": 1,
            }
            results = [
                {
                    "id": 0,
                    "errors": [
                        issue(
                            "Minor problem.",
                            severity="Minor",
                            needs_confirmation=True,
                        )
                    ],
                    "corrected": None,
                }
            ]
            lqe_io._build_xlsx(
                state,
                [{"iteration": 0, "errors": results}],
                99,
                98,
                output,
                reference_suggestions={0: "Must not be shown"},
            )

            workbook = openpyxl.load_workbook(output, data_only=True)
            try:
                sheet = workbook["LQE Results"]
                self.assertIsNone(sheet.cell(2, 4).value)
                self.assertEqual(
                    sheet.cell(2, 5).value,
                    "未生成建议，需人工处理",
                )
            finally:
                workbook.close()

    def test_reviewer_view_has_two_visible_sheets_and_exact_statuses(self):
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "review.xlsx"
            segments = [
                {"id": index, "source": f"Source {index}", "target": f"Target {index}"}
                for index in range(6)
            ]
            segments[4]["protected"] = True
            state = {
                "input_path": str(Path(tempdir) / "source.xlsx"),
                "headers": ["原文", "译文"],
                "rows_raw": [
                    [segment["source"], segment["target"]]
                    for segment in segments
                ],
                "target_col": 1,
                "segments": segments,
                "wordcount": 6,
            }
            safe_edit = {
                "from": "Target",
                "to": "Revised",
                "evidence": None,
            }
            results = [
                {
                    "id": 0,
                    "errors": [issue("Safe.", edit=safe_edit)],
                    "corrected": "Revised 0",
                },
                {
                    "id": 1,
                    "errors": [issue("Needs judgment.", needs_confirmation=True)],
                    "corrected": None,
                },
                {
                    "id": 2,
                    "errors": [
                        issue("Safe part.", edit=safe_edit),
                        issue("Open question.", needs_confirmation=True),
                    ],
                    "corrected": "Revised 2",
                },
                {
                    "id": 3,
                    "errors": [issue("No reliable rewrite.", needs_confirmation=True)],
                    "corrected": None,
                },
                {"id": 4, "errors": [], "corrected": None},
                {"id": 5, "errors": [], "corrected": None},
            ]
            history = [{"iteration": 0, "errors": results}]
            lqe_io._build_xlsx(
                state,
                history,
                99,
                98,
                output,
                reference_suggestions={1: "Reference 1"},
            )

            workbook = openpyxl.load_workbook(output, data_only=True)
            try:
                visible = [
                    sheet.title
                    for sheet in workbook.worksheets
                    if sheet.sheet_state == "visible"
                ]
                self.assertEqual(
                    visible,
                    ["说明·导读", "LQA Scorecard", "LQE Results"],
                )
                self.assertEqual(workbook.active.title, "说明·导读")
                guide = workbook["说明·导读"]
                guide_values = {
                    cell.value
                    for row in guide.iter_rows()
                    for cell in row
                    if cell.value is not None
                }
                for section in (
                    "三步读报告",
                    "LQA Scorecard 怎么读",
                    "审校区 10 列说明",
                    "建议状态说明",
                    "审校结论说明",
                    "阅读与交付提示",
                ):
                    self.assertIn(section, guide_values)

                results_sheet = workbook["LQE Results"]
                headers = [cell.value for cell in results_sheet[1]]
                self.assertEqual(
                    headers[:10],
                    [
                        "Segment ID",
                        "原文",
                        "原译",
                        "AI/建议译文",
                        "建议状态",
                        "错误类别",
                        "严重度",
                        "问题说明",
                        "审校结论",
                        "审校终稿或备注",
                    ],
                )
                rows = {
                    results_sheet.cell(row, 1).value: row
                    for row in range(2, results_sheet.max_row + 1)
                    if results_sheet.cell(row, 1).value is not None
                }
                expected_statuses = {
                    0: "可直接采用",
                    1: "建议待确认",
                    2: "部分修正，仍需确认",
                    3: "未生成建议，需人工处理",
                    4: "已保护",
                    5: None,
                }
                for segment_id, status in expected_statuses.items():
                    self.assertEqual(
                        results_sheet.cell(rows[segment_id], 5).value,
                        status,
                    )
                self.assertTrue(results_sheet.row_dimensions[rows[5]].hidden)
                hidden_issue_rows = [
                    row
                    for row in range(rows[2] + 1, rows[3])
                    if results_sheet.row_dimensions[row].hidden
                ]
                self.assertTrue(hidden_issue_rows)
                self.assertTrue(
                    all(
                        results_sheet.column_dimensions[
                            openpyxl.utils.get_column_letter(column)
                        ].hidden
                        for column in range(11, results_sheet.max_column + 1)
                    )
                )
                self.assertEqual(len(results_sheet.data_validations.dataValidation), 1)

                scorecard = workbook["LQA Scorecard"]
                detail_header = next(
                    row
                    for row in range(1, scorecard.max_row + 1)
                    if scorecard.cell(row, 1).value == "Segment ID"
                )
                self.assertFalse(scorecard.row_dimensions[detail_header].hidden)
                self.assertEqual(
                    [cell.value for cell in scorecard[detail_header]],
                    headers[:10],
                )
                self.assertFalse(
                    any(
                        scorecard.row_dimensions[row].hidden
                        for row in range(1, scorecard.max_row + 1)
                    )
                )
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
