import json
from pathlib import Path
import re
import subprocess
import sys
import tempfile
import unittest

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
IO_SCRIPT = SCRIPTS / "lqe_io.py"
CHUNK_SCRIPT = SCRIPTS / "lqe_chunk.py"
BATCH_SCRIPT = SCRIPTS / "lqe_batch.py"
CALC_SCRIPT = SCRIPTS / "lqe_calc.py"
FINALIZE_SCRIPT = SCRIPTS / "finalize_job.sh"

sys.path.insert(0, str(SCRIPTS))

from lqe_engine import (
    NO_TERMINOLOGY_REQUIRED_MODULES,
    OPTIONAL_MODULES,
    STANDARD_REQUIRED_MODULES,
    build_check_scope,
    disabled_modules,
    get_check_scope,
    load_terms,
    optional_modules,
    required_modules,
    scope_issue_problem,
    terminology_enabled,
    validate_scope_entries,
)


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def flatten_issues(entries):
    return [issue for entry in entries for issue in entry.get("issues", [])]


class NoTerminologyScopeTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.source = self.root / "source.csv"
        self.source.write_text("Source,Target\nA,Alpha\n", encoding="utf-8")
        self.terms = self.root / "project" / "terms.json"
        write_json(self.terms, [{"source": "A", "target": "Alpha"}])
        self.profile = self.root / "project" / "profile.json"
        write_json(
            self.profile,
            {
                "name": "fixture/zh-en",
                "language_pair": "zh-en",
                "source_lang": "zh",
                "target_lang": "en",
                "terminology": "terms.json",
            },
        )
        self.state = self.root / "job" / "state.json"

    def tearDown(self):
        self.tempdir.cleanup()

    def run_io(self, *args):
        return subprocess.run(
            [sys.executable, str(IO_SCRIPT), *map(str, args)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def test_build_check_scope_uses_fixed_module_contracts(self):
        self.assertEqual(
            build_check_scope(False),
            {
                "mode": "standard",
                "terminology_enabled": True,
                "enabled_modules": list(STANDARD_REQUIRED_MODULES),
                "disabled_modules": [],
                "source": "runtime",
            },
        )
        self.assertEqual(
            build_check_scope(True, "test"),
            {
                "mode": "no-terminology",
                "terminology_enabled": False,
                "enabled_modules": list(NO_TERMINOLOGY_REQUIRED_MODULES),
                "disabled_modules": ["terminology", "proper_names", "term_audit"],
                "source": "test",
            },
        )

    def test_profile_terms_are_not_loaded_when_disabled(self):
        result = self.run_io(
            "read",
            "--input",
            self.source,
            "--project",
            self.profile,
            "--source-col",
            "Source",
            "--target-col",
            "Target",
            "--no-terminology",
            "--out",
            self.state,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        state = read_json(self.state)
        self.assertEqual(state["terms_path"], "")
        self.assertEqual(state["terminology"], [])
        self.assertEqual(
            state["check_scope"]["enabled_modules"],
            ["precheck_review", "accuracy", "grammar", "naturalness"],
        )
        self.assertFalse((self.state.parent / "terms.json").exists())
        self.assertEqual(
            read_json(self.state.parent / "scope.json"), state["check_scope"]
        )


        self.assertIn(
            "profile terminology overridden by --no-terminology", result.stdout
        )

    def test_explicit_terms_conflict_with_no_terminology(self):
        result = self.run_io(
            "read",
            "--input",
            self.source,
            "--source-col",
            "Source",
            "--target-col",
            "Target",
            "--terminology",
            self.terms,
            "--no-terminology",
            "--out",
            self.state,
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("not allowed with argument", result.stderr)
        self.assertFalse(self.state.exists())

    def test_out_cannot_replace_reserved_scope_artifact(self):
        scope_path = self.state.parent / "scope.json"
        scope_path.parent.mkdir(parents=True)
        original = b'{"keep": "original"}\n'
        scope_path.write_bytes(original)

        result = self.run_io(
            "read",
            "--input",
            self.source,
            "--source-col",
            "Source",
            "--target-col",
            "Target",
            "--no-terminology",
            "--out",
            scope_path,
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn(
            "--out path conflicts with reserved scope artifact", result.stderr
        )
        self.assertEqual(scope_path.read_bytes(), original)
        self.assertEqual(
            {path.name for path in scope_path.parent.iterdir()}, {"scope.json"}
        )

    def test_out_case_variant_cannot_replace_reserved_scope_artifact(self):
        scope_path = self.state.parent / "scope.json"
        scope_path.parent.mkdir(parents=True)
        original = b'{"keep": "original"}\n'
        scope_path.write_bytes(original)

        result = self.run_io(
            "read",
            "--input",
            self.source,
            "--source-col",
            "Source",
            "--target-col",
            "Target",
            "--no-terminology",
            "--out",
            scope_path.with_name("SCOPE.JSON"),
        )

        self.assertEqual(result.returncode, 2)
        self.assertIn(
            "--out path conflicts with reserved scope artifact", result.stderr
        )
        self.assertEqual(scope_path.read_bytes(), original)
        self.assertEqual(
            {path.name for path in scope_path.parent.iterdir()}, {"scope.json"}
        )

    def test_load_terms_cannot_bypass_disabled_scope(self):
        state = {
            "check_scope": build_check_scope(True),
            "terms_path": str(self.terms),
            "terminology": [{"source": "A", "target": "B"}],
        }

        self.assertEqual(load_terms(state), [])

    def test_malformed_scope_fails_with_contract_error(self):
        with self.assertRaisesRegex(ValueError, "state must be an object"):
            get_check_scope([])
        with self.assertRaisesRegex(ValueError, "check_scope must be an object"):
            get_check_scope({"check_scope": []})

    def test_scope_mode_rejects_unknown_or_conflicting_values(self):
        with self.assertRaisesRegex(ValueError, "unsupported check_scope mode"):
            get_check_scope({"check_scope": {"mode": "custom"}})

        for key, value in (
            ("terminology_enabled", False),
            ("enabled_modules", ["accuracy"]),
            ("disabled_modules", ["terminology"]),
        ):
            with self.subTest(key=key):
                raw = build_check_scope(False)
                raw[key] = value
                with self.assertRaisesRegex(ValueError, f"check_scope {key} conflicts"):
                    get_check_scope({"check_scope": raw})

    def test_legacy_state_uses_standard_scope_without_rewrite(self):
        state = {"segments": []}

        self.assertEqual(
            required_modules(state),
            ("terminology", "accuracy", "grammar", "naturalness"),
        )
        self.assertEqual(optional_modules(state), OPTIONAL_MODULES)
        self.assertEqual(disabled_modules(state), ())
        self.assertTrue(terminology_enabled(state))
        self.assertEqual(get_check_scope(state)["source"], "legacy-default")
        self.assertNotIn("check_scope", state)

    def test_no_terminology_scope_accessors_are_consistent(self):
        state = {"check_scope": build_check_scope(True)}

        self.assertEqual(required_modules(state), NO_TERMINOLOGY_REQUIRED_MODULES)
        self.assertEqual(optional_modules(state), ())
        self.assertEqual(
            disabled_modules(state), ("terminology", "proper_names", "term_audit")
        )
        self.assertFalse(terminology_enabled(state))

    def test_scope_issue_problem_rejects_all_terminology_evidence(self):
        state = {"check_scope": build_check_scope(True)}
        cases = (
            (None, "issue must be an object"),
            ({"category": "Terminology"}, "Terminology issue is disabled"),
            ({"category": " terminology "}, "Terminology issue is disabled"),
            (
                {"category": "Other", "comment": "  TERM REVIEW: verify"},
                "TERM REVIEW evidence is disabled",
            ),
            (
                {"category": "Other", "comment": "  term review: verify"},
                "TERM REVIEW evidence is disabled",
            ),
            (
                {
                    "category": "Other",
                    "edit": {"evidence": {"type": "confirmed_term"}},
                },
                "confirmed_term edit evidence is disabled",
            ),
            (
                {
                    "category": "Other",
                    "edit": {"evidence": {"type": "Confirmed_Term"}},
                },
                "confirmed_term edit evidence is disabled",
            ),
            ({"category": "Unknown"}, "unsupported category"),
            (
                {"category": "Grammar", "severity": "Severe"},
                "unsupported severity",
            ),
        )
        for issue, message in cases:
            with self.subTest(message=message):
                self.assertIn(message, scope_issue_problem(state, issue))

        self.assertIsNone(
            scope_issue_problem(state, {"category": "Grammar", "comment": "Review"})
        )
        self.assertIsNone(scope_issue_problem({}, {"category": "Terminology"}))

    def test_validate_scope_entries_supports_check_and_final_shapes(self):
        state = {"check_scope": build_check_scope(True)}
        validate_scope_entries(
            state,
            [{"id": 0, "issues": [{"category": "Grammar"}]}],
            issues_key="issues",
            label="check",
        )

        for issues_key, issue in (
            ("issues", {"category": "Terminology"}),
            ("errors", {"comment": "TERM REVIEW: verify"}),
        ):
            with self.subTest(issues_key=issues_key):
                with self.assertRaisesRegex(
                    ValueError, rf"{issues_key}: scope conflict:"
                ):
                    validate_scope_entries(
                        state,
                        [{"id": 0, issues_key: [issue]}],
                        issues_key=issues_key,
                        label=issues_key,
                    )

        with self.assertRaisesRegex(ValueError, "unsupported issues key"):
            validate_scope_entries(state, [], issues_key="findings", label="bad")

    def test_standard_mode_still_loads_profile_terminology(self):
        result = self.run_io(
            "read",
            "--input",
            self.source,
            "--project",
            self.profile,
            "--source-col",
            "Source",
            "--target-col",
            "Target",
            "--out",
            self.state,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        state = read_json(self.state)
        self.assertTrue(terminology_enabled(state))
        self.assertTrue(load_terms(state))
        self.assertEqual(required_modules(state), STANDARD_REQUIRED_MODULES)
        self.assertEqual(state["check_scope"], build_check_scope(False))
        self.assertEqual(
            read_json(self.state.parent / "scope.json"), state["check_scope"]
        )


class NoTerminologyPrecheckTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.job = self.root / "job"
        self.state = self.job / "state.json"
        self.terms = self.job / "terms.json"
        self.read_no_term_job(
            [("生命值", "Health")],
            terms=[{"source": "生命值", "target": "HP", "confirmed": True}],
        )

    def tearDown(self):
        self.tempdir.cleanup()

    def read_no_term_job(self, rows, *, terms=None):
        terms = terms or []
        write_json(self.terms, terms)
        segments = [
            {"id": index, "source": source, "target": target}
            for index, (source, target) in enumerate(rows)
        ]
        write_json(
            self.state,
            {
                "source_lang": "zh",
                "target_lang": "en",
                "segments": segments,
                "check_scope": build_check_scope(True, "test"),
                "terms_path": str(self.terms),
                "terminology": terms,
            },
        )
        write_json(
            self.job / "errors_precheck.json",
            [{"id": segment["id"], "issues": []} for segment in segments],
        )

    def run_script(self, script, *args):
        return subprocess.run(
            [sys.executable, str(script), *map(str, args)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def run_precheck(self, *, out=None):
        return self.run_script(
            IO_SCRIPT,
            "pre-check",
            "--state",
            self.state,
            "--out",
            out or self.job / "errors_precheck.json",
        )

    def run_chunk(self, *args):
        return self.run_script(CHUNK_SCRIPT, *args)

    def run_batch(self, *args):
        return self.run_script(BATCH_SCRIPT, *args)

    def write_batch_eval(
        self,
        *,
        category="Grammar",
        severity=None,
        comment="check",
        edit=None,
        precheck_ref=None,
    ):
        segments = read_json(self.state)["segments"]
        rows = [{"id": segment["id"], "issues": []} for segment in segments]
        issue = {
            "category": category,
            "severity": severity or (
                "Major" if category == "Terminology" else "Minor"
            ),
            "comment": comment,
            "needs_confirmation": edit is None,
            "edit": edit,
        }
        if precheck_ref is not None:
            issue["precheck_ref"] = precheck_ref
        rows[0]["issues"] = [issue]
        write_json(self.job / "evals" / "eval_00.json", rows)

    def test_precheck_omits_terms_but_keeps_non_term_checks(self):
        self.read_no_term_job(
            [
                ("生命值", "Health"),
                ("<b>你好</b>", "Hello"),
                ("获得 100 金币 {0}", "Get 10 coins"),
                ("点击开始", "Click Start"),
                ("点击开始", "Tap Begin"),
            ],
            terms=[{"source": "生命值", "target": "HP", "confirmed": True}],
        )

        result = self.run_precheck()

        self.assertEqual(result.returncode, 0, result.stderr)
        issues = flatten_issues(read_json(self.job / "errors_precheck.json"))
        self.assertFalse(any(issue["category"] == "Terminology" for issue in issues))
        self.assertFalse(
            any(issue["comment"].startswith("TERM REVIEW:") for issue in issues)
        )
        self.assertTrue(any(issue["category"] == "Markup" for issue in issues))
        self.assertTrue(any(issue["category"] == "Inconsistency" for issue in issues))
        self.assertTrue(
            any("100" in issue["comment"] or "{0}" in issue["comment"] for issue in issues)
        )

    def test_precheck_rejects_scope_forbidden_custom_issue_before_write(self):
        checks = self.job / "checks.json"
        write_json(
            checks,
            {
                "custom": [
                    {
                        "id": "forbidden-term",
                        "pattern": "Health",
                        "where": "target",
                        "category": "Terminology",
                        "comment": "forbidden custom terminology issue",
                    }
                ]
            },
        )
        state = read_json(self.state)
        state["checks_path"] = str(checks)
        write_json(self.state, state)
        output = self.job / "forbidden_precheck.json"

        result = self.run_precheck(out=output)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)
        self.assertFalse(output.exists())

    def test_split_without_terms_writes_empty_term_context(self):
        result = self.run_chunk(
            "split",
            "--state",
            self.state,
            "--errors",
            self.job / "errors_precheck.json",
            "--outdir",
            self.job / "chunks",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        segment = read_json(self.job / "chunks" / "chunk_00.json")["segments"][0]
        self.assertEqual(segment["term_hits"], [])
        self.assertEqual(segment["term_near"], [])

    def test_split_and_batch_plan_assign_stable_precheck_reference(self):
        issue = {
            "category": "Markup",
            "severity": "Major",
            "comment": "machine finding",
            "needs_confirmation": True,
            "edit": None,
        }
        write_json(
            self.job / "errors_precheck.json",
            [{"id": 0, "issues": [issue]}],
        )

        refs = []
        for _ in range(2):
            result = self.run_chunk(
                "split",
                "--state",
                self.state,
                "--errors",
                self.job / "errors_precheck.json",
                "--outdir",
                self.job / "chunks",
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            reviewed = read_json(
                self.job / "chunks" / "chunk_00.json"
            )["segments"][0]["precheck"][0]
            refs.append(reviewed.get("precheck_ref"))
        self.assertIsInstance(refs[0], str)
        self.assertTrue(refs[0])
        self.assertEqual(refs[0], refs[1])

        result = self.run_batch("plan", "--job", self.job)
        self.assertEqual(result.returncode, 0, result.stderr)
        prompt = (self.job / "batches" / "batch_00.txt").read_text(
            encoding="utf-8"
        )
        self.assertIn(refs[0], prompt)

    def test_split_rejects_explicit_terms_in_disabled_scope(self):
        result = self.run_chunk(
            "split",
            "--state",
            self.state,
            "--errors",
            self.job / "errors_precheck.json",
            "--terms",
            self.terms,
            "--outdir",
            self.job / "chunks",
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)

    def test_batch_plan_ignores_residual_terms_file(self):
        write_json(self.job / "terms.json", [{"source": "生命值", "target": "HP"}])

        result = self.run_batch("plan", "--job", self.job)

        self.assertEqual(result.returncode, 0, result.stderr)
        prompt = (self.job / "batches" / "batch_00.txt").read_text(encoding="utf-8")
        self.assertNotIn("TERMS:", prompt)
        self.assertIn("Terminology check: disabled", prompt)

    def test_batch_merge_rejects_terminology_issue(self):
        self.write_batch_eval(category="Terminology", comment="forbidden")

        result = self.run_batch("merge", "--job", self.job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)
        self.assertFalse((self.job / "errors.json").exists())

    def test_batch_merge_rejects_other_disabled_term_evidence(self):
        cases = (
            {"comment": "TERM REVIEW: forbidden", "edit": None},
            {
                "comment": "forbidden evidence",
                "edit": {
                    "from": "Health",
                    "to": "HP",
                    "start": 0,
                    "end": 6,
                    "evidence": {
                        "type": "confirmed_term",
                        "source": "生命值",
                        "target": "HP",
                    },
                },
            },
        )
        for case in cases:
            with self.subTest(comment=case["comment"]):
                self.write_batch_eval(comment=case["comment"], edit=case["edit"])
                result = self.run_batch("merge", "--job", self.job)
                self.assertNotEqual(result.returncode, 0)
                self.assertIn("scope conflict", result.stderr)
                self.assertFalse((self.job / "errors.json").exists())

    def test_batch_merge_rejects_case_variants_and_unknown_schema_values(self):
        cases = (
            {"category": " terminology ", "comment": "forbidden"},
            {"category": "Grammar", "comment": " term review: forbidden"},
            {
                "category": "Grammar",
                "comment": "forbidden evidence",
                "edit": {
                    "from": "Health",
                    "to": "HP",
                    "start": 0,
                    "end": 6,
                    "evidence": {
                        "type": "Confirmed_Term",
                        "source": "生命值",
                        "target": "HP",
                    },
                },
            },
            {"category": "Unknown", "comment": "unknown category"},
            {
                "category": "Grammar",
                "severity": "Severe",
                "comment": "unknown severity",
            },
            {
                "category": "Grammar",
                "severity": "major",
                "comment": "non-canonical severity",
            },
        )
        for case in cases:
            with self.subTest(case=case):
                self.write_batch_eval(**case)
                result = self.run_batch("merge", "--job", self.job)
                self.assertNotEqual(result.returncode, 0)
                self.assertIn("scope conflict", result.stderr)
                self.assertFalse((self.job / "errors.json").exists())

    def test_split_does_not_deduplicate_different_per_id_prechecks(self):
        issue = {
            "category": "Length",
            "severity": "Major",
            "comment": "row-specific max length",
            "needs_confirmation": True,
            "edit": None,
        }
        for issue_id in (0, 1):
            with self.subTest(issue_id=issue_id):
                self.job = self.root / f"dedup-{issue_id}"
                self.state = self.job / "state.json"
                self.terms = self.job / "terms.json"
                self.read_no_term_job([("重复", "Same"), ("重复", "Same")])
                write_json(
                    self.job / "errors_precheck.json",
                    [
                        {
                            "id": segment_id,
                            "issues": [issue] if segment_id == issue_id else [],
                        }
                        for segment_id in (0, 1)
                    ],
                )

                result = self.run_chunk(
                    "split",
                    "--state",
                    self.state,
                    "--errors",
                    self.job / "errors_precheck.json",
                    "--outdir",
                    self.job / "chunks",
                )
                self.assertEqual(result.returncode, 0, result.stderr)
                chunk = read_json(self.job / "chunks" / "chunk_00.json")
                self.assertEqual(
                    [segment["id"] for segment in chunk["segments"]], [0, 1]
                )

                reviewed = next(
                    segment["precheck"][0]
                    for segment in chunk["segments"]
                    if segment["id"] == issue_id
                )
                for module in NO_TERMINOLOGY_REQUIRED_MODULES:
                    rows = [{"id": segment_id, "issues": []} for segment_id in (0, 1)]
                    if module == "precheck_review":
                        rows[issue_id]["issues"] = [reviewed]
                    write_json(
                        self.job / "chunks" / f"chunk_00.{module}.json", rows
                    )

                result = self.run_chunk("merge-checks", "--job", self.job)
                self.assertEqual(result.returncode, 0, result.stderr)
                result = self.run_chunk(
                    "merge",
                    "--state",
                    self.state,
                    "--errors",
                    self.job / "errors_precheck.json",
                    "--outdir",
                    self.job / "chunks",
                    "--out",
                    self.job / "errors.json",
                )
                self.assertEqual(result.returncode, 0, result.stderr)
                errors = {
                    entry["id"]: [item["category"] for item in entry["errors"]]
                    for entry in read_json(self.job / "errors.json")
                }
                self.assertEqual(errors[issue_id], ["Length"])
                self.assertEqual(errors[1 - issue_id], [])

    def test_batch_merge_rejects_unreferenced_precheck_issue(self):
        self.write_batch_eval(category="Markup", comment="invented")

        result = self.run_batch("merge", "--job", self.job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("precheck provenance", result.stderr)
        self.assertFalse((self.job / "errors.json").exists())

    def test_batch_merge_accepts_referenced_precheck_issue(self):
        write_json(
            self.job / "errors_precheck.json",
            [
                {
                    "id": 0,
                    "issues": [
                        {
                            "category": "Markup",
                            "severity": "Minor",
                            "comment": "machine finding",
                            "needs_confirmation": True,
                            "edit": None,
                        }
                    ],
                }
            ],
        )
        result = self.run_batch("plan", "--job", self.job)
        self.assertEqual(result.returncode, 0, result.stderr)
        prompt = (self.job / "batches" / "batch_00.txt").read_text(
            encoding="utf-8"
        )
        match = re.search(r"precheck_ref=([a-z0-9:_-]+)", prompt)
        self.assertIsNotNone(match, prompt)
        self.write_batch_eval(
            category="Markup",
            comment="confirmed and clarified",
            precheck_ref=match.group(1),
        )

        result = self.run_batch("merge", "--job", self.job)

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertTrue((self.job / "errors.json").exists())


class NoTerminologyModuleTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.job = self.root / "job"
        self.chunks = self.job / "chunks"
        self.chunks.mkdir(parents=True)

    def tearDown(self):
        self.tempdir.cleanup()

    def run_script(self, script, *args):
        return subprocess.run(
            [sys.executable, str(script), *map(str, args)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def run_chunk(self, *args):
        return self.run_script(CHUNK_SCRIPT, *args)

    def run_calc(self):
        return self.run_script(
            CALC_SCRIPT,
            "--state",
            self.job / "state.json",
            "--errors",
            self.job / "errors.json",
            "--json",
        )

    def make_chunk_job(self, *, no_terminology):
        state = {
            "wordcount": 1,
            "segments": [{"id": 0, "source": "A", "target": "A"}],
        }
        if no_terminology:
            state["check_scope"] = build_check_scope(True, "test")
        write_json(self.job / "state.json", state)
        write_json(
            self.job / "errors_precheck.json", [{"id": 0, "issues": []}]
        )
        write_json(
            self.chunks / "chunk_00.json",
            {
                "chunk_id": 0,
                "segments": [
                    {
                        "id": 0,
                        "source": "A",
                        "target": "A",
                        "kind": "name",
                        "term_hits": [],
                        "protected_texts": [],
                    }
                ],
            },
        )

    def make_no_term_chunk_job(self):
        self.make_chunk_job(no_terminology=True)

    def make_legacy_chunk_job(self):
        self.make_chunk_job(no_terminology=False)

    def write_modules(self, modules):
        for module in modules:
            write_json(
                self.chunks / f"chunk_00.{module}.json",
                [{"id": 0, "issues": []}],
            )

    def write_complete_no_term_errors(self, *, category, comment, edit):
        self.make_no_term_chunk_job()
        write_json(
            self.job / "errors.json",
            [
                {
                    "id": 0,
                    "errors": [
                        {
                            "category": category,
                            "severity": "Minor",
                            "comment": comment,
                            "needs_confirmation": False,
                            "edit": edit,
                        }
                    ],
                    "corrected": None,
                }
            ],
        )

    def test_validate_requires_precheck_review_not_terminology(self):
        self.make_no_term_chunk_job()
        self.write_modules(
            ("precheck_review", "accuracy", "grammar", "naturalness")
        )

        result = self.run_chunk("validate-checks", "--job", self.job)

        self.assertEqual(result.returncode, 0, result.stderr)

    def test_validate_fails_when_precheck_review_is_missing(self):
        self.make_no_term_chunk_job()
        self.write_modules(("accuracy", "grammar", "naturalness"))

        result = self.run_chunk("validate-checks", "--job", self.job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("precheck_review", result.stderr)

    def test_validate_rejects_disabled_module_file(self):
        self.make_no_term_chunk_job()
        self.write_modules(
            (
                "precheck_review",
                "accuracy",
                "grammar",
                "naturalness",
                "terminology",
            )
        )

        result = self.run_chunk("validate-checks", "--job", self.job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)

    def test_merge_checks_rejects_terminology_category(self):
        self.make_no_term_chunk_job()
        self.write_modules(
            ("precheck_review", "accuracy", "grammar", "naturalness")
        )
        write_json(
            self.chunks / "chunk_00.precheck_review.json",
            [
                {
                    "id": 0,
                    "issues": [
                        {
                            "category": "Terminology",
                            "severity": "Major",
                            "comment": "forbidden",
                            "needs_confirmation": True,
                            "edit": None,
                        }
                    ],
                }
            ],
        )

        result = self.run_chunk("merge-checks", "--job", self.job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)
        self.assertFalse((self.chunks / "chunk_00.out.json").exists())

    def test_merge_checks_enforces_precheck_review_category_ownership(self):
        self.make_no_term_chunk_job()
        self.write_modules(
            ("precheck_review", "accuracy", "grammar", "naturalness")
        )
        write_json(
            self.chunks / "chunk_00.precheck_review.json",
            [
                {
                    "id": 0,
                    "issues": [
                        {
                            "category": "Grammar",
                            "severity": "Minor",
                            "comment": "wrong owner",
                            "needs_confirmation": True,
                            "edit": None,
                        }
                    ],
                }
            ],
        )

        result = self.run_chunk("merge-checks", "--job", self.job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("precheck_review cannot own", result.stderr)
        self.assertFalse((self.chunks / "chunk_00.out.json").exists())

    def test_precheck_review_rejects_issue_when_chunk_precheck_is_empty(self):
        self.make_no_term_chunk_job()
        self.write_modules(
            ("precheck_review", "accuracy", "grammar", "naturalness")
        )
        write_json(
            self.chunks / "chunk_00.precheck_review.json",
            [
                {
                    "id": 0,
                    "issues": [
                        {
                            "category": "Markup",
                            "severity": "Major",
                            "comment": "invented issue",
                            "needs_confirmation": True,
                            "edit": None,
                        }
                    ],
                }
            ],
        )

        for command in ("validate-checks", "merge-checks"):
            with self.subTest(command=command):
                result = self.run_chunk(command, "--job", self.job)
                self.assertNotEqual(result.returncode, 0)
                self.assertIn("precheck provenance", result.stderr)
        self.assertFalse((self.chunks / "chunk_00.out.json").exists())

    def test_precheck_review_can_refine_existing_issue_without_new_edit(self):
        self.make_no_term_chunk_job()
        chunk_path = self.chunks / "chunk_00.json"
        chunk = read_json(chunk_path)
        chunk["segments"][0]["precheck"] = [
            {
                "precheck_ref": "precheck:0:test",
                "category": "Markup",
                "severity": "Minor",
                "comment": "machine finding",
                "needs_confirmation": True,
                "edit": None,
            }
        ]
        write_json(chunk_path, chunk)
        self.write_modules(
            ("precheck_review", "accuracy", "grammar", "naturalness")
        )
        write_json(
            self.chunks / "chunk_00.precheck_review.json",
            [
                {
                    "id": 0,
                    "issues": [
                        {
                            "precheck_ref": "precheck:0:test",
                            "category": "Markup",
                            "severity": "Major",
                            "comment": "confirmed and clarified",
                            "needs_confirmation": False,
                            "edit": None,
                        }
                    ],
                }
            ],
        )

        result = self.run_chunk("validate-checks", "--job", self.job)
        self.assertEqual(result.returncode, 0, result.stderr)
        result = self.run_chunk("merge-checks", "--job", self.job)
        self.assertEqual(result.returncode, 0, result.stderr)
        output = read_json(self.chunks / "chunk_00.out.json")
        self.assertEqual(output[0]["issues"][0]["comment"], "confirmed and clarified")

    def test_precheck_review_rejects_changed_edit_and_duplicate_claim(self):
        self.make_no_term_chunk_job()
        chunk_path = self.chunks / "chunk_00.json"
        chunk = read_json(chunk_path)
        chunk["segments"][0]["precheck"] = [
            {
                "precheck_ref": "precheck:0:test",
                "category": "Locale convention",
                "severity": "Minor",
                "comment": "machine finding",
                "needs_confirmation": False,
                "edit": {"from": "A", "to": "B", "evidence": None},
            }
        ]
        write_json(chunk_path, chunk)
        self.write_modules(
            ("precheck_review", "accuracy", "grammar", "naturalness")
        )
        cases = {
            "changed-edit": [
                {
                    "precheck_ref": "precheck:0:test",
                    "category": "Locale convention",
                    "severity": "Minor",
                    "comment": "changed correction",
                    "needs_confirmation": False,
                    "edit": {"from": "A", "to": "C", "evidence": None},
                }
            ],
            "duplicate-claim": [
                {
                    "precheck_ref": "precheck:0:test",
                    "category": "Locale convention",
                    "severity": "Minor",
                    "comment": comment,
                    "needs_confirmation": True,
                    "edit": None,
                }
                for comment in ("first", "second")
            ],
        }
        for name, issues in cases.items():
            with self.subTest(name=name):
                write_json(
                    self.chunks / "chunk_00.precheck_review.json",
                    [{"id": 0, "issues": issues}],
                )
                result = self.run_chunk("validate-checks", "--job", self.job)
                self.assertNotEqual(result.returncode, 0)
                self.assertIn("precheck provenance", result.stderr)

    def test_validate_enforces_all_module_category_ownership(self):
        cases = (
            ("grammar", "Markup"),
            ("naturalness", "Other"),
            ("accuracy", "Grammar"),
        )
        for module, category in cases:
            with self.subTest(module=module, category=category):
                self.make_no_term_chunk_job()
                self.write_modules(
                    ("precheck_review", "accuracy", "grammar", "naturalness")
                )
                write_json(
                    self.chunks / f"chunk_00.{module}.json",
                    [
                        {
                            "id": 0,
                            "issues": [
                                {
                                    "category": category,
                                    "severity": "Minor",
                                    "comment": "wrong owner",
                                    "needs_confirmation": True,
                                    "edit": None,
                                }
                            ],
                        }
                    ],
                )

                result = self.run_chunk("validate-checks", "--job", self.job)

                self.assertNotEqual(result.returncode, 0)
                self.assertIn(f"{module} cannot own", result.stderr)

    def test_validate_rejects_unknown_severity(self):
        self.make_no_term_chunk_job()
        self.write_modules(
            ("precheck_review", "accuracy", "grammar", "naturalness")
        )
        write_json(
            self.chunks / "chunk_00.grammar.json",
            [
                {
                    "id": 0,
                    "issues": [
                        {
                            "category": "Grammar",
                            "severity": "Severe",
                            "comment": "unknown severity",
                            "needs_confirmation": True,
                            "edit": None,
                        }
                    ],
                }
            ],
        )

        result = self.run_chunk("validate-checks", "--job", self.job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("unsupported severity", result.stderr)

    def test_legacy_state_still_requires_standard_four_modules(self):
        self.make_legacy_chunk_job()
        self.write_modules(("accuracy", "grammar", "naturalness"))

        result = self.run_chunk("validate-checks", "--job", self.job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("terminology", result.stderr)

    def test_merge_rejects_forbidden_module_issue_before_writing(self):
        self.make_no_term_chunk_job()
        write_json(
            self.chunks / "chunk_00.out.json",
            [
                {
                    "id": 0,
                    "issues": [
                        {
                            "category": "Other",
                            "severity": "Minor",
                            "comment": "forbidden",
                            "needs_confirmation": False,
                            "edit": {
                                "from": "A",
                                "to": "B",
                                "evidence": {
                                    "type": "confirmed_term",
                                    "source": "A",
                                    "target": "B",
                                },
                            },
                        }
                    ],
                }
            ],
        )
        output = self.job / "errors.json"

        result = self.run_chunk(
            "merge",
            "--state",
            self.job / "state.json",
            "--errors",
            self.job / "errors_precheck.json",
            "--outdir",
            self.chunks,
            "--out",
            output,
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)
        self.assertFalse(output.exists())

    def test_calc_rejects_forbidden_issue_before_scoring(self):
        self.write_complete_no_term_errors(
            category="Other",
            comment="forbidden",
            edit={
                "from": "A",
                "to": "B",
                "evidence": {"type": "confirmed_term"},
            },
        )

        result = self.run_calc()

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)

    def test_calc_zero_wordcount_with_penalized_error_is_not_full_score(self):
        self.make_no_term_chunk_job()
        state = read_json(self.job / "state.json")
        state["wordcount"] = 0
        state["segments"][0]["target"] = ""
        write_json(self.job / "state.json", state)
        write_json(
            self.job / "errors.json",
            [
                {
                    "id": 0,
                    "errors": [
                        {
                            "category": "Untranslated",
                            "severity": "Major",
                            "comment": "The target is empty.",
                            "needs_confirmation": True,
                            "edit": None,
                        }
                    ],
                    "corrected": None,
                }
            ],
        )

        result = self.run_calc()

        self.assertEqual(result.returncode, 0, result.stderr)
        score = json.loads(result.stdout)
        self.assertEqual(score["wordcount"], 0)
        self.assertEqual(score["errors"], 1)
        self.assertEqual(score["score"], 0)
        self.assertEqual(score["status"], "FAIL")


class NoTerminologyFinalizeTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.job = self.root / "job"
        self.state = self.job / "state.json"
        self.source = self.root / "source.xlsx"
        self.build_complete_no_term_job()

    def tearDown(self):
        self.tempdir.cleanup()

    def run_io(self, *args):
        return subprocess.run(
            [sys.executable, str(IO_SCRIPT), *map(str, args)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def run_finalize(self):
        return subprocess.run(
            ["bash", str(FINALIZE_SCRIPT), str(self.job), "1", "single"],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def build_complete_no_term_job(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Source", "Target"])
        sheet.append(["A", "Alpha"])
        workbook.save(self.source)
        workbook.close()

        result = self.run_io(
            "read",
            "--input",
            self.source,
            "--source-col",
            "Source",
            "--target-col",
            "Target",
            "--no-terminology",
            "--out",
            self.state,
        )
        self.assertEqual(result.returncode, 0, result.stderr)

        write_json(self.job / "errors_precheck.json", [{"id": 0, "issues": []}])
        write_json(
            self.job / "chunks" / "chunk_00.json",
            {
                "chunk_id": 0,
                "segments": [
                    {
                        "id": 0,
                        "source": "A",
                        "target": "Alpha",
                        "kind": "name",
                        "term_hits": [],
                        "protected_texts": [],
                    }
                ],
            },
        )
        for module in NO_TERMINOLOGY_REQUIRED_MODULES:
            write_json(
                self.job / "chunks" / f"chunk_00.{module}.json",
                [{"id": 0, "issues": []}],
            )

    def write_raw_checks(self, *, category="Terminology", comment="forbidden"):
        write_json(
            self.job / "checks.json",
            [
                {
                    "id": 0,
                    "issues": [
                        {
                            "category": category,
                            "severity": "Minor",
                            "comment": comment,
                            "needs_confirmation": True,
                            "edit": None,
                        }
                    ],
                }
            ],
        )

    def write_complete_no_term_errors(
        self,
        *,
        category="Terminology",
        comment="forbidden",
        edit=None,
        corrected=None,
    ):
        write_json(
            self.job / "errors.json",
            [
                {
                    "id": 0,
                    "errors": [
                        {
                            "category": category,
                            "severity": "Minor",
                            "comment": comment,
                            "needs_confirmation": edit is None,
                            "edit": edit,
                        }
                    ],
                    "corrected": corrected,
                }
            ],
        )

    def test_no_terminology_finalize_produces_report_without_term_files(self):
        result = self.run_finalize()

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertIn(
            "enabled modules: precheck_review, accuracy, grammar, naturalness",
            result.stdout,
        )
        self.assertTrue((self.job / "errors.json").exists())
        self.assertFalse(any(self.job.glob("chunks/*.terminology.json")))
        self.assertTrue(list(self.job.glob("*_lqe.xlsx")))
        self.assertTrue(list(self.job.glob("*_corrected.xlsx")))

    def test_report_records_disabled_terminology_and_enabled_modules(self):
        result = self.run_finalize()
        self.assertEqual(result.returncode, 0, result.stderr)
        report = next(self.job.glob("*_lqe.xlsx"))

        workbook = openpyxl.load_workbook(report, data_only=True)
        try:
            values = [
                str(cell.value)
                for sheet in workbook
                for row in sheet
                for cell in row
                if cell.value is not None
            ]
            scorecard = workbook["LQA Scorecard"]
            terminology_row = next(
                row
                for row in range(1, scorecard.max_row + 1)
                if scorecard.cell(row=row, column=1).value == "Terminology"
            )
            self.assertEqual(
                [
                    scorecard.cell(row=terminology_row, column=col).value
                    for col in range(3, 13)
                ],
                [0] * 10,
            )
        finally:
            workbook.close()

        joined = "\n".join(values)
        self.assertIn("Terminology check: Disabled by runtime request", joined)
        self.assertIn("precheck_review, accuracy, grammar, naturalness", joined)

    def test_standard_report_records_enabled_terminology(self):
        state = read_json(self.state)
        state["check_scope"] = build_check_scope(False, "test")
        write_json(self.state, state)
        write_json(
            self.job / "errors.json",
            [{"id": 0, "errors": [], "corrected": None}],
        )

        result = self.run_io(
            "write",
            "--state",
            self.state,
            "--errors",
            self.job / "errors.json",
            "--score",
            "100",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        workbook = openpyxl.load_workbook(
            next(self.job.glob("*_lqe.xlsx")), data_only=True
        )
        try:
            values = [
                str(cell.value)
                for sheet in workbook
                for row in sheet
                for cell in row
                if cell.value is not None
            ]
        finally:
            workbook.close()
        self.assertIn("Terminology check: Enabled", "\n".join(values))

    def test_write_rejects_forbidden_issue_before_state_or_workbook_write(self):
        self.write_complete_no_term_errors()
        original_state = self.state.read_bytes()

        result = self.run_io(
            "write",
            "--state",
            self.state,
            "--errors",
            self.job / "errors.json",
            "--score",
            "100",
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)
        self.assertEqual(self.state.read_bytes(), original_state)
        self.assertFalse(list(self.job.glob("*_lqe.xlsx")))

    def test_apply_fixes_rejects_forbidden_issue_before_state_or_workbook_write(self):
        self.write_complete_no_term_errors(
            edit={"from": "Alpha", "to": "Beta", "evidence": None},
            corrected="Beta",
        )
        original_state = self.state.read_bytes()

        result = self.run_io(
            "apply-fixes",
            "--state",
            self.state,
            "--errors",
            self.job / "errors.json",
            "--score",
            "90",
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)
        self.assertEqual(self.state.read_bytes(), original_state)
        self.assertFalse(list(self.job.glob("*_lqe_iter*.xlsx")))

    def test_build_results_rejects_forbidden_issue_before_output_write(self):
        self.write_raw_checks()
        output = self.job / "built-errors.json"

        result = self.run_io(
            "build-results",
            "--state",
            self.state,
            "--checks",
            self.job / "checks.json",
            "--out",
            output,
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)
        self.assertFalse(output.exists())

    def test_export_rejects_forbidden_issue_before_workbook_write(self):
        self.write_complete_no_term_errors()

        result = self.run_io(
            "export",
            "--state",
            self.state,
            "--errors",
            self.job / "errors.json",
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)
        self.assertFalse(list(self.job.glob("*_corrected.xlsx")))


if __name__ == "__main__":
    unittest.main()
