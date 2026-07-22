import json
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest
from unittest import mock

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
IO_SCRIPT = ROOT / "scripts" / "lqe_io.py"
CALC_SCRIPT = ROOT / "scripts" / "lqe_calc.py"
CHUNK_SCRIPT = ROOT / "scripts" / "lqe_chunk.py"
FINALIZE_SCRIPT = ROOT / "scripts" / "finalize_job.sh"
sys.path.insert(0, str(ROOT / "scripts"))

import lqe_paths
from lqe_scoring import resolve_scoring_policy, score_errors
from lqe_terms import TermContractError, canonicalize_terms


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def issue(
    *,
    severity: str = "Major",
    protected: object = None,
    with_edit: bool = False,
) -> dict:
    value = {
        "category": "Grammar",
        "severity": severity,
        "comment": "Hardening regression fixture.",
        "needs_confirmation": not with_edit,
        "edit": None,
    }
    if with_edit:
        value["edit"] = {
            "from": "Alpha",
            "to": "Beta",
            "start": 0,
            "end": 5,
            "evidence": None,
        }
    if protected is not None:
        value["protected"] = protected
    return value


class ReadInputAliasRegressionTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.source = self.root / "source.csv"
        self.source.write_text("Source,Target\nA,Alpha\n", encoding="utf-8")

    def tearDown(self):
        self.tempdir.cleanup()

    def run_read(self, out: Path, *extra: object) -> subprocess.CompletedProcess:
        return subprocess.run(
            [
                sys.executable,
                str(IO_SCRIPT),
                "read",
                "--input",
                str(self.source),
                "--source-col",
                "Source",
                "--target-col",
                "Target",
                "--out",
                str(out),
                *map(str, extra),
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def assert_rejected_without_publication(
        self,
        result: subprocess.CompletedProcess,
        input_path: Path,
        original: bytes,
    ) -> None:
        self.assertNotEqual(result.returncode, 0)
        self.assertEqual(input_path.read_bytes(), original)
        for name in ("scope.json", "sg.txt", "terms.json"):
            self.assertFalse((self.root / name).exists(), name)

    def test_read_rejects_out_aliasing_style_guide(self):
        guide = self.root / "guide.input.txt"
        guide.write_text("Original style guide.", encoding="utf-8")
        original = guide.read_bytes()

        result = self.run_read(
            guide,
            "--no-terminology",
            "--style-guide",
            guide,
        )

        self.assert_rejected_without_publication(result, guide, original)

    def test_read_rejects_out_aliasing_terminology(self):
        terms = self.root / "glossary.input.json"
        write_json(
            terms,
            [
                {
                    "source": "A",
                    "target": "Alpha",
                    "confirmed": True,
                    "protected": False,
                }
            ],
        )
        original = terms.read_bytes()

        result = self.run_read(terms, "--terminology", terms)

        self.assert_rejected_without_publication(result, terms, original)

    def test_read_rejects_out_aliasing_project_profile(self):
        profile = self.root / "profile.input.json"
        write_json(
            profile,
            {
                "name": "fixture/zh-en",
                "language_pair": "zh-en",
                "source_lang": "zh",
                "target_lang": "en",
            },
        )
        original = profile.read_bytes()

        result = self.run_read(
            profile,
            "--project",
            profile,
            "--no-terminology",
        )

        self.assert_rejected_without_publication(result, profile, original)


class PublicationTransactionRegressionTests(unittest.TestCase):
    def test_replace_then_base_exception_rolls_back_every_destination(self):
        with tempfile.TemporaryDirectory() as tempdir:
            root = Path(tempdir)
            replacements = []
            originals = {}
            for name in ("first.json", "second.json"):
                destination = root / name
                destination.write_text(f"old-{name}", encoding="utf-8")
                originals[destination] = destination.read_bytes()
                staged = root / f"{name}.staged"
                staged.write_text(f"new-{name}", encoding="utf-8")
                replacements.append((staged, destination))

            real_replace = lqe_paths._replace_staged

            def replace_then_interrupt(source: Path, destination: Path) -> None:
                real_replace(source, destination)
                if destination.name == "second.json":
                    raise KeyboardInterrupt("injected after successful replace")

            with mock.patch.object(
                lqe_paths,
                "_replace_staged",
                replace_then_interrupt,
            ):
                with self.assertRaises(KeyboardInterrupt):
                    lqe_paths.publish_replacement_transaction(replacements)

            for destination, original in originals.items():
                self.assertEqual(destination.read_bytes(), original)
            self.assertFalse(
                any(".rollback." in path.name for path in root.iterdir())
            )


class ArtifactBoundaryRegressionTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.state_path = self.root / "state.json"
        self.precheck_path = self.root / "errors_precheck.json"
        self.source_path = self.root / "source.xlsx"
        workbook = openpyxl.Workbook()
        workbook.active.append(["Source", "Target"])
        workbook.active.append(["A", "Alpha"])
        workbook.save(self.source_path)
        workbook.close()
        write_json(
            self.state_path,
            {
                "input_path": str(self.source_path),
                "source_col": "Source",
                "target_col": "Target",
                "headers": ["Source", "Target"],
                "segments": [
                    {"id": 0, "row_index": 0, "source": "A", "target": "Alpha"}
                ],
            },
        )

    def tearDown(self):
        self.tempdir.cleanup()

    def run_script(self, script: Path, *args: object) -> subprocess.CompletedProcess:
        return subprocess.run(
            [sys.executable, str(script), *map(str, args)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def test_precheck_cannot_overwrite_state(self):
        original = self.state_path.read_bytes()

        result = self.run_script(
            IO_SCRIPT,
            "pre-check",
            "--state",
            self.state_path,
            "--out",
            self.state_path,
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("conflicts with input", result.stderr)
        self.assertEqual(self.state_path.read_bytes(), original)

    def test_merge_cannot_overwrite_state_or_chunk_input(self):
        write_json(self.precheck_path, [{"id": 0, "issues": []}])
        split = self.run_script(
            CHUNK_SCRIPT,
            "split",
            "--state",
            self.state_path,
            "--errors",
            self.precheck_path,
            "--outdir",
            self.root / "chunks",
        )
        self.assertEqual(split.returncode, 0, split.stderr)
        write_json(
            self.root / "chunks" / "chunk_00.out.json",
            [{"id": 0, "issues": []}],
        )
        original = self.state_path.read_bytes()

        result = self.run_script(
            CHUNK_SCRIPT,
            "merge",
            "--state",
            self.state_path,
            "--errors",
            self.precheck_path,
            "--outdir",
            self.root / "chunks",
            "--out",
            self.state_path,
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("conflicts with input", result.stderr)
        self.assertEqual(self.state_path.read_bytes(), original)

    def test_protection_decision_cannot_overwrite_source(self):
        original = self.source_path.read_bytes()

        result = self.run_script(
            IO_SCRIPT,
            "protect-segments",
            "--state",
            self.state_path,
            "--protected-ids",
            "0",
            "--out",
            self.source_path,
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("conflicts with input", result.stderr)
        self.assertEqual(self.source_path.read_bytes(), original)

    def test_missing_chunks_still_enforces_name_edit_evidence(self):
        errors_path = self.root / "errors.json"
        write_json(
            errors_path,
            [
                {
                    "id": 0,
                    "errors": [issue(with_edit=True)],
                    "corrected": "Beta",
                }
            ],
        )

        result = self.run_script(
            IO_SCRIPT,
            "write",
            "--state",
            self.state_path,
            "--errors",
            errors_path,
            "--score",
            "99",
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("corrected mismatch", result.stderr)
        self.assertFalse(list(self.root.glob("*_lqe.xlsx")))

    def test_finalize_counts_chunk_indices_above_two_digits(self):
        chunks = self.root / "chunks"
        chunks.mkdir()
        for index in range(101):
            write_json(chunks / f"chunk_{index:02d}.json", {})
        (self.root / ".finalized").touch()

        result = subprocess.run(
            ["bash", str(FINALIZE_SCRIPT), str(self.root), "101", "single"],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertIn("ALREADY-FINALIZED", result.stdout)
        self.assertNotIn("INCOMPLETE", result.stdout)


class ScoringBoundaryRegressionTests(unittest.TestCase):
    def state(self, wordcount: object = 100) -> dict:
        return {
            "wordcount": wordcount,
            "segments": [{"id": 0, "source": "A", "target": "Alpha"}],
        }

    def errors(self, value: dict | None = None) -> list[dict]:
        return [
            {
                "id": 0,
                "errors": [] if value is None else [value],
                "corrected": None,
            }
        ]

    def test_nan_wordcount_is_rejected(self):
        state = self.state(float("nan"))

        with self.assertRaisesRegex(ValueError, "wordcount"):
            score_errors(
                state,
                self.errors(issue()),
                resolve_scoring_policy(state),
            )

    def test_infinite_wordcount_is_rejected(self):
        state = self.state(float("inf"))

        with self.assertRaisesRegex(ValueError, "wordcount"):
            score_errors(
                state,
                self.errors(issue()),
                resolve_scoring_policy(state),
            )

    def test_empty_list_scoring_policy_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "scoring_policy"):
            resolve_scoring_policy({"scoring_policy": []})

    def test_scalar_protected_statuses_is_rejected(self):
        with self.assertRaisesRegex(TermContractError, "protected_statuses"):
            canonicalize_terms(
                [
                    {
                        "source": "A",
                        "target": "Alpha",
                        "status": "Approved",
                    }
                ],
                term_status_map={"Approved": "confirmed"},
                protected_statuses="Approved",
            )


class ErrorLevelProtectionRegressionTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.state_path = self.root / "state.json"
        self.errors_path = self.root / "errors.json"
        self.source_path = self.root / "source.csv"
        self.source_path.write_text("Source,Target\nA,Alpha\n", encoding="utf-8")

    def tearDown(self):
        self.tempdir.cleanup()

    def state(self) -> dict:
        return {
            "input_path": str(self.source_path),
            "headers": ["Source", "Target"],
            "rows_raw": [["A", "Alpha"]],
            "source_col": "Source",
            "target_col": "Target",
            "source_lang": "en",
            "target_lang": "en",
            "wordcount": 100,
            "iteration": 0,
            "segments": [
                {
                    "id": 0,
                    "row_index": 0,
                    "source": "A",
                    "target": "Alpha",
                    "kind": "desc",
                    "term_hits": [],
                    "protected_texts": [],
                }
            ],
        }

    def errors(self, value: dict, corrected: str | None = None) -> list[dict]:
        return [{"id": 0, "errors": [value], "corrected": corrected}]

    def run_calc(self) -> subprocess.CompletedProcess:
        return subprocess.run(
            [
                sys.executable,
                str(CALC_SCRIPT),
                "--state",
                str(self.state_path),
                "--errors",
                str(self.errors_path),
                "--json",
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def run_io(self, command: str, *extra: object) -> subprocess.CompletedProcess:
        return subprocess.run(
            [
                sys.executable,
                str(IO_SCRIPT),
                command,
                "--state",
                str(self.state_path),
                "--errors",
                str(self.errors_path),
                *map(str, extra),
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def test_error_level_protected_segment_is_not_scored(self):
        write_json(self.state_path, self.state())
        write_json(
            self.errors_path,
            self.errors(issue(protected=True)),
        )

        result = self.run_calc()

        self.assertEqual(result.returncode, 0, result.stderr)
        score = json.loads(result.stdout)
        self.assertEqual(score["score"], 100.0)
        self.assertEqual(score["status"], "PASS")
        self.assertEqual(score["errors"], 0)
        self.assertEqual(score["critical"], 0)

    def test_error_level_protected_segment_is_excluded_from_report(self):
        write_json(self.state_path, self.state())
        write_json(
            self.errors_path,
            self.errors(issue(protected=True)),
        )

        result = self.run_io("write", "--score", 100)

        self.assertEqual(result.returncode, 0, result.stderr)
        report_path = next(self.root.glob("*_lqe.xlsx"))
        workbook = openpyxl.load_workbook(report_path, data_only=True)
        try:
            scorecard = workbook["LQA Scorecard"]
            self.assertEqual(scorecard["B10"].value, "PASS")
            self.assertEqual(scorecard["B11"].value, 100)
            total_row = next(
                row
                for row in range(1, scorecard.max_row + 1)
                if scorecard.cell(row=row, column=1).value == "TOTAL"
            )
            self.assertEqual(scorecard.cell(total_row, 11).value, 0)
            self.assertEqual(scorecard.cell(total_row, 12).value, 0)

            results = workbook["LQE Results"]
            headers = [cell.value for cell in results[1]]
            processing = results.cell(
                row=2,
                column=headers.index("处理方式") + 1,
            ).value
            self.assertEqual(processing, "已保护，不修改")
        finally:
            workbook.close()

    def test_error_level_protected_segment_is_not_corrected(self):
        write_json(self.state_path, self.state())
        write_json(
            self.errors_path,
            self.errors(
                issue(severity="Minor", protected=True, with_edit=True),
                corrected="Beta",
            ),
        )

        result = self.run_io("apply-fixes")

        self.assertEqual(result.returncode, 0, result.stderr)
        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        segment = state["segments"][0]
        self.assertEqual(segment["target"], "Alpha")
        self.assertNotIn("current_target", segment)
        self.assertEqual(state["iteration"], 0)
        self.assertNotIn("error_history", state)

    def test_non_boolean_error_level_protected_is_rejected(self):
        write_json(self.state_path, self.state())
        write_json(
            self.errors_path,
            self.errors(issue(protected="yes")),
        )

        result = self.run_calc()

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("protected", result.stderr.casefold())


if __name__ == "__main__":
    unittest.main()
