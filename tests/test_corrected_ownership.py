import json
import os
import subprocess
import sys
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from io import StringIO
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

import openpyxl
from openpyxl.styles import PatternFill


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
CHUNK_SCRIPT = SCRIPTS / "lqe_chunk.py"
IO_SCRIPT = SCRIPTS / "lqe_io.py"
BATCH_SCRIPT = SCRIPTS / "lqe_batch.py"
MASTERTB_SCRIPT = SCRIPTS / "mastertb_prep.py"
FINALIZE_SCRIPT = SCRIPTS / "finalize_job.sh"
AGGREGATE_SCRIPT = SCRIPTS / "aggregate_sheets.py"
REQUIRED_MODULES = ("terminology", "accuracy", "grammar", "naturalness")

sys.path.insert(0, str(SCRIPTS))

from lqe_checks import run_pre_check
import aggregate_sheets
import lqe_chunk
import lqe_io


def write_json(path, value):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


def check_issue(
    category="Grammar",
    severity="Minor",
    comment="check",
    *,
    needs_confirmation=False,
    edit=None,
):
    return {
        "category": category,
        "severity": severity,
        "comment": comment,
        "needs_confirmation": needs_confirmation,
        "edit": edit,
    }


def replacement(frm, to, *, evidence=None, start=None, end=None):
    value = {"from": frm, "to": to, "evidence": evidence}
    if start is not None or end is not None:
        value["start"] = start
        value["end"] = end
    return value


def confirmed_evidence(source, target):
    return {"type": "confirmed_term", "source": source, "target": target}


def call_quietly(function, *args, **kwargs):
    with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
        return function(*args, **kwargs)


class CorrectedOwnershipChunkTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.job = Path(self.tempdir.name) / "job"
        self.chunks = self.job / "chunks"
        self.chunks.mkdir(parents=True)

    def tearDown(self):
        self.tempdir.cleanup()

    def run_chunk(self, *args):
        return subprocess.run(
            [sys.executable, str(CHUNK_SCRIPT), *map(str, args)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def make_job(self, segments, *, representatives=None, dedup_map=None, pre=None):
        write_json(self.job / "state.json", {"segments": segments})
        if pre is None:
            pre = [{"id": segment["id"], "issues": []} for segment in segments]
        write_json(self.job / "errors_precheck.json", pre)

        chunk_segments = representatives if representatives is not None else segments
        write_json(
            self.chunks / "chunk_00.json",
            {"chunk_id": 0, "segments": chunk_segments},
        )
        if dedup_map is None:
            dedup_map = {str(segment["id"]): [segment["id"]] for segment in chunk_segments}
        write_json(self.chunks / "dedup_map.json", dedup_map)

    def write_modules(self, ids, *, issues_by_module=None):
        issues_by_module = issues_by_module or {}
        for module in REQUIRED_MODULES:
            rows = [
                {
                    "id": segment_id,
                    "issues": issues_by_module.get(module, {}).get(segment_id, []),
                }
                for segment_id in ids
            ]
            write_json(self.chunks / f"chunk_00.{module}.json", rows)

    def merge_checks(self):
        result = self.run_chunk("merge-checks", "--job", self.job)
        self.assertEqual(result.returncode, 0, result.stderr)
        return read_json(self.chunks / "chunk_00.out.json")

    def merge_final(self):
        output = self.job / "errors.json"
        result = self.run_chunk(
            "merge",
            "--state",
            self.job / "state.json",
            "--errors",
            self.job / "errors_precheck.json",
            "--outdir",
            self.chunks,
            "--out",
            output,
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        return read_json(output)

    def test_validate_checks_rejects_top_level_corrected(self):
        self.make_job([{"id": 0, "source": "x", "target": "x"}])
        self.write_modules([0])
        path = self.chunks / "chunk_00.terminology.json"
        write_json(path, [{"id": 0, "issues": [], "corrected": "model text"}])

        result = self.run_chunk("validate-checks", "--job", self.job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("corrected is not allowed", result.stderr)

    def test_ckpt_append_rejects_top_level_corrected(self):
        checkpoint = self.chunks / "chunk_00.terminology.ckpt.jsonl"
        entry = json.dumps(
            {"id": 0, "issues": [], "corrected": "model text"},
            ensure_ascii=False,
        )

        result = self.run_chunk(
            "ckpt-append", "--file", checkpoint, "--entry", entry
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("corrected is not allowed", result.stderr)
        self.assertFalse(checkpoint.exists())

    def test_merge_checks_rejects_top_level_corrected(self):
        self.make_job([{"id": 0, "source": "x", "target": "x"}])
        self.write_modules([0])
        path = self.chunks / "chunk_00.naturalness.json"
        write_json(path, [{"id": 0, "issues": [], "corrected": "model text"}])

        result = self.run_chunk("merge-checks", "--job", self.job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("corrected is not allowed", result.stderr)
        self.assertFalse((self.chunks / "chunk_00.out.json").exists())

    def test_validate_checks_requires_modules_and_complete_ids(self):
        segments = [
            {"id": 0, "source": "a", "target": "a"},
            {"id": 1, "source": "b", "target": "b"},
        ]
        self.make_job(segments)
        self.write_modules([0, 1])
        (self.chunks / "chunk_00.grammar.json").unlink()
        write_json(
            self.chunks / "chunk_00.terminology.json",
            [{"id": 0, "issues": []}],
        )

        result = self.run_chunk("validate-checks", "--job", self.job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("grammar", result.stderr)
        self.assertIn("1", result.stderr)

    def test_merge_checks_unions_deduplicates_and_filters_accuracy_owned_issues(self):
        self.make_job([{"id": 0, "source": "x", "target": "x"}])
        punctuation = check_issue(
            category="Punctuation",
            comment="same",
            edit=replacement("x", "X"),
        )
        unconfirmed_accuracy = check_issue(
            category="Mistranslation",
            severity="Major",
            comment="terminology guessed",
            needs_confirmation=True,
        )
        accuracy = check_issue(
            category="Omission",
            severity="Major",
            comment="accuracy confirmed",
            needs_confirmation=True,
        )
        self.write_modules(
            [0],
            issues_by_module={
                "terminology": {0: [punctuation, unconfirmed_accuracy]},
                "accuracy": {0: [accuracy]},
                "grammar": {0: [punctuation]},
            },
        )

        merged = self.merge_checks()

        self.assertEqual(merged, [{"id": 0, "issues": [punctuation, accuracy]}])
        self.assertNotIn("corrected", merged[0])
        self.assertNotIn("errors", merged[0])

    def test_accuracy_owned_issue_uses_only_accuracy_module_edit(self):
        segment = {
            "id": 0,
            "source": "source",
            "target": "bad",
            "kind": "desc",
            "term_hits": [],
            "protected_texts": [],
        }
        terminology_issue = check_issue(
            category="Mistranslation",
            severity="Major",
            comment="terminology rewrite",
            edit=replacement("bad", "term guess"),
        )
        accuracy_issue = check_issue(
            category="Mistranslation",
            severity="Major",
            comment="accuracy rewrite",
            edit=replacement("bad", "accurate"),
        )
        self.make_job([segment])
        self.write_modules(
            [0],
            issues_by_module={
                "terminology": {0: [terminology_issue]},
                "accuracy": {0: [accuracy_issue]},
            },
        )

        checks = self.merge_checks()
        merged = self.merge_final()

        self.assertEqual(checks, [{"id": 0, "issues": [accuracy_issue]}])
        self.assertEqual(merged[0]["errors"], [accuracy_issue])
        self.assertEqual(merged[0]["corrected"], "accurate")

    def test_precheck_emits_issues_and_a_local_deterministic_edit(self):
        state_path = self.job / "state.json"
        output_path = self.job / "errors_precheck.json"
        write_json(
            state_path,
            {
                "source_lang": "en",
                "target_lang": "en",
                "segments": [{"id": 0, "source": "A-B", "target": "A—B"}],
            },
        )

        call_quietly(run_pre_check, state_path, output_path)
        result = read_json(output_path)

        self.assertEqual(set(result[0]), {"id", "issues"})
        em_dash = next(
            issue for issue in result[0]["issues"] if "Em dash" in issue["comment"]
        )
        self.assertFalse(em_dash["needs_confirmation"])
        self.assertEqual(
            em_dash["edit"],
            {
                "from": "—",
                "to": " - ",
                "start": 1,
                "end": 2,
                "evidence": None,
            },
        )

    def test_em_dash_edit_absorbs_adjacent_spaces(self):
        variants = ("A—B", "A — B", "A— B", "A —B")
        for target in variants:
            with self.subTest(target=target):
                state_path = self.job / "state.json"
                output_path = self.job / "errors_precheck.json"
                write_json(
                    state_path,
                    {
                        "source_lang": "en",
                        "target_lang": "en",
                        "segments": [
                            {"id": 0, "source": "A-B", "target": target}
                        ],
                    },
                )
                call_quietly(run_pre_check, state_path, output_path)
                issues = read_json(output_path)[0]["issues"]

                result = lqe_chunk.build_segment_result(
                    {
                        "id": 0,
                        "target": target,
                        "kind": "desc",
                        "term_hits": [],
                        "protected_texts": [],
                    },
                    issues,
                )

                self.assertEqual(result["corrected"], "A - B")

    def test_precheck_issue_schema_is_complete_for_all_output_modes(self):
        cases = (
            ("empty", "Text", ""),
            ("reminder", "Hello {name}", "Hello"),
            ("double_space", "A B", "A  B"),
            ("fullwidth", "A,B", "A，B"),
        )
        required = {
            "category",
            "severity",
            "comment",
            "needs_confirmation",
            "edit",
        }
        for label, source, target in cases:
            with self.subTest(case=label):
                state_path = self.job / "state.json"
                output_path = self.job / "errors_precheck.json"
                write_json(
                    state_path,
                    {
                        "source_lang": "en",
                        "target_lang": "en",
                        "segments": [
                            {"id": 0, "source": source, "target": target}
                        ],
                    },
                )

                call_quietly(run_pre_check, state_path, output_path)
                issues = read_json(output_path)[0]["issues"]

                self.assertTrue(issues)
                for issue in issues:
                    self.assertTrue(required.issubset(issue))
                    self.assertTrue(
                        issue["edit"] is None or isinstance(issue["edit"], dict)
                    )

    def test_locked_segment_still_uses_correction_builder(self):
        segment = {
            "id": 0,
            "source": "protected",
            "target": "original",
            "locked": True,
            "kind": "desc",
            "term_hits": [],
            "protected_texts": ["original"],
        }
        self.make_job([segment])
        output = self.job / "errors.json"

        with mock.patch.object(
            lqe_chunk,
            "build_segment_result",
            wraps=lqe_chunk.build_segment_result,
        ) as builder:
            call_quietly(
                lqe_chunk.cmd_merge,
                SimpleNamespace(
                    state=str(self.job / "state.json"),
                    errors=str(self.job / "errors_precheck.json"),
                    outdir=str(self.chunks),
                    out=str(output),
                ),
            )

        builder.assert_called_once()
        called_segment, called_issues = builder.call_args.args
        self.assertEqual(called_segment["id"], 0)
        self.assertEqual(called_issues, [])
        self.assertEqual(
            read_json(output),
            [{"id": 0, "errors": [], "corrected": None}],
        )

    def test_term_hits_flatten_senses_and_preserve_explicit_confirmation(self):
        from lqe_chunk import _term_hits

        hits = _term_hits(
            "魔草巫灵",
            [
                (
                    "魔草巫灵",
                    [
                        {"target": "Verdrel", "confirmed": True},
                        {"target": "Reference", "confirmed": False},
                    ],
                )
            ],
        )

        self.assertEqual(
            hits,
            [
                {"source": "魔草巫灵", "target": "Verdrel", "confirmed": True},
                {
                    "source": "魔草巫灵",
                    "target": "Reference",
                    "confirmed": False,
                },
            ],
        )

    def test_unconfirmed_names_do_not_generate_new_translations(self):
        cases = [
            ("深蓝鲸", "ธาลูน", "Thalune"),
            ("伊里斯", "Iris", "Irys"),
            ("星光狮", "Old Lion", "Blitzmane"),
        ]
        segments = []
        terminology = {}
        for segment_id, (source, original, proposed) in enumerate(cases):
            segments.append(
                {
                    "id": segment_id,
                    "source": source,
                    "target": original,
                    "kind": "name",
                    "term_hits": [
                        {
                            "source": source,
                            "target": proposed,
                            "confirmed": False,
                        }
                    ],
                    "protected_texts": [],
                }
            )
            terminology[segment_id] = [
                check_issue(
                    category="Terminology",
                    severity="Major",
                    comment=f"{source} proposed name",
                    edit=replacement(
                        original,
                        proposed,
                        evidence=confirmed_evidence(source, proposed),
                    ),
                )
            ]
        self.make_job(segments)
        self.write_modules(range(len(segments)), issues_by_module={"terminology": terminology})

        self.merge_checks()
        merged = self.merge_final()

        self.assertEqual([row["corrected"] for row in merged], [None, None, None])
        for row in merged:
            self.assertTrue(row["errors"][0]["needs_confirmation"])
            self.assertIsNone(row["errors"][0]["edit"])

    def test_term_replacement_requires_confirmed_true(self):
        terms = [
            ("魔草巫灵", "Old Verdrel", "Verdrel"),
            ("奇丽花", "Old Florora", "Florora"),
        ]
        segments = []
        terminology = {}
        expected = []
        segment_id = 0
        for source, original, target in terms:
            for confirmed in (False, True):
                segments.append(
                    {
                        "id": segment_id,
                        "source": source,
                        "target": original,
                        "kind": "name",
                        "term_hits": [
                            {
                                "source": source,
                                "target": target,
                                "confirmed": confirmed,
                            }
                        ],
                        "protected_texts": [],
                    }
                )
                terminology[segment_id] = [
                    check_issue(
                        category="Terminology",
                        severity="Major",
                        comment="use confirmed term",
                        edit=replacement(
                            original,
                            target,
                            evidence=confirmed_evidence(source, target),
                        ),
                    )
                ]
                expected.append(target if confirmed else None)
                segment_id += 1
        self.make_job(segments)
        self.write_modules(range(len(segments)), issues_by_module={"terminology": terminology})

        self.merge_checks()
        merged = self.merge_final()

        self.assertEqual([row["corrected"] for row in merged], expected)

    def test_conflicting_flower_butterfly_edits_do_not_land(self):
        source = "花衣蝶登场"
        original = "Old Butterfly appears"
        first = "Sprigella"
        second = "Hanagoromia"
        segment = {
            "id": 0,
            "source": source,
            "target": original,
            "kind": "desc",
            "term_hits": [
                {
                    "source": "花衣蝶",
                    "target": first,
                    "confirmed": True,
                    "matched_text": "Old Butterfly",
                },
                {
                    "source": "花衣蝶",
                    "target": second,
                    "confirmed": True,
                    "matched_text": "Old Butterfly",
                },
            ],
            "protected_texts": [],
        }
        first_issue = check_issue(
            category="Terminology",
            severity="Major",
            comment="first name",
            edit=replacement(
                "Old Butterfly",
                first,
                evidence=confirmed_evidence("花衣蝶", first),
            ),
        )
        second_issue = check_issue(
            category="Terminology",
            severity="Major",
            comment="second name",
            edit=replacement(
                "Old Butterfly",
                second,
                evidence=confirmed_evidence("花衣蝶", second),
            ),
        )
        self.make_job([segment])
        self.write_modules(
            [0],
            issues_by_module={
                "terminology": {0: [first_issue]},
                "proper_names": {0: [second_issue]},
            },
        )
        write_json(
            self.chunks / "chunk_00.proper_names.json",
            [{"id": 0, "issues": [second_issue]}],
        )

        self.merge_checks()
        merged = self.merge_final()

        self.assertIsNone(merged[0]["corrected"])
        self.assertEqual(len(merged[0]["errors"]), 2)
        self.assertTrue(
            all(issue["needs_confirmation"] for issue in merged[0]["errors"])
        )
        self.assertTrue(all(issue["edit"] is None for issue in merged[0]["errors"]))

    def test_mixed_row_only_applies_safe_edit(self):
        original = "สไตล์ฟลอโรรา- รองเท้า"
        segment = {
            "id": 0,
            "source": "奇丽花印象-鞋子",
            "target": original,
            "kind": "desc",
            "term_hits": [
                {"source": "奇丽花", "target": "Florora", "confirmed": False}
            ],
            "protected_texts": [],
        }
        pending_name = check_issue(
            category="Terminology",
            severity="Major",
            comment="proper name needs confirmation",
            needs_confirmation=True,
        )
        spacing = check_issue(
            category="Punctuation",
            comment="space before hyphen",
            edit=replacement("-", " -"),
        )
        self.make_job([segment])
        self.write_modules(
            [0],
            issues_by_module={"terminology": {0: [pending_name]}, "grammar": {0: [spacing]}},
        )

        self.merge_checks()
        merged = self.merge_final()

        self.assertEqual(merged[0]["corrected"], "สไตล์ฟลอโรรา - รองเท้า")
        self.assertTrue(
            any(issue["needs_confirmation"] for issue in merged[0]["errors"])
        )

    def test_duplicate_broadcast_keeps_errors_and_generated_corrected(self):
        segments = [
            {"id": 0, "source": "same", "target": "bad-text"},
            {"id": 1, "source": "same", "target": "bad-text"},
        ]
        representative = {
            "id": 0,
            "source": "same",
            "target": "bad-text",
            "kind": "desc",
            "term_hits": [],
            "protected_texts": [],
        }
        safe = check_issue(
            category="Punctuation",
            comment="replace hyphen",
            edit=replacement("-", " "),
        )
        self.make_job(
            segments,
            representatives=[representative],
            dedup_map={"0": [0, 1]},
        )
        self.write_modules([0], issues_by_module={"grammar": {0: [safe]}})

        self.merge_checks()
        merged = self.merge_final()

        self.assertEqual([row["id"] for row in merged], [0, 1])
        self.assertEqual([row["corrected"] for row in merged], ["bad text", "bad text"])
        self.assertEqual(merged[0]["errors"], merged[1]["errors"])

    def test_reconcile_writes_only_id_and_issues(self):
        self.make_job([{"id": 0, "source": "x", "target": "x"}])
        accuracy_issue = check_issue(
            category="Mistranslation",
            severity="Major",
            comment="confirmed by accuracy",
            edit=replacement("x", "accurate"),
        )
        dropped_issue = check_issue(
            category="Mistranslation",
            severity="Major",
            comment="other module guess",
            edit=replacement("x", "guess"),
        )
        self.write_modules([0], issues_by_module={"accuracy": {0: [accuracy_issue]}})
        write_json(
            self.chunks / "chunk_00.out.json",
            [{"id": 0, "issues": [accuracy_issue, dropped_issue]}],
        )

        result = self.run_chunk("reconcile", "--job", self.job)

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(
            read_json(self.chunks / "chunk_00.out.json"),
            [{"id": 0, "issues": [accuracy_issue]}],
        )

    def test_split_half_uses_module_checkpoint_name(self):
        self.make_job(
            [
                {"id": 0, "source": "a", "target": "a"},
                {"id": 1, "source": "b", "target": "b"},
            ]
        )
        checkpoint = self.chunks / "chunk_00.grammar.ckpt.jsonl"
        checkpoint.write_text(
            json.dumps({"id": 0, "issues": []})
            + "\n"
            + json.dumps({"id": 1, "issues": []})
            + "\n",
            encoding="utf-8",
        )

        result = self.run_chunk(
            "split-half",
            "--job",
            self.job,
            "--chunk",
            "0",
            "--module",
            "grammar",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertTrue((self.chunks / "chunk_00_p1.grammar.ckpt.jsonl").exists())
        self.assertTrue((self.chunks / "chunk_00_p2.grammar.ckpt.jsonl").exists())

    def test_old_check_command_names_and_lens_flag_are_removed(self):
        source = CHUNK_SCRIPT.read_text(encoding="utf-8")

        self.assertNotIn("merge-lenses", source)
        self.assertNotIn("validate-lenses", source)
        self.assertNotIn('"--lens"', source)


class CorrectedOwnershipPipelineTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def run_script(self, script, *args):
        return subprocess.run(
            [sys.executable, str(script), *map(str, args)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def make_aggregate_job(self, *, errors, workbook=None, segments=None, name="multi"):
        job = self.root / name
        sheet_job = job / "Sheet1"
        source = self.root / f"{name}-source.xlsx"
        workbook = workbook or openpyxl.Workbook()
        worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
        worksheet.title = "Sheet1"
        if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None:
            worksheet["A1"] = "Source"
            worksheet["B1"] = "Target"
            worksheet["A2"] = "原文"
            worksheet["B2"] = "原译"
        workbook.save(source)
        segments = segments or [
            {
                "id": 0,
                "row_index": 0,
                "source": "原文",
                "target": "原译",
                "kind": "desc",
                "term_hits": [],
                "protected_texts": [],
            }
        ]
        write_json(
            sheet_job / "state.json",
            {
                "input_path": str(source),
                "target_col": 1,
                "headers": ["Source", "Target"],
                "wordcount": 1,
                "segments": segments,
            },
        )
        write_json(sheet_job / "errors.json", errors)
        return job, source

    def run_aggregate(self, job):
        return self.run_script(AGGREGATE_SCRIPT, "--job", job)

    def test_build_results_is_required_before_apply(self):
        state = self.root / "state.json"
        checks = self.root / "checks.json"
        errors = self.root / "errors.json"
        write_json(
            state,
            {
                "segments": [
                    {
                        "id": 0,
                        "source": "ABC",
                        "target": "A“B”C",
                        "kind": "desc",
                        "term_hits": [],
                        "protected_texts": [],
                    }
                ]
            },
        )
        write_json(
            checks,
            [
                {
                    "id": 0,
                    "issues": [
                        check_issue(edit=replacement("“B”", '\"B\"'))
                    ],
                }
            ],
        )

        result = self.run_script(
            IO_SCRIPT,
            "build-results",
            "--state",
            state,
            "--checks",
            checks,
            "--out",
            errors,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(read_json(errors)[0]["corrected"], 'A"B"C')

    def test_build_results_rejects_missing_and_extra_check_ids(self):
        state = self.root / "state.json"
        write_json(
            state,
            {
                "segments": [
                    {"id": 0, "source": "a", "target": "a"},
                    {"id": 1, "source": "b", "target": "b"},
                ]
            },
        )
        cases = {
            "missing": ([{"id": 0, "issues": []}], "missing=[1]"),
            "extra": (
                [
                    {"id": 0, "issues": []},
                    {"id": 1, "issues": []},
                    {"id": 2, "issues": []},
                ],
                "extra=[2]",
            ),
        }
        for name, (entries, message) in cases.items():
            with self.subTest(name=name):
                checks = self.root / f"{name}.json"
                errors = self.root / f"{name}-errors.json"
                write_json(checks, entries)

                result = self.run_script(
                    IO_SCRIPT,
                    "build-results",
                    "--state",
                    state,
                    "--checks",
                    checks,
                    "--out",
                    errors,
                )

                self.assertNotEqual(result.returncode, 0)
                self.assertIn(message, result.stderr)
                self.assertFalse(errors.exists())

    def test_batch_merge_rejects_model_corrected(self):
        job = self.root / "batch"
        write_json(
            job / "state.json",
            {"segments": [{"id": 0, "source": "x", "target": "x"}]},
        )
        write_json(
            job / "evals" / "eval_00.json",
            [{"id": 0, "issues": [], "corrected": "model text"}],
        )

        result = self.run_script(BATCH_SCRIPT, "merge", "--job", job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("corrected", result.stderr)

    def test_mastertb_merge_rejects_model_corrected(self):
        job = self.root / "mastertb"
        write_json(
            job / "state.json",
            {"segments": [{"id": 0, "source": "x", "target": "x"}]},
        )
        write_json(job / "chunks" / "chunk_00.json", {"terms": []})
        write_json(
            job / "chunks" / "chunk_00.out.json",
            {
                "findings": [
                    {"id": 0, "issues": [], "corrected": "model text"}
                ]
            },
        )

        result = self.run_script(
            MASTERTB_SCRIPT,
            "merge",
            "--job-dir",
            job,
            "--no-consistency",
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("corrected", result.stderr)

    def test_aggregate_applies_only_non_null_program_result(self):
        job, _ = self.make_aggregate_job(
            errors=[{"id": 0, "errors": [], "corrected": None}],
            segments=[
                {
                    "id": 0,
                    "row_index": 0,
                    "source": "原文",
                    "target": "原译",
                    "corrected": "旧状态修正",
                    "kind": "desc",
                    "term_hits": [],
                    "protected_texts": [],
                }
            ],
        )

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        output = openpyxl.load_workbook(job / f"{job.name}_corrected.xlsx")
        try:
            self.assertEqual(output["Sheet1"]["B2"].value, "原译")
        finally:
            output.close()

    def test_aggregate_rejects_forged_corrected_before_writing_outputs(self):
        job, _ = self.make_aggregate_job(
            errors=[{"id": 0, "errors": [], "corrected": "伪造修正"}]
        )
        outputs = [job / "multi_corrected.xlsx", job / "multi_LQE报告.xlsx"]
        for output in outputs:
            output.write_bytes(b"existing output")

        result = self.run_aggregate(job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("corrected", result.stderr)
        for output in outputs:
            self.assertEqual(output.read_bytes(), b"existing output")

    def test_aggregate_requires_errors_field(self):
        job, _ = self.make_aggregate_job(
            errors=[{"id": 0, "corrected": None}],
        )

        result = self.run_aggregate(job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("errors", result.stderr)
        self.assertFalse((job / "multi_corrected.xlsx").exists())
        self.assertFalse((job / "multi_LQE报告.xlsx").exists())

    def test_aggregate_requires_exact_result_id_coverage(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["Source", "Target"])
        sheet.append(["one", "one"])
        sheet.append(["two", "two"])
        segments = [
            {"id": 0, "row_index": 0, "source": "one", "target": "one"},
            {"id": 1, "row_index": 1, "source": "two", "target": "two"},
        ]
        cases = {
            "missing": (
                [{"id": 0, "errors": [], "corrected": None}],
                "missing=[1]",
            ),
            "extra": (
                [
                    {"id": 0, "errors": [], "corrected": None},
                    {"id": 1, "errors": [], "corrected": None},
                    {"id": 2, "errors": [], "corrected": None},
                ],
                "extra=[2]",
            ),
        }
        for name, (errors, message) in cases.items():
            with self.subTest(name=name):
                job, _ = self.make_aggregate_job(
                    errors=errors,
                    workbook=workbook,
                    segments=segments,
                    name=name,
                )

                result = self.run_aggregate(job)

                self.assertNotEqual(result.returncode, 0)
                self.assertIn(message, result.stderr)
                self.assertFalse((job / f"{name}_corrected.xlsx").exists())
                self.assertFalse((job / f"{name}_LQE报告.xlsx").exists())

    def test_aggregate_preserves_complete_source_workbook(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["Source", "Target", "Formula", "Merged", None])
        sheet.append(["原文", "原译", "=LEN(B2)", "keep", None])
        sheet.merge_cells("D1:E1")
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 24
        sheet["B2"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
        notes = workbook.create_sheet("Notes")
        notes["A1"] = "untouched"
        error = check_issue(
            edit=replacement("原", "新"),
        )
        job, source = self.make_aggregate_job(
            errors=[{"id": 0, "errors": [error], "corrected": "新译"}],
            workbook=workbook,
        )

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        original = openpyxl.load_workbook(source, data_only=False)
        output = openpyxl.load_workbook(job / "multi_corrected.xlsx", data_only=False)
        try:
            self.assertEqual(output.sheetnames, original.sheetnames)
            self.assertEqual(output["Sheet1"]["B2"].value, "新译")
            self.assertEqual(
                output["Sheet1"]["B2"].style_id,
                original["Sheet1"]["B2"].style_id,
            )
            self.assertEqual(output["Sheet1"]["C2"].value, "=LEN(B2)")
            self.assertEqual(
                list(output["Sheet1"].merged_cells.ranges),
                list(original["Sheet1"].merged_cells.ranges),
            )
            self.assertEqual(output["Sheet1"].column_dimensions["A"].width, 18)
            self.assertEqual(output["Sheet1"].column_dimensions["B"].width, 24)
            self.assertEqual(output["Notes"]["A1"].value, "untouched")
        finally:
            original.close()
            output.close()

    def test_finalize_uses_only_new_check_commands(self):
        syntax = subprocess.run(
            ["bash", "-n", str(FINALIZE_SCRIPT)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )
        self.assertEqual(syntax.returncode, 0, syntax.stderr)

        job = self.root / "finalize"
        (job / "chunks").mkdir(parents=True)
        write_json(job / "state.json", {})
        write_json(job / "chunks" / "chunk_00.json", {})
        log = self.root / "python-calls.log"
        bin_dir = self.root / "bin"
        bin_dir.mkdir()
        python_stub = bin_dir / "python3"
        python_stub.write_text(
            """#!/bin/sh
printf '%s\\n' "$*" >> "$CALL_LOG"
case "$1" in
  -) printf '98\\n' ;;
  *lqe_calc.py) printf '{"score":100,"status":"PASS","errors":0,"wordcount":1,"critical":0,"npt":0}\\n' ;;
  -c)
    case "$2" in
      *score*) printf '100\\n' ;;
      *status*) printf 'PASS\\n' ;;
    esac
    ;;
esac
""",
            encoding="utf-8",
        )
        python_stub.chmod(0o755)
        env = dict(os.environ)
        env["CALL_LOG"] = str(log)
        env["PATH"] = f"{bin_dir}{os.pathsep}{env['PATH']}"

        result = subprocess.run(
            ["bash", str(FINALIZE_SCRIPT), str(job), "1", "single"],
            cwd=ROOT,
            text=True,
            capture_output=True,
            env=env,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        calls = [
            line.split()
            for line in log.read_text(encoding="utf-8").splitlines()
            if "lqe_chunk.py" in line
        ]
        self.assertEqual(
            [call[1] for call in calls],
            ["validate-checks", "merge-checks", "reconcile", "merge"],
        )
        for call in calls[:3]:
            self.assertEqual(call[2:4], ["--job", str(job)])
        self.assertEqual(
            calls[3][2:],
            [
                "--state", str(job / "state.json"),
                "--errors", str(job / "errors_precheck.json"),
                "--outdir", str(job / "chunks"),
                "--out", str(job / "errors.json"),
            ],
        )


class CorrectedOwnershipOutputTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def run_io(self, *args):
        return subprocess.run(
            [sys.executable, str(IO_SCRIPT), *map(str, args)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def make_export_fixture(self, name, segments, errors):
        job = self.root / name
        job.mkdir()
        source = job / "source.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Source", "Target"])
        for segment in segments:
            sheet.append([segment["source"], segment["target"]])
        workbook.save(source)
        workbook.close()
        state = job / "state.json"
        results = job / "errors.json"
        write_json(
            state,
            {
                "input_path": str(source),
                "source_col": "Source",
                "target_col": 1,
                "headers": ["Source", "Target"],
                "segments": segments,
            },
        )
        write_json(results, errors)
        return state, results, job / f"{name}_corrected.xlsx"

    def test_export_rejects_forged_corrected_before_touching_output(self):
        segments = [
            {"id": 0, "row_index": 0, "source": "Source", "target": "原译"}
        ]
        state, errors, output = self.make_export_fixture(
            "forged-export",
            segments,
            [{"id": 0, "errors": [], "corrected": "伪造修正"}],
        )
        output.write_bytes(b"existing output")

        result = self.run_io("export", "--state", state, "--errors", errors)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("corrected", result.stderr)
        self.assertEqual(output.read_bytes(), b"existing output")

    def test_export_requires_exact_result_id_coverage_before_touching_output(self):
        segments = [
            {"id": 0, "row_index": 0, "source": "one", "target": "one"},
            {"id": 1, "row_index": 1, "source": "two", "target": "two"},
        ]
        cases = {
            "missing": (
                [{"id": 0, "errors": [], "corrected": None}],
                "missing=[1]",
            ),
            "extra": (
                [
                    {"id": 0, "errors": [], "corrected": None},
                    {"id": 1, "errors": [], "corrected": None},
                    {"id": 2, "errors": [], "corrected": None},
                ],
                "extra=[2]",
            ),
        }
        for name, (entries, message) in cases.items():
            with self.subTest(name=name):
                state, errors, output = self.make_export_fixture(
                    f"{name}-export", segments, entries
                )
                output.write_bytes(b"existing output")

                result = self.run_io("export", "--state", state, "--errors", errors)

                self.assertNotEqual(result.returncode, 0)
                self.assertIn(message, result.stderr)
                self.assertEqual(output.read_bytes(), b"existing output")

    def test_apply_fixes_does_not_apply_or_advance_for_protected_row(self):
        protected = check_issue(edit=replacement("原译", "篡改"))
        protected["protected"] = True
        segments = [
            {
                "id": 0,
                "row_index": 0,
                "source": "Source",
                "target": "原译",
                "kind": "desc",
                "term_hits": [],
                "protected_texts": [],
            }
        ]
        state, errors, _ = self.make_export_fixture(
            "protected-apply",
            segments,
            [{"id": 0, "errors": [protected], "corrected": "篡改"}],
        )
        state_data = read_json(state)
        state_data["rows_raw"] = [["Source", "原译"]]
        state_data["wordcount"] = 1
        state_data["iteration"] = 0
        write_json(state, state_data)

        result = self.run_io("apply-fixes", "--state", state, "--errors", errors)

        self.assertEqual(result.returncode, 0, result.stderr)
        updated = read_json(state)
        self.assertIsNone(updated["segments"][0].get("corrected"))
        self.assertEqual(updated["iteration"], 0)

    def test_export_protects_rows_but_applies_mixed_safe_correction(self):
        protected = check_issue(edit=replacement("原译一", "篡改"))
        protected["protected"] = True
        confirmation = check_issue(needs_confirmation=True)
        safe = check_issue(edit=replacement("原译二", "安全修改"))
        segments = [
            {"id": 0, "row_index": 0, "source": "one", "target": "原译一"},
            {"id": 1, "row_index": 1, "source": "two", "target": "原译二"},
        ]
        state, errors, output = self.make_export_fixture(
            "protected-export",
            segments,
            [
                {"id": 0, "errors": [protected], "corrected": "篡改"},
                {
                    "id": 1,
                    "errors": [confirmation, safe],
                    "corrected": "安全修改",
                },
            ],
        )

        result = self.run_io("export", "--state", state, "--errors", errors)

        self.assertEqual(result.returncode, 0, result.stderr)
        workbook = openpyxl.load_workbook(output)
        try:
            self.assertEqual(workbook.active["B2"].value, "原译一")
            self.assertEqual(workbook.active["B3"].value, "安全修改")
        finally:
            workbook.close()

    @staticmethod
    def workbook_signature(path):
        workbook = openpyxl.load_workbook(path, data_only=False)
        try:
            return {
                "sheetnames": workbook.sheetnames,
                "dimensions": {
                    sheet.title: sheet.calculate_dimension() for sheet in workbook
                },
                "headers": {
                    sheet.title: [cell.value for cell in sheet[1]]
                    for sheet in workbook
                },
            }
        finally:
            workbook.close()

    def test_report_uses_plain_processing_labels(self):
        output = self.root / "sample_lqe.xlsx"
        targets = ["原译一", "原译二", "原译三", "原译四", "原译五"]
        state = {
            "input_path": str(self.root / "sample.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [
                [f"Source {index}", target]
                for index, target in enumerate(targets)
            ],
            "segments": [
                {
                    "id": index,
                    "source": f"Source {index}",
                    "target": target,
                    "kind": "desc",
                }
                for index, target in enumerate(targets)
            ],
            "wordcount": len(targets),
        }
        protected = check_issue()
        protected["protected"] = True
        history = [
            {
                "iteration": 0,
                "errors": [
                    {
                        "id": 0,
                        "errors": [check_issue()],
                        "corrected": "建议译文一",
                    },
                    {
                        "id": 1,
                        "errors": [check_issue(needs_confirmation=True)],
                        "corrected": "安全修改二",
                    },
                    {"id": 2, "errors": [check_issue()], "corrected": None},
                    {"id": 3, "errors": [], "corrected": None},
                    {"id": 4, "errors": [protected], "corrected": None},
                ],
            }
        ]

        call_quietly(lqe_io._build_xlsx, state, history, 99, 98, output)

        workbook = openpyxl.load_workbook(output)
        try:
            headers = [cell.value for cell in workbook["LQE Results"][1]]
            self.assertIn("原译", headers)
            self.assertIn("建议译文", headers)
            self.assertIn("处理方式", headers)
            processing_col = headers.index("处理方式") + 1
            labels = [
                workbook["LQE Results"].cell(row=row, column=processing_col).value
                for row in range(2, 7)
            ]
            self.assertEqual(
                labels,
                ["建议修改", "需要人工确认", "仅提醒", "无需修改", "已保护，不修改"],
            )
        finally:
            workbook.close()

    def test_processing_label_uses_only_errors_and_corrected(self):
        processing_label = getattr(lqe_io, "_processing_label", None)
        self.assertIsNotNone(processing_label)
        protected = check_issue()
        protected["protected"] = True
        cases = [
            ({"errors": [protected], "corrected": "ignored"}, "已保护，不修改"),
            (
                {
                    "errors": [check_issue(needs_confirmation=True)],
                    "corrected": "安全修改",
                },
                "需要人工确认",
            ),
            ({"errors": [check_issue()], "corrected": "建议译文"}, "建议修改"),
            ({"errors": [check_issue()], "corrected": None}, "仅提醒"),
            ({"errors": [], "corrected": None}, "无需修改"),
        ]
        for entry, expected in cases:
            with self.subTest(expected=expected):
                self.assertEqual(processing_label(entry), expected)

    def test_errors_without_corrected_are_valid(self):
        issues = lqe_io._validate_errors(
            [{"id": 0, "errors": [check_issue()], "corrected": None}],
            {0},
        )

        self.assertFalse(
            any("corrected=null" in issue for issue in issues),
            issues,
        )

    def test_corrected_workbook_has_source_structure(self):
        source = self.root / "sample.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Content"
        sheet.append(["Source", "Target", "Formula"])
        sheet.append(["Source", "原译", "=LEN(B2)"])
        sheet.append(["Reminder", "保留原译", "=LEN(B3)"])
        notes = workbook.create_sheet("Notes")
        notes["A1"] = "untouched"
        workbook.save(source)
        workbook.close()
        state_path = self.root / "state.json"
        errors_path = self.root / "errors.json"
        write_json(
            state_path,
            {
                "input_path": str(source),
                "source_col": "Source",
                "target_col": 1,
                "headers": ["Source", "Target", "Formula"],
                "segments": [
                    {
                        "id": 0,
                        "row_index": 0,
                        "source": "Source",
                        "target": "原译",
                    },
                    {
                        "id": 1,
                        "row_index": 1,
                        "source": "Reminder",
                        "target": "保留原译",
                    },
                ],
            },
        )
        write_json(
            errors_path,
            [
                {
                    "id": 0,
                    "errors": [
                        check_issue(edit=replacement("原译", "建议译文"))
                    ],
                    "corrected": "建议译文",
                },
                {"id": 1, "errors": [check_issue()], "corrected": None},
            ],
        )
        before = self.workbook_signature(source)

        call_quietly(
            lqe_io.cmd_export,
            SimpleNamespace(state=str(state_path), errors=str(errors_path)),
        )

        output = next(self.root.glob("*_corrected.xlsx"))
        after = self.workbook_signature(output)
        self.assertEqual(after["sheetnames"], before["sheetnames"])
        self.assertEqual(after["dimensions"], before["dimensions"])
        self.assertEqual(after["headers"], before["headers"])
        output_workbook = openpyxl.load_workbook(output, data_only=False)
        try:
            self.assertEqual(output_workbook["Content"]["B2"].value, "建议译文")
            self.assertEqual(output_workbook["Content"]["B3"].value, "保留原译")
        finally:
            output_workbook.close()

    def test_export_prints_only_four_plain_counts(self):
        source = self.root / "counts.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Source", "Target"])
        for index in range(4):
            sheet.append([f"Source {index}", f"原译{index}"])
        workbook.save(source)
        workbook.close()
        state_path = self.root / "counts-state.json"
        errors_path = self.root / "counts-errors.json"
        segments = [
            {
                "id": index,
                "row_index": index,
                "source": f"Source {index}",
                "target": f"原译{index}",
            }
            for index in range(4)
        ]
        segments[3]["locked"] = True
        write_json(
            state_path,
            {
                "input_path": str(source),
                "source_col": "Source",
                "target_col": 1,
                "headers": ["Source", "Target"],
                "segments": segments,
            },
        )
        write_json(
            errors_path,
            [
                {
                    "id": 0,
                    "errors": [
                        check_issue(edit=replacement("原译0", "建议译文"))
                    ],
                    "corrected": "建议译文",
                },
                {
                    "id": 1,
                    "errors": [
                        check_issue(needs_confirmation=True),
                        check_issue(edit=replacement("原译1", "安全修改")),
                    ],
                    "corrected": "安全修改",
                },
                {"id": 2, "errors": [], "corrected": None},
                {"id": 3, "errors": [], "corrected": None},
            ],
        )
        stdout = StringIO()

        with redirect_stdout(stdout):
            lqe_io.cmd_export(
                SimpleNamespace(state=str(state_path), errors=str(errors_path))
            )

        summary = stdout.getvalue()
        for expected in [
            "建议修改 1",
            "需要人工确认 1",
            "保持原译 1",
            "已保护 1",
        ]:
            self.assertIn(expected, summary)
        for legacy in ["AI修正", "人工批准", "待人工裁决", "未改", "TM保护"]:
            self.assertNotIn(legacy, summary)


if __name__ == "__main__":
    unittest.main()
