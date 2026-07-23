import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "aggregate_sheets.py"
CHUNK_SCRIPT = ROOT / "scripts" / "lqe_chunk.py"
sys.path.insert(0, str(ROOT / "scripts"))

import aggregate_sheets
import lqe_io
from lqe_report_contract import attach_report_contract
from lqe_paths import file_sha256
from lqe_result_contract import build_result_contract


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False), encoding="utf-8")


def issue(*, severity: str, category: str = "Grammar") -> dict:
    return {
        "category": category,
        "severity": severity,
        "comment": "Scoring fixture.",
        "needs_confirmation": True,
        "edit": None,
    }


class AggregateScoringPolicyTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def make_job(self, policies, child_issues):
        job = self.root / "multi"
        source = self.root / "source.xlsx"
        workbook = openpyxl.Workbook()
        for index, name in enumerate(("Sheet1", "Sheet2")):
            sheet = workbook.active if index == 0 else workbook.create_sheet()
            sheet.title = name
            sheet.append(["Source", "Target"])
            sheet.append([f"Source {index}", f"Target {index}"])
        workbook.save(source)
        workbook.close()

        for index, name in enumerate(("Sheet1", "Sheet2")):
            child = job / name
            state = {
                "input_path": str(source),
                "sheet_name": name,
                "target_col": 1,
                "headers": ["Source", "Target"],
                "wordcount": 1000,
                "scoring_policy": policies[index],
                "segments": [
                    {
                        "id": 0,
                        "row_index": 0,
                        "source": f"Source {index}",
                        "target": f"Target {index}",
                        "kind": "desc",
                        "term_hits": [],
                        "protected_texts": [],
                    }
                ],
            }
            errors = [
                {
                    "id": 0,
                    "errors": child_issues[index],
                    "corrected": None,
                }
            ]
            write_json(child / "state.json", state)
            write_json(child / "errors.json", errors)
            report = openpyxl.Workbook()
            results = report.active
            results.title = "LQE Results"
            results.append(["原文", "原译", "建议译文", "处理方式"])
            results.append(
                [f"Source {index}", f"Target {index}", None, "无需修改"]
            )
            attach_report_contract(report, state, errors)
            report.save(child / f"{name}_lqe.xlsx")
            report.close()
            self.split_child(child)
        return job

    def refresh_report_contracts(self, job):
        for child in job.iterdir():
            if not child.is_dir():
                continue
            self.split_child(child)
            report_path = child / f"{child.name}_lqe.xlsx"
            if not report_path.is_file():
                continue
            try:
                workbook = openpyxl.load_workbook(report_path)
            except Exception:
                continue
            try:
                attach_report_contract(
                    workbook,
                    json.loads((child / "state.json").read_text(encoding="utf-8")),
                    json.loads((child / "errors.json").read_text(encoding="utf-8")),
                )
                workbook.save(report_path)
            finally:
                workbook.close()

    def run_aggregate(self, job, *extra, refresh_contracts=True):
        if refresh_contracts:
            self.refresh_report_contracts(job)
        return subprocess.run(
            [sys.executable, str(SCRIPT), "--job", str(job), *map(str, extra)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def split_child(self, child):
        write_json(child / "errors_precheck.json", [{"id": 0, "issues": []}])
        result = subprocess.run(
            [
                sys.executable,
                str(CHUNK_SCRIPT),
                "split",
                "--state",
                str(child / "state.json"),
                "--errors",
                str(child / "errors_precheck.json"),
                "--outdir",
                str(child / "chunks"),
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )
        self.assertEqual(result.returncode, 0, result.stderr)

    def summary_rows(self, job):
        workbook = openpyxl.load_workbook(job / "multi_lqe.xlsx", data_only=True)
        try:
            return [
                tuple(cell.value for cell in row)
                for row in workbook["汇总"].iter_rows()
            ]
        finally:
            workbook.close()

    def test_omitted_threshold_inherits_each_child_policy(self):
        policy = {
            "threshold": 99.95,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[issue(severity="Minor")], []])

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        rows = self.summary_rows(job)
        sheet1 = next(row for row in rows if row[0] == "Sheet1")
        overall = next(row for row in rows if row[0] == "合计")
        self.assertEqual(sheet1[6], "FAIL")
        self.assertEqual(overall[6], "FAIL")
        self.assertIn("阈值 99.95", rows[1][0])

    def test_any_critical_gate_child_failure_forces_overall_failure(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": True,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[issue(severity="Critical")], []])

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        rows = self.summary_rows(job)
        overall = next(row for row in rows if row[0] == "合计")
        self.assertGreaterEqual(overall[5], 98)
        self.assertEqual(overall[6], "FAIL")
        self.assertIn("(FAIL)", result.stdout)

    def test_non_threshold_policy_mismatch_fails_closed(self):
        lisa = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        mqm = {**lisa, "severity_scale": "mqm"}
        job = self.make_job([lisa, mqm], [[], []])

        result = self.run_aggregate(job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scoring policy mismatch", result.stderr)
        self.assertFalse((job / "multi_corrected.xlsx").exists())
        self.assertFalse((job / "multi_lqe.xlsx").exists())

    def test_explicit_threshold_overrides_threshold_but_keeps_critical_gate(self):
        policy = {
            "threshold": 100,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": True,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[issue(severity="Critical")], []])

        result = self.run_aggregate(job, "--threshold", "98")

        self.assertEqual(result.returncode, 0, result.stderr)
        rows = self.summary_rows(job)
        sheet1 = next(row for row in rows if row[0] == "Sheet1")
        overall = next(row for row in rows if row[0] == "合计")
        self.assertEqual(sheet1[6], "FAIL")
        self.assertEqual(overall[6], "FAIL")
        self.assertIn("阈值 98", rows[1][0])

    def test_aggregate_delivery_keeps_accumulated_current_target(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        state_path = job / "Sheet1" / "state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        state["segments"][0]["current_target"] = "Applied target"
        write_json(state_path, state)

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        workbook = openpyxl.load_workbook(job / "multi_corrected.xlsx")
        try:
            self.assertEqual(workbook["Sheet1"]["B2"].value, "Applied target")
        finally:
            workbook.close()

    def test_aggregate_reuses_verified_chunk_terminology_evidence(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        child = job / "Sheet1"
        state_path = child / "state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        state["terminology"] = [
            {
                "source": "Source 0",
                "target": "Canonical target",
                "confirmed": True,
                "protected": False,
            }
        ]
        write_json(state_path, state)
        self.split_child(child)
        write_json(
            child / "errors.json",
            [
                {
                    "id": 0,
                    "errors": [
                        {
                            "category": "Terminology",
                            "severity": "Major",
                            "comment": "Use the confirmed term.",
                            "needs_confirmation": False,
                            "edit": {
                                "from": "Target 0",
                                "to": "Canonical target",
                                "evidence": {
                                    "type": "confirmed_term",
                                    "source": "Source 0",
                                    "target": "Canonical target",
                                },
                            },
                        }
                    ],
                    "corrected": "Canonical target",
                }
            ],
        )

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        workbook = openpyxl.load_workbook(job / "multi_corrected.xlsx")
        try:
            self.assertEqual(
                workbook["Sheet1"]["B2"].value,
                "Canonical target",
            )
        finally:
            workbook.close()

    def test_aggregate_uses_exact_weighted_penalty_without_npt_round_trip(self):
        policy = {
            "threshold": 99.99505,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job(
            [policy, policy],
            [[issue(severity="Minor")], []],
        )
        first_state = job / "Sheet1" / "state.json"
        state = json.loads(first_state.read_text(encoding="utf-8"))
        state["wordcount"] = 30600
        write_json(first_state, state)
        second_state = job / "Sheet2" / "state.json"
        state = json.loads(second_state.read_text(encoding="utf-8"))
        state["wordcount"] = 0
        write_json(second_state, state)

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        rows = self.summary_rows(job)
        child = next(row for row in rows if row[0] == "Sheet1")
        overall = next(row for row in rows if row[0] == "合计")
        self.assertEqual(child[6], "PASS")
        self.assertEqual(overall[6], "PASS")

    def test_zero_wordcount_without_penalty_is_100_pass(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        for name in ("Sheet1", "Sheet2"):
            state_path = job / name / "state.json"
            state = json.loads(state_path.read_text(encoding="utf-8"))
            state["wordcount"] = 0
            write_json(state_path, state)

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        overall = next(
            row for row in self.summary_rows(job) if row[0] == "合计"
        )
        self.assertEqual(overall[5], 100)
        self.assertEqual(overall[6], "PASS")

    def test_zero_wordcount_with_penalty_is_0_fail(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job(
            [policy, policy],
            [[issue(severity="Minor")], []],
        )
        for name in ("Sheet1", "Sheet2"):
            state_path = job / name / "state.json"
            state = json.loads(state_path.read_text(encoding="utf-8"))
            state["wordcount"] = 0
            write_json(state_path, state)

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        overall = next(
            row for row in self.summary_rows(job) if row[0] == "合计"
        )
        self.assertEqual(overall[5], 0)
        self.assertEqual(overall[6], "FAIL")

    def test_score_equal_to_threshold_passes(self):
        policy = {
            "threshold": 99.85,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job(
            [policy, policy],
            [[issue(severity="Minor")], []],
        )
        second_state = job / "Sheet2" / "state.json"
        state = json.loads(second_state.read_text(encoding="utf-8"))
        state["wordcount"] = 0
        write_json(second_state, state)

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        rows = self.summary_rows(job)
        child = next(row for row in rows if row[0] == "Sheet1")
        overall = next(row for row in rows if row[0] == "合计")
        self.assertEqual(child[5], 99.85)
        self.assertEqual(child[6], "PASS")
        self.assertEqual(overall[6], "PASS")

    def test_scope_conflict_fails_before_outputs(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job(
            [policy, policy],
            [[issue(severity="Major", category="Terminology")], []],
        )
        state_path = job / "Sheet1" / "state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        state["check_scope"] = {
            "mode": "no-terminology",
            "terminology_enabled": False,
            "enabled_modules": [
                "precheck_review",
                "accuracy",
                "grammar",
                "naturalness",
            ],
            "disabled_modules": ["terminology", "proper_names", "term_audit"],
            "source": "test",
        }
        write_json(state_path, state)

        result = self.run_aggregate(job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("scope conflict", result.stderr)
        self.assertFalse((job / "multi_corrected.xlsx").exists())
        self.assertFalse((job / "multi_lqe.xlsx").exists())

    def test_missing_child_report_fails_closed(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        (job / "Sheet2" / "Sheet2_lqe.xlsx").unlink()

        result = self.run_aggregate(job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("missing child report", result.stderr)
        self.assertFalse((job / "multi_corrected.xlsx").exists())
        self.assertFalse((job / "multi_lqe.xlsx").exists())

    def test_stale_child_report_fails_closed(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        write_json(
            job / "Sheet1" / "errors.json",
            [
                {
                    "id": 0,
                    "errors": [
                        {
                            "category": "Grammar",
                            "severity": "Minor",
                            "comment": "Use the revised target.",
                            "needs_confirmation": False,
                            "edit": {
                                "from": "Target 0",
                                "to": "New target",
                                "evidence": None,
                            },
                        }
                    ],
                    "corrected": "New target",
                }
            ],
        )

        result = self.run_aggregate(job, refresh_contracts=False)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("stale", result.stderr)
        self.assertFalse((job / "multi_corrected.xlsx").exists())
        self.assertFalse((job / "multi_lqe.xlsx").exists())

    def test_visible_child_report_tamper_fails_closed(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        report_path = job / "Sheet1" / "Sheet1_lqe.xlsx"
        workbook = openpyxl.load_workbook(report_path)
        try:
            workbook["LQE Results"]["B2"] = "Tampered target"
            workbook.save(report_path)
        finally:
            workbook.close()

        result = self.run_aggregate(job, refresh_contracts=False)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("stale", result.stderr)
        self.assertFalse((job / "multi_corrected.xlsx").exists())
        self.assertFalse((job / "multi_lqe.xlsx").exists())

    def test_source_mutation_after_last_source_check_is_not_rebaselined(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        source = self.root / "source.xlsx"
        source_digest = file_sha256(source)
        for name in ("Sheet1", "Sheet2"):
            state_path = job / name / "state.json"
            state = json.loads(state_path.read_text(encoding="utf-8"))
            state["input_sha256"] = source_digest
            write_json(state_path, state)
        self.refresh_report_contracts(job)

        original_child_report_path = aggregate_sheets._child_report_path

        def mutate_after_last_child(child):
            report_path = original_child_report_path(child)
            if child.name == "Sheet2":
                workbook = openpyxl.load_workbook(source)
                try:
                    workbook["Sheet1"]["C1"] = "mutated after validation"
                    workbook.save(source)
                finally:
                    workbook.close()
            return report_path

        with patch.object(
            aggregate_sheets,
            "_child_report_path",
            side_effect=mutate_after_last_child,
        ):
            with self.assertRaisesRegex(ValueError, "changed during aggregate"):
                aggregate_sheets._aggregate(
                    SimpleNamespace(job=str(job), sheets=None, threshold=None)
                )

        self.assertFalse((job / "multi_corrected.xlsx").exists())
        self.assertFalse((job / "multi_lqe.xlsx").exists())

    def test_headerless_child_writes_first_data_row(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        source = self.root / "source.xlsx"
        workbook = openpyxl.load_workbook(source)
        try:
            workbook["Sheet1"].delete_rows(1)
            workbook.save(source)
        finally:
            workbook.close()
        state_path = job / "Sheet1" / "state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        state["source_col"] = 0
        state["target_col"] = 1
        write_json(state_path, state)
        write_json(
            job / "Sheet1" / "errors.json",
            [
                {
                    "id": 0,
                    "errors": [
                        {
                            "category": "Grammar",
                            "severity": "Minor",
                            "comment": "Use the revised target.",
                            "needs_confirmation": False,
                            "edit": {
                                "from": "Target 0",
                                "to": "New target",
                                "evidence": None,
                            },
                        }
                    ],
                    "corrected": "New target",
                }
            ],
        )

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        workbook = openpyxl.load_workbook(job / "multi_corrected.xlsx")
        try:
            self.assertEqual(workbook["Sheet1"]["B1"].value, "New target")
            self.assertIsNone(workbook["Sheet1"]["B2"].value)
        finally:
            workbook.close()

    def test_numeric_header_names_preserve_header_and_edit_data_row(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        source = self.root / "source.xlsx"
        workbook = openpyxl.load_workbook(source)
        try:
            workbook["Sheet1"]["A1"] = "0"
            workbook["Sheet1"]["B1"] = "1"
            workbook.save(source)
        finally:
            workbook.close()
        state_path = job / "Sheet1" / "state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        state.update(
            {
                "source_col": "0",
                "target_col": "1",
                "headers": ["0", "1"],
                "no_header": False,
            }
        )
        write_json(state_path, state)
        write_json(
            job / "Sheet1" / "errors.json",
            [
                {
                    "id": 0,
                    "errors": [
                        {
                            "category": "Grammar",
                            "severity": "Minor",
                            "comment": "Use the revised target.",
                            "needs_confirmation": False,
                            "edit": {
                                "from": "Target 0",
                                "to": "New target",
                                "evidence": None,
                            },
                        }
                    ],
                    "corrected": "New target",
                }
            ],
        )

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        workbook = openpyxl.load_workbook(job / "multi_corrected.xlsx")
        try:
            self.assertEqual(
                [workbook["Sheet1"]["A1"].value, workbook["Sheet1"]["B1"].value],
                ["0", "1"],
            )
            self.assertEqual(workbook["Sheet1"]["B2"].value, "New target")
        finally:
            workbook.close()

    def test_legacy_source_row_drift_fails_closed(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        source = self.root / "source.xlsx"
        workbook = openpyxl.load_workbook(source)
        try:
            workbook["Sheet1"]["A2"] = "Changed source"
            workbook.save(source)
        finally:
            workbook.close()

        result = self.run_aggregate(job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("changed after read", result.stderr)
        self.assertFalse((job / "multi_corrected.xlsx").exists())
        self.assertFalse((job / "multi_lqe.xlsx").exists())

    def test_corrupt_child_report_leaves_existing_outputs_unchanged(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        (job / "Sheet2" / "Sheet2_lqe.xlsx").write_bytes(b"not-an-xlsx")
        corrected = job / "multi_corrected.xlsx"
        report = job / "multi_lqe.xlsx"
        corrected.write_bytes(b"old-corrected")
        report.write_bytes(b"old-report")

        result = self.run_aggregate(job)

        self.assertNotEqual(result.returncode, 0)
        self.assertEqual(corrected.read_bytes(), b"old-corrected")
        self.assertEqual(report.read_bytes(), b"old-report")

    def test_aggregate_does_not_mutate_child_errors(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        state_path = job / "Sheet1" / "state.json"
        errors_path = job / "Sheet1" / "errors.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        state["segments"].append(
            {
                **state["segments"][0],
                "id": 1,
                "row_index": 1,
            }
        )
        write_json(state_path, state)
        source = self.root / "source.xlsx"
        workbook = openpyxl.load_workbook(source)
        try:
            workbook["Sheet1"].append(["Source 0", "Target 0"])
            workbook.save(source)
        finally:
            workbook.close()
        duplicate_issue = issue(severity="Minor")
        write_json(
            errors_path,
            [
                {"id": 0, "errors": [duplicate_issue], "corrected": None},
                {"id": 1, "errors": [duplicate_issue], "corrected": None},
            ],
        )
        original = errors_path.read_bytes()

        result = self.run_aggregate(job)

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(errors_path.read_bytes(), original)

    def test_outputs_cannot_alias_source_workbook(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        source = self.root / "source.xlsx"
        corrected = job / "multi_corrected.xlsx"
        source.replace(corrected)
        original = corrected.read_bytes()
        for name in ("Sheet1", "Sheet2"):
            state_path = job / name / "state.json"
            state = json.loads(state_path.read_text(encoding="utf-8"))
            state["input_path"] = str(corrected)
            write_json(state_path, state)

        result = self.run_aggregate(job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("conflicts with input", result.stderr)
        self.assertEqual(corrected.read_bytes(), original)
        self.assertFalse((job / "multi_lqe.xlsx").exists())

    def test_outputs_cannot_alias_child_state_or_errors(self):
        policy = {
            "threshold": 98,
            "scorecard_profile": "legacy",
            "severity_scale": "lisa",
            "critical_gate": False,
            "repeat_dedup": True,
        }
        job = self.make_job([policy, policy], [[], []])
        state_path = job / "Sheet1" / "state.json"
        errors_path = job / "Sheet2" / "errors.json"
        corrected = job / "multi_corrected.xlsx"
        report = job / "multi_lqe.xlsx"
        os.link(state_path, corrected)
        os.link(errors_path, report)
        original_state = state_path.read_bytes()
        original_errors = errors_path.read_bytes()

        result = self.run_aggregate(job)

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("conflicts with input", result.stderr)
        self.assertEqual(state_path.read_bytes(), original_state)
        self.assertEqual(errors_path.read_bytes(), original_errors)

    def test_current_aggregate_preserves_audit_rows_history_and_long_row_merges(self):
        job = self.root / "current"
        child = job / "Sheet1"
        source = self.root / "current-source.xlsx"
        workbook = openpyxl.Workbook()
        workbook.active.title = "Sheet1"
        workbook.active.append(["Source", "Target"])
        workbook.active.append(["Long source", "Target"])
        workbook.save(source)
        workbook.close()

        state = {
            "artifact_contract_version": 1,
            "input_path": str(source),
            "input_format": "tabular",
            "sheet_name": "Sheet1",
            "headers": ["Source", "Target"],
            "rows_raw": [["Long source", "Target"]],
            "source_col": 0,
            "target_col": 1,
            "no_header": False,
            "wordcount": 1000,
            "segments": [
                {
                    "id": 0,
                    "row_index": 0,
                    "source": "Long source",
                    "target": "Target",
                    "kind": "desc",
                    "term_hits": [],
                    "protected_texts": [],
                }
            ],
        }
        reviewed = {
            "category": "Grammar",
            "severity": "Minor",
            "comment": "Long reviewed issue. " * 320,
            "needs_confirmation": True,
            "edit": None,
            "review_provenance": {
                "finding_origin": "ai_module",
                "ai_reviewed": True,
                "ai_edited": False,
                "review_module": "grammar",
                "reviewed_segment_id": 0,
                "edit_origin": None,
            },
        }
        errors = [{"id": 0, "errors": [reviewed], "corrected": None}]
        write_json(child / "state.json", state)
        write_json(child / "errors.json", errors)
        self.split_child(child)
        manifest = json.loads(
            (child / "chunks" / "split_manifest.json").read_text(
                encoding="utf-8"
            )
        )
        write_json(
            child / "errors.contract.json",
            build_result_contract(manifest, errors),
        )
        lqe_io._build_xlsx(
            state,
            [{"iteration": 0, "errors": errors}],
            99,
            98,
            child / "Sheet1_lqe.xlsx",
            report_contract_results=errors,
        )

        result = self.run_aggregate(job, refresh_contracts=False)
        self.assertEqual(result.returncode, 0, result.stderr)
        aggregate_report = job / "current_lqe.xlsx"
        child_report = child / "Sheet1_lqe.xlsx"
        aggregate = openpyxl.load_workbook(aggregate_report, data_only=True)
        child_book = openpyxl.load_workbook(child_report, data_only=True)
        try:
            self.assertIn("Sheet1 Scorecard", aggregate.sheetnames)
            copied = aggregate["Sheet1 Results"]
            original = child_book["LQE Results"]
            headers = [cell.value for cell in copied[1]]
            review_column = headers.index("LQE AI 复核状态") + 1
            self.assertEqual(
                copied.cell(2, review_column).value,
                "已复核（AI 模块记录）",
            )
            self.assertEqual(
                sorted(str(value) for value in copied.merged_cells.ranges),
                sorted(str(value) for value in original.merged_cells.ranges),
            )
        finally:
            aggregate.close()
            child_book.close()


if __name__ == "__main__":
    unittest.main()
