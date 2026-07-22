import csv
import io
import json
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
IO_SCRIPT = SCRIPTS / "lqe_io.py"
sys.path.insert(0, str(SCRIPTS))

from lqe_corrections import verify_results


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def grammar_issue(*, replacement: str | None = None) -> dict:
    return {
        "category": "Grammar",
        "severity": "Minor",
        "comment": "Verified artifact regression.",
        "needs_confirmation": replacement is None,
        "edit": (
            None
            if replacement is None
            else {
                "from": "Alpha",
                "to": replacement,
                "start": 0,
                "end": 5,
                "evidence": None,
            }
        ),
    }


class VerifiedArtifactContractTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def run_io(self, command: str, *args: object) -> subprocess.CompletedProcess:
        return subprocess.run(
            [sys.executable, str(IO_SCRIPT), command, *map(str, args)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def csv_job(self, name: str = "job", rows=None) -> tuple[Path, Path, Path]:
        job = self.root / name
        source = self.root / f"{name}-source.csv"
        rows = rows or [("A", "Alpha")]
        with source.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Source", "Target"])
            writer.writerows(rows)
        state_path = job / "state.json"
        errors_path = job / "errors.json"
        write_json(
            state_path,
            {
                "input_format": "tabular",
                "input_path": str(source),
                "headers": ["Source", "Target"],
                "rows_raw": [list(row) for row in rows],
                "source_col": "Source",
                "target_col": "Target",
                "wordcount": 100,
                "iteration": 0,
                "segments": [
                    {
                        "id": index,
                        "row_index": index,
                        "source": source_text,
                        "target": target_text,
                        "kind": "desc",
                        "term_hits": [],
                        "protected_texts": [],
                    }
                    for index, (source_text, target_text) in enumerate(rows)
                ],
            },
        )
        return source, state_path, errors_path

    def test_write_rejects_forged_corrected_and_preserves_state(self):
        _, state_path, errors_path = self.csv_job("forged")
        write_json(
            errors_path,
            [{"id": 0, "errors": [], "corrected": "FORGED"}],
        )
        before = state_path.read_bytes()

        result = self.run_io(
            "write", "--state", state_path, "--errors", errors_path, "--score", 100
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("corrected mismatch", result.stderr)
        self.assertEqual(state_path.read_bytes(), before)
        self.assertFalse((state_path.parent / "forged_lqe.xlsx").exists())

    def test_write_requires_complete_id_coverage(self):
        _, state_path, errors_path = self.csv_job(
            "coverage", [("A", "Alpha"), ("B", "Bravo")]
        )
        write_json(errors_path, [{"id": 0, "errors": [], "corrected": None}])

        result = self.run_io(
            "write", "--state", state_path, "--errors", errors_path, "--score", 100
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("missing=[1]", result.stderr)

    def test_empty_string_correction_is_applied_and_exported(self):
        _, state_path, errors_path = self.csv_job("delete")
        write_json(
            errors_path,
            [
                {
                    "id": 0,
                    "errors": [grammar_issue(replacement="")],
                    "corrected": "",
                }
            ],
        )

        applied = self.run_io(
            "apply-fixes", "--state", state_path, "--errors", errors_path
        )

        self.assertEqual(applied.returncode, 0, applied.stderr)
        transition = json.loads(applied.stdout.strip().splitlines()[-1])
        self.assertEqual(transition["applied_count"], 1)
        state = json.loads(state_path.read_text(encoding="utf-8"))
        self.assertEqual(state["segments"][0]["current_target"], "")
        self.assertEqual(state["iteration"], 1)

        exported = self.run_io("export", "--state", state_path)
        self.assertEqual(exported.returncode, 0, exported.stderr)
        output = state_path.parent / "delete_corrected.csv"
        rows = list(
            csv.reader(io.StringIO(output.read_bytes().decode("utf-8-sig")))
        )
        self.assertEqual(rows[1], ["A", ""])

    def test_write_scorecard_totals_use_only_current_iteration(self):
        _, state_path, errors_path = self.csv_job("history")
        state = json.loads(state_path.read_text(encoding="utf-8"))
        state["iteration"] = 1
        state["segments"][0]["current_target"] = "Beta"
        state["error_history"] = [
            {
                "iteration": 0,
                "score": 98.5,
                "status": "PASS",
                "errors": [
                    {
                        "id": 0,
                        "errors": [grammar_issue()],
                        "corrected": None,
                    }
                ],
                "protected_ids": [],
            }
        ]
        write_json(state_path, state)
        write_json(errors_path, [{"id": 0, "errors": [], "corrected": None}])

        result = self.run_io(
            "write", "--state", state_path, "--errors", errors_path, "--score", 100
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        workbook = openpyxl.load_workbook(
            state_path.parent / "history_lqe.xlsx", data_only=True
        )
        try:
            sheet = workbook["LQA Scorecard"]
            total_row = next(
                row
                for row in range(1, sheet.max_row + 1)
                if sheet.cell(row=row, column=1).value == "TOTAL"
            )
            self.assertEqual(sheet.cell(total_row, 4).value, 0)
            self.assertEqual(sheet.cell(total_row, 11).value, 0)
            self.assertEqual(sheet.cell(total_row, 12).value, 0)
        finally:
            workbook.close()

    def test_write_report_cannot_alias_source_workbook(self):
        job = self.root / "collision"
        source = job / "collision_lqe.xlsx"
        source.parent.mkdir(parents=True)
        workbook = openpyxl.Workbook()
        workbook.active.append(["Source", "Target"])
        workbook.active.append(["A", "Alpha"])
        workbook.save(source)
        workbook.close()
        original = source.read_bytes()
        state_path = job / "state.json"
        errors_path = job / "errors.json"
        write_json(
            state_path,
            {
                "input_format": "tabular",
                "input_path": str(source),
                "headers": ["Source", "Target"],
                "rows_raw": [["A", "Alpha"]],
                "source_col": "Source",
                "target_col": "Target",
                "wordcount": 1,
                "segments": [
                    {"id": 0, "row_index": 0, "source": "A", "target": "Alpha"}
                ],
            },
        )
        write_json(errors_path, [{"id": 0, "errors": [], "corrected": None}])

        result = self.run_io(
            "write", "--state", state_path, "--errors", errors_path, "--score", 100
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("conflicts with input", result.stderr)
        self.assertEqual(source.read_bytes(), original)

    def test_export_rejects_source_alias_and_source_drift(self):
        job = self.root / "alias"
        source = job / "alias_corrected.csv"
        source.parent.mkdir(parents=True)
        source.write_text("Source,Target\nA,Alpha\n", encoding="utf-8")
        original = source.read_bytes()
        state_path = job / "state.json"
        write_json(
            state_path,
            {
                "input_format": "tabular",
                "input_path": str(source),
                "headers": ["Source", "Target"],
                "source_col": "Source",
                "target_col": "Target",
                "segments": [
                    {"id": 0, "row_index": 0, "source": "A", "target": "Alpha"}
                ],
            },
        )

        alias_result = self.run_io("export", "--state", state_path)
        self.assertNotEqual(alias_result.returncode, 0)
        self.assertEqual(source.read_bytes(), original)

        drift_source, drift_state, _ = self.csv_job(
            "drift", [("A", "Alpha"), ("B", "Bravo")]
        )
        drift_source.write_text(
            "Source,Target\nB,Bravo\nA,Alpha\n", encoding="utf-8"
        )
        drift_result = self.run_io("export", "--state", drift_state)
        self.assertNotEqual(drift_result.returncode, 0)
        self.assertIn("changed after read", drift_result.stderr)
        self.assertFalse((drift_state.parent / "drift_corrected.csv").exists())

    def test_numeric_header_names_are_not_treated_as_headerless(self):
        source = self.root / "numeric-headers.csv"
        source.write_text("0,1\nA,Alpha\n", encoding="utf-8")
        state_path = self.root / "numeric-headers" / "state.json"

        read_result = self.run_io(
            "read",
            "--input",
            source,
            "--source-col",
            "0",
            "--target-col",
            "1",
            "--no-terminology",
            "--out",
            state_path,
        )

        self.assertEqual(read_result.returncode, 0, read_result.stderr)
        state = json.loads(state_path.read_text(encoding="utf-8"))
        self.assertIs(state["no_header"], False)
        state["segments"][0]["current_target"] = "Beta"
        write_json(state_path, state)

        export_result = self.run_io("export", "--state", state_path)

        self.assertEqual(export_result.returncode, 0, export_result.stderr)
        output = state_path.parent / "numeric-headers_corrected.csv"
        rows = list(
            csv.reader(io.StringIO(output.read_bytes().decode("utf-8-sig")))
        )
        self.assertEqual(rows, [["0", "1"], ["A", "Beta"]])

    def test_verify_results_returns_safety_normalized_issues(self):
        segment = {
            "id": 0,
            "source": "A",
            "target": "Alpha",
            "kind": "name",
            "term_hits": [],
        }
        unsafe = grammar_issue(replacement="Beta")
        unsafe["edit"]["evidence"] = {
            "type": "confirmed_term",
            "source": "A",
            "target": "Beta",
        }

        verified = verify_results(
            [segment],
            [{"id": 0, "errors": [unsafe], "corrected": None}],
            "fixture",
        )

        self.assertTrue(verified[0]["errors"][0]["needs_confirmation"])
        self.assertIsNone(verified[0]["errors"][0]["edit"])


if __name__ == "__main__":
    unittest.main()
