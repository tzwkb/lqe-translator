import csv
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"


class PendingAdjudicationTests(unittest.TestCase):
    def setUp(self):
        self._tempdir = tempfile.TemporaryDirectory(prefix="lqe_pending_")
        self.tmp = Path(self._tempdir.name)
        self.job = self.tmp / "job"
        self.job.mkdir()
        self.input_path = self.tmp / "source.xlsx"
        self.state_path = self.job / "state.json"
        self.errors_path = self.job / "errors.json"

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["原文", "译文"])
        sheet.append(["原文A", "原译A"])
        sheet.append(["原文B", "原译B"])
        workbook.save(self.input_path)
        workbook.close()

        result = self._run_cli(
            "lqe_io.py",
            "read",
            "--input",
            self.input_path,
            "--source-col",
            "原文",
            "--target-col",
            "译文",
            "--source-lang",
            "zh",
            "--target-lang",
            "en",
            "--out",
            self.state_path,
        )
        self._assert_success(result)

    def tearDown(self):
        self._tempdir.cleanup()

    def _run_cli(self, script, *args):
        return subprocess.run(
            [sys.executable, str(SCRIPTS / script), *(str(arg) for arg in args)],
            cwd=ROOT,
            capture_output=True,
            text=True,
        )

    def _assert_success(self, result):
        self.assertEqual(
            result.returncode,
            0,
            msg=f"stdout:\n{result.stdout}\nstderr:\n{result.stderr}",
        )

    def _assert_export_counts(self, result, ai, approved, pending):
        with self.subTest(surface="stdout-counts"):
            self.assertIn(
                f"AI修正 {ai} / 人工批准 {approved} / 待人工裁决 {pending} /",
                result.stdout,
            )

    def _assert_invalid_status_warning(self, result, operation):
        with self.subTest(operation=operation, surface="validation-warning"):
            self.assertIn(
                "非法 correction_status: 'unexpected_status'",
                result.stdout + result.stderr,
            )

    def _errors(self, pending_status="pending_adjudication", pending_candidate="候选A"):
        return [
            {
                "id": 0,
                "errors": [
                    {
                        "category": "Grammar",
                        "severity": "Minor",
                        "comment": "ordinary correction in the same segment",
                    },
                    {
                        "category": "Terminology",
                        "severity": "Major",
                        "comment": "PM/TB adjudication",
                    },
                ],
                "corrected": pending_candidate,
                "correction_status": pending_status,
            },
            {
                "id": 1,
                "errors": [
                    {
                        "category": "Grammar",
                        "severity": "Minor",
                        "comment": "ordinary suggested correction",
                    }
                ],
                "corrected": "建议B",
                "correction_status": "suggested",
            },
        ]

    def _write_json(self, path, value):
        path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")

    def _make_status_job(self, name, correction_status):
        job = self.tmp / name
        job.mkdir()
        state_path = job / "state.json"
        errors_path = job / "errors.json"
        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        self._write_json(state_path, state)
        self._write_json(
            errors_path,
            [self._errors(pending_status=correction_status)[0]],
        )
        return job, state_path, errors_path

    def _assert_non_scalar_status_fails_closed(self, correction_status, status_type):
        job, state_path, errors_path = self._make_status_job(
            f"{status_type}_write", correction_status
        )
        write_result = self._run_cli(
            "lqe_io.py",
            "write",
            "--state",
            state_path,
            "--errors",
            errors_path,
            "--score",
            "80",
        )
        with self.subTest(status_type=status_type, operation="write", surface="returncode"):
            self.assertEqual(
                write_result.returncode,
                0,
                msg=write_result.stdout + write_result.stderr,
            )
        with self.subTest(status_type=status_type, operation="write", surface="warning"):
            self.assertIn(
                "非法 correction_status",
                write_result.stdout + write_result.stderr,
            )
        if write_result.returncode == 0:
            report = openpyxl.load_workbook(job / f"{job.name}_lqe.xlsx")
            results = report["LQE Results"]
            headers = {
                cell.value: cell.column for cell in results[1] if cell.value is not None
            }
            with self.subTest(
                status_type=status_type,
                operation="write",
                surface="pending-status",
            ):
                self.assertEqual(
                    results.cell(2, headers["LQE_Status"]).value,
                    "Pending Adjudication",
                )
            report.close()

        job, state_path, errors_path = self._make_status_job(
            f"{status_type}_apply", correction_status
        )
        apply_result = self._run_cli(
            "lqe_io.py",
            "apply-fixes",
            "--state",
            state_path,
            "--errors",
            errors_path,
            "--score",
            "80",
        )
        with self.subTest(
            status_type=status_type,
            operation="apply-fixes",
            surface="returncode",
        ):
            self.assertEqual(
                apply_result.returncode,
                0,
                msg=apply_result.stdout + apply_result.stderr,
            )
        with self.subTest(
            status_type=status_type,
            operation="apply-fixes",
            surface="warning",
        ):
            self.assertIn(
                "非法 correction_status",
                apply_result.stdout + apply_result.stderr,
            )
        if apply_result.returncode == 0:
            state = json.loads(state_path.read_text(encoding="utf-8"))
            segments = {segment["id"]: segment for segment in state["segments"]}
            with self.subTest(
                status_type=status_type,
                operation="apply-fixes",
                surface="iteration",
            ):
                self.assertEqual(state.get("iteration", 0), 0)
            with self.subTest(
                status_type=status_type,
                operation="apply-fixes",
                surface="candidate",
            ):
                self.assertIsNone(segments[0]["corrected"])
            latest = state.get("error_history", [{}])[-1]
            skipped = {
                entry.get("id"): entry
                for entry in latest.get("skipped_corrections", [])
            }
            with self.subTest(
                status_type=status_type,
                operation="apply-fixes",
                surface="skip-reason",
            ):
                self.assertEqual(
                    skipped.get(0, {}).get("reason"),
                    "PENDING_ADJUDICATION",
                )

        job, state_path, errors_path = self._make_status_job(
            f"{status_type}_export", correction_status
        )
        export_result = self._run_cli(
            "lqe_io.py",
            "export",
            "--state",
            state_path,
            "--errors",
            errors_path,
        )
        with self.subTest(status_type=status_type, operation="export", surface="returncode"):
            self.assertEqual(
                export_result.returncode,
                0,
                msg=export_result.stdout + export_result.stderr,
            )
        with self.subTest(status_type=status_type, operation="export", surface="warning"):
            self.assertIn(
                "非法 correction_status",
                export_result.stdout + export_result.stderr,
            )
        if export_result.returncode == 0:
            with self.subTest(
                status_type=status_type,
                operation="export",
                surface="counts",
            ):
                self.assertIn(
                    "AI修正 0 / 人工批准 0 / 待人工裁决 1 /",
                    export_result.stdout,
                )
            corrected = openpyxl.load_workbook(job / f"{job.name}_corrected.xlsx")
            sheet = corrected.active
            with self.subTest(
                status_type=status_type,
                operation="export",
                surface="translation",
            ):
                self.assertEqual(sheet.cell(2, 2).value, "原译A")
            with self.subTest(
                status_type=status_type,
                operation="export",
                surface="status",
            ):
                self.assertEqual(sheet.cell(2, 3).value, "待人工裁决")
            corrected.close()

        aggregate_job = self.tmp / f"{status_type}_aggregate"
        sheet_job = aggregate_job / "sheet_a"
        sheet_job.mkdir(parents=True)
        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        self._write_json(sheet_job / "state.json", state)
        self._write_json(
            sheet_job / "errors.json",
            [self._errors(pending_status=correction_status)[0]],
        )
        aggregate_result = self._run_cli(
            "aggregate_sheets.py", "--job", aggregate_job
        )
        with self.subTest(
            status_type=status_type,
            operation="aggregate",
            surface="returncode",
        ):
            self.assertEqual(
                aggregate_result.returncode,
                0,
                msg=aggregate_result.stdout + aggregate_result.stderr,
            )
        with self.subTest(
            status_type=status_type,
            operation="aggregate",
            surface="warning",
        ):
            self.assertIn(
                "非法 correction_status",
                aggregate_result.stdout + aggregate_result.stderr,
            )
        if aggregate_result.returncode == 0:
            corrected = openpyxl.load_workbook(
                aggregate_job / f"{aggregate_job.name}_corrected.xlsx"
            )
            sheet = corrected.active
            with self.subTest(
                status_type=status_type,
                operation="aggregate",
                surface="translation",
            ):
                self.assertEqual(sheet.cell(2, 2).value, "原译A")
            corrected.close()
            report = openpyxl.load_workbook(
                aggregate_job / f"{aggregate_job.name}_LQE报告.xlsx"
            )
            summary = report["汇总"]
            headers = {
                cell.value: cell.column for cell in summary[4] if cell.value is not None
            }
            with self.subTest(
                status_type=status_type,
                operation="aggregate",
                surface="correction-count",
            ):
                self.assertEqual(summary.cell(5, headers["建议修正数"]).value, 0)
            report.close()

    def test_export_keeps_pending_segment_unchanged(self):
        self._write_json(self.errors_path, self._errors())

        result = self._run_cli(
            "lqe_io.py", "export", "--state", self.state_path, "--errors", self.errors_path
        )
        self._assert_success(result)
        self._assert_export_counts(result, ai=1, approved=0, pending=1)

        workbook = openpyxl.load_workbook(self.job / "job_corrected.xlsx")
        sheet = workbook.active
        expected = {
            (2, 2): "原译A",
            (2, 3): "待人工裁决",
            (3, 2): "建议B",
            (3, 3): "AI修正",
        }
        for cell, value in expected.items():
            with self.subTest(cell=cell):
                self.assertEqual(sheet.cell(*cell).value, value)
        workbook.close()

    def test_export_preserves_exact_source_cell_for_pending_segment(self):
        workbook = openpyxl.load_workbook(self.input_path)
        workbook.active.cell(2, 2).value = " 原译A "
        workbook.save(self.input_path)
        workbook.close()
        read_result = self._run_cli(
            "lqe_io.py",
            "read",
            "--input",
            self.input_path,
            "--source-col",
            "原文",
            "--target-col",
            "译文",
            "--source-lang",
            "zh",
            "--target-lang",
            "en",
            "--out",
            self.state_path,
        )
        self._assert_success(read_result)
        self._write_json(self.errors_path, self._errors())

        result = self._run_cli(
            "lqe_io.py", "export", "--state", self.state_path, "--errors", self.errors_path
        )
        self._assert_success(result)

        workbook = openpyxl.load_workbook(self.job / "job_corrected.xlsx")
        self.assertEqual(workbook.active.cell(2, 2).value, " 原译A ")
        workbook.close()

    def test_pending_status_columns_are_wide_enough(self):
        self._write_json(self.errors_path, self._errors())
        write_result = self._run_cli(
            "lqe_io.py",
            "write",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
            "--score",
            "80",
        )
        self._assert_success(write_result)
        export_result = self._run_cli(
            "lqe_io.py", "export", "--state", self.state_path, "--errors", self.errors_path
        )
        self._assert_success(export_result)

        report = openpyxl.load_workbook(self.job / "job_lqe.xlsx")
        results = report["LQE Results"]
        result_headers = {
            cell.value: cell.column_letter for cell in results[1] if cell.value is not None
        }
        self.assertGreaterEqual(
            results.column_dimensions[result_headers["LQE_Status"]].width,
            20,
        )
        report.close()

        corrected = openpyxl.load_workbook(self.job / "job_corrected.xlsx")
        status_column = corrected.active.max_column
        self.assertGreaterEqual(
            corrected.active.column_dimensions[
                openpyxl.utils.get_column_letter(status_column)
            ].width,
            18,
        )
        corrected.close()

    def test_export_keeps_existing_corrected_baseline_for_pending_segment(self):
        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        state["iteration"] = 1
        state["segments"][0]["corrected"] = "已生效基线A"
        state["segments"][0]["iter"] = 1
        self._write_json(self.state_path, state)
        self._write_json(
            self.errors_path,
            self._errors(pending_candidate="待裁决新候选A"),
        )

        result = self._run_cli(
            "lqe_io.py", "export", "--state", self.state_path, "--errors", self.errors_path
        )
        self._assert_success(result)
        self._assert_export_counts(result, ai=1, approved=0, pending=1)

        workbook = openpyxl.load_workbook(self.job / "job_corrected.xlsx")
        sheet = workbook.active
        with self.subTest(field="translation"):
            self.assertEqual(sheet.cell(2, 2).value, "已生效基线A")
        with self.subTest(field="status"):
            self.assertEqual(sheet.cell(2, 3).value, "待人工裁决")
        workbook.close()

    def test_export_applies_approved_candidate_with_approved_status(self):
        errors = self._errors(pending_status="approved")
        self._write_json(self.errors_path, errors)

        result = self._run_cli(
            "lqe_io.py", "export", "--state", self.state_path, "--errors", self.errors_path
        )
        self._assert_success(result)
        self._assert_export_counts(result, ai=1, approved=1, pending=0)

        workbook = openpyxl.load_workbook(self.job / "job_corrected.xlsx")
        sheet = workbook.active
        with self.subTest(field="translation"):
            self.assertEqual(sheet.cell(2, 2).value, "候选A")
        with self.subTest(field="status"):
            self.assertEqual(sheet.cell(2, 3).value, "人工批准")
        workbook.close()

    def test_csv_export_keeps_pending_segment_unchanged(self):
        csv_path = self.tmp / "source.csv"
        with csv_path.open("w", encoding="utf-8-sig", newline="") as stream:
            csv.writer(stream).writerows(
                [["原文", "译文"], ["原文A", "原译A"], ["原文B", "原译B"]]
            )
        csv_job = self.tmp / "csv_job"
        csv_job.mkdir()
        csv_state = csv_job / "state.json"
        csv_errors = csv_job / "errors.json"
        read_result = self._run_cli(
            "lqe_io.py",
            "read",
            "--input",
            csv_path,
            "--source-col",
            "原文",
            "--target-col",
            "译文",
            "--source-lang",
            "zh",
            "--target-lang",
            "en",
            "--out",
            csv_state,
        )
        self._assert_success(read_result)
        self._write_json(csv_errors, self._errors())

        export_result = self._run_cli(
            "lqe_io.py", "export", "--state", csv_state, "--errors", csv_errors
        )
        self._assert_success(export_result)
        self._assert_export_counts(export_result, ai=1, approved=0, pending=1)

        with (csv_job / "csv_job_corrected.csv").open(
            encoding="utf-8-sig", newline=""
        ) as stream:
            rows = list(csv.reader(stream))
        expected = {
            (1, 1): "原译A",
            (1, 2): "待人工裁决",
            (2, 1): "建议B",
            (2, 2): "AI修正",
        }
        for (row, column), value in expected.items():
            with self.subTest(row=row, column=column):
                self.assertEqual(rows[row][column], value)

    def test_csv_export_preserves_exact_source_field_for_pending_segment(self):
        csv_path = self.tmp / "source_whitespace.csv"
        with csv_path.open("w", encoding="utf-8-sig", newline="") as stream:
            csv.writer(stream).writerows(
                [["原文", "译文"], ["原文A", " 原译A "], ["原文B", "原译B"]]
            )
        csv_job = self.tmp / "csv_whitespace_job"
        csv_job.mkdir()
        csv_state = csv_job / "state.json"
        csv_errors = csv_job / "errors.json"
        read_result = self._run_cli(
            "lqe_io.py",
            "read",
            "--input",
            csv_path,
            "--source-col",
            "原文",
            "--target-col",
            "译文",
            "--source-lang",
            "zh",
            "--target-lang",
            "en",
            "--out",
            csv_state,
        )
        self._assert_success(read_result)
        self._write_json(csv_errors, self._errors())

        export_result = self._run_cli(
            "lqe_io.py", "export", "--state", csv_state, "--errors", csv_errors
        )
        self._assert_success(export_result)

        with (csv_job / "csv_whitespace_job_corrected.csv").open(
            encoding="utf-8-sig", newline=""
        ) as stream:
            rows = list(csv.reader(stream))
        self.assertEqual(rows[1][1], " 原译A ")

    def test_apply_fixes_skips_entire_pending_segment(self):
        self._write_json(self.errors_path, self._errors())

        result = self._run_cli(
            "lqe_io.py",
            "apply-fixes",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
            "--score",
            "80",
        )
        self._assert_success(result)

        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        segments = {segment["id"]: segment for segment in state["segments"]}
        with self.subTest(segment=0):
            self.assertIsNone(segments[0]["corrected"])
        with self.subTest(segment=1):
            self.assertEqual(segments[1]["corrected"], "建议B")

    def test_pending_only_apply_fixes_records_skip_without_advancing_iteration(self):
        self._write_json(self.errors_path, [self._errors()[0]])
        before = json.loads(self.state_path.read_text(encoding="utf-8"))

        result = self._run_cli(
            "lqe_io.py",
            "apply-fixes",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
            "--score",
            "80",
        )
        self._assert_success(result)

        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        segments = {segment["id"]: segment for segment in state["segments"]}
        with self.subTest(field="iteration"):
            self.assertEqual(state.get("iteration", 0), before.get("iteration", 0))
        with self.subTest(field="corrected"):
            self.assertIsNone(segments[0]["corrected"])
        history = state.get("error_history", [])
        latest = history[-1] if history else {}
        skipped = {entry.get("id"): entry for entry in latest.get("skipped_corrections", [])}
        with self.subTest(field="skip-reason"):
            self.assertEqual(
                skipped.get(0, {}).get("reason"),
                "PENDING_ADJUDICATION",
            )

    def test_apply_fixes_persists_approved_status_for_export_without_errors(self):
        self._write_json(self.errors_path, self._errors(pending_status="approved"))

        apply_result = self._run_cli(
            "lqe_io.py",
            "apply-fixes",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
            "--score",
            "80",
        )
        self._assert_success(apply_result)

        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        segments = {segment["id"]: segment for segment in state["segments"]}
        with self.subTest(surface="state-candidate"):
            self.assertEqual(segments[0]["corrected"], "候选A")
        with self.subTest(surface="state-status"):
            self.assertEqual(segments[0].get("correction_status"), "approved")

        export_result = self._run_cli(
            "lqe_io.py",
            "export",
            "--state",
            self.state_path,
        )
        self._assert_success(export_result)
        self._assert_export_counts(export_result, ai=1, approved=1, pending=0)

        workbook = openpyxl.load_workbook(self.job / "job_corrected.xlsx")
        sheet = workbook.active
        with self.subTest(surface="export-candidate"):
            self.assertEqual(sheet.cell(2, 2).value, "候选A")
        with self.subTest(surface="export-status"):
            self.assertEqual(sheet.cell(2, 3).value, "人工批准")
        workbook.close()

    def test_invalid_correction_status_fails_closed_across_io(self):
        invalid_entry = self._errors(pending_status="unexpected_status")[0]
        self._write_json(self.errors_path, [invalid_entry])

        write_result = self._run_cli(
            "lqe_io.py",
            "write",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
            "--score",
            "80",
        )
        self._assert_success(write_result)
        self._assert_invalid_status_warning(write_result, "write")

        report = openpyxl.load_workbook(self.job / "job_lqe.xlsx")
        results = report["LQE Results"]
        result_headers = {
            cell.value: cell.column for cell in results[1] if cell.value is not None
        }
        with self.subTest(operation="write", surface="suggestion"):
            self.assertEqual(
                results.cell(2, result_headers["Suggest translation"]).value,
                "候选A",
            )
        with self.subTest(operation="write", surface="result-status"):
            self.assertEqual(
                results.cell(2, result_headers["LQE_Status"]).value,
                "Pending Adjudication",
            )
        scorecard = report["LQA Scorecard"]
        header_row = next(
            row
            for row in range(1, scorecard.max_row + 1)
            if scorecard.cell(row, 2).value == "Segment #"
        )
        score_headers = {
            scorecard.cell(header_row, column).value: column
            for column in range(1, scorecard.max_column + 1)
        }
        invalid_row = next(
            row
            for row in range(header_row + 1, scorecard.max_row + 1)
            if scorecard.cell(row, score_headers["Segment #"]).value == 0
        )
        with self.subTest(operation="write", surface="scorecard"):
            self.assertEqual(
                scorecard.cell(invalid_row, score_headers["Fixed"]).value,
                "Pending",
            )
        report.close()

        before_apply = json.loads(self.state_path.read_text(encoding="utf-8"))
        apply_result = self._run_cli(
            "lqe_io.py",
            "apply-fixes",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
            "--score",
            "80",
        )
        self._assert_success(apply_result)
        self._assert_invalid_status_warning(apply_result, "apply-fixes")

        after_apply = json.loads(self.state_path.read_text(encoding="utf-8"))
        segments = {segment["id"]: segment for segment in after_apply["segments"]}
        with self.subTest(operation="apply-fixes", surface="iteration"):
            self.assertEqual(
                after_apply.get("iteration", 0),
                before_apply.get("iteration", 0),
            )
        with self.subTest(operation="apply-fixes", surface="candidate"):
            self.assertIsNone(segments[0]["corrected"])
        latest = after_apply.get("error_history", [{}])[-1]
        skipped = {entry.get("id"): entry for entry in latest.get("skipped_corrections", [])}
        with self.subTest(operation="apply-fixes", surface="skip-reason"):
            self.assertEqual(
                skipped.get(0, {}).get("reason"),
                "PENDING_ADJUDICATION",
            )

        export_result = self._run_cli(
            "lqe_io.py",
            "export",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
        )
        self._assert_success(export_result)
        self._assert_invalid_status_warning(export_result, "export")
        self._assert_export_counts(export_result, ai=0, approved=0, pending=1)

        corrected = openpyxl.load_workbook(self.job / "job_corrected.xlsx")
        corrected_sheet = corrected.active
        with self.subTest(operation="export", surface="translation"):
            self.assertEqual(corrected_sheet.cell(2, 2).value, "原译A")
        with self.subTest(operation="export", surface="status"):
            self.assertEqual(corrected_sheet.cell(2, 3).value, "待人工裁决")
        corrected.close()

    def test_list_correction_status_fails_closed_across_all_entrypoints(self):
        self._assert_non_scalar_status_fails_closed(["approved"], "list_status")

    def test_dict_correction_status_fails_closed_across_all_entrypoints(self):
        self._assert_non_scalar_status_fails_closed(
            {"status": "approved"},
            "dict_status",
        )

    def test_same_iteration_write_replaces_history_and_reports_pending(self):
        old_errors = self._errors(pending_status="suggested", pending_candidate="旧建议A")
        self._write_json(self.errors_path, old_errors)
        first = self._run_cli(
            "lqe_io.py",
            "write",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
            "--score",
            "80",
        )
        self._assert_success(first)

        self._write_json(self.errors_path, self._errors())
        second = self._run_cli(
            "lqe_io.py",
            "write",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
            "--score",
            "80",
        )
        self._assert_success(second)

        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        current_iteration = state.get("iteration", 0)
        current_history = [
            entry
            for entry in state["error_history"]
            if entry.get("iteration") == current_iteration
        ]
        with self.subTest(surface="history-count"):
            self.assertEqual(len(current_history), 1)
        current_errors = {
            entry["id"]: entry
            for entry in (current_history[-1].get("errors", []) if current_history else [])
        }
        pending_entry = current_errors.get(0, {})
        with self.subTest(surface="history-candidate"):
            self.assertEqual(pending_entry.get("corrected"), "候选A")
        with self.subTest(surface="history-status"):
            self.assertEqual(
                pending_entry.get("correction_status"),
                "pending_adjudication",
            )

        workbook = openpyxl.load_workbook(self.job / "job_lqe.xlsx")
        results = workbook["LQE Results"]
        result_headers = {
            cell.value: cell.column for cell in results[1] if cell.value is not None
        }
        with self.subTest(surface="suggestion"):
            self.assertEqual(results.cell(2, result_headers["Suggest translation"]).value, "候选A")
        with self.subTest(surface="result-status"):
            self.assertEqual(
                results.cell(2, result_headers["LQE_Status"]).value,
                "Pending Adjudication",
            )

        scorecard = workbook["LQA Scorecard"]
        header_row = next(
            row
            for row in range(1, scorecard.max_row + 1)
            if scorecard.cell(row, 2).value == "Segment #"
        )
        score_headers = {
            scorecard.cell(header_row, column).value: column
            for column in range(1, scorecard.max_column + 1)
        }
        pending_row = next(
            row
            for row in range(header_row + 1, scorecard.max_row + 1)
            if scorecard.cell(row, score_headers["Segment #"]).value == 0
        )
        with self.subTest(surface="scorecard"):
            self.assertEqual(scorecard.cell(pending_row, score_headers["Fixed"]).value, "Pending")
        workbook.close()

    def test_pending_apply_then_write_collapses_history_and_preserves_skip_audit(self):
        old_entry = self._errors(
            pending_status="suggested",
            pending_candidate="旧建议A",
        )[0]
        pending_entry = self._errors()[0]
        self._write_json(self.errors_path, [old_entry])
        first_write = self._run_cli(
            "lqe_io.py",
            "write",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
            "--score",
            "80",
        )
        self._assert_success(first_write)

        self._write_json(self.errors_path, [pending_entry])
        apply_result = self._run_cli(
            "lqe_io.py",
            "apply-fixes",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
            "--score",
            "80",
        )
        self._assert_success(apply_result)

        final_write = self._run_cli(
            "lqe_io.py",
            "write",
            "--state",
            self.state_path,
            "--errors",
            self.errors_path,
            "--score",
            "80",
        )
        self._assert_success(final_write)

        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        current_iteration = state.get("iteration", 0)
        current_history = [
            entry
            for entry in state.get("error_history", [])
            if entry.get("iteration") == current_iteration
        ]
        with self.subTest(surface="history-count"):
            self.assertEqual(len(current_history), 1)
        latest = current_history[-1] if current_history else {}
        skipped = {entry.get("id"): entry for entry in latest.get("skipped_corrections", [])}
        with self.subTest(surface="skip-reason"):
            self.assertEqual(
                skipped.get(0, {}).get("reason"),
                "PENDING_ADJUDICATION",
            )

        workbook = openpyxl.load_workbook(self.job / "job_lqe.xlsx")
        scorecard = workbook["LQA Scorecard"]
        header_row = next(
            row
            for row in range(1, scorecard.max_row + 1)
            if scorecard.cell(row, 2).value == "Segment #"
        )
        headers = {
            scorecard.cell(header_row, column).value: column
            for column in range(1, scorecard.max_column + 1)
        }
        detail_rows = [
            row
            for row in range(header_row + 1, scorecard.max_row + 1)
            if scorecard.cell(row, headers["Segment #"]).value == 0
        ]
        with self.subTest(surface="scorecard-row-count"):
            self.assertEqual(len(detail_rows), len(pending_entry["errors"]))
        with self.subTest(surface="scorecard-status"):
            self.assertEqual(
                {
                    scorecard.cell(row, headers["Fixed"]).value
                    for row in detail_rows
                },
                {"Pending"},
            )
        workbook.close()

    def test_single_lens_pending_status_survives_both_merges(self):
        chunks = self.job / "chunks"
        chunks.mkdir()
        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        base_segments = [
            {"id": segment["id"], "source": segment["source"], "target": segment["target"]}
            for segment in state["segments"]
        ]
        self._write_json(chunks / "chunk_00.json", {"chunk_id": 0, "segments": base_segments})
        pending = self._errors()[0]
        self._write_json(
            chunks / "chunk_00.T.json",
            [pending, {"id": 1, "errors": [], "corrected": None}],
        )

        merged_lenses = self._run_cli("lqe_chunk.py", "merge-lenses", "--outdir", chunks)
        self._assert_success(merged_lenses)
        lens_entry = json.loads((chunks / "chunk_00.out.json").read_text(encoding="utf-8"))[0]
        with self.subTest(stage="merge-lenses"):
            self.assertEqual(
                {"id": lens_entry["id"], "correction_status": lens_entry.get("correction_status")},
                {"id": 0, "correction_status": "pending_adjudication"},
            )

        baseline = self.job / "errors_precheck.json"
        final_errors = self.job / "merged_errors.json"
        self._write_json(
            baseline,
            [{"id": 0, "errors": [], "corrected": None}, {"id": 1, "errors": [], "corrected": None}],
        )
        merged = self._run_cli(
            "lqe_chunk.py",
            "merge",
            "--state",
            self.state_path,
            "--errors",
            baseline,
            "--outdir",
            chunks,
            "--out",
            final_errors,
        )
        self._assert_success(merged)
        final_entry = json.loads(final_errors.read_text(encoding="utf-8"))[0]
        with self.subTest(stage="merge"):
            self.assertEqual(
                {"id": final_entry["id"], "correction_status": final_entry.get("correction_status")},
                {"id": 0, "correction_status": "pending_adjudication"},
            )

    def test_distinct_multi_lens_candidates_default_to_pending(self):
        chunks = self.job / "chunks"
        chunks.mkdir()
        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        base_segments = [
            {"id": segment["id"], "source": segment["source"], "target": segment["target"]}
            for segment in state["segments"]
        ]
        self._write_json(chunks / "chunk_00.json", {"chunk_id": 0, "segments": base_segments})
        self._write_json(
            chunks / "chunk_00.T.json",
            [
                {
                    "id": 0,
                    "errors": [
                        {"category": "Terminology", "severity": "Major", "comment": "T finding"}
                    ],
                    "corrected": "候选T",
                    "correction_status": "suggested",
                },
                {"id": 1, "errors": [], "corrected": None},
            ],
        )
        self._write_json(
            chunks / "chunk_00.A.json",
            [
                {
                    "id": 0,
                    "errors": [
                        {
                            "category": "Mistranslation",
                            "severity": "Major",
                            "comment": "A finding",
                        }
                    ],
                    "corrected": "候选A",
                    "correction_status": "suggested",
                }
            ],
        )

        result = self._run_cli("lqe_chunk.py", "merge-lenses", "--outdir", chunks)
        self._assert_success(result)
        entry = json.loads((chunks / "chunk_00.out.json").read_text(encoding="utf-8"))[0]
        self.assertEqual(entry.get("correction_status"), "pending_adjudication")

    def test_finalize_accepts_absolute_job_directory(self):
        (self.job / "chunks").mkdir(exist_ok=True)
        (self.job / "chunks" / "chunk_00.T.json").write_text("{}", encoding="utf-8")
        (self.job / ".finalized").write_text("", encoding="utf-8")

        result = subprocess.run(
            ["bash", str(SCRIPTS / "finalize_job.sh"), str(self.job), "1", "single"],
            cwd=ROOT,
            capture_output=True,
            text=True,
        )

        self._assert_success(result)
        self.assertIn("ALREADY-FINALIZED", result.stdout)
        self.assertNotIn("INCOMPLETE", result.stdout)

    def test_flat_lens_duplicate_promotes_pending_status(self):
        chunks = self.job / "flat_chunks"
        chunks.mkdir()
        self._write_json(
            chunks / "chunk_00.json",
            {
                "chunk_id": 0,
                "segments": [
                    {"id": 0, "source": "原文A", "target": "原译A"},
                ],
            },
        )
        self._write_json(
            chunks / "chunk_00.T.json",
            [
                {
                    "id": 0,
                    "category": "Terminology",
                    "severity": "Major",
                    "comment": "first flat finding",
                    "corrected": "候选A",
                    "correction_status": "suggested",
                },
                {
                    "id": 0,
                    "category": "Grammar",
                    "severity": "Minor",
                    "comment": "pending flat finding",
                    "corrected": "候选A",
                    "correction_status": "pending_adjudication",
                },
            ],
        )

        result = self._run_cli("lqe_chunk.py", "merge-lenses", "--outdir", chunks)
        self._assert_success(result)

        entry = json.loads(
            (chunks / "chunk_00.out.json").read_text(encoding="utf-8")
        )[0]
        self.assertEqual(entry.get("correction_status"), "pending_adjudication")

    def test_final_merge_dedup_broadcast_preserves_pending_status(self):
        chunks = self.job / "dedup_chunks"
        chunks.mkdir()
        self._write_json(
            chunks / "chunk_00.out.json",
            [
                {
                    "id": 0,
                    "errors": [
                        {
                            "category": "Terminology",
                            "severity": "Major",
                            "comment": "pending representative",
                        }
                    ],
                    "corrected": "候选A",
                    "correction_status": "pending_adjudication",
                }
            ],
        )
        self._write_json(chunks / "dedup_map.json", {"0": [0, 1]})
        baseline = self.job / "dedup_errors_precheck.json"
        final_errors = self.job / "dedup_merged_errors.json"
        self._write_json(
            baseline,
            [
                {"id": 0, "errors": [], "corrected": None},
                {"id": 1, "errors": [], "corrected": None},
            ],
        )

        result = self._run_cli(
            "lqe_chunk.py",
            "merge",
            "--state",
            self.state_path,
            "--errors",
            baseline,
            "--outdir",
            chunks,
            "--out",
            final_errors,
        )
        self._assert_success(result)

        entries = {
            entry["id"]: entry
            for entry in json.loads(final_errors.read_text(encoding="utf-8"))
        }
        for segment_id in (0, 1):
            with self.subTest(segment_id=segment_id):
                self.assertEqual(
                    entries[segment_id].get("correction_status"),
                    "pending_adjudication",
                )

    def test_reconcile_clears_pending_metadata_only_when_all_errors_are_dropped(self):
        chunks = self.job / "reconcile_chunks"
        chunks.mkdir()
        self._write_json(
            chunks / "chunk_00.json",
            {
                "chunk_id": 0,
                "segments": [
                    {"id": 0, "source": "原文A", "target": "原译A"},
                    {"id": 1, "source": "原文B", "target": "原译B"},
                ],
            },
        )
        self._write_json(
            chunks / "chunk_00.out.json",
            [
                {
                    "id": 0,
                    "errors": [
                        {
                            "category": "Mistranslation",
                            "severity": "Major",
                            "comment": "unconfirmed A-owned error",
                        }
                    ],
                    "corrected": "候选A",
                    "correction_status": "pending_adjudication",
                },
                {
                    "id": 1,
                    "errors": [
                        {
                            "category": "Mistranslation",
                            "severity": "Major",
                            "comment": "unconfirmed A-owned error",
                        },
                        {
                            "category": "Grammar",
                            "severity": "Minor",
                            "comment": "non-A-owned error remains",
                        },
                    ],
                    "corrected": "候选B",
                    "correction_status": "pending_adjudication",
                },
            ],
        )

        result = self._run_cli("lqe_chunk.py", "reconcile", "--outdir", chunks)
        self._assert_success(result)

        entries = {
            entry["id"]: entry
            for entry in json.loads(
                (chunks / "chunk_00.out.json").read_text(encoding="utf-8")
            )
        }
        with self.subTest(segment_id=0, branch="all-errors-dropped"):
            self.assertEqual(entries[0]["errors"], [])
            self.assertIsNone(entries[0]["corrected"])
            self.assertNotIn("correction_status", entries[0])
        with self.subTest(segment_id=1, branch="errors-remain"):
            self.assertEqual(
                [error["category"] for error in entries[1]["errors"]],
                ["Grammar"],
            )
            self.assertEqual(entries[1]["corrected"], "候选B")
            self.assertEqual(
                entries[1].get("correction_status"),
                "pending_adjudication",
            )

    def test_aggregate_sheets_does_not_apply_pending_candidate(self):
        aggregate_job = self.tmp / "aggregate_job"
        sheet_job = aggregate_job / "sheet_a"
        sheet_job.mkdir(parents=True)
        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        self._write_json(sheet_job / "state.json", state)
        self._write_json(sheet_job / "errors.json", self._errors())

        result = self._run_cli("aggregate_sheets.py", "--job", aggregate_job)
        self._assert_success(result)

        corrected = openpyxl.load_workbook(
            aggregate_job / "aggregate_job_corrected.xlsx"
        )
        sheet = corrected.active
        with self.subTest(surface="pending-translation"):
            self.assertEqual(sheet.cell(2, 2).value, "原译A")
        with self.subTest(surface="suggested-control"):
            self.assertEqual(sheet.cell(3, 2).value, "建议B")
        corrected.close()

        report = openpyxl.load_workbook(
            aggregate_job / "aggregate_job_LQE报告.xlsx"
        )
        summary = report["汇总"]
        headers = {
            cell.value: cell.column for cell in summary[4] if cell.value is not None
        }
        with self.subTest(surface="correction-count"):
            self.assertEqual(summary.cell(5, headers["建议修正数"]).value, 1)
        report.close()

    def test_aggregate_sheets_preserves_existing_baseline_for_pending_candidate(self):
        aggregate_job = self.tmp / "aggregate_baseline_job"
        sheet_job = aggregate_job / "sheet_a"
        sheet_job.mkdir(parents=True)
        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        state["iteration"] = 1
        state["segments"][0]["corrected"] = "已生效基线A"
        state["segments"][0]["correction_status"] = "approved"
        state["segments"][0]["iter"] = 1
        self._write_json(sheet_job / "state.json", state)
        self._write_json(
            sheet_job / "errors.json",
            self._errors(pending_candidate="待裁决新候选A"),
        )

        result = self._run_cli("aggregate_sheets.py", "--job", aggregate_job)
        self._assert_success(result)

        corrected = openpyxl.load_workbook(
            aggregate_job / "aggregate_baseline_job_corrected.xlsx"
        )
        sheet = corrected.active
        with self.subTest(surface="pending-baseline"):
            self.assertEqual(sheet.cell(2, 2).value, "已生效基线A")
        with self.subTest(surface="suggested-control"):
            self.assertEqual(sheet.cell(3, 2).value, "建议B")
        corrected.close()

    def test_aggregate_sheets_delivers_state_baselines_when_current_corrections_are_null(self):
        source_path = self.tmp / "aggregate_null_source.xlsx"
        source = openpyxl.Workbook()
        source_sheet = source.active
        source_sheet.append(["原文", "译文"])
        source_sheet.append(["原文A", "原译A"])
        source_sheet.append(["原文B", "原译B"])
        source_sheet.append(["原文C", "原译C"])
        source.save(source_path)
        source.close()

        aggregate_job = self.tmp / "aggregate_null_job"
        sheet_job = aggregate_job / "sheet_a"
        sheet_job.mkdir(parents=True)
        state_path = sheet_job / "state.json"
        read_result = self._run_cli(
            "lqe_io.py",
            "read",
            "--input",
            source_path,
            "--source-col",
            "原文",
            "--target-col",
            "译文",
            "--source-lang",
            "zh",
            "--target-lang",
            "en",
            "--out",
            state_path,
        )
        self._assert_success(read_result)

        state = json.loads(state_path.read_text(encoding="utf-8"))
        state["iteration"] = 1
        baselines = [
            ("建议基线A", "suggested"),
            ("批准基线B", "approved"),
            ("已生效基线C", "approved"),
        ]
        for segment, (corrected, correction_status) in zip(
            state["segments"], baselines
        ):
            segment["corrected"] = corrected
            segment["correction_status"] = correction_status
            segment["iter"] = 1
        self._write_json(state_path, state)
        self._write_json(
            sheet_job / "errors.json",
            [
                {
                    "id": 0,
                    "errors": [],
                    "corrected": None,
                    "correction_status": "suggested",
                },
                {
                    "id": 1,
                    "errors": [],
                    "corrected": None,
                    "correction_status": "approved",
                },
                {
                    "id": 2,
                    "errors": [],
                    "corrected": None,
                    "correction_status": "pending_adjudication",
                },
            ],
        )

        result = self._run_cli("aggregate_sheets.py", "--job", aggregate_job)
        self._assert_success(result)

        corrected = openpyxl.load_workbook(
            aggregate_job / "aggregate_null_job_corrected.xlsx"
        )
        sheet = corrected.active
        for row, expected in enumerate(
            ["建议基线A", "批准基线B", "已生效基线C"],
            start=2,
        ):
            with self.subTest(surface="delivery", row=row):
                self.assertEqual(sheet.cell(row, 2).value, expected)
        corrected.close()

        report = openpyxl.load_workbook(
            aggregate_job / "aggregate_null_job_LQE报告.xlsx"
        )
        summary = report["汇总"]
        headers = {
            cell.value: cell.column for cell in summary[4] if cell.value is not None
        }
        with self.subTest(surface="correction-count"):
            self.assertEqual(summary.cell(5, headers["建议修正数"]).value, 3)
        report.close()


if __name__ == "__main__":
    unittest.main()
