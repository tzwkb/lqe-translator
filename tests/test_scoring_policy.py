import json
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
CALC_SCRIPT = ROOT / "scripts" / "lqe_calc.py"
IO_SCRIPT = ROOT / "scripts" / "lqe_io.py"


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def issue(category="Grammar", severity="Minor") -> dict:
    return {
        "category": category,
        "severity": severity,
        "comment": "fixture issue",
        "needs_confirmation": True,
        "edit": None,
    }


class ScoringPolicyTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.state_path = self.root / "state.json"
        self.errors_path = self.root / "errors.json"

    def tearDown(self):
        self.tempdir.cleanup()

    def run_calc(self, *extra: object):
        result = subprocess.run(
            [
                sys.executable,
                str(CALC_SCRIPT),
                "--state",
                str(self.state_path),
                "--errors",
                str(self.errors_path),
                "--json",
                *map(str, extra),
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        return json.loads(result.stdout)

    def write_duplicate_fixture(self):
        write_json(
            self.state_path,
            {
                "wordcount": 100,
                "segments": [
                    {"id": 0, "source": "Same", "target": "Same target"},
                    {"id": 1, "source": "Same", "target": "Same target"},
                ],
            },
        )
        write_json(
            self.errors_path,
            [
                {"id": 0, "errors": [issue()], "corrected": None},
                {"id": 1, "errors": [issue()], "corrected": None},
            ],
        )

    def test_no_repeat_dedup_clears_stale_repeated_annotations(self):
        self.write_duplicate_fixture()
        first = self.run_calc()
        self.assertEqual(first["repeated"], 1)

        second = self.run_calc("--no-repeat-dedup")

        self.assertEqual(second["repeated"], 0)
        self.assertEqual(second["errors"], 2)
        saved = json.loads(self.errors_path.read_text(encoding="utf-8"))
        self.assertTrue(
            all("repeated" not in item for entry in saved for item in entry["errors"])
        )

    def test_repeat_recalculation_clears_marker_when_current_targets_diverge(self):
        self.write_duplicate_fixture()
        self.run_calc()
        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        state["segments"][1]["corrected"] = "Different target"
        write_json(self.state_path, state)

        result = self.run_calc()

        self.assertEqual(result["repeated"], 0)
        self.assertEqual(result["errors"], 2)
        saved = json.loads(self.errors_path.read_text(encoding="utf-8"))
        self.assertNotIn("repeated", saved[1]["errors"][0])

    def test_state_scoring_policy_is_default_and_cli_can_override_it(self):
        write_json(
            self.state_path,
            {
                "wordcount": 100,
                "scoring_policy": {
                    "threshold": 99,
                    "scorecard_profile": "lqe_2026",
                    "severity_scale": "lisa",
                    "critical_gate": False,
                    "repeat_dedup": True,
                },
                "segments": [{"id": 0, "source": "A", "target": "Alpha"}],
            },
        )
        write_json(
            self.errors_path,
            [{"id": 0, "errors": [issue()], "corrected": None}],
        )

        inherited = self.run_calc()
        overridden = self.run_calc(
            "--scorecard-profile",
            "legacy",
            "--threshold",
            "98",
        )

        self.assertEqual(inherited["score"], 99.0)
        self.assertEqual(inherited["status"], "PASS")
        self.assertEqual(overridden["score"], 98.5)
        self.assertEqual(overridden["status"], "PASS")

    def test_state_critical_gate_fails_high_score(self):
        write_json(
            self.state_path,
            {
                "wordcount": 10000,
                "scoring_policy": {
                    "threshold": 98,
                    "scorecard_profile": "legacy",
                    "severity_scale": "lisa",
                    "critical_gate": True,
                    "repeat_dedup": True,
                },
                "segments": [{"id": 0, "source": "A", "target": "Alpha"}],
            },
        )
        write_json(
            self.errors_path,
            [{"id": 0, "errors": [issue("Grammar", "Critical")], "corrected": None}],
        )

        result = self.run_calc()

        self.assertGreater(result["score"], 98)
        self.assertEqual(result["status"], "FAIL")
        self.assertTrue(result["critical_gate"])

    def test_read_persists_resolved_scoring_policy(self):
        source = self.root / "source.csv"
        source.write_text("Source,Target\nA,Alpha\n", encoding="utf-8")
        project = self.root / "project"
        profile = project / "profile.json"
        write_json(
            profile,
            {
                "name": "fixture/zh-en",
                "language_pair": "zh-en",
                "source_lang": "zh",
                "target_lang": "en",
                "scoring_policy": {
                    "threshold": 99,
                    "scorecard_profile": "lqe_2026",
                    "severity_scale": "mqm",
                    "critical_gate": True,
                    "repeat_dedup": False,
                },
            },
        )
        result = subprocess.run(
            [
                sys.executable,
                str(IO_SCRIPT),
                "read",
                "--input",
                str(source),
                "--project",
                str(profile),
                "--source-col",
                "Source",
                "--target-col",
                "Target",
                "--out",
                str(self.state_path),
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        self.assertEqual(state["threshold"], 99)
        self.assertEqual(
            state["scoring_policy"],
            {
                "threshold": 99.0,
                "scorecard_profile": "lqe_2026",
                "severity_scale": "mqm",
                "critical_gate": True,
                "repeat_dedup": False,
            },
        )

    def test_write_recomputes_score_with_state_policy(self):
        source = self.root / "report-source.csv"
        source.write_text("Source,Target\nA,Alpha\n", encoding="utf-8")
        project = self.root / "report-project"
        profile = project / "profile.json"
        write_json(
            profile,
            {
                "name": "fixture/zh-en",
                "language_pair": "zh-en",
                "source_lang": "zh",
                "target_lang": "en",
                "scoring_policy": {
                    "threshold": 99,
                    "scorecard_profile": "lqe_2026",
                    "severity_scale": "lisa",
                    "critical_gate": False,
                    "repeat_dedup": True,
                },
            },
        )
        read = subprocess.run(
            [
                sys.executable,
                str(IO_SCRIPT),
                "read",
                "--input",
                str(source),
                "--project",
                str(profile),
                "--source-col",
                "Source",
                "--target-col",
                "Target",
                "--out",
                str(self.state_path),
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )
        self.assertEqual(read.returncode, 0, read.stderr)
        state = json.loads(self.state_path.read_text(encoding="utf-8"))
        state.pop("artifact_contract_version", None)
        state["wordcount"] = 100
        write_json(self.state_path, state)
        write_json(
            self.errors_path,
            [{"id": 0, "errors": [issue()], "corrected": None}],
        )

        written = subprocess.run(
            [
                sys.executable,
                str(IO_SCRIPT),
                "write",
                "--state",
                str(self.state_path),
                "--errors",
                str(self.errors_path),
                "--score",
                "98.5",
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

        self.assertEqual(written.returncode, 0, written.stderr)
        saved_state = json.loads(self.state_path.read_text(encoding="utf-8"))
        self.assertEqual(saved_state["error_history"][-1]["score"], 99.0)
        self.assertIn("supplied score 98.5 differs", written.stderr)

    def test_report_uses_state_severity_scale_and_critical_gate(self):
        source = self.root / "critical-source.csv"
        source.write_text("Source,Target\nA,Alpha\n", encoding="utf-8")
        write_json(
            self.state_path,
            {
                "input_path": str(source),
                "headers": ["Source", "Target"],
                "rows_raw": [["A", "Alpha"]],
                "source_col": "Source",
                "target_col": "Target",
                "source_lang": "en",
                "target_lang": "en",
                "wordcount": 10000,
                "iteration": 0,
                "scoring_policy": {
                    "threshold": 98,
                    "scorecard_profile": "legacy",
                    "severity_scale": "mqm",
                    "critical_gate": True,
                    "repeat_dedup": True,
                },
                "segments": [
                    {"id": 0, "source": "A", "target": "Alpha"}
                ],
            },
        )
        write_json(
            self.errors_path,
            [
                {
                    "id": 0,
                    "errors": [issue("Grammar", "Critical")],
                    "corrected": None,
                }
            ],
        )

        written = subprocess.run(
            [
                sys.executable,
                str(IO_SCRIPT),
                "write",
                "--state",
                str(self.state_path),
                "--errors",
                str(self.errors_path),
                "--score",
                "99.75",
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

        self.assertEqual(written.returncode, 0, written.stderr)
        report = next(self.root.glob("*_lqe.xlsx"))
        workbook = openpyxl.load_workbook(report, data_only=True)
        try:
            scorecard = workbook["LQA Scorecard"]
            self.assertEqual(scorecard["B10"].value, "FAIL")
            total_row = next(
                row
                for row in range(1, scorecard.max_row + 1)
                if scorecard.cell(row=row, column=1).value == "TOTAL"
            )
            self.assertEqual(scorecard.cell(row=total_row, column=11).value, 25)
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
