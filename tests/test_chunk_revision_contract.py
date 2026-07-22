import copy
import json
import os
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest
from unittest import mock


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
CHUNK_SCRIPT = SCRIPTS / "lqe_chunk.py"
CALC_SCRIPT = SCRIPTS / "lqe_calc.py"
IO_SCRIPT = SCRIPTS / "lqe_io.py"
sys.path.insert(0, str(SCRIPTS))

import lqe_split_contract as split_contract
from lqe_chunk import build_module_output, load_verification_segments
from lqe_engine import build_check_scope
from lqe_split_contract import (
    SplitContractError,
    add_chunk_payload_digest,
    build_split_manifest,
    build_split_revision,
    publish_generation,
    validate_chunk_payload,
    validate_dedup_payload,
    validate_live_manifest,
)


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def issue() -> dict:
    return {
        "category": "Grammar",
        "severity": "Minor",
        "comment": "fixture issue",
        "needs_confirmation": True,
        "edit": None,
    }


class SplitContractUnitTests(unittest.TestCase):
    def setUp(self):
        self.state = {
            "iteration": 0,
            "source_lang": "zh",
            "target_lang": "en",
            "check_scope": build_check_scope(False, "test"),
            "segments": [
                {
                    "id": 0,
                    "source": "世界",
                    "target": "World",
                    "protected_texts": ["{name}"],
                }
            ],
        }
        self.precheck = [{"id": 0, "issues": []}]
        self.terms = [
            {
                "source": "世界",
                "target": "World",
                "confirmed": True,
                "protected": False,
            }
        ]
        self.scope = self.state["check_scope"]
        self.revision = build_split_revision(
            self.state,
            self.precheck,
            self.terms,
            self.scope,
            size=100,
            char_budget=0,
        )
        self.chunk = add_chunk_payload_digest(
            {
                "chunk_id": 0,
                "iteration": 0,
                "state_fingerprint": self.revision["state_fingerprint"],
                "split_fingerprint": self.revision["split_fingerprint"],
                "segments": [
                    {
                        "id": 0,
                        "source": "世界",
                        "target": "World",
                        "kind": "name",
                        "term_hits": self.terms,
                        "protected_texts": ["{name}"],
                    }
                ],
            }
        )
        self.dedup = {"0": [0]}
        self.manifest = build_split_manifest(
            self.revision,
            chunks=[self.chunk],
            dedup_map=self.dedup,
            input_references={
                "state": {"base": "job", "path": "state.json"},
                "precheck": {
                    "base": "job",
                    "path": "errors_precheck.json",
                },
                "terms": None,
                "terms_mode": "state",
            },
        )

    def manifest_for_state(self, state: dict) -> dict:
        revision = build_split_revision(
            state,
            self.precheck,
            self.terms,
            self.scope,
            size=100,
            char_budget=0,
        )
        chunk = copy.deepcopy(self.chunk)
        chunk["state_fingerprint"] = revision["state_fingerprint"]
        chunk["split_fingerprint"] = revision["split_fingerprint"]
        chunk = add_chunk_payload_digest(chunk)
        return build_split_manifest(
            revision,
            chunks=[chunk],
            dedup_map=self.dedup,
            input_references=copy.deepcopy(self.manifest["inputs"]),
        )

    def test_manifest_recomputes_every_live_input(self):
        validate_live_manifest(
            self.manifest,
            self.state,
            self.precheck,
            self.terms,
            self.scope,
        )
        mutations = [
            ("state", {**self.state, "segments": [{**self.state["segments"][0], "protected_texts": []}]}),
            ("precheck", [{"id": 0, "issues": [issue()]}]),
            ("terms", [{**self.terms[0], "target": "Realm"}]),
            ("scope", build_check_scope(True, "test")),
        ]
        for label, replacement in mutations:
            with self.subTest(label=label):
                values = {
                    "state": self.state,
                    "precheck": self.precheck,
                    "terms": self.terms,
                    "scope": self.scope,
                }
                values[label] = replacement
                with self.assertRaises(SplitContractError):
                    validate_live_manifest(
                        self.manifest,
                        values["state"],
                        values["precheck"],
                        values["terms"],
                        values["scope"],
                    )

        for key, value in (("size", 99), ("char_budget", 10)):
            changed = copy.deepcopy(self.manifest)
            changed["revision"][key] = value
            with self.subTest(label=key), self.assertRaises(SplitContractError):
                validate_live_manifest(
                    changed,
                    self.state,
                    self.precheck,
                    self.terms,
                    self.scope,
                )

    def test_chunk_and_dedup_payloads_are_content_bound(self):
        validate_chunk_payload(self.manifest, self.chunk)
        validate_dedup_payload(self.manifest, self.dedup)

        tampered_chunk = copy.deepcopy(self.chunk)
        tampered_chunk["segments"][0]["kind"] = "desc"
        with self.assertRaises(SplitContractError):
            validate_chunk_payload(self.manifest, tampered_chunk)

        with self.assertRaises(SplitContractError):
            validate_dedup_payload(self.manifest, {"0": [0, 1]})

    def test_context_asset_content_changes_stale_existing_generation(self):
        fields = (
            "sg_path",
            "checks_path",
            "confirmed_rules_path",
            "lang_notes_path",
            "background_path",
        )
        with tempfile.TemporaryDirectory() as tempdir:
            root = Path(tempdir)
            for field in fields:
                with self.subTest(field=field):
                    asset = root / f"{field}.txt"
                    asset.write_text("first revision", encoding="utf-8")
                    state = copy.deepcopy(self.state)
                    state[field] = str(asset)
                    manifest = self.manifest_for_state(state)

                    validate_live_manifest(
                        manifest,
                        state,
                        self.precheck,
                        self.terms,
                        self.scope,
                    )
                    asset.write_text("second revision", encoding="utf-8")
                    with self.assertRaisesRegex(
                        SplitContractError, "stale split manifest"
                    ):
                        validate_live_manifest(
                            manifest,
                            state,
                            self.precheck,
                            self.terms,
                            self.scope,
                        )

    def test_context_asset_status_changes_stale_existing_generation(self):
        with tempfile.TemporaryDirectory() as tempdir:
            asset = Path(tempdir) / "background.md"
            state = copy.deepcopy(self.state)
            state["background_path"] = str(asset)
            manifest = self.manifest_for_state(state)

            validate_live_manifest(
                manifest,
                state,
                self.precheck,
                self.terms,
                self.scope,
            )
            asset.write_text("now present", encoding="utf-8")
            with self.assertRaisesRegex(
                SplitContractError, "stale split manifest"
            ):
                validate_live_manifest(
                    manifest,
                    state,
                    self.precheck,
                    self.terms,
                    self.scope,
                )

    def test_generation_lock_uses_exclusive_msvcrt_backend_on_windows(self):
        fake_msvcrt = mock.Mock()
        fake_msvcrt.LK_LOCK = 10
        fake_msvcrt.LK_UNLCK = 20
        with tempfile.TemporaryDirectory() as tempdir:
            active = Path(tempdir) / "chunks"
            with mock.patch.object(split_contract, "fcntl", None), mock.patch.object(
                split_contract, "msvcrt", fake_msvcrt
            ):
                with split_contract.generation_lock(active, exclusive=False):
                    self.assertEqual(fake_msvcrt.locking.call_count, 1)

            calls = fake_msvcrt.locking.call_args_list
            self.assertEqual([call.args[1:] for call in calls], [(10, 1), (20, 1)])
            self.assertEqual(calls[0].args[0], calls[1].args[0])
            self.assertEqual((Path(tempdir) / ".chunks.lock").read_bytes(), b"\0")

    def test_generation_lock_fails_closed_without_supported_backend(self):
        with tempfile.TemporaryDirectory() as tempdir, mock.patch.object(
            split_contract, "fcntl", None
        ), mock.patch.object(split_contract, "msvcrt", None):
            with self.assertRaisesRegex(
                SplitContractError, "locking is unsupported"
            ):
                with split_contract.generation_lock(
                    Path(tempdir) / "chunks", exclusive=True
                ):
                    self.fail("unsupported lock backend entered critical section")

    def test_missing_manifest_or_fingerprints_fail_closed(self):
        with self.assertRaises(SplitContractError):
            validate_chunk_payload(None, self.chunk)
        for field in ("state_fingerprint", "split_fingerprint", "payload_digest"):
            chunk = copy.deepcopy(self.chunk)
            chunk.pop(field)
            with self.subTest(field=field), self.assertRaises(SplitContractError):
                validate_chunk_payload(self.manifest, chunk)

    def test_generation_swap_restores_whole_old_directory_on_failure(self):
        with tempfile.TemporaryDirectory() as tempdir:
            root = Path(tempdir)
            active = root / "chunks"
            staging = root / ".chunks.generation"
            active.mkdir()
            staging.mkdir()
            originals = {
                "chunk_00.json": b"old chunk",
                "chunk_00.grammar.json": b"old module",
                "split_manifest.json": b"old manifest",
            }
            for name, content in originals.items():
                (active / name).write_bytes(content)
            (staging / "chunk_00.json").write_bytes(b"new chunk")

            real_replace = Path.replace

            def fail_new_generation(path: Path, destination: Path):
                if path == staging:
                    raise KeyboardInterrupt("injected generation switch failure")
                return real_replace(path, destination)

            with mock.patch.object(Path, "replace", fail_new_generation):
                with self.assertRaises(KeyboardInterrupt):
                    publish_generation(staging, active, archive_label="old")

            self.assertTrue(active.is_dir())
            self.assertEqual(
                {path.name: path.read_bytes() for path in active.iterdir()},
                originals,
            )
            archive_root = root / "chunks_archive"
            self.assertFalse(archive_root.exists() and any(archive_root.iterdir()))


class SplitContractCommandTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.job = self.root / "job"
        self.state_path = self.job / "state.json"
        self.precheck_path = self.job / "errors_precheck.json"
        self.chunks = self.job / "chunks"
        self.state = {
            "iteration": 0,
            "source_lang": "en",
            "target_lang": "en",
            "check_scope": build_check_scope(True, "test"),
            "segments": [
                {"id": 0, "source": "One", "target": "First"},
                {"id": 1, "source": "Two", "target": "Second"},
            ],
        }
        write_json(self.state_path, self.state)
        write_json(
            self.precheck_path,
            [{"id": 0, "issues": []}, {"id": 1, "issues": []}],
        )

    def tearDown(self):
        self.tempdir.cleanup()

    def run_chunk(self, *arguments: object) -> subprocess.CompletedProcess:
        return subprocess.run(
            [sys.executable, str(CHUNK_SCRIPT), *map(str, arguments)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def split(self) -> subprocess.CompletedProcess:
        return self.run_chunk(
            "split",
            "--state",
            self.state_path,
            "--errors",
            self.precheck_path,
            "--outdir",
            self.chunks,
            "--size",
            100,
        )

    def test_validate_recomputes_live_precheck_and_rejects_chunk_tampering(self):
        created = self.split()
        self.assertEqual(created.returncode, 0, created.stderr)
        manifest = json.loads(
            (self.chunks / "split_manifest.json").read_text(encoding="utf-8")
        )
        self.assertIn("revision", manifest)
        self.assertIn("chunk_digests", manifest)
        self.assertIn("dedup_map_digest", manifest)

        write_json(
            self.precheck_path,
            [{"id": 0, "issues": [issue()]}, {"id": 1, "issues": []}],
        )
        stale = self.run_chunk("validate-checks", "--job", self.job)
        self.assertNotEqual(stale.returncode, 0)
        self.assertIn("stale split manifest", stale.stderr.lower())

        write_json(
            self.precheck_path,
            [{"id": 0, "issues": []}, {"id": 1, "issues": []}],
        )
        chunk_path = self.chunks / "chunk_00.json"
        chunk = json.loads(chunk_path.read_text(encoding="utf-8"))
        chunk["segments"][0]["kind"] = (
            "desc" if chunk["segments"][0]["kind"] == "name" else "name"
        )
        write_json(chunk_path, chunk)
        tampered = self.run_chunk("validate-checks", "--job", self.job)
        self.assertNotEqual(tampered.returncode, 0)
        self.assertIn("payload digest", tampered.stderr.lower())

    def test_resplit_archives_old_generation_as_one_directory(self):
        first = self.split()
        self.assertEqual(first.returncode, 0, first.stderr)
        old_module = self.chunks / "chunk_00.grammar.json"
        old_module.write_text("old module", encoding="utf-8")

        self.state["segments"][0]["current_target"] = "Changed"
        write_json(self.state_path, self.state)
        second = self.split()

        self.assertEqual(second.returncode, 0, second.stderr)
        self.assertFalse(old_module.exists())
        archives = list((self.job / "chunks_archive").iterdir())
        self.assertEqual(len(archives), 1)
        self.assertTrue((archives[0] / "chunk_00.json").is_file())
        self.assertEqual(
            (archives[0] / "chunk_00.grammar.json").read_text(encoding="utf-8"),
            "old module",
        )

    def test_merge_failure_does_not_replace_existing_output(self):
        created = self.split()
        self.assertEqual(created.returncode, 0, created.stderr)
        write_json(
            self.chunks / "chunk_00.out.json",
            [{"id": 0, "issues": []}],
        )
        output = self.job / "errors.json"
        original = b"existing result"
        output.write_bytes(original)

        merged = self.run_chunk(
            "merge",
            "--state",
            self.state_path,
            "--errors",
            self.precheck_path,
            "--outdir",
            self.chunks,
            "--out",
            output,
        )

        self.assertNotEqual(merged.returncode, 0)
        self.assertEqual(output.read_bytes(), original)

    def test_bound_module_outputs_cannot_cross_generations(self):
        self.state["artifact_contract_version"] = 1
        write_json(self.state_path, self.state)
        created = self.split()
        self.assertEqual(created.returncode, 0, created.stderr)
        old_base = json.loads(
            (self.chunks / "chunk_00.json").read_text(encoding="utf-8")
        )
        modules = ("precheck_review", "accuracy", "grammar", "naturalness")
        old_payloads = {}
        for module in modules:
            draft = self.job / f"{module}.draft.json"
            write_json(
                draft,
                [{"id": 0, "issues": []}, {"id": 1, "issues": []}],
            )
            published = self.run_chunk(
                "publish-module",
                "--job",
                self.job,
                "--chunk",
                0,
                "--module",
                module,
                "--input",
                draft,
                "--split-fingerprint",
                old_base["split_fingerprint"],
                "--chunk-payload-digest",
                old_base["payload_digest"],
            )
            self.assertEqual(published.returncode, 0, published.stderr)
            old_payloads[module] = json.loads(
                (
                    self.chunks / f"chunk_00.{module}.json"
                ).read_text(encoding="utf-8")
            )
        valid = self.run_chunk("validate-checks", "--job", self.job)
        self.assertEqual(valid.returncode, 0, valid.stderr)

        self.state["segments"][0]["current_target"] = "Changed"
        write_json(self.state_path, self.state)
        replaced = self.split()
        self.assertEqual(replaced.returncode, 0, replaced.stderr)
        for module, payload in old_payloads.items():
            write_json(self.chunks / f"chunk_00.{module}.json", payload)

        stale = self.run_chunk("validate-checks", "--job", self.job)
        self.assertNotEqual(stale.returncode, 0)
        self.assertIn("stale module output", stale.stderr)

        draft = self.job / "grammar.draft.json"
        write_json(
            draft,
            [{"id": 0, "issues": []}, {"id": 1, "issues": []}],
        )
        late_publish = self.run_chunk(
            "publish-module",
            "--job",
            self.job,
            "--chunk",
            0,
            "--module",
            "grammar",
            "--input",
            draft,
            "--split-fingerprint",
            old_base["split_fingerprint"],
            "--chunk-payload-digest",
            old_base["payload_digest"],
        )
        self.assertNotEqual(late_publish.returncode, 0)
        self.assertIn("stale task", late_publish.stderr)

    def test_current_state_requires_verified_generation(self):
        self.state["artifact_contract_version"] = 1
        write_json(self.state_path, self.state)
        with self.assertRaisesRegex(
            SplitContractError, "verified chunk generation is required"
        ):
            load_verification_segments(self.state_path)

    def test_configured_missing_terminology_fails_state_fallback(self):
        self.state["check_scope"] = build_check_scope(False, "test")
        self.state["terms_path"] = str(self.job / "missing-terms.json")
        write_json(self.state_path, self.state)
        with self.assertRaisesRegex(ValueError, "terminology file is missing"):
            load_verification_segments(self.state_path)

    def test_provenance_is_derived_from_bound_modules_and_survives_calc(self):
        source_path = self.job / "source.csv"
        source_path.parent.mkdir(parents=True, exist_ok=True)
        source_path.write_text(
            "Source,Target\nSame text,bad\nSame text,bad\nRaw text,Raw text\n",
            encoding="utf-8",
        )
        self.state.update(
            {
                "artifact_contract_version": 1,
                "input_path": str(source_path),
                "input_format": "tabular",
                "headers": ["Source", "Target"],
                "rows_raw": [
                    ["Same text", "bad"],
                    ["Same text", "bad"],
                    ["Raw text", "Raw text"],
                ],
                "source_col": 0,
                "target_col": 1,
                "no_header": False,
                "segments": [
                    {"id": 0, "source": "Same text", "target": "bad"},
                    {"id": 1, "source": "Same text", "target": "bad"},
                    {"id": 2, "source": "Raw text", "target": "Raw text"},
                ],
                "wordcount": 3,
            }
        )
        write_json(self.state_path, self.state)
        machine_issue = {
            "category": "Untranslated",
            "severity": "Major",
            "comment": "Source and target are identical.",
            "needs_confirmation": True,
            "edit": None,
        }
        write_json(
            self.precheck_path,
            [
                {"id": 0, "issues": []},
                {"id": 1, "issues": []},
                {"id": 2, "issues": [machine_issue]},
            ],
        )
        created = self.split()
        self.assertEqual(created.returncode, 0, created.stderr)
        base = json.loads(
            (self.chunks / "chunk_00.json").read_text(encoding="utf-8")
        )
        grammar_issue = {
            "category": "Grammar",
            "severity": "Minor",
            "comment": "Use the corrected form.",
            "needs_confirmation": False,
            "edit": {
                "from": "bad",
                "to": "good",
                "evidence": None,
            },
            "review_provenance": {
                "finding_origin": "machine_precheck",
                "ai_reviewed": False,
                "ai_edited": True,
                "review_module": None,
                "reviewed_segment_id": None,
                "edit_origin": "machine_precheck",
            },
        }
        entries_by_module = {
            "precheck_review": [
                {"id": 0, "issues": []},
                {"id": 2, "issues": []},
            ],
            "accuracy": [
                {"id": 0, "issues": []},
                {"id": 2, "issues": []},
            ],
            "grammar": [
                {"id": 0, "issues": [grammar_issue]},
                {"id": 2, "issues": []},
            ],
            "naturalness": [
                {"id": 0, "issues": []},
                {"id": 2, "issues": []},
            ],
        }
        for module, entries in entries_by_module.items():
            draft = self.job / f"{module}.draft.json"
            write_json(draft, entries)
            published = self.run_chunk(
                "publish-module",
                "--job",
                self.job,
                "--chunk",
                0,
                "--module",
                module,
                "--input",
                draft,
                "--split-fingerprint",
                base["split_fingerprint"],
                "--chunk-payload-digest",
                base["payload_digest"],
            )
            self.assertEqual(published.returncode, 0, published.stderr)

        formal_path = self.chunks / "chunk_00.grammar.json"
        formal = json.loads(formal_path.read_text(encoding="utf-8"))
        formal["entries"][0]["issues"][0]["comment"] = "Injected after publish."
        formal["entries_digest"] = split_contract.canonical_digest(
            formal["entries"]
        )
        write_json(formal_path, formal)
        tampered = self.run_chunk("validate-checks", "--job", self.job)
        self.assertNotEqual(tampered.returncode, 0)
        self.assertIn("publication receipt mismatch", tampered.stderr)
        republished = self.run_chunk(
            "publish-module",
            "--job",
            self.job,
            "--chunk",
            0,
            "--module",
            "grammar",
            "--input",
            self.job / "grammar.draft.json",
            "--split-fingerprint",
            base["split_fingerprint"],
            "--chunk-payload-digest",
            base["payload_digest"],
        )
        self.assertEqual(republished.returncode, 0, republished.stderr)

        merged_checks = self.run_chunk("merge-checks", "--job", self.job)
        self.assertEqual(merged_checks.returncode, 0, merged_checks.stderr)
        merged_path = self.chunks / "chunk_00.out.json"
        forged = json.loads(merged_path.read_text(encoding="utf-8"))
        forged["entries"][0]["issues"][0]["review_provenance"][
            "review_module"
        ] = "accuracy"
        forged["entries_digest"] = split_contract.canonical_digest(
            forged["entries"]
        )
        write_json(merged_path, forged)
        output = self.job / "errors.json"
        rejected = self.run_chunk(
            "merge",
            "--state",
            self.state_path,
            "--errors",
            self.precheck_path,
            "--outdir",
            self.chunks,
            "--out",
            output,
        )
        self.assertNotEqual(rejected.returncode, 0)
        self.assertIn("differs from current bound module outputs", rejected.stderr)

        rebuilt = self.run_chunk("merge-checks", "--job", self.job)
        self.assertEqual(rebuilt.returncode, 0, rebuilt.stderr)
        merged = self.run_chunk(
            "merge",
            "--state",
            self.state_path,
            "--errors",
            self.precheck_path,
            "--outdir",
            self.chunks,
            "--out",
            output,
        )
        self.assertEqual(merged.returncode, 0, merged.stderr)
        results = json.loads(output.read_text(encoding="utf-8"))
        by_id = {entry["id"]: entry for entry in results}
        for segment_id in (0, 1):
            provenance = by_id[segment_id]["errors"][0]["review_provenance"]
            self.assertEqual(provenance["finding_origin"], "ai_module")
            self.assertEqual(provenance["review_module"], "grammar")
            self.assertEqual(provenance["reviewed_segment_id"], 0)
            self.assertTrue(provenance["ai_edited"])
            self.assertEqual(by_id[segment_id]["corrected"], "good")
        machine = by_id[2]["errors"][0]["review_provenance"]
        self.assertEqual(machine["finding_origin"], "machine_precheck")
        self.assertFalse(machine["ai_reviewed"])
        self.assertFalse(machine["ai_edited"])

        calculated = subprocess.run(
            [
                sys.executable,
                str(CALC_SCRIPT),
                "--state",
                str(self.state_path),
                "--errors",
                str(output),
                "--json",
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )
        self.assertEqual(calculated.returncode, 0, calculated.stderr)
        after_calc = json.loads(output.read_text(encoding="utf-8"))
        self.assertTrue(
            after_calc[0]["errors"][0]["review_provenance"]["ai_edited"]
        )
        score = json.loads(calculated.stdout)["score"]
        written = subprocess.run(
            [
                sys.executable,
                str(IO_SCRIPT),
                "write",
                "--state",
                str(self.state_path),
                "--errors",
                str(output),
                "--score",
                str(score),
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )
        self.assertEqual(written.returncode, 0, written.stderr)
        report = self.job / "job_lqe.xlsx"
        self.assertTrue(report.is_file())
        import openpyxl

        workbook = openpyxl.load_workbook(report, data_only=True)
        try:
            sheet = workbook["LQE Results"]
            headers = [cell.value for cell in sheet[1]]
            segment_column = headers.index("LQE Segment ID") + 1
            review_column = headers.index("LQE AI 复核状态") + 1
            rows = {
                sheet.cell(row, segment_column).value: sheet.cell(
                    row,
                    review_column,
                ).value
                for row in range(2, sheet.max_row + 1)
            }
            self.assertEqual(rows[0], "已复核（AI 模块记录）")
            self.assertEqual(
                rows[1],
                "已复核（复用 Segment 0；AI 模块记录）",
            )
            self.assertEqual(rows[2], "未确认（机器预检保留）")
        finally:
            workbook.close()

    def test_machine_precheck_can_be_explicitly_ai_reviewed_without_ai_edit(self):
        self.state["artifact_contract_version"] = 1
        write_json(self.state_path, self.state)
        precheck_issue = {
            "category": "Markup",
            "severity": "Major",
            "comment": "Machine-detected tag mismatch.",
            "needs_confirmation": True,
            "edit": None,
        }
        write_json(
            self.precheck_path,
            [
                {"id": 0, "issues": [precheck_issue]},
                {"id": 1, "issues": []},
            ],
        )
        created = self.split()
        self.assertEqual(created.returncode, 0, created.stderr)
        base = json.loads(
            (self.chunks / "chunk_00.json").read_text(encoding="utf-8")
        )
        precheck_ref = base["segments"][0]["precheck"][0]["precheck_ref"]
        reviewed_issue = {
            **precheck_issue,
            "comment": "AI confirmed the tag mismatch.",
            "precheck_ref": precheck_ref,
        }
        for module in (
            "precheck_review",
            "accuracy",
            "grammar",
            "naturalness",
        ):
            entries = [
                {
                    "id": 0,
                    "issues": [reviewed_issue]
                    if module == "precheck_review"
                    else [],
                },
                {"id": 1, "issues": []},
            ]
            draft = self.job / f"{module}.draft.json"
            write_json(draft, entries)
            published = self.run_chunk(
                "publish-module",
                "--job",
                self.job,
                "--chunk",
                0,
                "--module",
                module,
                "--input",
                draft,
                "--split-fingerprint",
                base["split_fingerprint"],
                "--chunk-payload-digest",
                base["payload_digest"],
            )
            self.assertEqual(published.returncode, 0, published.stderr)
        self.assertEqual(
            self.run_chunk("merge-checks", "--job", self.job).returncode,
            0,
        )
        output = self.job / "errors.json"
        merged = self.run_chunk(
            "merge",
            "--state",
            self.state_path,
            "--errors",
            self.precheck_path,
            "--outdir",
            self.chunks,
            "--out",
            output,
        )
        self.assertEqual(merged.returncode, 0, merged.stderr)
        provenance = json.loads(output.read_text(encoding="utf-8"))[0][
            "errors"
        ][0]["review_provenance"]
        self.assertEqual(provenance["finding_origin"], "machine_precheck")
        self.assertTrue(provenance["ai_reviewed"])
        self.assertEqual(provenance["review_module"], "precheck_review")
        self.assertEqual(provenance["reviewed_segment_id"], 0)
        self.assertIsNone(provenance["edit_origin"])
        self.assertFalse(provenance["ai_edited"])


if __name__ == "__main__":
    unittest.main()
