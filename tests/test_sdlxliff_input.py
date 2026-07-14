import hashlib
import json
from pathlib import Path
import shutil
import subprocess
import sys
import tempfile
import unittest
from unittest import mock
from xml.etree import ElementTree as ET

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
IO_SCRIPT = SCRIPTS / "lqe_io.py"
CHUNK_SCRIPT = SCRIPTS / "lqe_chunk.py"
BATCH_SCRIPT = SCRIPTS / "lqe_batch.py"
FIXTURES = ROOT / "tests" / "fixtures" / "sdlxliff"
RECURSIVE_FIXTURES = FIXTURES / "recursive"

sys.path.insert(0, str(SCRIPTS))

from lqe_inputs import detect_input_format
import lqe_engine
import lqe_io
from lqe_inputs.sdlxliff import (
    SDLXLIFFImportError,
    SDLXLIFFOptions,
    SerializedMixedContent,
    is_exact_tm,
    match_content_type,
    match_exclusions,
    read_sdlxliff,
    serialize_mixed,
    validate_options,
)


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


XLIFF_NS = "urn:oasis:names:tc:xliff:document:1.2"
SDL_NS = "http://sdl.com/FileTypes/SdlXliff/1.0"
XLIFF_QNAME = "{" + XLIFF_NS + "}"
EXPECTED_SIGNATURE = (
    (XLIFF_QNAME + "g", (("id", "g1"),)),
    (XLIFF_QNAME + "x", (("id", "x1"),)),
    (XLIFF_QNAME + "bx", (("id", "bx1"),)),
    (XLIFF_QNAME + "ex", (("id", "ex1"),)),
    (XLIFF_QNAME + "ph", (("id", "ph1"),)),
    (XLIFF_QNAME + "bpt", (("id", "bpt1"),)),
    (XLIFF_QNAME + "ept", (("id", "ept1"),)),
    (XLIFF_QNAME + "it", (("id", "it1"),)),
    (XLIFF_QNAME + "sub", ()),
    (XLIFF_QNAME + "mrk", (("mid", "inner1"), ("mtype", "protected"))),
)
EXPECTED_SOURCE_PLAIN = " 前组尾子内值尾末 后 "
SOURCE_REF_KEYS = (
    "relative_path",
    "file_index",
    "tu_id",
    "tu_index",
    "sdl_segment_id",
    "segment_index",
)


class SDLXLIFFFormatDetectionTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def touch(self, relative_path: str) -> Path:
        path = self.root / relative_path
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("fixture", encoding="utf-8")
        return path

    def copy_sdl(self, relative_path: str) -> Path:
        path = self.root / relative_path
        path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(FIXTURES / "multi_segment.sdlxliff", path)
        return path

    def test_single_file_auto_and_explicit_formats(self):
        sdl = self.copy_sdl("sample.sdlxliff")
        csv = self.touch("sample.csv")

        self.assertEqual(detect_input_format(sdl, "auto"), "sdlxliff")
        self.assertEqual(detect_input_format(csv, "auto"), "tabular")
        self.assertEqual(detect_input_format(sdl, "sdlxliff"), "sdlxliff")
        self.assertEqual(detect_input_format(csv, "tabular"), "tabular")

    def test_invalid_request_unknown_suffix_and_explicit_mismatch_fail(self):
        sdl = self.copy_sdl("sample.sdlxliff")
        unknown = self.touch("sample.txt")

        with self.assertRaisesRegex(ValueError, "requested"):
            detect_input_format(sdl, "xml")
        with self.assertRaisesRegex(ValueError, "unsupported"):
            detect_input_format(unknown, "auto")
        with self.assertRaisesRegex(ValueError, "does not match"):
            detect_input_format(sdl, "tabular")

    def test_auto_directory_requires_only_sdlxliff(self):
        self.copy_sdl("a/dialogs.sdlxliff")
        self.copy_sdl("b.sdlxliff")

        self.assertEqual(detect_input_format(self.root, "auto"), "sdlxliff")

        self.touch("notes.csv")
        with self.assertRaisesRegex(ValueError, "mixed"):
            detect_input_format(self.root, "auto")

    def test_empty_and_tabular_directories_are_not_supported(self):
        with self.assertRaisesRegex(ValueError, "no supported"):
            detect_input_format(self.root, "auto")

        self.touch("only.csv")
        with self.assertRaisesRegex(ValueError, "tabular director"):
            detect_input_format(self.root, "auto")
        with self.assertRaisesRegex(ValueError, "tabular director"):
            detect_input_format(self.root, "tabular")

    def test_explicit_sdl_directory_allows_unselected_supported_files(self):
        self.copy_sdl("a/dialogs.sdlxliff")
        self.touch("notes.csv")

        self.assertEqual(detect_input_format(self.root, "sdlxliff"), "sdlxliff")
        result = read_sdlxliff(self.root, options=SDLXLIFFOptions())
        self.assertEqual(result.manifest["unselected_supported_files"], ["notes.csv"])


class SDLXLIFFParserTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def temp_fixture(self, name: str, content: str) -> Path:
        path = self.root / name
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(content, encoding="utf-8")
        return path

    def document(self, body: str, *, version: str = "1.2", namespace: str = XLIFF_NS) -> str:
        return (
            f'<xliff version="{version}" xmlns="{namespace}" xmlns:sdl="{SDL_NS}">'
            f'<file original="anonymous.xml" source-language="zh-CN" target-language="en-US">'
            f"<body>{body}</body></file></xliff>"
        )

    def test_imports_all_files_and_multiple_mrk_in_order(self):
        result = read_sdlxliff(
            FIXTURES / "multi_segment.sdlxliff", options=SDLXLIFFOptions()
        )

        self.assertEqual([segment["id"] for segment in result.segments], [0, 1, 2])
        self.assertEqual(
            [segment["source_ref"]["sdl_segment_id"] for segment in result.segments],
            ["1", "2", "3"],
        )
        self.assertEqual(
            result.segments[0]["source_ref"]["relative_path"],
            "multi_segment.sdlxliff",
        )
        self.assertEqual(
            [segment["metadata"]["sdlxliff"]["file_original"] for segment in result.segments],
            ["dialogs.xml", "dialogs.xml", "ui.xml"],
        )
        self.assertEqual(result.source_lang, "zh-CN")
        self.assertEqual(result.target_lang, "en-US")
        self.assertEqual(
            result.headers,
            ["来源文件", "TU ID", "SDL Segment ID", "原文", "译文"],
        )
        self.assertEqual(len(result.rows_raw), 3)

    def test_recursive_directory_order_and_missing_ids_are_stable(self):
        first = read_sdlxliff(RECURSIVE_FIXTURES, options=SDLXLIFFOptions())
        second = read_sdlxliff(RECURSIVE_FIXTURES, options=SDLXLIFFOptions())
        first_refs = [segment["source_ref"] for segment in first.segments]
        second_refs = [segment["source_ref"] for segment in second.segments]

        self.assertEqual(first_refs, second_refs)
        relative_paths = [ref["relative_path"] for ref in first_refs]
        self.assertEqual(relative_paths, sorted(relative_paths))
        self.assertEqual(len(first_refs), len({tuple(ref.values()) for ref in first_refs}))
        self.assertTrue(any(ref["tu_id"] is None for ref in first_refs))
        self.assertTrue(any(ref["sdl_segment_id"] is None for ref in first_refs))
        self.assertTrue(all(tuple(ref) == SOURCE_REF_KEYS for ref in first_refs))
        self.assertTrue(all(isinstance(ref["tu_index"], int) for ref in first_refs))
        self.assertEqual(
            first.input_paths,
            [
                str((RECURSIVE_FIXTURES / "a" / "dialogs.sdlxliff").resolve()),
                str((RECURSIVE_FIXTURES / "b.sdlxliff").resolve()),
            ],
        )

    def test_missing_tu_ids_use_document_order_for_duplicate_ownership(self):
        fixture = FIXTURES / "missing_tu_ids.sdlxliff"
        try:
            first = read_sdlxliff(fixture, options=SDLXLIFFOptions())
            second = read_sdlxliff(fixture, options=SDLXLIFFOptions())
        except SDLXLIFFImportError as exc:
            self.fail(f"missing TU IDs must use document-order ownership: {exc}")

        self.assertEqual(
            [segment["source_plain"] for segment in first.segments], ["甲", "乙"]
        )
        refs = [segment["source_ref"] for segment in first.segments]
        self.assertEqual(refs, [segment["source_ref"] for segment in second.segments])
        self.assertEqual([ref["tu_id"] for ref in refs], [None, None])
        self.assertEqual([ref["tu_index"] for ref in refs], [0, 1])
        self.assertEqual(len(refs), len({tuple(ref.values()) for ref in refs}))

    def test_missing_tu_id_fallback_is_owned_by_each_input_file(self):
        fixture = FIXTURES / "missing_tu_ids.sdlxliff"
        inputs = self.root / "missing-tu-id-directory"
        (inputs / "nested").mkdir(parents=True)
        shutil.copy2(fixture, inputs / "a.sdlxliff")
        shutil.copy2(fixture, inputs / "nested" / "b.sdlxliff")

        result = read_sdlxliff(inputs, options=SDLXLIFFOptions())

        refs = [segment["source_ref"] for segment in result.segments]
        self.assertEqual(
            [ref["relative_path"] for ref in refs],
            [
                "a.sdlxliff",
                "a.sdlxliff",
                "nested/b.sdlxliff",
                "nested/b.sdlxliff",
            ],
        )
        self.assertEqual([ref["tu_index"] for ref in refs], [0, 1, 0, 1])
        self.assertEqual(len(refs), len({tuple(ref.values()) for ref in refs}))

    def test_rows_raw_uses_empty_strings_for_missing_business_ids(self):
        result = read_sdlxliff(RECURSIVE_FIXTURES, options=SDLXLIFFOptions())
        missing_id_indexes = [
            index
            for index, segment in enumerate(result.segments)
            if segment["source_ref"]["tu_id"] is None
            or segment["source_ref"]["sdl_segment_id"] is None
        ]

        self.assertTrue(missing_id_indexes)
        self.assertTrue(
            all(isinstance(cell, str) for row in result.rows_raw for cell in row)
        )
        for index in missing_id_indexes:
            source_ref = result.segments[index]["source_ref"]
            row = result.rows_raw[index]
            self.assertEqual(row[1], source_ref["tu_id"] or "")
            self.assertEqual(row[2], source_ref["sdl_segment_id"] or "")

    def test_reads_status_comment_modifier_empty_target_and_tm_metadata(self):
        result = read_sdlxliff(RECURSIVE_FIXTURES, options=SDLXLIFFOptions())
        by_tu = {
            (segment["source_ref"]["relative_path"], segment["source_ref"]["tu_id"]): segment
            for segment in result.segments
        }
        metadata = by_tu[("a/dialogs.sdlxliff", "shared-tu")]["metadata"]["sdlxliff"]

        self.assertEqual(metadata["confirmation"], "Translated")
        self.assertEqual(metadata["comment"], "Anonymous review note")
        self.assertEqual(metadata["last_modified_by"], "AnonymousUser")
        self.assertEqual(by_tu[("b.sdlxliff", "empty-target")]["target_plain"], "")
        self.assertNotIn(("b.sdlxliff", "blank-both"), by_tu)
        self.assertEqual(
            result.manifest["excluded"][0]["source_ref"]["tu_id"], "blank-both"
        )
        self.assertEqual(
            by_tu[("b.sdlxliff", "tm-negative-origin")]["metadata"]["sdlxliff"]["origin"],
            "machine",
        )
        self.assertEqual(
            by_tu[("b.sdlxliff", "tm-negative-percent")]["metadata"]["sdlxliff"]["match_percent"],
            "99",
        )
        self.assertEqual(
            by_tu[("b.sdlxliff", "tm-negative-match")]["metadata"]["sdlxliff"]["text_match"],
            "Fuzzy",
        )

    def test_tu_direct_comment_references_are_strictly_scoped(self):
        result = read_sdlxliff(
            FIXTURES / "tu_direct_comments.sdlxliff",
            options=SDLXLIFFOptions(),
        )
        comments = {
            (
                segment["source_ref"]["file_index"],
                segment["source_ref"]["tu_id"],
                segment["source_ref"]["sdl_segment_id"],
            ): segment["metadata"]["sdlxliff"]["comment"]
            for segment in result.segments
        }

        self.assertEqual(
            comments[(0, "direct-comments", "1")],
            "TU direct note\nFirst segment note",
        )
        self.assertEqual(
            comments[(0, "direct-comments", "2")], "TU direct note"
        )
        self.assertIsNone(comments[(0, "unknown-comment", "3")])
        self.assertEqual(
            comments[(1, "file-scoped-comment", "4")], "Second file note"
        )
        self.assertEqual(
            comments[(2, "root-scoped-comment", "5")], "Root fallback note"
        )

    def test_missing_target_node_produces_an_empty_target(self):
        fixture = self.temp_fixture(
            "missing-target.sdlxliff",
            self.document(
                '<trans-unit id="tu"><source>甲</source>'
                '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs></trans-unit>'
            ),
        )

        result = read_sdlxliff(fixture, options=SDLXLIFFOptions())

        self.assertEqual(result.segments[0]["target"], "")
        self.assertEqual(result.segments[0]["target_plain"], "")

    def test_empty_source_seg_source_and_target_are_excluded_as_blank(self):
        fixture = self.temp_fixture(
            "empty-seg-source.sdlxliff",
            self.document(
                '<trans-unit id="blank"><source/><seg-source/><target/>'
                '</trans-unit><trans-unit id="keep"><source>甲</source>'
                '<target>Alpha</target><sdl:seg-defs><sdl:seg id="1"/>'
                '</sdl:seg-defs></trans-unit>'
            ),
        )

        try:
            result = read_sdlxliff(fixture, options=SDLXLIFFOptions())
        except SDLXLIFFImportError as exc:
            self.fail(f"unambiguous blank TU must reach blank exclusion: {exc}")

        self.assertEqual(
            [segment["source_ref"]["tu_id"] for segment in result.segments],
            ["keep"],
        )
        self.assertEqual(result.manifest["counts"]["parsed_segments"], 2)
        self.assertEqual(result.manifest["counts"]["excluded_segments"], 1)
        blank = result.manifest["excluded"][0]
        self.assertEqual(blank["source_ref"]["tu_id"], "blank")
        self.assertEqual(blank["rule_ids"], ["blank-both-sides"])

    def test_unknown_namespaced_attributes_are_evidence_not_sdl_semantics(self):
        result = read_sdlxliff(
            FIXTURES / "unknown_namespaced_attributes.sdlxliff",
            options=SDLXLIFFOptions(
                tm_protection="protect-exact-source-and-target"
            ),
        )

        segment, sdl_control = result.segments
        metadata = segment["metadata"]["sdlxliff"]
        self.assertFalse(segment.get("protected", False))
        self.assertEqual(result.tm_candidates["candidate_ids"], [sdl_control["id"]])
        self.assertEqual(sdl_control["protected_reason"], "SOURCE_LOCKED")
        self.assertEqual(
            sdl_control["metadata"]["sdlxliff"]["origin"], "tm"
        )
        self.assertEqual(
            sdl_control["metadata"]["sdlxliff"]["comment"], "SDL control note"
        )
        self.assertEqual(
            sdl_control["metadata"]["sdlxliff"]["last_modified_by"], "SDLUser"
        )
        self.assertIsNone(metadata["origin"])
        self.assertIsNone(metadata["match_percent"])
        self.assertIsNone(metadata["text_match"])
        self.assertFalse(metadata["locked"])
        self.assertIsNone(metadata["comment"])
        self.assertIsNone(metadata["last_modified_by"])
        self.assertIn(
            "urn:vendor:attributes", "\n".join(metadata["extension_xml"])
        )
        evidence = {
            (item["element"], item["name"], item["value"])
            for item in metadata["extension_attributes"]
        }
        vendor_attribute_names = {
            name
            for _, name, _ in evidence
            if name.startswith("{urn:vendor:attributes}")
        }
        self.assertTrue(
            {
                "{urn:vendor:attributes}locked",
                "{urn:vendor:attributes}origin",
                "{urn:vendor:attributes}percent",
                "{urn:vendor:attributes}match-percent",
                "{urn:vendor:attributes}text-match",
                "{urn:vendor:attributes}source-meta",
                "{urn:vendor:attributes}target-meta",
            }.issubset(vendor_attribute_names)
        )
        self.assertIn(
            (
                XLIFF_QNAME + "trans-unit",
                "{urn:vendor:attributes}locked",
                "true",
            ),
            evidence,
        )
        self.assertIn(
            (
                "{" + SDL_NS + "}seg",
                "{urn:vendor:attributes}match-percent",
                "100",
            ),
            evidence,
        )
        self.assertIn(
            "urn:vendor:attributes", result.manifest["extension_namespaces"]
        )
        self.assertEqual(
            sdl_control["metadata"]["sdlxliff"]["extension_attributes"], []
        )

    def test_mixed_content_preserves_tags_tails_whitespace_and_extensions(self):
        result = read_sdlxliff(
            FIXTURES / "extensions.sdlxliff", options=SDLXLIFFOptions()
        )
        segment = result.segments[0]
        metadata = segment["metadata"]["sdlxliff"]

        self.assertTrue(segment["source"].startswith(" "))
        self.assertTrue(segment["source"].endswith(" "))
        self.assertIn('<g id="g1">', segment["source"])
        self.assertIn("尾</g>", segment["source"])
        self.assertEqual(segment["source_plain"], EXPECTED_SOURCE_PLAIN)
        self.assertEqual(metadata["source_tag_signature"], EXPECTED_SIGNATURE)
        self.assertIn("urn:vendor:test", metadata["source_raw_xml"])
        self.assertIn("urn:vendor:test", metadata["extension_xml"][0])
        self.assertIn("urn:vendor:test", result.manifest["extension_namespaces"])
        self.assertNotIn(
            "QUJDREVGR0g=",
            json.dumps(
                {"manifest": result.manifest, "segments": result.segments},
                ensure_ascii=False,
            ),
        )
        self.assertEqual(
            result.manifest["files"][0]["internal_file"],
            {"present": True, "size": 12},
        )

    def test_public_serializer_has_literal_display_plain_raw_and_signature(self):
        element = ET.fromstring(
            f'<source xmlns="{XLIFF_NS}" xmlns:v="urn:vendor:test"> '
            '<g id="g1">甲<x id="x1"/>乙</g><v:meta code="m">丙</v:meta> '
            "</source>"
        )

        serialized = serialize_mixed(
            element,
            {"": XLIFF_NS, "v": "urn:vendor:test"},
        )

        self.assertIsInstance(serialized, SerializedMixedContent)
        self.assertEqual(
            serialized.display,
            ' <g id="g1">甲<x id="x1"/>乙</g><v:meta code="m">丙</v:meta> ',
        )
        self.assertEqual(serialized.plain, " 甲乙丙 ")
        self.assertIn('xmlns="urn:oasis:names:tc:xliff:document:1.2"', serialized.raw_xml)
        self.assertIn('xmlns:v="urn:vendor:test"', serialized.raw_xml)
        self.assertEqual(
            serialized.tag_signature,
            (
                (XLIFF_QNAME + "g", (("id", "g1"),)),
                (XLIFF_QNAME + "x", (("id", "x1"),)),
            ),
        )

    def test_plain_excludes_native_inline_code_but_keeps_subflow_text(self):
        element = ET.fromstring(
            f'<source xmlns="{XLIFF_NS}">甲'
            '<bpt id="b1">&lt;b&gt;</bpt>乙<sub>丙</sub>'
            '<ept id="e1">&lt;/b&gt;</ept>丁<ph id="p1">{{0}}</ph>戊'
            "</source>"
        )

        serialized = serialize_mixed(element, {"": XLIFF_NS})

        self.assertEqual(serialized.plain, "甲乙丙丁戊")

    def test_xliff2_and_ambiguous_mid_fail_with_file_context(self):
        xliff2 = self.temp_fixture(
            "xliff2.xlf",
            self.document(
                '<trans-unit id="tu"><source>A</source><target>B</target></trans-unit>',
                version="2.0",
                namespace="urn:oasis:names:tc:xliff:document:2.0",
            ),
        )
        bad_mid = self.temp_fixture(
            "bad_mid.sdlxliff",
            self.document(
                '<trans-unit id="tu"><seg-source><mrk mtype="seg" mid="1">A</mrk></seg-source>'
                '<target><mrk mtype="seg" mid="2">B</mrk></target>'
                '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs></trans-unit>'
            ),
        )

        for fixture, message in ((xliff2, "XLIFF 1.2"), (bad_mid, "mid")):
            with self.subTest(fixture=fixture.name):
                with self.assertRaisesRegex(SDLXLIFFImportError, message) as caught:
                    read_sdlxliff(fixture, options=SDLXLIFFOptions())
                self.assertIn(fixture.name, str(caught.exception))

    def test_invalid_xml_namespace_and_missing_source_fail_fast(self):
        invalid = self.temp_fixture("invalid.sdlxliff", "<xliff>")
        fake = self.temp_fixture(
            "fake.sdlxliff",
            '<xliff version="1.2" xmlns:sdl="http://example.invalid/sdl"/>',
        )
        missing = self.temp_fixture(
            "missing-source.sdlxliff",
            self.document('<trans-unit id="tu"><target>B</target></trans-unit>'),
        )

        cases = (
            (invalid, "line"),
            (fake, "XLIFF 1.2"),
            (missing, "source"),
        )
        for fixture, message in cases:
            with self.subTest(fixture=fixture.name):
                with self.assertRaisesRegex(SDLXLIFFImportError, message) as caught:
                    read_sdlxliff(fixture, options=SDLXLIFFOptions())
                self.assertIn(fixture.name, str(caught.exception))

    def test_doctype_entities_are_rejected_for_utf8_and_utf16(self):
        for encoding in ("utf-8", "utf-16"):
            with self.subTest(encoding=encoding):
                fixture = self.root / f"doctype-{encoding}.sdlxliff"
                document = (
                    f'<?xml version="1.0" encoding="{encoding}"?>\n'
                    '<!DOCTYPE xliff [<!ENTITY injected "Expanded vendor text">]>\n'
                    f'<xliff version="1.2" xmlns="{XLIFF_NS}" xmlns:sdl="{SDL_NS}">'
                    '<file original="anonymous.xml" source-language="zh-CN" '
                    'target-language="en-US"><body><trans-unit id="tu">'
                    '<source>&injected;</source><target>Target</target>'
                    '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs>'
                    '</trans-unit></body></file></xliff>'
                )
                fixture.write_bytes(document.encode(encoding))

                with self.assertRaisesRegex(
                    SDLXLIFFImportError, "DOCTYPE|entity"
                ) as caught:
                    read_sdlxliff(fixture, options=SDLXLIFFOptions())
                self.assertIn(fixture.name, str(caught.exception))

    def test_duplicate_business_key_duplicate_mid_and_unbounded_seg_defs_fail(self):
        duplicate_key = self.temp_fixture(
            "duplicate-key.sdlxliff",
            self.document(
                '<trans-unit id="same"><source>A</source><target>B</target>'
                '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs></trans-unit>'
                '<trans-unit id="same"><source>C</source><target>D</target>'
                '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs></trans-unit>'
            ),
        )
        duplicate_mid = self.temp_fixture(
            "duplicate-mid.sdlxliff",
            self.document(
                '<trans-unit id="tu"><seg-source>'
                '<mrk mtype="seg" mid="1">A</mrk><mrk mtype="seg" mid="1">B</mrk>'
                '</seg-source><target><mrk mtype="seg" mid="1">C</mrk></target>'
                '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs></trans-unit>'
            ),
        )
        unbounded = self.temp_fixture(
            "unbounded.sdlxliff",
            self.document(
                '<trans-unit id="tu"><source>A</source><target>B</target><sdl:seg-defs>'
                '<sdl:seg id="1"/><sdl:seg id="2"/></sdl:seg-defs></trans-unit>'
            ),
        )

        for fixture, message in (
            (duplicate_key, "duplicate"),
            (duplicate_mid, "duplicate mid"),
            (unbounded, "seg-def"),
        ):
            with self.subTest(fixture=fixture.name):
                with self.assertRaisesRegex(SDLXLIFFImportError, message):
                    read_sdlxliff(fixture, options=SDLXLIFFOptions())

    def test_segmented_content_requires_matching_seg_definitions(self):
        fixture = self.temp_fixture(
            "missing-seg-defs.sdlxliff",
            self.document(
                '<trans-unit id="tu"><seg-source>'
                '<mrk mtype="seg" mid="1">A</mrk><mrk mtype="seg" mid="2">B</mrk>'
                '</seg-source><target><mrk mtype="seg" mid="1">C</mrk>'
                '<mrk mtype="seg" mid="2">D</mrk></target></trans-unit>'
            ),
        )

        with self.assertRaisesRegex(SDLXLIFFImportError, "seg-def"):
            read_sdlxliff(fixture, options=SDLXLIFFOptions())

    def test_xliff_g_wrappers_preserve_segments_inline_tags_and_ownership(self):
        fixture = FIXTURES / "g_wrapped_segments.sdlxliff"
        try:
            result = read_sdlxliff(fixture, options=SDLXLIFFOptions())
        except SDLXLIFFImportError as exc:
            self.fail(f"XLIFF g-wrapped segmentation must be supported: {exc}")

        first, second = result.segments
        self.assertEqual(
            [first["source_ref"]["sdl_segment_id"], second["source_ref"]["sdl_segment_id"]],
            ["1", "2"],
        )
        self.assertEqual(first["source_plain"], "甲尾")
        self.assertEqual(second["source_plain"], "乙")
        self.assertEqual(
            first["source"],
            '<g id="outer"><g id="inner">甲<x id="x1"/>尾</g></g>',
        )
        self.assertEqual(
            first["metadata"]["sdlxliff"]["source_tag_signature"],
            (
                (XLIFF_QNAME + "g", (("id", "outer"),)),
                (XLIFF_QNAME + "g", (("id", "inner"),)),
                (XLIFF_QNAME + "x", (("id", "x1"),)),
            ),
        )
        self.assertEqual(
            second["target"],
            '<g id="outer"><g id="inner">Beta</g></g>',
        )
        first_extensions = "\n".join(
            first["metadata"]["sdlxliff"]["extension_xml"]
        )
        second_extensions = "\n".join(
            second["metadata"]["sdlxliff"]["extension_xml"]
        )
        self.assertIn("First wrapper segment", first_extensions)
        self.assertNotIn("Second wrapper segment", first_extensions)
        self.assertIn("Second wrapper segment", second_extensions)
        self.assertNotIn("First wrapper segment", second_extensions)
        first_attributes = {
            (item["name"], item["value"])
            for item in first["metadata"]["sdlxliff"]["extension_attributes"]
        }
        second_attributes = {
            (item["name"], item["value"])
            for item in second["metadata"]["sdlxliff"]["extension_attributes"]
        }
        attribute_name = "{urn:vendor:g-wrapper}segment-meta"
        self.assertEqual(
            {value for name, value in first_attributes if name == attribute_name},
            {"source-one", "target-one"},
        )
        self.assertEqual(
            {value for name, value in second_attributes if name == attribute_name},
            {"source-two", "target-two"},
        )

    def test_xliff_g_wrapper_rejects_text_outside_segment_boundaries(self):
        fixture = self.temp_fixture(
            "ambiguous-g-wrapper.sdlxliff",
            self.document(
                '<trans-unit id="g"><seg-source><g id="g1">Outside'
                '<mrk mtype="seg" mid="1">A</mrk></g></seg-source>'
                '<target><g id="g1"><mrk mtype="seg" mid="1">B</mrk>'
                '</g></target><sdl:seg-defs><sdl:seg id="1"/>'
                '</sdl:seg-defs></trans-unit>'
            ),
        )

        with self.assertRaisesRegex(
            SDLXLIFFImportError, "text outside segmentation"
        ):
            read_sdlxliff(fixture, options=SDLXLIFFOptions())

    def test_tu_extension_cannot_hide_a_second_structural_boundary(self):
        fixture = self.temp_fixture(
            "hidden-tu-boundary.sdlxliff",
            f'<xliff version="1.2" xmlns="{XLIFF_NS}" xmlns:sdl="{SDL_NS}" '
            'xmlns:v="urn:vendor:ambiguous"><file original="anonymous.xml"><body>'
            '<trans-unit id="outer"><source>A</source><target>B</target>'
            '<v:wrapper><source>Hidden</source></v:wrapper>'
            '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs>'
            "</trans-unit></body></file></xliff>",
        )

        with self.assertRaisesRegex(
            SDLXLIFFImportError, "urn:vendor:ambiguous"
        ) as caught:
            read_sdlxliff(fixture, options=SDLXLIFFOptions())
        self.assertIn("hidden-tu-boundary.sdlxliff", str(caught.exception))
        self.assertIn("TU 'outer'", str(caught.exception))

    def test_vendor_inline_cannot_hide_nested_segmentation_boundary(self):
        fixture = self.temp_fixture(
            "nested-segment-boundary.sdlxliff",
            f'<xliff version="1.2" xmlns="{XLIFF_NS}" xmlns:sdl="{SDL_NS}" '
            'xmlns:v="urn:vendor:inline"><file original="anonymous.xml"><body>'
            '<trans-unit id="outer"><seg-source><mrk mtype="seg" mid="1">'
            'A<v:inline><mrk mtype="seg" mid="hidden">B</mrk></v:inline>'
            '</mrk></seg-source><target><mrk mtype="seg" mid="1">C</mrk></target>'
            '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs>'
            "</trans-unit></body></file></xliff>",
        )

        with self.assertRaisesRegex(SDLXLIFFImportError, "nested.*segmentation") as caught:
            read_sdlxliff(fixture, options=SDLXLIFFOptions())
        self.assertIn("nested-segment-boundary.sdlxliff", str(caught.exception))
        self.assertIn("TU 'outer'", str(caught.exception))

    def test_nested_vendor_metadata_is_preserved_once_outside_mixed_content(self):
        fixture = self.temp_fixture(
            "nested-metadata.sdlxliff",
            f'<xliff version="1.2" xmlns="{XLIFF_NS}" xmlns:sdl="{SDL_NS}" '
            'xmlns:v="urn:vendor:metadata"><file original="anonymous.xml"><body>'
            '<trans-unit id="tu"><source>A<v:inline>Visible</v:inline></source>'
            '<target>B</target><sdl:seg-defs><sdl:seg id="1"><sdl:values>'
            '<v:note code="anonymous"><v:item>Nested anonymous metadata</v:item></v:note>'
            '</sdl:values></sdl:seg></sdl:seg-defs></trans-unit>'
            "</body></file></xliff>",
        )

        result = read_sdlxliff(fixture, options=SDLXLIFFOptions())
        extension_xml = result.segments[0]["metadata"]["sdlxliff"]["extension_xml"]

        self.assertEqual(len(extension_xml), 1)
        self.assertIn("urn:vendor:metadata", extension_xml[0])
        self.assertEqual(extension_xml[0].count("Nested anonymous metadata"), 1)
        self.assertNotIn("v:inline", extension_xml[0])
        self.assertIn("v:inline", result.segments[0]["source"])

    def test_segment_extensions_keep_seg_def_ownership(self):
        result = read_sdlxliff(
            FIXTURES / "segmented_extensions.sdlxliff",
            options=SDLXLIFFOptions(),
        )
        first = "\n".join(
            result.segments[0]["metadata"]["sdlxliff"]["extension_xml"]
        )
        second = "\n".join(
            result.segments[1]["metadata"]["sdlxliff"]["extension_xml"]
        )

        self.assertIn("Shared anonymous metadata", first)
        self.assertIn("First segment metadata", first)
        self.assertNotIn("Second segment metadata", first)
        self.assertIn("Shared anonymous metadata", second)
        self.assertIn("Second segment metadata", second)
        self.assertNotIn("First segment metadata", second)

    def test_direct_boundaries_are_unique_but_source_and_seg_source_can_coexist(self):
        legal = self.temp_fixture(
            "source-and-seg-source.sdlxliff",
            self.document(
                '<trans-unit id="legal"><source>Full source</source><seg-source>'
                '<mrk mtype="seg" mid="1">Segment source</mrk></seg-source>'
                '<target><mrk mtype="seg" mid="1">Target</mrk></target>'
                '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs></trans-unit>'
            ),
        )
        legal_result = read_sdlxliff(legal, options=SDLXLIFFOptions())
        self.assertEqual(legal_result.segments[0]["source_plain"], "Segment source")

        duplicate_nodes = {
            "source": (
                '<source>A</source><source>B</source><target>C</target>'
                '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs>'
            ),
            "target": (
                '<source>A</source><target>B</target><target>C</target>'
                '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs>'
            ),
            "seg-source": (
                '<seg-source><mrk mtype="seg" mid="1">A</mrk></seg-source>'
                '<seg-source><mrk mtype="seg" mid="2">B</mrk></seg-source>'
                '<target><mrk mtype="seg" mid="1">C</mrk></target>'
                '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs>'
            ),
        }
        for name, content in duplicate_nodes.items():
            with self.subTest(name=name):
                fixture = self.temp_fixture(
                    f"duplicate-{name}.sdlxliff",
                    self.document(f'<trans-unit id="duplicate">{content}</trans-unit>'),
                )
                with self.assertRaisesRegex(
                    SDLXLIFFImportError, f"multiple direct {name}"
                ) as caught:
                    read_sdlxliff(fixture, options=SDLXLIFFOptions())
                self.assertIn(fixture.name, str(caught.exception))
                self.assertIn("TU 'duplicate'", str(caught.exception))

    def test_structural_unknown_extension_fails_instead_of_guessing_boundaries(self):
        fixture = self.temp_fixture(
            "ambiguous-extension.sdlxliff",
            f'<xliff version="1.2" xmlns="{XLIFF_NS}" xmlns:sdl="{SDL_NS}" '
            'xmlns:v="urn:vendor:ambiguous"><file original="anonymous.xml"><body>'
            '<trans-unit id="tu"><seg-source><v:segments>'
            '<mrk mtype="seg" mid="1">A</mrk></v:segments></seg-source>'
            '<target><mrk mtype="seg" mid="1">B</mrk></target>'
            '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs>'
            "</trans-unit></body></file></xliff>",
        )

        with self.assertRaisesRegex(SDLXLIFFImportError, "urn:vendor:ambiguous"):
            read_sdlxliff(fixture, options=SDLXLIFFOptions())


class SDLXLIFFRuleTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def temp_fixture(self, relative_path: str, body: str) -> Path:
        path = self.root / relative_path
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            f'<xliff version="1.2" xmlns="{XLIFF_NS}" xmlns:sdl="{SDL_NS}">'
            '<file original="anonymous.xml" source-language="zh-CN" target-language="en-US">'
            f"<body>{body}</body></file></xliff>",
            encoding="utf-8",
        )
        return path

    def copy_fixture(self, source: Path, relative_path: str) -> Path:
        target = self.root / relative_path
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, target)
        return target

    def test_profile_rules_are_strict_and_auditable(self):
        self.temp_fixture(
            "story/dialog_main.sdlxliff",
            '<trans-unit id="keep"><source>甲</source><target>Alpha</target>'
            '<sdl:seg-defs><sdl:seg id="1" conf="Translated"/></sdl:seg-defs>'
            '</trans-unit><trans-unit id="drop"><source>乙</source><target>Beta</target>'
            '<sdl:seg-defs><sdl:seg id="2" conf="Rejected"/></sdl:seg-defs></trans-unit>',
        )
        options = validate_options(
            {
                "tm_protection": "candidate-only",
                "content_type_rules": [
                    {
                        "id": "dialog",
                        "glob": "**/dialog*.sdlxliff",
                        "content_type": "剧情/对话",
                    }
                ],
                "exclude_rules": [
                    {
                        "id": "rejected",
                        "field": "confirmation",
                        "equals": "Rejected",
                        "reason": "Client excluded",
                    }
                ],
            }
        )

        result = read_sdlxliff(self.root, options=options)

        self.assertEqual([segment["id"] for segment in result.segments], [0])
        self.assertEqual(result.segments[0]["content_type"], "剧情/对话")
        self.assertEqual(result.segments[0]["content_type_rule_id"], "dialog")
        self.assertEqual(result.manifest["excluded"][0]["rule_ids"], ["rejected"])
        self.assertEqual(result.manifest["excluded"][0]["reasons"], ["Client excluded"])
        self.assertEqual(result.manifest["counts"]["included_segments"], 1)
        self.assertEqual(result.manifest["counts"]["excluded_segments"], 1)

    def test_options_schema_rejects_unknown_invalid_and_ambiguous_rules(self):
        cases = (
            ([], "object"),
            ({"unknown": True}, "unknown"),
            ({"tm_protection": "always"}, "tm_protection"),
            ({"tm_protection": []}, "tm_protection"),
            ({"content_type_rules": {}}, "content_type_rules"),
            ({"content_type_rules": ["not-an-object"]}, "object"),
            (
                {"content_type_rules": [{"id": " ", "glob": "*.sdlxliff", "content_type": "UI"}]},
                "id",
            ),
            (
                {"content_type_rules": [{"id": "x", "glob": "*.sdlxliff", "content_type": "UI", "extra": True}]},
                "extra",
            ),
            (
                {"content_type_rules": [{"id": "x", "glob": "[bad", "content_type": "UI"}]},
                "glob",
            ),
            (
                {"content_type_rules": [{"id": "x", "glob": "*.sdlxliff", "content_type": ""}]},
                "content_type",
            ),
            (
                {"exclude_rules": [{"id": "bad", "field": "segment.id", "equals": 1, "reason": "unstable"}]},
                "segment.id",
            ),
            (
                {"exclude_rules": [{"id": "x", "field": "source", "reason": "missing matcher"}]},
                "exactly one",
            ),
            (
                {"exclude_rules": [{"id": "x", "field": "source", "equals": "A", "regex": "A", "reason": "two"}]},
                "exactly one",
            ),
            (
                {"exclude_rules": [{"id": "x", "field": "source", "regex": "(", "reason": "bad regex"}]},
                "regex",
            ),
            (
                {"exclude_rules": [{"id": "x", "field": "source", "equals": "A", "reason": " "}]},
                "reason",
            ),
            (
                {"exclude_rules": [{"id": "x", "field": "source", "equals": float("nan"), "reason": "non-finite"}]},
                "equals",
            ),
            (
                {"exclude_rules": [{"id": "x", "field": "source", "equals": float("inf"), "reason": "non-finite"}]},
                "equals",
            ),
            (
                {
                    "content_type_rules": [{"id": "same", "glob": "*.sdlxliff", "content_type": "UI"}],
                    "exclude_rules": [{"id": "same", "field": "source", "equals": "A", "reason": "duplicate"}],
                },
                "same",
            ),
            (
                {"exclude_rules": [{"id": "blank-both-sides", "field": "source", "equals": "", "reason": "collision"}]},
                "blank-both-sides",
            ),
        )
        for raw, message in cases:
            with self.subTest(raw=raw):
                with self.assertRaisesRegex(SDLXLIFFImportError, message):
                    validate_options(raw)

    def test_cli_exact_tm_flag_overrides_profile_policy(self):
        options = validate_options(
            {"tm_protection": "candidate-only"}, cli_protect_exact_tm=True
        )

        self.assertEqual(options.tm_protection, "protect-exact-source-and-target")

    def test_rule_matchers_are_ordered_case_sensitive_and_complete(self):
        content_rules = validate_options(
            {
                "content_type_rules": [
                    {"id": "first", "glob": "**/dialog*.sdlxliff", "content_type": "Dialogue"},
                    {"id": "second", "glob": "**/*.sdlxliff", "content_type": "Fallback"},
                ]
            }
        ).content_type_rules
        self.assertEqual(
            match_content_type("story/dialog_main.sdlxliff", content_rules),
            ("Dialogue", "first"),
        )
        self.assertEqual(
            match_content_type("dialog_root.sdlxliff", content_rules),
            ("Dialogue", "first"),
        )
        self.assertEqual(
            match_content_type("story/Dialog_main.sdlxliff", content_rules),
            ("Fallback", "second"),
        )

        exclusion_rules = validate_options(
            {
                "exclude_rules": [
                    {"id": "locked", "field": "locked", "equals": True, "reason": "Locked"},
                    {"id": "target", "field": "target", "regex": "^Skip", "glob": "**/*.sdlxliff", "reason": "Target"},
                    {"id": "other", "field": "target", "equals": "Skip me", "glob": "other/**", "reason": "Other"},
                ]
            }
        ).exclude_rules
        matches = match_exclusions(
            {
                "relative_path": "story/dialog_main.sdlxliff",
                "file_original": "dialog.xml",
                "confirmation": "Translated",
                "origin": None,
                "locked": True,
                "source": "甲",
                "target": "Skip me",
            },
            exclusion_rules,
        )
        self.assertEqual([match["id"] for match in matches], ["locked", "target"])

    def test_exact_tm_requires_all_three_case_insensitive_conditions(self):
        self.assertTrue(
            is_exact_tm(
                {"origin": "TM", "match_percent": "100.0", "text_match": "sourceandtarget"}
            )
        )
        for metadata in (
            {"origin": "tm", "match_percent": "100"},
            {"origin": "machine", "match_percent": "100", "text_match": "SourceAndTarget"},
            {"origin": "tm", "match_percent": "99.999", "text_match": "SourceAndTarget"},
            {"origin": "tm", "match_percent": "not-a-number", "text_match": "SourceAndTarget"},
            {"origin": "tm", "match_percent": "100", "text_match": "Fuzzy"},
        ):
            with self.subTest(metadata=metadata):
                self.assertFalse(is_exact_tm(metadata))

    def test_tm_candidate_and_locked_protection_are_separate(self):
        default = read_sdlxliff(
            FIXTURES / "multi_segment.sdlxliff", options=SDLXLIFFOptions()
        )
        self.assertFalse(default.segments[0].get("protected", False))
        self.assertEqual(default.tm_candidates["candidate_ids"], [0])
        self.assertEqual(default.segments[1]["protected_reason"], "SOURCE_LOCKED")
        self.assertNotIn("protected_ids", default.tm_candidates)

        strict = read_sdlxliff(
            FIXTURES / "multi_segment.sdlxliff",
            options=SDLXLIFFOptions(
                tm_protection="protect-exact-source-and-target"
            ),
        )
        self.assertEqual(strict.segments[0]["protected_reason"], "TM_100_MATCH")
        self.assertEqual(strict.segments[1]["protected_reason"], "SOURCE_LOCKED")

        filtered_fixture = self.temp_fixture(
            "filtered-candidate.sdlxliff",
            '<trans-unit id="blank"><source/><target/>'
            '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs></trans-unit>'
            '<trans-unit id="candidate"><source>甲</source><target>Alpha</target>'
            '<sdl:seg-defs><sdl:seg id="2" origin="tm" percent="100" '
            'text-match="SourceAndTarget"/></sdl:seg-defs></trans-unit>',
        )
        filtered = read_sdlxliff(filtered_fixture, options=SDLXLIFFOptions())
        self.assertEqual(filtered.tm_candidates["candidate_ids"], [0])
        self.assertEqual(filtered.segments[0]["source_ref"]["tu_id"], "candidate")

    def test_locked_wins_but_exact_tm_evidence_remains_candidate(self):
        fixture = self.temp_fixture(
            "both.sdlxliff",
            '<trans-unit id="both"><source>甲</source><target>Alpha</target>'
            '<sdl:seg-defs><sdl:seg id="1" locked="true" origin="tm" percent="100" '
            'text-match="SourceAndTarget"/></sdl:seg-defs></trans-unit>',
        )
        result = read_sdlxliff(
            fixture,
            options=SDLXLIFFOptions(
                tm_protection="protect-exact-source-and-target"
            ),
        )

        self.assertEqual(result.segments[0]["protected_reason"], "SOURCE_LOCKED")
        self.assertEqual(result.tm_candidates["candidate_ids"], [0])
        evidence = result.manifest["protection_evidence"][0]
        self.assertTrue(evidence["locked"]["matched"])
        self.assertTrue(evidence["tm"]["candidate"])
        self.assertTrue(evidence["tm"]["protected_by_policy"])
        self.assertEqual(evidence["effective_reason"], "SOURCE_LOCKED")

    def test_100_percent_alone_is_not_an_exact_tm_candidate(self):
        fixture = self.temp_fixture(
            "negative.sdlxliff",
            '<trans-unit id="negative"><source>甲</source><target>Alpha</target>'
            '<sdl:seg-defs><sdl:seg id="1" origin="machine" percent="100" '
            'text-match="SourceAndTarget"/></sdl:seg-defs></trans-unit>',
        )
        result = read_sdlxliff(
            fixture,
            options=SDLXLIFFOptions(
                tm_protection="protect-exact-source-and-target"
            ),
        )

        self.assertEqual(result.tm_candidates["candidate_ids"], [])
        self.assertFalse(any(segment.get("protected") for segment in result.segments))

    def test_blank_both_is_excluded_but_single_blank_is_kept(self):
        result = read_sdlxliff(RECURSIVE_FIXTURES, options=SDLXLIFFOptions())
        by_tu = {segment["source_ref"]["tu_id"]: segment for segment in result.segments}

        self.assertIn("empty-target", by_tu)
        self.assertNotIn("blank-both", by_tu)
        blank = next(
            item
            for item in result.manifest["excluded"]
            if item["source_ref"]["tu_id"] == "blank-both"
        )
        self.assertEqual(blank["rule_ids"], ["blank-both-sides"])
        self.assertEqual([segment["id"] for segment in result.segments], list(range(len(result.segments))))
        self.assertEqual(len(result.rows_raw), len(result.segments))
        self.assertEqual(
            by_tu["tm-negative-origin"]["source_ref"]["tu_index"], 3
        )

    def test_inline_only_segment_is_not_treated_as_blank(self):
        fixture = self.temp_fixture(
            "inline-only.sdlxliff",
            '<trans-unit id="inline"><source><ph id="source-placeholder"/></source>'
            '<target><ph id="target-placeholder"/></target>'
            '<sdl:seg-defs><sdl:seg id="1"/></sdl:seg-defs></trans-unit>',
        )

        result = read_sdlxliff(fixture, options=SDLXLIFFOptions())

        self.assertEqual(len(result.segments), 1)
        self.assertEqual(result.manifest["excluded"], [])

    def test_exclusion_preserves_segment_extension_ownership(self):
        options = validate_options(
            {
                "exclude_rules": [
                    {
                        "id": "drop-first",
                        "field": "source",
                        "equals": "甲",
                        "reason": "Drop first segment",
                    }
                ]
            }
        )

        result = read_sdlxliff(
            FIXTURES / "segmented_extensions.sdlxliff", options=options
        )

        self.assertEqual(len(result.segments), 1)
        extension_xml = "\n".join(
            result.segments[0]["metadata"]["sdlxliff"]["extension_xml"]
        )
        self.assertIn("Shared anonymous metadata", extension_xml)
        self.assertIn("Second segment metadata", extension_xml)
        self.assertNotIn("First segment metadata", extension_xml)

    def test_no_filename_or_directory_content_type_inference(self):
        self.copy_fixture(
            FIXTURES / "multi_segment.sdlxliff",
            "CC/FF/dialogs.sdlxliff",
        )

        result = read_sdlxliff(self.root, options=SDLXLIFFOptions())

        self.assertTrue(all(not segment.get("content_type") for segment in result.segments))
        self.assertEqual(result.manifest["content_type_matches"], [])

    def test_manifest_is_json_serializable_and_has_file_hashes(self):
        selected = self.copy_fixture(
            FIXTURES / "extensions.sdlxliff", "nested/extensions.sdlxliff"
        )
        (self.root / "notes.csv").write_text("source,target\n", encoding="utf-8")

        result = read_sdlxliff(self.root, options=SDLXLIFFOptions())
        encoded = json.dumps(
            result.manifest, ensure_ascii=False, sort_keys=True, allow_nan=False
        )
        manifest_file = result.manifest["files"][0]

        self.assertEqual(result.manifest["schema"], "lqe.sdlxliff.import-manifest")
        self.assertEqual(result.manifest["version"], 1)
        self.assertEqual(manifest_file["relative_path"], "nested/extensions.sdlxliff")
        self.assertEqual(
            manifest_file["sha256"], hashlib.sha256(selected.read_bytes()).hexdigest()
        )
        self.assertEqual(len(manifest_file["sha256"]), 64)
        self.assertEqual(result.manifest["tm_protection"], "candidate-only")
        self.assertEqual(result.manifest["unselected_supported_files"], ["notes.csv"])
        self.assertIn("urn:vendor:test", result.manifest["extension_namespaces"])
        self.assertNotIn("QUJDREVGR0g=", encoded)
        self.assertEqual(manifest_file["internal_file"], {"present": True, "size": 12})
        self.assertEqual(
            result.manifest["counts"],
            {
                "selected_files": 1,
                "unselected_supported_files": 1,
                "parsed_segments": 1,
                "included_segments": 1,
                "excluded_segments": 0,
                "content_type_matches": 0,
                "tm_candidates": 0,
                "locked_segments": 0,
                "protected_segments": 0,
            },
        )
        self.assertEqual(len(result.manifest["protection_evidence"]), 1)
        evidence = result.manifest["protection_evidence"][0]
        self.assertTrue(evidence["included"])
        self.assertFalse(evidence["locked"]["matched"])
        self.assertFalse(evidence["tm"]["candidate"])


class SDLXLIFFIntegrationTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.job = self.root / "job"
        self.fixture_dir = self.root / "sdl"
        shutil.copytree(RECURSIVE_FIXTURES, self.fixture_dir)

        self.mixed_input_dir = self.root / "mixed-input"
        shutil.copytree(RECURSIVE_FIXTURES, self.mixed_input_dir)
        (self.mixed_input_dir / "notes.csv").write_text(
            "Source,Target\n甲,Alpha\n", encoding="utf-8"
        )

        self.csv = self.root / "source.csv"
        self.csv.write_text("Source,Target\n甲,Alpha\n乙,Beta\n", encoding="utf-8")
        self.tsv = self.root / "source.tsv"
        self.tsv.write_text("Source\tTarget\n甲\tAlpha\n乙\tBeta\n", encoding="utf-8")
        self.xlsx = self.root / "source.xlsx"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["Source", "Target"])
        worksheet.append(["甲", "Alpha"])
        worksheet.append(["乙", "Beta"])
        workbook.save(self.xlsx)
        workbook.close()

        self.candidate_profile = self.root / "profile" / "profile.json"
        write_json(
            self.candidate_profile,
            {
                "name": "anonymous/zh-en",
                "language_pair": "zh-en",
                "source_lang": "zh",
                "target_lang": "en",
                "sdlxliff": {
                    "tm_protection": "candidate-only",
                    "content_type_rules": [
                        {
                            "id": "all-dialogue",
                            "glob": "*.sdlxliff",
                            "content_type": "Dialogue",
                        }
                    ],
                },
            },
        )

    def tearDown(self):
        self.tempdir.cleanup()

    def run_script(self, script: Path, *args):
        return subprocess.run(
            [sys.executable, str(script), *map(str, args)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    def run_io(self, *args):
        return self.run_script(IO_SCRIPT, *args)

    def make_sdl(
        self,
        path: Path,
        *,
        source_language: str = "zh-CN",
        target_language: str = "en-US",
        source: str = "甲",
        target: str = "Alpha",
        seg_attributes: str = "",
    ) -> Path:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            f'<xliff version="1.2" xmlns="{XLIFF_NS}" xmlns:sdl="{SDL_NS}">'
            f'<file original="anonymous.xml" source-language="{source_language}" '
            f'target-language="{target_language}"><body><trans-unit id="tu">'
            f'<source>{source}</source><target>{target}</target>'
            f'<sdl:seg-defs><sdl:seg id="1" {seg_attributes}/></sdl:seg-defs>'
            "</trans-unit></body></file></xliff>",
            encoding="utf-8",
        )
        return path

    def read_tabular(self, source: Path) -> dict:
        state_path = self.root / f"job-{source.suffix[1:]}" / "state.json"
        result = self.run_io(
            "read",
            "--input",
            source,
            "--source-col",
            "Source",
            "--target-col",
            "Target",
            "--out",
            state_path,
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        return read_json(state_path)

    def test_language_tag_normalization_and_base_matching(self):
        self.assertEqual(
            lqe_engine.normalize_language_tag(" ZH_hans_cn_x-private "),
            "zh-Hans-CN-x-private",
        )
        self.assertEqual(
            lqe_engine.normalize_language_tag("en-latn-us-x-ab"),
            "en-Latn-US-x-ab",
        )
        self.assertTrue(lqe_engine.language_tags_match("zh", "zh-CN"))
        self.assertTrue(lqe_engine.language_tags_match("en_us", "en-US"))
        self.assertFalse(lqe_engine.language_tags_match("zh-TW", "zh-CN"))
        self.assertFalse(lqe_engine.language_tags_match("", "zh-CN"))
        self.assertTrue(lqe_engine._load_lang("th"))
        self.assertEqual(
            lqe_engine._load_lang("th-TH"),
            lqe_engine._load_lang("th"),
        )

    def test_regional_target_uses_base_language_attributes_and_notes(self):
        state_path = self.job / "state.json"

        result = self.run_io(
            "read",
            "--input",
            self.fixture_dir,
            "--target-lang",
            "en-US",
            "--out",
            state_path,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        state = read_json(state_path)
        notes_path = self.job / "lang_notes.md"
        self.assertEqual(state["lang_notes_path"], str(notes_path))
        self.assertEqual(
            notes_path.read_text(encoding="utf-8"),
            (ROOT / "target_languages" / "en" / "eval_notes.md").read_text(
                encoding="utf-8"
            ),
        )
        self.assertIn("target_languages/en/attributes.json", result.stdout)

    def test_cli_reads_directory_without_source_target_columns(self):
        state_path = self.job / "state.json"
        result = self.run_io(
            "read",
            "--input",
            self.fixture_dir,
            "--input-format",
            "sdlxliff",
            "--source-lang",
            "zh",
            "--target-lang",
            "en",
            "--out",
            state_path,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        state = read_json(state_path)
        self.assertEqual(state["input_format"], "sdlxliff")
        self.assertEqual(
            state["headers"],
            ["来源文件", "TU ID", "SDL Segment ID", "原文", "译文"],
        )
        self.assertEqual(state["source_lang"], "zh-CN")
        self.assertEqual(state["target_lang"], "en-US")
        self.assertEqual(
            state["input_paths"],
            [
                str((self.fixture_dir / "a" / "dialogs.sdlxliff").resolve()),
                str((self.fixture_dir / "b.sdlxliff").resolve()),
            ],
        )
        self.assertEqual(
            state["source_manifest_path"],
            str(self.job / "source_manifest.json"),
        )
        self.assertEqual(
            state["tm_candidates_path"], str(self.job / "tm_candidates.json")
        )
        commented = next(
            segment
            for segment in state["segments"]
            if segment["source_ref"]["tu_id"] == "shared-tu"
            and segment["source_ref"]["relative_path"] == "a/dialogs.sdlxliff"
        )
        self.assertEqual(commented["context_note"], "Anonymous review note")
        self.assertEqual(
            commented["metadata"]["sdlxliff"]["comment"],
            "Anonymous review note",
        )
        self.assertIn("source_plain", commented)
        self.assertIn("target_plain", commented)
        self.assertEqual(
            read_json(self.job / "scope.json"), state["check_scope"]
        )

    def test_tu_direct_comment_populates_segment_context_note(self):
        state_path = self.job / "state.json"
        result = self.run_io(
            "read",
            "--input",
            FIXTURES / "tu_direct_comments.sdlxliff",
            "--out",
            state_path,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        state = read_json(state_path)
        by_source = {segment["source_plain"]: segment for segment in state["segments"]}
        self.assertEqual(
            by_source["甲"]["context_note"],
            "TU direct note\nFirst segment note",
        )
        self.assertEqual(by_source["乙"]["context_note"], "TU direct note")
        self.assertIsNone(by_source["丙"]["context_note"])
        self.assertEqual(by_source["丁"]["context_note"], "Second file note")
        self.assertEqual(by_source["戊"]["context_note"], "Root fallback note")

    def test_language_and_structure_errors_are_atomic(self):
        mixed_language_dir = self.root / "mixed-language"
        self.make_sdl(mixed_language_dir / "a.sdlxliff")
        self.make_sdl(
            mixed_language_dir / "b.sdlxliff",
            source_language="ja-JP",
            target_language="en-US",
        )
        state_path = self.job / "state.json"

        result = self.run_io(
            "read", "--input", mixed_language_dir, "--out", state_path
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("language", result.stderr.lower())
        for name in (
            "state.json",
            "source_manifest.json",
            "tm_candidates.json",
            "scope.json",
        ):
            self.assertFalse((self.job / name).exists(), name)

        invalid_inputs = {
            "invalid-xml.sdlxliff": "<xliff>",
            "missing-source.sdlxliff": (
                f'<xliff version="1.2" xmlns="{XLIFF_NS}" xmlns:sdl="{SDL_NS}">'
                '<file source-language="zh-CN" target-language="en-US"><body>'
                '<trans-unit id="tu"><target>Alpha</target></trans-unit>'
                "</body></file></xliff>"
            ),
        }
        for index, (name, content) in enumerate(invalid_inputs.items()):
            with self.subTest(structure=name):
                source = self.root / name
                source.write_text(content, encoding="utf-8")
                job = self.root / f"invalid-job-{index}"
                result = self.run_io(
                    "read", "--input", source, "--out", job / "state.json"
                )
                self.assertNotEqual(result.returncode, 0)
                for artifact in (
                    "state.json",
                    "source_manifest.json",
                    "tm_candidates.json",
                    "scope.json",
                ):
                    self.assertFalse((job / artifact).exists(), artifact)

    def test_configured_language_mismatch_is_atomic(self):
        state_path = self.job / "state.json"
        result = self.run_io(
            "read",
            "--input",
            self.fixture_dir,
            "--source-lang",
            "ja",
            "--target-lang",
            "en",
            "--out",
            state_path,
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("source language", result.stderr.lower())
        self.assertFalse(self.job.exists())

    def test_cli_base_language_cannot_mask_profile_region_mismatch(self):
        profile = self.root / "mismatched-profile" / "profile.json"
        write_json(
            profile,
            {
                "name": "anonymous/zh-tw-en",
                "language_pair": "zh-TW-en",
                "source_lang": "zh-TW",
                "target_lang": "en",
            },
        )

        result = self.run_io(
            "read",
            "--input",
            self.fixture_dir,
            "--project",
            profile,
            "--source-lang",
            "zh",
            "--out",
            self.job / "state.json",
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("profile source language", result.stderr.lower())
        self.assertFalse(self.job.exists())

    def test_existing_sdl_job_artifact_is_not_overwritten(self):
        for artifact_name in (
            "state.json",
            "source_manifest.json",
            "tm_candidates.json",
            "scope.json",
        ):
            with self.subTest(artifact=artifact_name):
                job = self.root / f"existing-{artifact_name}"
                state_path = job / "state.json"
                artifact = job / artifact_name
                artifact.parent.mkdir(parents=True)
                artifact.write_text("sentinel", encoding="utf-8")

                result = self.run_io(
                    "read", "--input", self.fixture_dir, "--out", state_path
                )

                self.assertNotEqual(result.returncode, 0)
                self.assertIn("already exists", result.stderr)
                self.assertEqual(artifact.read_text(encoding="utf-8"), "sentinel")
                expected = {
                    "state.json",
                    "source_manifest.json",
                    "tm_candidates.json",
                    "scope.json",
                }
                self.assertEqual(
                    {path.name for path in job.iterdir()}, {artifact_name}
                )

    def test_existing_generated_asset_is_not_overwritten(self):
        style_guide = self.root / "style.md"
        style_guide.write_text("# Style\n", encoding="utf-8")
        terminology = self.root / "source-terms.json"
        write_json(terminology, [{"source": "甲", "target": "Alpha"}])
        cases = (
            ("sg.txt", ("--style-guide", style_guide)),
            ("terms.json", ("--terminology", terminology)),
        )

        for asset_name, extra_args in cases:
            with self.subTest(asset=asset_name):
                job = self.root / f"existing-{asset_name}"
                job.mkdir()
                asset = job / asset_name
                asset.write_text("sentinel", encoding="utf-8")

                result = self.run_io(
                    "read",
                    "--input",
                    self.fixture_dir,
                    *extra_args,
                    "--out",
                    job / "state.json",
                )

                self.assertNotEqual(result.returncode, 0)
                self.assertIn("already exists", result.stderr)
                self.assertEqual(asset.read_text(encoding="utf-8"), "sentinel")
                self.assertEqual({path.name for path in job.iterdir()}, {asset_name})

    def test_dangling_asset_symlink_is_not_replaced(self):
        style_guide = self.root / "style.md"
        style_guide.write_text("# Style\n", encoding="utf-8")
        self.job.mkdir()
        asset = self.job / "sg.txt"
        link_target = self.job / "missing-style.txt"
        asset.symlink_to(link_target)

        result = self.run_io(
            "read",
            "--input",
            self.fixture_dir,
            "--style-guide",
            style_guide,
            "--out",
            self.job / "state.json",
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("already exists", result.stderr)
        self.assertTrue(asset.is_symlink())
        self.assertEqual(asset.readlink(), link_target)
        self.assertEqual({path.name for path in self.job.iterdir()}, {"sg.txt"})

    def test_sdl_state_path_cannot_alias_a_helper_artifact(self):
        state_path = self.job / "source_manifest.json"

        result = self.run_io(
            "read", "--input", self.fixture_dir, "--out", state_path
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("reserved", result.stderr)
        self.assertFalse(self.job.exists())

    def test_sdl_state_path_case_variant_cannot_alias_helper_artifact(self):
        state_path = self.job / "SOURCE_MANIFEST.JSON"

        result = self.run_io(
            "read", "--input", self.fixture_dir, "--out", state_path
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("reserved", result.stderr)
        self.assertFalse(self.job.exists())

    def test_publish_interruption_removes_state_helpers_and_staging_files(self):
        state_path = self.job / "state.json"
        staged_asset = self.root / "staged" / "sg.txt"
        staged_asset.parent.mkdir()
        staged_asset.write_text("# Style\n", encoding="utf-8")
        original_publish = lqe_io._publish_staged_file

        def interrupt_before_publish(source, target):
            if target.name == "tm_candidates.json":
                raise KeyboardInterrupt("simulated publication interruption")
            return original_publish(source, target)

        with mock.patch.object(
            lqe_io, "_publish_staged_file", interrupt_before_publish
        ):
            with self.assertRaises(KeyboardInterrupt):
                lqe_io._publish_sdlxliff_job(
                    state_path,
                    manifest={"kind": "manifest"},
                    tm_candidates={"candidate_ids": []},
                    scope={"mode": "standard"},
                    state={"input_format": "sdlxliff"},
                    staged_assets={self.job / "sg.txt": staged_asset},
                )

        self.assertTrue(self.job.exists())
        self.assertEqual(list(self.job.iterdir()), [])

    def test_publish_race_preserves_competing_json_and_rolls_back_owned_files(self):
        state_path = self.job / "state.json"
        staged_asset = self.root / "staged-race" / "sg.txt"
        staged_asset.parent.mkdir()
        staged_asset.write_text("# Style\n", encoding="utf-8")
        original_publish = lqe_io._publish_staged_file

        def create_competing_json(source, target):
            if target.name == "tm_candidates.json":
                target.write_text("competitor", encoding="utf-8")
            return original_publish(source, target)

        with mock.patch.object(
            lqe_io, "_publish_staged_file", create_competing_json
        ):
            with self.assertRaises(FileExistsError):
                lqe_io._publish_sdlxliff_job(
                    state_path,
                    manifest={"kind": "manifest"},
                    tm_candidates={"candidate_ids": []},
                    scope={"mode": "standard"},
                    state={"input_format": "sdlxliff"},
                    staged_assets={self.job / "sg.txt": staged_asset},
                )

        competitor = self.job / "tm_candidates.json"
        self.assertEqual(competitor.read_text(encoding="utf-8"), "competitor")
        self.assertEqual({path.name for path in self.job.iterdir()}, {competitor.name})

    def test_rollback_preserves_competing_replacement_of_published_asset(self):
        state_path = self.job / "state.json"
        staged_asset = self.root / "staged-replacement" / "sg.txt"
        staged_asset.parent.mkdir()
        staged_asset.write_text("# Style\n", encoding="utf-8")
        original_publish = lqe_io._publish_staged_file

        def replace_asset_then_interrupt(source, target):
            if target.name == "source_manifest.json":
                published_asset = self.job / "sg.txt"
                published_asset.unlink()
                published_asset.write_text("competitor", encoding="utf-8")
                raise KeyboardInterrupt("simulated competing replacement")
            return original_publish(source, target)

        with mock.patch.object(
            lqe_io, "_publish_staged_file", replace_asset_then_interrupt
        ):
            with self.assertRaises(KeyboardInterrupt):
                lqe_io._publish_sdlxliff_job(
                    state_path,
                    manifest={"kind": "manifest"},
                    tm_candidates={"candidate_ids": []},
                    scope={"mode": "standard"},
                    state={"input_format": "sdlxliff"},
                    staged_assets={self.job / "sg.txt": staged_asset},
                )

        competitor = self.job / "sg.txt"
        self.assertEqual(competitor.read_text(encoding="utf-8"), "competitor")
        self.assertEqual({path.name for path in self.job.iterdir()}, {competitor.name})

    def test_explicit_sdl_directory_ignores_but_records_other_supported_files(self):
        state_path = self.job / "state.json"
        result = self.run_io(
            "read",
            "--input",
            self.mixed_input_dir,
            "--input-format",
            "sdlxliff",
            "--out",
            state_path,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        manifest = read_json(self.job / "source_manifest.json")
        self.assertEqual(manifest["unselected_supported_files"], ["notes.csv"])

    def test_tabular_read_still_requires_columns(self):
        result = self.run_io(
            "read", "--input", self.csv, "--out", self.job / "state.json"
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("--source-col", result.stderr)
        self.assertFalse(self.job.exists())

    def test_tabular_csv_tsv_and_xlsx_success_paths_are_unchanged(self):
        for source in (self.csv, self.tsv, self.xlsx):
            with self.subTest(source=source.name):
                state = self.read_tabular(source)
                self.assertEqual(state["input_format"], "tabular")
                self.assertEqual(
                    [segment["source"] for segment in state["segments"]],
                    ["甲", "乙"],
                )

    def test_protect_exact_tm_is_rejected_for_tabular_input(self):
        result = self.run_io(
            "read",
            "--input",
            self.csv,
            "--source-col",
            "Source",
            "--target-col",
            "Target",
            "--protect-exact-tm",
            "--out",
            self.job / "state.json",
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("--protect-exact-tm", result.stderr)
        self.assertFalse(self.job.exists())

    def test_cli_exact_tm_overrides_candidate_only_profile(self):
        state_path = self.job / "state.json"
        result = self.run_io(
            "read",
            "--input",
            FIXTURES / "multi_segment.sdlxliff",
            "--project",
            self.candidate_profile,
            "--protect-exact-tm",
            "--out",
            state_path,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        state = read_json(state_path)
        self.assertEqual(state["segments"][0]["protected_reason"], "TM_100_MATCH")
        self.assertEqual(state["segments"][0]["content_type"], "Dialogue")
        self.assertEqual(
            state["segments"][0]["metadata"]["sdlxliff"]["content_type"],
            "Dialogue",
        )

    def test_candidate_file_requires_explicit_protect_segments_decision(self):
        state_path = self.job / "state.json"
        result = self.run_io(
            "read",
            "--input",
            FIXTURES / "multi_segment.sdlxliff",
            "--project",
            self.candidate_profile,
            "--out",
            state_path,
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        state = read_json(state_path)
        self.assertFalse(state["segments"][0].get("protected", False))
        candidates = read_json(self.job / "tm_candidates.json")
        self.assertEqual(candidates["candidate_ids"], [0])
        self.assertEqual(candidates["segments"][0]["id"], 0)

        result = self.run_io(
            "protect-segments",
            "--state",
            state_path,
            "--protected-file",
            self.job / "tm_candidates.json",
            "--reason",
            "TM_100_MATCH",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        protected = read_json(state_path)["segments"][0]
        self.assertTrue(protected["protected"])
        self.assertEqual(protected["protected_reason"], "TM_100_MATCH")

    def test_candidate_ids_are_not_generic_apply_fixes_protected_ids(self):
        candidate_file = self.root / "tm_candidates.json"
        write_json(candidate_file, {"candidate_ids": [0]})
        args = type(
            "ProtectionArgs",
            (),
            {"protected_ids": None, "protected_file": candidate_file},
        )()

        self.assertEqual(lqe_io._protected_ids(args), set())
        self.assertEqual(
            lqe_io._protected_ids(args, allow_candidates=True),
            {0},
        )

    def test_explicit_candidate_protection_preserves_source_locked_reason(self):
        source = self.make_sdl(
            self.root / "locked-exact.sdlxliff",
            seg_attributes=(
                'locked="true" origin="tm" percent="100" '
                'text-match="SourceAndTarget"'
            ),
        )
        state_path = self.job / "state.json"
        result = self.run_io(
            "read",
            "--input",
            source,
            "--project",
            self.candidate_profile,
            "--out",
            state_path,
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(
            read_json(state_path)["segments"][0]["protected_reason"],
            "SOURCE_LOCKED",
        )

        result = self.run_io(
            "protect-segments",
            "--state",
            state_path,
            "--protected-file",
            self.job / "tm_candidates.json",
            "--reason",
            "TM_100_MATCH",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(
            read_json(state_path)["segments"][0]["protected_reason"],
            "SOURCE_LOCKED",
        )

    def test_wordcount_uses_plain_text_not_inline_attributes(self):
        source = self.make_sdl(
            self.root / "inline.sdlxliff",
            source='甲<x id="attribute_should_not_count"/>乙 AB',
        )
        state_path = self.job / "state.json"
        result = self.run_io(
            "read",
            "--input",
            source,
            "--wordcount-basis",
            "source-chars",
            "--out",
            state_path,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(read_json(state_path)["wordcount"], 3)

        target_source = self.make_sdl(
            self.root / "inline-target.sdlxliff",
            target='Alpha<x id="attribute should not count"/> Beta',
        )
        target_state = self.root / "target-wordcount-job" / "state.json"
        result = self.run_io(
            "read",
            "--input",
            target_source,
            "--wordcount-basis",
            "target-words",
            "--out",
            target_state,
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(read_json(target_state)["wordcount"], 2)

    def test_chunk_dedup_and_batch_prompt_keep_review_context(self):
        state_path = self.job / "state.json"
        errors_path = self.job / "errors_precheck.json"
        base = {
            "source": "Same",
            "target": "相同",
            "protected": False,
            "content_type": "UI",
            "text_type_context": "Menu",
            "context_note": "Button label",
        }
        segments = [
            {
                "id": 0,
                **base,
                "source_ref": {"relative_path": "base.sdlxliff"},
            },
            {
                "id": 1,
                **base,
                "source_ref": {"relative_path": "duplicate.sdlxliff"},
            },
            {
                "id": 2,
                **base,
                "protected": True,
                "source_ref": {"relative_path": "protected.sdlxliff"},
            },
            {
                "id": 3,
                **base,
                "content_type": "Dialogue",
                "source_ref": {"relative_path": "content.sdlxliff"},
            },
            {
                "id": 4,
                **base,
                "text_type_context": "Story",
                "source_ref": {"relative_path": "text-context.sdlxliff"},
            },
            {
                "id": 5,
                **base,
                "context_note": "Spoken line",
                "source_ref": {"relative_path": "note.sdlxliff"},
            },
        ]
        write_json(
            state_path,
            {
                "segments": segments,
                "check_scope": {
                    "mode": "no-terminology",
                    "terminology_enabled": False,
                    "enabled_modules": [
                        "precheck_review",
                        "accuracy",
                        "grammar",
                        "naturalness",
                    ],
                    "disabled_modules": [
                        "terminology",
                        "proper_names",
                        "term_audit",
                    ],
                    "source": "test",
                },
            },
        )
        write_json(errors_path, [])
        chunks = self.job / "chunks"

        result = self.run_script(
            CHUNK_SCRIPT,
            "split",
            "--state",
            state_path,
            "--errors",
            errors_path,
            "--outdir",
            chunks,
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        chunk_rows = read_json(chunks / "chunk_00.json")["segments"]
        self.assertEqual([row["id"] for row in chunk_rows], [0, 2, 3, 4, 5])
        self.assertEqual(
            read_json(chunks / "dedup_map.json")["0"],
            [0, 1],
        )
        self.assertEqual(chunk_rows[0]["context_note"], "Button label")
        self.assertEqual(
            chunk_rows[2]["source_ref"], {"relative_path": "content.sdlxliff"}
        )

        result = self.run_script(BATCH_SCRIPT, "plan", "--job", self.job)
        self.assertEqual(result.returncode, 0, result.stderr)
        prompt = (self.job / "batches" / "batch_00.txt").read_text(encoding="utf-8")
        self.assertIn("CONTENT_TYPE: UI", prompt)
        self.assertIn("CONTEXT: Menu | Button label", prompt)
        self.assertIn("CONTENT_TYPE: Dialogue", prompt)
        self.assertIn("CONTEXT: Menu | Spoken line", prompt)


class SDLXLIFFOutputTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.job = self.root / "job"
        self.source = self.root / "source.sdlxliff"
        shutil.copy2(FIXTURES / "multi_segment.sdlxliff", self.source)
        self.input_files = [self.source]

    def tearDown(self):
        self.tempdir.cleanup()

    def run_io(self, *args):
        return subprocess.run(
            [sys.executable, str(IO_SCRIPT), *map(str, args)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )

    @staticmethod
    def sha256(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    @staticmethod
    def issue(*, edit=None):
        return {
            "category": "Grammar",
            "severity": "Minor",
            "comment": "Output fixture issue",
            "needs_confirmation": False,
            "edit": edit,
        }

    def read_job(self):
        state_path = self.job / "state.json"
        result = self.run_io("read", "--input", self.source, "--out", state_path)
        self.assertEqual(result.returncode, 0, result.stderr)
        return state_path

    def finish_sdl_job(self):
        state_path = self.read_job()
        errors_path = self.job / "errors.json"
        write_json(
            errors_path,
            [
                {"id": 0, "errors": [self.issue()], "corrected": None},
                {"id": 1, "errors": [], "corrected": None},
                {"id": 2, "errors": [self.issue()], "corrected": None},
            ],
        )
        result = self.run_io(
            "write",
            "--state",
            state_path,
            "--errors",
            errors_path,
            "--score",
            "99",
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        return state_path, errors_path

    def test_report_has_fixed_eleven_columns_and_per_file_names(self):
        self.finish_sdl_job()
        report = next(self.job.glob("*_lqe.xlsx"))
        workbook = openpyxl.load_workbook(report, data_only=True)
        try:
            sheet = workbook["LQE Results"]
            self.assertEqual(
                [cell.value for cell in sheet[1]],
                [
                    "来源文件",
                    "TU ID",
                    "SDL Segment ID",
                    "原文",
                    "原译",
                    "建议译文",
                    "处理方式",
                    "错误详情",
                    "LQE_Iter",
                    "Protected",
                    "Protection Evidence",
                ],
            )
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            self.assertEqual([row[0] for row in rows], ["dialogs.xml", "dialogs.xml", "ui.xml"])
            self.assertEqual(rows[1][9], "Yes")
            self.assertIn("SOURCE_LOCKED", rows[1][10])
            self.assertIn('"locked"', rows[1][10])
            scorecard_values = [
                cell.value for row in workbook["LQA Scorecard"] for cell in row
            ]
            self.assertIn("dialogs.xml", scorecard_values)
            self.assertIn("ui.xml", scorecard_values)
            self.assertIn("Protected", scorecard_values)
            self.assertIn("Protection Evidence", scorecard_values)
        finally:
            workbook.close()

    def test_report_source_table_rejects_misaligned_tabular_rows(self):
        state = {
            "input_format": "tabular",
            "headers": ["Source", "Target"],
            "rows_raw": [["one", "one"]],
            "segments": [
                {"id": 0, "source": "one", "target": "one"},
                {"id": 1, "source": "two", "target": "two"},
            ],
        }

        with self.assertRaisesRegex(ValueError, "rows_raw"):
            lqe_io._report_source_table(state)

    def test_segment_filename_uses_file_original_then_relative_path(self):
        state = {"input_format": "sdlxliff"}
        segment = {
            "metadata": {"sdlxliff": {"file_original": "dialogs.xml"}},
            "source_ref": {"relative_path": "nested/source.sdlxliff"},
        }
        self.assertEqual(lqe_io._segment_filename(state, segment), "dialogs.xml")
        segment["metadata"]["sdlxliff"]["file_original"] = "  "
        self.assertEqual(
            lqe_io._segment_filename(state, segment), "nested/source.sdlxliff"
        )

    def test_export_creates_corrected_xlsx_without_touching_xml(self):
        state_path = self.read_job()
        errors_path = self.job / "errors.json"
        write_json(
            errors_path,
            [
                {"id": 0, "errors": [], "corrected": None},
                {"id": 1, "errors": [], "corrected": None},
                {
                    "id": 2,
                    "errors": [
                        self.issue(
                            edit={
                                "from": "Start",
                                "to": "Begin",
                                "evidence": None,
                            }
                        )
                    ],
                    "corrected": "Begin",
                },
            ],
        )
        before = {path: self.sha256(path) for path in self.input_files}

        result = self.run_io(
            "export", "--state", state_path, "--errors", errors_path
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        output = next(self.job.glob("*_corrected.xlsx"))
        workbook = openpyxl.load_workbook(output, data_only=True)
        try:
            self.assertEqual(
                [cell.value for cell in workbook.active[1]],
                ["来源文件", "TU ID", "SDL Segment ID", "原文", "译文"],
            )
            rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
            self.assertEqual(rows[1][4], "Second")
            self.assertEqual(rows[2][4], "Begin")
        finally:
            workbook.close()
        self.assertEqual(before, {path: self.sha256(path) for path in self.input_files})

    def test_apply_fixes_preserves_source_locked_reason_and_evidence(self):
        state_path = self.read_job()
        before = read_json(state_path)["segments"][1]["protection_evidence"]
        errors_path = self.job / "errors.json"
        write_json(
            errors_path,
            [
                {
                    "id": 1,
                    "errors": [
                        self.issue(
                            edit={
                                "from": "Second",
                                "to": "Changed",
                                "evidence": None,
                            }
                        )
                    ],
                    "corrected": "Changed",
                }
            ],
        )

        result = self.run_io(
            "apply-fixes", "--state", state_path, "--errors", errors_path
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        updated = read_json(state_path)
        segment = updated["segments"][1]
        self.assertEqual(segment["protected_reason"], "SOURCE_LOCKED")
        self.assertEqual(segment["protection_evidence"], before)
        self.assertIsNone(segment.get("corrected"))
        self.assertEqual(
            updated["error_history"][-1]["skipped_corrections"],
            [
                {
                    "id": 1,
                    "reason": "SOURCE_LOCKED",
                    "evidence": before,
                    "attempted": "Changed",
                }
            ],
        )
        scrubbed = read_json(errors_path)[0]
        self.assertEqual(scrubbed["errors"], [])
        self.assertIsNone(scrubbed["corrected"])


if __name__ == "__main__":
    unittest.main()
