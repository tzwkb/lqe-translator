import json
from difflib import SequenceMatcher as StandardSequenceMatcher
from pathlib import Path
import subprocess
import sys
import tempfile
from types import SimpleNamespace
import unittest
from unittest.mock import patch

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from lqe_excel_diff import build_rich_diff
import lqe_io


def changed_blocks(value):
    return [run for run in value if isinstance(run, TextBlock)]


def issue(*, protected=False, needs_confirmation=False):
    value = {
        "category": "Grammar",
        "severity": "Minor",
        "comment": "Replace the outdated word.",
        "needs_confirmation": needs_confirmation,
        "edit": None,
    }
    if protected:
        value["protected"] = True
    return value


class RichDiffUnitTests(unittest.TestCase):
    def assert_red(self, block, *, strike):
        self.assertEqual(block.font.color.type, "rgb")
        self.assertEqual(block.font.color.rgb, "FFFF0000")
        self.assertEqual(bool(block.font.strike), strike)

    def test_replace_marks_both_sides(self):
        original, suggested = build_rich_diff("Save old file", "Save new file")
        self.assertIsInstance(original, CellRichText)
        self.assertIsInstance(suggested, CellRichText)
        self.assertEqual(str(original), "Save old file")
        self.assertEqual(str(suggested), "Save new file")
        self.assertEqual([block.text for block in changed_blocks(original)], ["old"])
        self.assertEqual([block.text for block in changed_blocks(suggested)], ["new"])
        self.assert_red(changed_blocks(original)[0], strike=True)
        self.assert_red(changed_blocks(suggested)[0], strike=False)

    def test_insert_marks_only_suggestion(self):
        original, suggested = build_rich_diff("Save file", "Save new file")
        self.assertEqual(original, "Save file")
        self.assertEqual([block.text for block in changed_blocks(suggested)], ["new "])
        self.assert_red(changed_blocks(suggested)[0], strike=False)

    def test_delete_marks_only_original(self):
        original, suggested = build_rich_diff("Save old file", "Save file")
        self.assertEqual([block.text for block in changed_blocks(original)], ["old "])
        self.assertEqual(suggested, "Save file")
        self.assert_red(changed_blocks(original)[0], strike=True)

    def test_equal_text_stays_plain(self):
        self.assertEqual(build_rich_diff("same", "same"), ("same", "same"))

    def test_multilingual_and_combining_text_preserves_exact_strings(self):
        cases = [
            ("保存旧文件", "保存新文件"),
            ("บันทึกไฟล์เก่า", "บันทึกไฟล์ใหม่"),
            ("Cafe\u0301 old", "Cafe\u0301 new"),
        ]
        for original_text, suggested_text in cases:
            with self.subTest(original=original_text):
                original, suggested = build_rich_diff(original_text, suggested_text)
                self.assertEqual(str(original), original_text)
                self.assertEqual(str(suggested), suggested_text)
                self.assertTrue(changed_blocks(original))
                self.assertTrue(changed_blocks(suggested))

    def test_wrapped_line_count_keeps_wide_characters_atomic(self):
        for character_count, expected_lines in ((39, 3), (58, 4)):
            with self.subTest(character_count=character_count):
                self.assertEqual(
                    lqe_io._wrapped_line_count("汉" * character_count, 45),
                    expected_lines,
                )

    def test_wrapped_line_count_counts_edge_tabs(self):
        for position, value in (
            ("leading", "\t" + "汉" * 18),
            ("trailing", "汉" * 18 + "\t"),
        ):
            with self.subTest(position=position):
                self.assertEqual(lqe_io._wrapped_line_count(value, 45), 2)

    def test_display_units_count_extended_graphemes_once(self):
        expected_units = {
            "e\u0301": 1,
            "汉": 2,
            "\ufe0f": 0,
            "\u200d": 0,
            "❤\ufe0e": 1,
            "❤️": 2,
            "1️⃣": 2,
            "👨‍👩‍👧‍👦": 2,
            "👍🏽": 2,
            "🇺🇸": 2,
        }
        for value, expected in expected_units.items():
            with self.subTest(value=value):
                self.assertEqual(lqe_io._display_units(value), expected)

    def test_wrapped_line_count_keeps_emoji_graphemes_atomic(self):
        self.assertEqual(
            lqe_io._wrapped_line_count("👨‍👩‍👧‍👦" * 20, 45),
            2,
        )

    def test_wrapped_line_count_handles_line_breaks_and_long_tokens(self):
        self.assertEqual(lqe_io._wrapped_line_count("one\r\n\r\ntwo\rthree", 45), 4)
        self.assertEqual(lqe_io._wrapped_line_count("x" * 40, 45), 2)

    def test_wrapped_row_height_rejects_excel_limit_overflow(self):
        with self.assertRaisesRegex(
            ValueError,
            r"LQE Results row 2.*30 wrapped lines.*497.*409",
        ):
            lqe_io._wrapped_row_height(
                [("\n".join(f"line {index}" for index in range(30)), 45)],
                context="LQE Results row 2",
            )

    def test_write_json_atomic_cleans_staging_after_replace_failure(self):
        with tempfile.TemporaryDirectory() as tempdir:
            root = Path(tempdir)
            path = root / "state.json"
            path.write_bytes(b"original")
            real_replace = lqe_io.os.replace

            def fail_replace(source, destination):
                if Path(destination) == path:
                    raise OSError("injected JSON replace failure")
                return real_replace(source, destination)

            with patch.object(lqe_io.os, "replace", side_effect=fail_replace):
                with self.assertRaisesRegex(
                    OSError,
                    "injected JSON replace failure",
                ):
                    lqe_io._write_json_atomic(path, {"updated": True})

            self.assertEqual(path.read_bytes(), b"original")
            self.assertFalse(list(root.glob(".*.tmp")))

    def test_long_repeated_text_marks_exact_replacement(self):
        original_text = "x" * 250 + "OLD" + "x" * 250
        suggested_text = "x" * 250 + "NEW" + "x" * 250
        original, suggested = build_rich_diff(original_text, suggested_text)
        self.assertEqual(str(original), original_text)
        self.assertEqual(str(suggested), suggested_text)
        self.assertEqual([block.text for block in changed_blocks(original)], ["OLD"])
        self.assertEqual([block.text for block in changed_blocks(suggested)], ["NEW"])

    def test_repeated_text_strips_common_edges_before_matching(self):
        original_text = "x" * 4000 + "OLD" + "x" * 4000
        suggested_text = "x" * 4000 + "NEW" + "x" * 4000
        with patch(
            "lqe_excel_diff.SequenceMatcher",
            wraps=StandardSequenceMatcher,
        ) as matcher:
            original, suggested = build_rich_diff(original_text, suggested_text)

        matcher_original = matcher.call_args.args[1]
        matcher_suggested = matcher.call_args.args[2]
        self.assertLessEqual(len(matcher_original), 3)
        self.assertLessEqual(len(matcher_suggested), 3)
        self.assertFalse(matcher.call_args.kwargs["autojunk"])
        self.assertEqual(str(original), original_text)
        self.assertEqual(str(suggested), suggested_text)

    def test_shifted_repeated_text_uses_bounded_diff(self):
        original_text = "x" * 8000 + "a"
        suggested_text = "a" + "x" * 8000
        with patch(
            "lqe_excel_diff.SequenceMatcher",
            side_effect=AssertionError("large middle must use bounded diff"),
        ):
            original, suggested = build_rich_diff(original_text, suggested_text)
        self.assertEqual([block.text for block in changed_blocks(original)], ["a"])
        self.assertEqual([block.text for block in changed_blocks(suggested)], ["a"])
        self.assertEqual(str(original), original_text)
        self.assertEqual(str(suggested), suggested_text)

    def test_large_unrelated_text_falls_back_to_one_changed_run(self):
        original_text = "a" * 3000
        suggested_text = "b" * 3000
        with patch(
            "lqe_excel_diff.SequenceMatcher",
            side_effect=AssertionError("large middle must use bounded diff"),
        ):
            original, suggested = build_rich_diff(original_text, suggested_text)
        self.assertEqual([block.text for block in changed_blocks(original)], [original_text])
        self.assertEqual([block.text for block in changed_blocks(suggested)], [suggested_text])

    def test_grapheme_clusters_are_not_split_across_runs(self):
        cases = [
            ("Cafe\u0301", "Cafe", ["e\u0301"], ["e"]),
            ("ก่", "ก้", ["ก่"], ["ก้"]),
            ("👩\u200d💻", "👩\u200d🔬", ["👩\u200d💻"], ["👩\u200d🔬"]),
            ("👍🏽", "👍🏻", ["👍🏽"], ["👍🏻"]),
            ("🇺🇸", "🇺🇳", ["🇺🇸"], ["🇺🇳"]),
            ("가", "개", ["가"], ["개"]),
            ("क्ष", "क्य", ["क्ष"], ["क्य"]),
        ]
        for original_text, suggested_text, original_changed, suggested_changed in cases:
            with self.subTest(original=original_text, suggested=suggested_text):
                original, suggested = build_rich_diff(original_text, suggested_text)
                self.assertEqual(
                    [block.text for block in changed_blocks(original)],
                    original_changed,
                )
                self.assertEqual(
                    [block.text for block in changed_blocks(suggested)],
                    suggested_changed,
                )

    def test_over_excel_cell_limit_stays_plain(self):
        original_text = "x" * 32768 + "OLD"
        suggested_text = "x" * 32768 + "NEW"
        with patch("lqe_excel_diff.SequenceMatcher") as matcher:
            original, suggested = build_rich_diff(original_text, suggested_text)
        matcher.assert_not_called()
        self.assertIsInstance(original, str)
        self.assertIsInstance(suggested, str)
        self.assertEqual(original, original_text)
        self.assertEqual(suggested, suggested_text)

    def test_at_excel_cell_limit_still_marks_diff(self):
        original_text = "x" * 32764 + "OLD"
        suggested_text = "x" * 32764 + "NEW"
        original, suggested = build_rich_diff(original_text, suggested_text)
        self.assertEqual([block.text for block in changed_blocks(original)], ["OLD"])
        self.assertEqual([block.text for block in changed_blocks(suggested)], ["NEW"])

    def test_formula_like_rich_text_round_trips_as_text(self):
        with tempfile.TemporaryDirectory() as tempdir:
            path = Path(tempdir) / "formula.xlsx"
            original, suggested = build_rich_diff("=SUM(1,2)", "=SUM(1,3)")
            workbook = openpyxl.Workbook()
            workbook.active["A1"] = original
            workbook.active["B1"] = suggested
            workbook.save(path)
            workbook.close()

            loaded = openpyxl.load_workbook(path, rich_text=True, data_only=False)
            try:
                self.assertEqual(str(loaded.active["A1"].value), "=SUM(1,2)")
                self.assertEqual(str(loaded.active["B1"].value), "=SUM(1,3)")
                self.assertEqual(loaded.active["A1"].data_type, "s")
                self.assertEqual(loaded.active["B1"].data_type, "s")
            finally:
                loaded.close()

    def test_formula_like_plain_side_round_trips_as_text(self):
        with tempfile.TemporaryDirectory() as tempdir:
            path = Path(tempdir) / "formula-insert.xlsx"
            original, suggested = build_rich_diff("=A1", "=A1+1")
            self.assertEqual(str(original), "=A1")
            self.assertEqual(str(suggested), "=A1+1")
            self.assertEqual(changed_blocks(original), [])

            workbook = openpyxl.Workbook()
            workbook.active["A1"] = original
            workbook.active["B1"] = suggested
            workbook.save(path)
            workbook.close()

            loaded = openpyxl.load_workbook(path, rich_text=True, data_only=False)
            try:
                self.assertEqual(str(loaded.active["A1"].value), "=A1")
                self.assertEqual(str(loaded.active["B1"].value), "=A1+1")
                self.assertEqual(loaded.active["A1"].data_type, "s")
                self.assertEqual(loaded.active["B1"].data_type, "s")
            finally:
                loaded.close()


class RichDiffReportTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def build_report(self):
        output = self.root / "sample_lqe.xlsx"
        state = {
            "input_path": str(self.root / "sample.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [
                ["Source 0", "Save old file"],
                ["Source 1", "Keep text"],
                ["Source 2", "Protected text"],
            ],
            "target_col": 1,
            "segments": [
                {"id": 0, "source": "Source 0", "target": "Save old file", "kind": "desc"},
                {"id": 1, "source": "Source 1", "target": "Keep text", "kind": "desc"},
                {"id": 2, "source": "Source 2", "target": "Protected text", "kind": "desc"},
            ],
            "wordcount": 3,
        }
        protected_issue = issue(protected=True)
        history = [{
            "iteration": 0,
            "errors": [
                {"id": 0, "errors": [issue()], "corrected": "Save new file"},
                {"id": 1, "errors": [issue(needs_confirmation=True)], "corrected": None},
                {"id": 2, "errors": [protected_issue], "corrected": None},
            ],
        }]
        lqe_io._build_xlsx(state, history, 99, 98, output)
        return output

    def test_long_wrapped_text_rows_expand_on_both_report_sheets(self):
        output = self.root / "long-wrapped-text_lqe.xlsx"
        short_original = "Save old file"
        short_suggested = "Save new file"
        long_original = "\n".join((
            "Line 1 stays unchanged.",
            "Line 2 stays unchanged.",
            "Line 3 stays unchanged.",
            "Line 4 uses the OLD token.",
            "Line 5 stays unchanged.",
            "Line 6 stays unchanged.",
            "Line 7 stays unchanged.",
            "Line 8 stays unchanged.",
        ))
        long_suggested = "\n".join((
            "Line 1 stays unchanged.",
            "Line 2 stays unchanged.",
            "Line 3 stays unchanged.",
            "Line 4 uses the NEW token.",
            "Line 5 stays unchanged.",
            "Line 6 stays unchanged.",
            "Line 7 stays unchanged.",
            "Line 8 stays unchanged.",
        ))
        state = {
            "input_path": str(self.root / "long-wrapped-text.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [
                ["Short source", short_original],
                ["Long source", long_original],
            ],
            "target_col": 1,
            "segments": [
                {"id": 0, "source": "Short source", "target": short_original, "kind": "desc"},
                {"id": 1, "source": "Long source", "target": long_original, "kind": "desc"},
            ],
            "wordcount": 42,
        }
        history = [{
            "iteration": 0,
            "errors": [
                {"id": 0, "errors": [issue()], "corrected": short_suggested},
                {"id": 1, "errors": [issue()], "corrected": long_suggested},
            ],
        }]
        lqe_io._build_xlsx(state, history, 99, 98, output)

        workbook = openpyxl.load_workbook(output, rich_text=True)
        try:
            scorecard = workbook["LQA Scorecard"]
            score_header_row = next(
                row for row in range(1, scorecard.max_row + 1)
                if scorecard.cell(row=row, column=1).value == "File name"
            )
            score_rows = {
                scorecard.cell(row=row, column=2).value: row
                for row in range(score_header_row + 1, scorecard.max_row + 1)
                if scorecard.cell(row=row, column=2).value in {0, 1}
            }

            results = workbook["LQE Results"]
            result_headers = [cell.value for cell in results[1]]
            result_original_col = result_headers.index("原译") + 1
            result_suggested_col = result_headers.index("建议译文") + 1
            result_rows = {
                str(results.cell(row=row, column=result_original_col).value): row
                for row in range(2, results.max_row + 1)
            }

            sheet_rows = (
                ("LQA Scorecard", scorecard, score_rows[0], score_rows[1], 4, 5),
                (
                    "LQE Results",
                    results,
                    result_rows[short_original],
                    result_rows[long_original],
                    result_original_col,
                    result_suggested_col,
                ),
            )
            for (
                sheet_name,
                sheet,
                short_row,
                long_row,
                original_col,
                suggested_col,
            ) in sheet_rows:
                with self.subTest(sheet=sheet_name):
                    tested_cells = (
                        sheet.cell(short_row, original_col),
                        sheet.cell(short_row, suggested_col),
                        sheet.cell(long_row, original_col),
                        sheet.cell(long_row, suggested_col),
                    )
                    for cell in tested_cells:
                        self.assertTrue(cell.alignment.wrap_text)

                    long_original_cell = sheet.cell(long_row, original_col)
                    long_suggested_cell = sheet.cell(long_row, suggested_col)
                    self.assertIsInstance(long_original_cell.value, CellRichText)
                    self.assertIsInstance(long_suggested_cell.value, CellRichText)
                    self.assertEqual(
                        [block.text for block in changed_blocks(long_original_cell.value)],
                        ["OLD"],
                    )
                    self.assertEqual(
                        [block.text for block in changed_blocks(long_suggested_cell.value)],
                        ["NEW"],
                    )

                    short_height = sheet.row_dimensions[short_row].height
                    long_height = sheet.row_dimensions[long_row].height
                    self.assertIsNotNone(short_height)
                    self.assertIsNotNone(long_height)
                    self.assertGreater(long_height, short_height)
                    self.assertGreaterEqual(long_height, 120)
                    self.assertLessEqual(short_height, 409)
                    self.assertLessEqual(long_height, 409)
        finally:
            workbook.close()

    def test_report_generation_rejects_rows_above_excel_height_limit(self):
        output = self.root / "too-tall_lqe.xlsx"
        original = "\n".join(f"Line {index} uses OLD." for index in range(30))
        suggested = "\n".join(f"Line {index} uses NEW." for index in range(30))
        state = {
            "input_path": str(self.root / "too-tall.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [["Source", original]],
            "target_col": 1,
            "segments": [
                {"id": 0, "source": "Source", "target": original, "kind": "desc"},
            ],
            "wordcount": 120,
        }
        history = [{
            "iteration": 0,
            "errors": [
                {"id": 0, "errors": [issue()], "corrected": suggested},
            ],
        }]

        with self.assertRaisesRegex(
            ValueError,
            r"LQA Scorecard row \d+.*30 wrapped lines.*409",
        ):
            lqe_io._build_xlsx(state, history, 99, 98, output)
        self.assertFalse(output.exists())

    def test_report_generation_accommodates_long_wrapped_legal_copy(self):
        output = self.root / "long-legal-copy_lqe.xlsx"
        paragraph = (
            "This paragraph contains detailed age-rating, privacy, payment, "
            "and account-management information for players and guardians."
        )
        target = "\n".join(f"({index}) {paragraph}" for index in range(1, 10))
        state = {
            "input_path": str(self.root / "long-legal-copy.xlsx"),
            "headers": ["##var", "keyword", "_explanation", "ZH", "EN"],
            "rows_raw": [["", "legal_copy", "", "合规说明" * 80, target]],
            "source_col": "ZH",
            "target_col": "EN",
            "segments": [
                {"id": 0, "source": "合规说明" * 80, "target": target, "kind": "desc"},
            ],
            "wordcount": 120,
        }
        issues = []
        for index in range(8):
            current_issue = issue()
            current_issue["comment"] = (
                f"Glossary candidate {index} needs a contextual terminology review."
            )
            issues.append(current_issue)
        history = [{
            "iteration": 0,
            "errors": [{"id": 0, "errors": issues, "corrected": None}],
        }]

        lqe_io._build_xlsx(state, history, 99, 98, output)

        workbook = openpyxl.load_workbook(output)
        try:
            self.assertLessEqual(
                max(
                    row.height or 0
                    for sheet in (workbook["LQA Scorecard"], workbook["LQE Results"])
                    for row in sheet.row_dimensions.values()
                ),
                409,
            )
        finally:
            workbook.close()

    def test_results_error_details_reject_rows_above_excel_height_limit(self):
        output = self.root / "too-many-errors_lqe.xlsx"
        state = {
            "input_path": str(self.root / "too-many-errors.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [["Source", "Target"]],
            "target_col": 1,
            "segments": [
                {"id": 0, "source": "Source", "target": "Target", "kind": "desc"},
            ],
            "wordcount": 1,
        }
        errors = []
        for index in range(25):
            current_issue = issue()
            current_issue["comment"] = f"Issue {index}"
            errors.append(current_issue)
        history = [{
            "iteration": 0,
            "errors": [{"id": 0, "errors": errors, "corrected": None}],
        }]

        with self.assertRaisesRegex(
            ValueError,
            r"LQE Results row 2.*25 wrapped lines.*414\.5.*409",
        ):
            lqe_io._build_xlsx(state, history, 99, 98, output)
        self.assertFalse(output.exists())

    def test_cmd_write_layout_failure_is_atomic(self):
        job = self.root / "atomic-job"
        job.mkdir()
        state_path = job / "state.json"
        errors_path = job / "errors.json"
        output = job / "atomic-job_lqe.xlsx"
        original = "\n".join(f"Line {index} uses OLD." for index in range(30))
        suggested = "\n".join(f"Line {index} uses NEW." for index in range(30))
        state = {
            "input_path": str(job / "source.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [
                ["Long source", original],
                ["Protected source", "Protected target"],
            ],
            "target_col": 1,
            "segments": [
                {"id": 0, "source": "Long source", "target": original, "kind": "desc"},
                {
                    "id": 1,
                    "source": "Protected source",
                    "target": "Protected target",
                    "kind": "desc",
                    "protected": True,
                },
            ],
            "protected_ids": [1],
            "wordcount": 120,
            "iteration": 0,
            "error_history": [{"iteration": 99, "errors": []}],
        }
        errors = [
            {"id": 0, "errors": [issue()], "corrected": suggested},
            {"id": 1, "errors": [issue(protected=True)], "corrected": None},
        ]
        state_path.write_text(
            json.dumps(state, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        errors_path.write_text(
            json.dumps(errors, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        output.write_bytes(b"existing workbook sentinel")
        state_before = state_path.read_bytes()
        errors_before = errors_path.read_bytes()
        output_before = output.read_bytes()
        args = SimpleNamespace(
            state=str(state_path),
            errors=str(errors_path),
            score="99",
            threshold=98,
            scorecard_profile="legacy",
        )

        with patch.object(lqe_io, "_validate_scope_or_exit"):
            with self.assertRaisesRegex(
                SystemExit,
                r"\[write\].*LQA Scorecard row \d+.*409",
            ):
                lqe_io.cmd_write(args)

        self.assertEqual(state_path.read_bytes(), state_before)
        self.assertEqual(errors_path.read_bytes(), errors_before)
        self.assertEqual(output.read_bytes(), output_before)
        self.assertFalse(list(job.glob(".*.xlsx")))

    def test_cmd_write_publication_failure_rolls_back_all_artifacts(self):
        for failure_target in ("state", "output"):
            with self.subTest(failure_target=failure_target):
                job = self.root / f"publish-{failure_target}"
                job.mkdir()
                state_path = job / "state.json"
                errors_path = job / "errors.json"
                output = job / f"publish-{failure_target}_lqe.xlsx"
                state = {
                    "input_path": str(job / "source.xlsx"),
                    "headers": ["原文", "译文"],
                    "rows_raw": [
                        ["Source", "Old target"],
                        ["Protected source", "Protected target"],
                    ],
                    "target_col": 1,
                    "segments": [
                        {
                            "id": 0,
                            "source": "Source",
                            "target": "Old target",
                            "kind": "desc",
                        },
                        {
                            "id": 1,
                            "source": "Protected source",
                            "target": "Protected target",
                            "kind": "desc",
                            "protected": True,
                        },
                    ],
                    "protected_ids": [1],
                    "wordcount": 2,
                    "iteration": 0,
                    "error_history": [{"iteration": 99, "errors": []}],
                }
                errors = [
                    {
                        "id": 0,
                        "errors": [issue()],
                        "corrected": "New target",
                    },
                    {
                        "id": 1,
                        "errors": [issue(protected=True)],
                        "corrected": None,
                    },
                ]
                state_path.write_text(
                    json.dumps(state, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                errors_path.write_text(
                    json.dumps(errors, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                output.write_bytes(b"existing workbook sentinel")
                originals = {
                    state_path: state_path.read_bytes(),
                    errors_path: errors_path.read_bytes(),
                    output: output.read_bytes(),
                }
                args = SimpleNamespace(
                    state=str(state_path),
                    errors=str(errors_path),
                    score="99",
                    threshold=98,
                    scorecard_profile="legacy",
                )
                target_path = state_path if failure_target == "state" else output
                real_replace = lqe_io.os.replace
                failed = False

                def fail_once(source, destination):
                    nonlocal failed
                    if Path(destination) == target_path and not failed:
                        failed = True
                        raise OSError(f"injected {failure_target} publish failure")
                    return real_replace(source, destination)

                with patch.object(lqe_io, "_validate_scope_or_exit"):
                    with patch.object(lqe_io.os, "replace", side_effect=fail_once):
                        with self.assertRaisesRegex(
                            OSError,
                            f"injected {failure_target} publish failure",
                        ):
                            lqe_io.cmd_write(args)

                for path, original in originals.items():
                    self.assertEqual(path.read_bytes(), original)
                self.assertFalse(list(job.glob(".*.tmp")))
                self.assertFalse(list(job.glob(".*.xlsx")))

    def test_cmd_write_success_publishes_scrubbed_state_errors_and_workbook(self):
        job = self.root / "publish-success"
        job.mkdir()
        state_path = job / "state.json"
        errors_path = job / "errors.json"
        output = job / "publish-success_lqe.xlsx"
        state = {
            "input_path": str(job / "source.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [["Protected source", "Protected target"]],
            "target_col": 1,
            "segments": [
                {
                    "id": 0,
                    "source": "Protected source",
                    "target": "Protected target",
                    "kind": "desc",
                    "protected": True,
                },
            ],
            "protected_ids": [0],
            "wordcount": 1,
            "iteration": 0,
            "error_history": [],
        }
        errors = [
            {
                "id": 0,
                "errors": [issue(protected=True)],
                "corrected": "Must be scrubbed",
            },
        ]
        state_path.write_text(
            json.dumps(state, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        errors_path.write_text(
            json.dumps(errors, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        output.write_bytes(b"existing workbook sentinel")
        args = SimpleNamespace(
            state=str(state_path),
            errors=str(errors_path),
            score="99",
            threshold=98,
            scorecard_profile="legacy",
        )

        with patch.object(lqe_io, "_validate_scope_or_exit"):
            lqe_io.cmd_write(args)

        persisted_errors = json.loads(errors_path.read_text(encoding="utf-8"))
        persisted_state = json.loads(state_path.read_text(encoding="utf-8"))
        self.assertEqual(persisted_errors[0]["errors"], [])
        self.assertIsNone(persisted_errors[0]["corrected"])
        self.assertEqual(
            persisted_state["error_history"][-1]["errors"],
            persisted_errors,
        )
        workbook = openpyxl.load_workbook(output, rich_text=True)
        try:
            self.assertEqual(
                workbook.sheetnames,
                ["说明·导读", "LQA Scorecard", "LQE Results"],
            )
        finally:
            workbook.close()
        self.assertFalse(list(job.glob(".*.tmp")))
        self.assertFalse(list(job.glob(".*.xlsx")))

    def test_both_report_sheets_render_paired_rich_diffs(self):
        output = self.build_report()
        workbook = openpyxl.load_workbook(output, rich_text=True)
        try:
            scorecard = workbook["LQA Scorecard"]
            header_row = next(
                row for row in range(1, scorecard.max_row + 1)
                if scorecard.cell(row=row, column=1).value == "File name"
            )
            score_original = scorecard.cell(header_row + 1, 4).value
            score_suggested = scorecard.cell(header_row + 1, 5).value
            self.assertEqual([block.text for block in changed_blocks(score_original)], ["old"])
            self.assertEqual([block.text for block in changed_blocks(score_suggested)], ["new"])

            results = workbook["LQE Results"]
            headers = [cell.value for cell in results[1]]
            original_col = headers.index("原译") + 1
            suggested_col = headers.index("建议译文") + 1
            result_original = results.cell(2, original_col).value
            result_suggested = results.cell(2, suggested_col).value
            self.assertEqual([block.text for block in changed_blocks(result_original)], ["old"])
            self.assertEqual([block.text for block in changed_blocks(result_suggested)], ["new"])
            self.assertEqual(results.cell(2, original_col).fill.fgColor.rgb, "00FCE5CD")
            self.assertEqual(results.cell(2, suggested_col).fill.fgColor.rgb, "00FCE5CD")
            self.assertIsInstance(results.cell(3, original_col).value, str)
            self.assertNotIsInstance(results.cell(3, suggested_col).value, CellRichText)
            self.assertIsInstance(results.cell(4, original_col).value, str)
            self.assertNotIsInstance(results.cell(4, suggested_col).value, CellRichText)
        finally:
            workbook.close()

    def test_normal_load_keeps_complete_plain_values(self):
        output = self.build_report()
        workbook = openpyxl.load_workbook(output, rich_text=False, data_only=False)
        try:
            results = workbook["LQE Results"]
            headers = [cell.value for cell in results[1]]
            self.assertEqual(results.cell(2, headers.index("原译") + 1).value, "Save old file")
            self.assertEqual(results.cell(2, headers.index("建议译文") + 1).value, "Save new file")
        finally:
            workbook.close()

    def test_missing_target_col_uses_translation_header_for_rich_diff(self):
        output = self.root / "missing-target-col_lqe.xlsx"
        state = {
            "input_path": str(self.root / "sample.xlsx"),
            "headers": ["来源相对路径", "原文", "译文"],
            "rows_raw": [["ui/dialog.xlsx", "Source text", "Save old file"]],
            "segments": [
                {"id": 0, "source": "Source text", "target": "Save old file", "kind": "desc"},
            ],
            "wordcount": 1,
        }
        history = [{
            "iteration": 0,
            "errors": [
                {"id": 0, "errors": [issue()], "corrected": "Save new file"},
            ],
        }]
        lqe_io._build_xlsx(state, history, 99, 98, output)

        workbook = openpyxl.load_workbook(output, rich_text=True)
        try:
            results = workbook["LQE Results"]
            headers = [cell.value for cell in results[1]]
            self.assertEqual(headers[:4], ["来源相对路径", "原文", "原译", "建议译文"])
            self.assertEqual(results.cell(2, 2).value, "Source text")
            original = results.cell(2, 3).value
            suggested = results.cell(2, 4).value
            self.assertEqual(str(original), "Save old file")
            self.assertEqual(str(suggested), "Save new file")
            self.assertEqual([block.text for block in changed_blocks(original)], ["old"])
            self.assertEqual([block.text for block in changed_blocks(suggested)], ["new"])
        finally:
            workbook.close()

    def test_protected_issue_with_corrected_is_hidden_on_both_report_sheets(self):
        output = self.root / "protected_lqe.xlsx"
        state = {
            "input_path": str(self.root / "protected.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [["Source", "Save old file"]],
            "target_col": 1,
            "segments": [{
                "id": 0,
                "source": "Source",
                "target": "Save old file",
                "kind": "desc",
            }],
            "wordcount": 1,
        }
        history = [{
            "iteration": 0,
            "errors": [{
                "id": 0,
                "errors": [issue(protected=True)],
                "corrected": "Save new file",
            }],
        }]
        lqe_io._build_xlsx(state, history, 99, 98, output)

        workbook = openpyxl.load_workbook(output, rich_text=True)
        try:
            scorecard = workbook["LQA Scorecard"]
            header_row = next(
                row for row in range(1, scorecard.max_row + 1)
                if scorecard.cell(row=row, column=1).value == "File name"
            )
            detail_ids = [
                scorecard.cell(row=row, column=2).value
                for row in range(header_row + 1, scorecard.max_row + 1)
            ]
            self.assertNotIn(0, detail_ids)

            results = workbook["LQE Results"]
            headers = [cell.value for cell in results[1]]
            original = results.cell(2, headers.index("原译") + 1).value
            suggested = results.cell(2, headers.index("建议译文") + 1).value
            processing = results.cell(2, headers.index("处理方式") + 1).value
            self.assertIsInstance(original, str)
            self.assertIsNone(suggested)
            self.assertEqual(processing, "已保护，不修改")
        finally:
            workbook.close()


class RichDiffAggregateTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def assert_red(self, block, *, strike):
        self.assertEqual(block.font.color.type, "rgb")
        self.assertEqual(block.font.color.rgb, "FFFF0000")
        self.assertEqual(bool(block.font.strike), strike)

    def build_aggregate(self):
        job = self.root / "multi"
        source = self.root / "source.xlsx"
        workbook = openpyxl.Workbook()
        for index, (sheet_name, verb) in enumerate((("Sheet1", "Save"), ("Sheet2", "Open"))):
            sheet = workbook.active if index == 0 else workbook.create_sheet()
            sheet.title = sheet_name
            sheet.append(["Source", "Target"])
            sheet.append([verb, f"{verb} old file"])
            sheet.append([f"{verb} formula", "=OLD()"])
            sheet["B3"].data_type = "s"
            sheet.append([f"{verb} protected segment", "Protected original"])
            sheet.append([f"{verb} protected issue", "Guarded original"])
        workbook.save(source)
        workbook.close()

        for sheet_name, verb in (("Sheet1", "Save"), ("Sheet2", "Open")):
            child = job / sheet_name
            child.mkdir(parents=True)
            state = {
                "input_path": str(source),
                "sheet_name": sheet_name,
                "target_col": 1,
                "headers": ["Source", "Target"],
                "wordcount": 6,
                "segments": [
                    {
                        "id": 0,
                        "row_index": 0,
                        "source": verb,
                        "target": f"{verb} old file",
                        "kind": "desc",
                        "term_hits": [],
                        "protected_texts": [],
                    },
                    {
                        "id": 1,
                        "row_index": 1,
                        "source": f"{verb} formula",
                        "target": "=OLD()",
                        "kind": "desc",
                        "term_hits": [],
                        "protected_texts": [],
                    },
                    {
                        "id": 2,
                        "row_index": 2,
                        "source": f"{verb} protected segment",
                        "target": "Protected original",
                        "kind": "desc",
                        "protected": True,
                        "term_hits": [],
                        "protected_texts": [],
                    },
                    {
                        "id": 3,
                        "row_index": 3,
                        "source": f"{verb} protected issue",
                        "target": "Guarded original",
                        "kind": "desc",
                        "term_hits": [],
                        "protected_texts": [],
                    },
                ],
            }
            errors = [
                {
                    "id": 0,
                    "errors": [{
                        "category": "Grammar",
                        "severity": "Minor",
                        "comment": "Replace the outdated word.",
                        "needs_confirmation": False,
                        "edit": {
                            "from": "old",
                            "to": "new",
                            "start": 5,
                            "end": 8,
                            "evidence": None,
                        },
                    }],
                    "corrected": f"{verb} new file",
                },
                {
                    "id": 1,
                    "errors": [{
                        "category": "Grammar",
                        "severity": "Minor",
                        "comment": "Keep formula-like text literal.",
                        "needs_confirmation": False,
                        "edit": {
                            "from": "OLD",
                            "to": "NEW",
                            "start": 1,
                            "end": 4,
                            "evidence": None,
                        },
                    }],
                    "corrected": "=NEW()",
                },
                {
                    "id": 2,
                    "errors": [{
                        "category": "Grammar",
                        "severity": "Minor",
                        "comment": "Protected segments must not be changed.",
                        "needs_confirmation": False,
                        "edit": {
                            "from": "original",
                            "to": "tampered",
                            "start": 10,
                            "end": 18,
                            "evidence": None,
                        },
                    }],
                    "corrected": "Protected tampered",
                },
                {
                    "id": 3,
                    "errors": [{
                        "category": "Grammar",
                        "severity": "Minor",
                        "comment": "Issue-level protection must win.",
                        "needs_confirmation": False,
                        "protected": True,
                        "edit": {
                            "from": "original",
                            "to": "tampered",
                            "start": 8,
                            "end": 16,
                            "evidence": None,
                        },
                    }],
                    "corrected": "Guarded tampered",
                },
            ]
            (child / "state.json").write_text(json.dumps(state), encoding="utf-8")
            (child / "errors.json").write_text(json.dumps(errors), encoding="utf-8")

            child_report = openpyxl.Workbook()
            results = child_report.active
            results.title = "LQE Results"
            results.append(["段号", "原译", "上下文", "建议译文", "状态"])
            results.append([0, f"{verb} old file", f"{verb} context", f"{verb} new file", "changed"])
            results.cell(2, 3).value = CellRichText([
                f"{verb} ",
                TextBlock(InlineFont(color="FF0000FF"), "context"),
            ])
            results.cell(1, 1).fill = PatternFill("solid", fgColor="FF112233")
            results.cell(2, 1).fill = PatternFill("solid", fgColor="FF22AA44")
            results.cell(2, 3).alignment = Alignment(wrap_text=True, vertical="top")
            results.cell(2, 3).font = Font(name="Arial", bold=True)
            results.cell(2, 3).border = Border(
                left=Side(style="thin", color="FF334455")
            )
            results.cell(2, 3).number_format = "@"
            results.cell(2, 3).protection = Protection(locked=False, hidden=True)
            results.row_dimensions[2].height = 31
            results.column_dimensions["C"].width = 37
            results.freeze_panes = "A2"
            results.append([1, "=NO_SUGGESTION()", "=1+1", None, "plain"])
            results.cell(3, 2).data_type = "s"
            results.append([2, "=SAME()", "context", "=SAME()", "plain"])
            results.cell(4, 2).data_type = "s"
            results.cell(4, 4).data_type = "s"
            results.append([3, None, "context", f"{verb} added", "changed"])
            child_report.save(child / f"{sheet_name}_lqe.xlsx")
            child_report.close()

        result = subprocess.run(
            [sys.executable, str(SCRIPTS / "aggregate_sheets.py"), "--job", str(job)],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        return job

    def test_aggregate_report_rebuilds_diff_by_headers_on_every_sheet(self):
        job = self.build_aggregate()
        report = openpyxl.load_workbook(job / "multi_lqe.xlsx", rich_text=True)
        try:
            for sheet_name, verb in (("Sheet1", "Save"), ("Sheet2", "Open")):
                with self.subTest(sheet=sheet_name):
                    sheet = report[f"{sheet_name} Results"]
                    headers = [cell.value for cell in sheet[1]]
                    self.assertEqual(headers, ["段号", "原译", "上下文", "建议译文", "状态"])
                    original_column = headers.index("原译") + 1
                    suggested_column = headers.index("建议译文") + 1
                    self.assertEqual(sheet.cell(2, 1).value, 0)
                    self.assertEqual(str(sheet.cell(2, 3).value), f"{verb} context")
                    self.assertEqual(sheet.cell(2, 5).value, "changed")
                    original_blocks = changed_blocks(sheet.cell(2, original_column).value)
                    suggested_blocks = changed_blocks(sheet.cell(2, suggested_column).value)
                    self.assertEqual([block.text for block in original_blocks], ["old"])
                    self.assertEqual([block.text for block in suggested_blocks], ["new"])
                    self.assert_red(original_blocks[0], strike=True)
                    self.assert_red(suggested_blocks[0], strike=False)
        finally:
            report.close()

        corrected = openpyxl.load_workbook(job / "multi_corrected.xlsx", rich_text=True)
        try:
            for sheet_name, verb in (("Sheet1", "Save"), ("Sheet2", "Open")):
                self.assertEqual(corrected[sheet_name]["B2"].value, f"{verb} new file")
                self.assertIsInstance(corrected[sheet_name]["B2"].value, str)
                self.assertEqual(corrected[sheet_name]["B3"].value, "=NEW()")
                self.assertEqual(corrected[sheet_name]["B3"].data_type, "s")
        finally:
            corrected.close()

        corrected_values = openpyxl.load_workbook(
            job / "multi_corrected.xlsx",
            data_only=True,
        )
        try:
            for sheet_name in ("Sheet1", "Sheet2"):
                self.assertEqual(corrected_values[sheet_name]["B3"].value, "=NEW()")
        finally:
            corrected_values.close()

    def test_aggregate_report_preserves_child_styles_and_rich_text(self):
        job = self.build_aggregate()
        report = openpyxl.load_workbook(job / "multi_lqe.xlsx", rich_text=True)
        try:
            for sheet_name in ("Sheet1", "Sheet2"):
                with self.subTest(sheet=sheet_name):
                    sheet = report[f"{sheet_name} Results"]
                    self.assertEqual(sheet["A1"].fill.fgColor.rgb, "FF112233")
                    self.assertEqual(sheet["A2"].fill.fgColor.rgb, "FF22AA44")
                    self.assertTrue(sheet["C2"].alignment.wrap_text)
                    self.assertEqual(sheet["C2"].alignment.vertical, "top")
                    self.assertEqual(sheet["C2"].font.name, "Arial")
                    self.assertTrue(sheet["C2"].font.bold)
                    self.assertEqual(sheet["C2"].border.left.style, "thin")
                    self.assertEqual(sheet["C2"].border.left.color.rgb, "FF334455")
                    self.assertEqual(sheet["C2"].number_format, "@")
                    self.assertFalse(sheet["C2"].protection.locked)
                    self.assertTrue(sheet["C2"].protection.hidden)
                    self.assertEqual(sheet.row_dimensions[2].height, 31)
                    self.assertEqual(sheet.column_dimensions["C"].width, 37)
                    self.assertEqual(sheet.freeze_panes, "A2")
                    self.assertIsInstance(sheet["C2"].value, CellRichText)
                    context_blocks = changed_blocks(sheet["C2"].value)
                    self.assertEqual([block.text for block in context_blocks], ["context"])
                    self.assertEqual(context_blocks[0].font.color.rgb, "FF0000FF")
        finally:
            report.close()

    def test_aggregate_corrected_skips_all_protected_results(self):
        job = self.build_aggregate()
        corrected = openpyxl.load_workbook(job / "multi_corrected.xlsx")
        try:
            for sheet_name in ("Sheet1", "Sheet2"):
                with self.subTest(sheet=sheet_name):
                    self.assertEqual(corrected[sheet_name]["B4"].value, "Protected original")
                    self.assertEqual(corrected[sheet_name]["B5"].value, "Guarded original")
        finally:
            corrected.close()

    def test_aggregate_report_preserves_formula_like_plain_text(self):
        job = self.build_aggregate()
        report = openpyxl.load_workbook(
            job / "multi_lqe.xlsx",
            rich_text=True,
            data_only=False,
        )
        try:
            for sheet_name in ("Sheet1", "Sheet2"):
                with self.subTest(sheet=sheet_name):
                    sheet = report[f"{sheet_name} Results"]
                    headers = [cell.value for cell in sheet[1]]
                    original_column = headers.index("原译") + 1
                    suggested_column = headers.index("建议译文") + 1
                    no_suggestion = sheet.cell(3, original_column)
                    self.assertEqual(no_suggestion.value, "=NO_SUGGESTION()")
                    self.assertEqual(no_suggestion.data_type, "s")
                    formula = sheet.cell(3, 3)
                    self.assertEqual(formula.value, "=1+1")
                    self.assertEqual(formula.data_type, "f")
                    same_original = sheet.cell(4, original_column)
                    same_suggested = sheet.cell(4, suggested_column)
                    self.assertEqual(same_original.value, "=SAME()")
                    self.assertEqual(same_suggested.value, "=SAME()")
                    self.assertEqual(same_original.data_type, "s")
                    self.assertEqual(same_suggested.data_type, "s")
        finally:
            report.close()

    def test_aggregate_report_marks_insert_after_blank_original(self):
        job = self.build_aggregate()
        report = openpyxl.load_workbook(job / "multi_lqe.xlsx", rich_text=True)
        try:
            for sheet_name, verb in (("Sheet1", "Save"), ("Sheet2", "Open")):
                with self.subTest(sheet=sheet_name):
                    sheet = report[f"{sheet_name} Results"]
                    headers = [cell.value for cell in sheet[1]]
                    original_column = headers.index("原译") + 1
                    suggested_column = headers.index("建议译文") + 1
                    self.assertIsNone(sheet.cell(5, original_column).value)
                    suggested_blocks = changed_blocks(sheet.cell(5, suggested_column).value)
                    self.assertEqual([block.text for block in suggested_blocks], [f"{verb} added"])
                    self.assert_red(suggested_blocks[0], strike=False)
        finally:
            report.close()


if __name__ == "__main__":
    unittest.main()
