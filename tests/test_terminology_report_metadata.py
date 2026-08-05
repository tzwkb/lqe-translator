import json
from pathlib import Path
import sys
import tempfile
import unittest
from unittest import mock

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from lqe_checks import run_pre_check
from lqe_chunk import _mark_ai_reviewed
from lqe_corrections import CheckFormatError, normalize_check_entries
from lqe_engine import build_check_scope
from lqe_io import _build_xlsx
from lqe_terms import terminology_issue_fields
import lqe_paths


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def red_blocks(value: object) -> list[TextBlock]:
    if not isinstance(value, CellRichText):
        return []
    return [run for run in value if isinstance(run, TextBlock)]


class TerminologyReportMetadataTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def test_precheck_emits_full_structured_term_with_apostrophe(self):
        state_path = self.root / "state.json"
        output_path = self.root / "errors_precheck.json"
        write_json(
            state_path,
            {
                "source_lang": "zh",
                "target_lang": "en",
                "check_scope": build_check_scope(False, "test"),
                "segments": [
                    {
                        "id": 0,
                        "source": "查看督管案台",
                        "target": "Check the control desk",
                    }
                ],
                "terminology": [
                    {
                        "source": "督管案台",
                        "target": "Supervisor's Counter",
                        "confirmed": True,
                        "protected": False,
                    }
                ],
            },
        )

        run_pre_check(state_path, output_path)

        entries = json.loads(output_path.read_text(encoding="utf-8"))
        issue = next(
            issue
            for issue in entries[0]["issues"]
            if issue["category"] == "Terminology"
        )
        self.assertEqual(issue["term_source"], "督管案台")
        self.assertEqual(issue["expected_targets"], ["Supervisor's Counter"])
        self.assertEqual(
            issue["term_spans"],
            {
                "source": [{"start": 2, "end": 6, "text": "督管案台"}],
                "target": [],
            },
        )
        normalized = normalize_check_entries(entries, label="precheck")
        self.assertEqual(
            normalized[0]["issues"][0]["expected_targets"],
            ["Supervisor's Counter"],
        )

    def test_comment_only_legacy_issue_is_not_parsed(self):
        self.assertIsNone(
            terminology_issue_fields(
                {
                    "category": "Terminology",
                    "comment": "'督管案台' → expected 'Supervisor's Counter'",
                }
            )
        )

    def test_structured_fields_are_the_only_report_source(self):
        parsed = terminology_issue_fields(
            {
                "category": "Terminology",
                "comment": "'督管案台' → expected 'Wrong'",
                "term_source": "督管案台",
                "expected_targets": ["Supervisor's Counter"],
                "term_spans": {
                    "source": [{"start": 2, "end": 6, "text": "督管案台"}],
                    "target": [
                        {"start": 10, "end": 22, "text": "control desk"}
                    ],
                },
            }
        )
        self.assertEqual(parsed["expected_targets"], ["Supervisor's Counter"])
        self.assertEqual(parsed["term_spans"]["target"][0]["text"], "control desk")

    def test_issue_contract_rejects_partial_term_metadata(self):
        entries = [
            {
                "id": 0,
                "issues": [
                    {
                        "category": "Terminology",
                        "severity": "Major",
                        "comment": "Mismatch",
                        "term_source": "督管案台",
                        "needs_confirmation": True,
                        "edit": None,
                    }
                ],
            }
        ]
        with self.assertRaisesRegex(
            CheckFormatError,
            "requires term_source, expected_targets, and term_spans",
        ):
            normalize_check_entries(entries, label="precheck")

    def test_ai_review_preserves_machine_term_metadata(self):
        original = {
            "category": "Terminology",
            "severity": "Major",
            "comment": "'督管案台' → expected 'Supervisor's Counter'",
            "term_source": "督管案台",
            "expected_targets": ["Supervisor's Counter"],
            "term_spans": {
                "source": [{"start": 2, "end": 6, "text": "督管案台"}],
                "target": [],
            },
            "needs_confirmation": True,
            "edit": None,
            "precheck_ref": "precheck:0:test",
        }
        reviewed = {
            "category": "Terminology",
            "severity": "Major",
            "comment": "Confirmed terminology mismatch.",
            "term_source": "督管案台",
            "expected_targets": ["Supervisor's Counter"],
            "term_spans": {
                "source": [{"start": 2, "end": 6, "text": "督管案台"}],
                "target": [
                    {"start": 10, "end": 22, "text": "control desk"}
                ],
            },
            "needs_confirmation": True,
            "edit": None,
            "precheck_ref": "precheck:0:test",
        }

        result = _mark_ai_reviewed(reviewed, "terminology", 0, [original])

        self.assertEqual(result["term_source"], "督管案台")
        self.assertEqual(result["expected_targets"], ["Supervisor's Counter"])
        self.assertEqual(result["term_spans"]["source"], original["term_spans"]["source"])
        self.assertEqual(result["term_spans"]["target"], reviewed["term_spans"]["target"])

    def test_report_preserves_structured_metadata_and_highlights_terms(self):
        output = self.root / "term-report_lqe.xlsx"
        state = {
            "input_path": str(self.root / "term-report.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [["查看督管案台", "Check the control desk"]],
            "source_col": 0,
            "target_col": 1,
            "source_lang": "zh",
            "target_lang": "en",
            "check_scope": build_check_scope(False, "test"),
            "segments": [
                {
                    "id": 0,
                    "source": "查看督管案台",
                    "target": "Check the control desk",
                    "kind": "desc",
                }
            ],
            "wordcount": 4,
        }
        issue = {
            "category": "Terminology",
            "severity": "Major",
            "comment": "'督管案台' → expected 'Supervisor's Counter'",
            "term_source": "督管案台",
            "expected_targets": ["Supervisor's Counter"],
            "term_spans": {
                "source": [{"start": 2, "end": 6, "text": "督管案台"}],
                "target": [
                    {"start": 10, "end": 22, "text": "control desk"}
                ],
            },
            "needs_confirmation": True,
            "edit": None,
        }
        history = [
            {
                "iteration": 0,
                "errors": [{"id": 0, "errors": [issue], "corrected": None}],
                "review_targets": {"0": "Check the control desk"},
            }
        ]

        _build_xlsx(state, history, 99, 98, output, announce=False)

        workbook = openpyxl.load_workbook(output, rich_text=True, data_only=False)
        try:
            results = workbook["LQE Results"]
            headers = [cell.value for cell in results[1]]
            source_column = headers.index("术语原文（结构化）") + 1
            target_column = headers.index("术语库译文（结构化）") + 1
            self.assertEqual(results.cell(2, source_column).value, "督管案台")
            self.assertEqual(
                results.cell(2, target_column).value,
                "Supervisor's Counter",
            )
            self.assertTrue(
                results.column_dimensions[
                    results.cell(1, target_column).column_letter
                ].hidden
            )
            source_value = results.cell(2, headers.index("原文") + 1).value
            target_value = results.cell(2, headers.index("原译") + 1).value
            self.assertEqual([block.text for block in red_blocks(source_value)], ["督管案台"])
            self.assertEqual([block.text for block in red_blocks(target_value)], ["control desk"])
            for block in red_blocks(source_value) + red_blocks(target_value):
                self.assertEqual(block.font.color.rgb, "FFFF0000")
                self.assertFalse(bool(block.font.strike))
        finally:
            workbook.close()

    def test_report_rejects_incompatible_term_history_before_render(self):
        state = {
            "input_path": str(self.root / "history.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [["查看督管案台", "Check the control desk"]],
            "source_col": 0,
            "target_col": 1,
            "source_lang": "zh",
            "target_lang": "en",
            "check_scope": build_check_scope(False, "test"),
            "segments": [
                {
                    "id": 0,
                    "source": "查看督管案台",
                    "target": "Check the control desk",
                    "kind": "desc",
                }
            ],
            "wordcount": 4,
        }
        valid_issue = {
            "category": "Terminology",
            "severity": "Major",
            "comment": "Use the project term.",
            "term_source": "督管案台",
            "expected_targets": ["Supervisor's Counter"],
            "term_spans": {
                "source": [{"start": 2, "end": 6, "text": "督管案台"}],
                "target": [
                    {"start": 10, "end": 22, "text": "control desk"}
                ],
            },
            "needs_confirmation": True,
            "edit": None,
        }
        legacy_issue = {
            "category": "Terminology",
            "severity": "Major",
            "comment": "Legacy term issue.",
            "needs_confirmation": True,
            "edit": None,
        }
        source_drift = json.loads(json.dumps(valid_issue, ensure_ascii=False))
        source_drift["term_spans"]["source"] = [
            {"start": 0, "end": 4, "text": "督管案台"}
        ]
        cases = (
            (
                "legacy-fields",
                legacy_issue,
                {"0": "Check the control desk"},
                "requires term_source",
            ),
            ("missing-snapshot", valid_issue, None, "review_targets is required"),
            (
                "target-drift",
                valid_issue,
                {"0": "Different historical target"},
                "does not match segment target",
            ),
            (
                "source-drift",
                source_drift,
                {"0": "Check the control desk"},
                "does not match segment source",
            ),
        )
        for name, issue, snapshots, message in cases:
            with self.subTest(name=name):
                entry = {
                    "iteration": 0,
                    "errors": [
                        {"id": 0, "errors": [issue], "corrected": None}
                    ],
                }
                if snapshots is not None:
                    entry["review_targets"] = snapshots
                history = [entry, {"iteration": 1, "errors": []}]
                output = self.root / f"{name}.xlsx"

                with self.assertRaisesRegex(CheckFormatError, message):
                    _build_xlsx(state, history, 99, 98, output, announce=False)

                self.assertFalse(output.exists())

    def test_precheck_clears_incompatible_term_history_and_rebuilds(self):
        valid_issue = {
            "category": "Terminology",
            "severity": "Major",
            "comment": "Use the project term.",
            "term_source": "督管案台",
            "expected_targets": ["Supervisor's Counter"],
            "term_spans": {
                "source": [{"start": 2, "end": 6, "text": "督管案台"}],
                "target": [
                    {"start": 10, "end": 22, "text": "control desk"}
                ],
            },
            "needs_confirmation": True,
            "edit": None,
        }
        legacy_issue = {
            "category": "Terminology",
            "severity": "Major",
            "comment": "Legacy term issue.",
            "needs_confirmation": True,
            "edit": None,
        }
        cases = (
            ("legacy-fields", legacy_issue, {"0": "Check the control desk"}),
            ("missing-snapshot", valid_issue, None),
            ("target-drift", valid_issue, {"0": "Different historical target"}),
        )
        for name, issue, snapshots in cases:
            with self.subTest(name=name):
                job = self.root / name
                state_path = job / "state.json"
                output_path = job / "errors_precheck.json"
                history_entry = {
                    "iteration": 0,
                    "errors": [
                        {"id": 0, "errors": [issue], "corrected": None}
                    ],
                }
                if snapshots is not None:
                    history_entry["review_targets"] = snapshots
                write_json(
                    state_path,
                    {
                        "source_lang": "zh",
                        "target_lang": "en",
                        "check_scope": build_check_scope(False, "test"),
                        "segments": [
                            {
                                "id": 0,
                                "source": "查看督管案台",
                                "target": "Check the control desk",
                            }
                        ],
                        "terminology": [
                            {
                                "source": "督管案台",
                                "target": "Supervisor's Counter",
                                "confirmed": True,
                                "protected": False,
                            }
                        ],
                        "error_history": [history_entry],
                    },
                )

                run_pre_check(state_path, output_path)

                saved_state = json.loads(state_path.read_text(encoding="utf-8"))
                self.assertEqual(saved_state["error_history"], [])
                entries = json.loads(output_path.read_text(encoding="utf-8"))
                rebuilt = next(
                    issue
                    for issue in entries[0]["issues"]
                    if issue["category"] == "Terminology"
                )
                self.assertEqual(
                    rebuilt["term_spans"]["source"],
                    [{"start": 2, "end": 6, "text": "督管案台"}],
                )
                self.assertEqual(rebuilt["term_spans"]["target"], [])

    def test_precheck_history_cleanup_publishes_state_and_results_atomically(self):
        state_path = self.root / "atomic" / "state.json"
        output_path = self.root / "atomic" / "errors_precheck.json"
        write_json(
            state_path,
            {
                "source_lang": "zh",
                "target_lang": "en",
                "check_scope": build_check_scope(False, "test"),
                "segments": [
                    {
                        "id": 0,
                        "source": "查看督管案台",
                        "target": "Check the control desk",
                    }
                ],
                "terminology": [
                    {
                        "source": "督管案台",
                        "target": "Supervisor's Counter",
                        "confirmed": True,
                        "protected": False,
                    }
                ],
                "error_history": [
                    {
                        "iteration": 0,
                        "errors": [
                            {
                                "id": 0,
                                "errors": [
                                    {
                                        "category": "Terminology",
                                        "severity": "Major",
                                        "comment": "Legacy term issue.",
                                        "needs_confirmation": True,
                                        "edit": None,
                                    }
                                ],
                                "corrected": None,
                            }
                        ],
                    }
                ],
            },
        )
        output_path.write_bytes(b"existing precheck sentinel")
        state_before = state_path.read_bytes()
        output_before = output_path.read_bytes()
        real_replace = lqe_paths.os.replace
        failed = False

        def fail_state_publish(source, destination):
            nonlocal failed
            if Path(destination) == state_path and not failed:
                failed = True
                raise OSError("injected state publish failure")
            return real_replace(source, destination)

        with mock.patch.object(
            lqe_paths.os,
            "replace",
            side_effect=fail_state_publish,
        ):
            with self.assertRaisesRegex(OSError, "injected state publish failure"):
                run_pre_check(state_path, output_path)

        self.assertEqual(state_path.read_bytes(), state_before)
        self.assertEqual(output_path.read_bytes(), output_before)
        self.assertFalse(list(state_path.parent.glob(".*.precheck.*")))

    def test_second_iteration_uses_snapshotted_current_target_for_spans(self):
        output = self.root / "iteration-term-report_lqe.xlsx"
        initial_target = "Check the control desk"
        current = "Updated prefix: Check the control desk"
        target_start = current.index("control desk")
        state = {
            "input_path": str(self.root / "iteration-term-report.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [["查看督管案台", initial_target]],
            "source_col": 0,
            "target_col": 1,
            "source_lang": "zh",
            "target_lang": "en",
            "check_scope": build_check_scope(False, "test"),
            "segments": [
                {
                    "id": 0,
                    "source": "查看督管案台",
                    "target": initial_target,
                    "current_target": current,
                    "kind": "desc",
                }
            ],
            "wordcount": 6,
            "iteration": 1,
        }
        issue = {
            "category": "Terminology",
            "severity": "Major",
            "comment": "Use the approved term.",
            "term_source": "督管案台",
            "expected_targets": ["Supervisor's Counter"],
            "term_spans": {
                "source": [{"start": 2, "end": 6, "text": "督管案台"}],
                "target": [
                    {
                        "start": target_start,
                        "end": target_start + len("control desk"),
                        "text": "control desk",
                    }
                ],
            },
            "needs_confirmation": True,
            "edit": None,
        }
        history = [
            {
                "iteration": 0,
                "errors": [
                    {"id": 0, "errors": [], "corrected": current}
                ],
                "review_targets": {"0": initial_target},
            },
            {
                "iteration": 1,
                "errors": [
                    {"id": 0, "errors": [issue], "corrected": None}
                ],
                "review_targets": {"0": current},
            },
        ]

        _build_xlsx(state, history, 99, 98, output, announce=False)

        workbook = openpyxl.load_workbook(
            output,
            rich_text=True,
            data_only=False,
        )
        try:
            scorecard = workbook["LQA Scorecard"]
            header_row = next(
                row
                for row in range(1, scorecard.max_row + 1)
                if scorecard.cell(row, 1).value == "Segment ID"
            )
            score_original = scorecard.cell(header_row + 1, 3).value
            results = workbook["LQE Results"]
            result_original = results.cell(2, 3).value
            for value in (score_original, result_original):
                self.assertEqual(str(value), current)
                self.assertEqual(
                    [block.text for block in red_blocks(value)],
                    ["control desk"],
                )
        finally:
            workbook.close()

    def test_term_history_without_review_target_snapshot_fails_closed(self):
        output = self.root / "missing-history-target_lqe.xlsx"
        target = "Check the control desk"
        state = {
            "input_path": str(self.root / "missing-history-target.xlsx"),
            "headers": ["原文", "译文"],
            "rows_raw": [["查看督管案台", target]],
            "source_col": 0,
            "target_col": 1,
            "check_scope": build_check_scope(False, "test"),
            "segments": [
                {"id": 0, "source": "查看督管案台", "target": target}
            ],
            "wordcount": 4,
        }
        issue = {
            "category": "Terminology",
            "severity": "Major",
            "comment": "Use the approved term.",
            "term_source": "督管案台",
            "expected_targets": ["Supervisor's Counter"],
            "term_spans": {
                "source": [{"start": 2, "end": 6, "text": "督管案台"}],
                "target": [
                    {"start": 10, "end": 22, "text": "control desk"}
                ],
            },
            "needs_confirmation": True,
            "edit": None,
        }
        history = [
            {
                "iteration": 0,
                "errors": [
                    {"id": 0, "errors": [issue], "corrected": None}
                ],
            }
        ]

        with self.assertRaisesRegex(ValueError, "review_targets"):
            _build_xlsx(state, history, 99, 98, output, announce=False)
